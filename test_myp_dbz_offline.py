"""Suíte offline do scanner MYP → DRAGON BALL (myp_dbz_scanner + summary).

Sem rede / sem Cloudflare / sem credencial — fixtures sintéticas calcadas nos
fatos PROVADOS pela sonda probe-myp-dragonball (PR #95, 2026-08-03):
  - título de edição concatenado ("Supreme RivalryB1307/05/2021");
  - h1 dbsfusion "Bardock - FB11-112" (+ "[Pré-Venda]"), h1 dbsmasters sem
    código; campo "Código" = dbsm_bt1-073_spr / dbsf_fb01_fb01-139p1;
  - tcgcsv com colisão base × "(Alternate Art)"/"(SPR)" no mesmo Number e
    abbreviation duplicada real (FB03_AR em 2 grupos).

Run: python test_myp_dbz_offline.py   (ou via pytest, coletado na raiz)
"""
import sys
import tempfile
from pathlib import Path

if (sys.stdout.encoding or "").lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from myp_dbz_scanner import (
    DbzCardData,
    MYPDbzScraper,
    SemRefRow,
    _base_name,
    _clean_card_name,
    _norm_name,
    canonical_set_code,
    extract_myp_product_code,
    find_reference,
    generate_dbz_xlsx,
    normalize_code,
    parse_myp_product_code,
    parse_threshold,
    pick_subtype,
    ref_volatile,
    resolve_edition_group,
    split_edition_title,
    split_myp_title,
    tcg_dbz_search_url,
)
import myp_dbz_summary


# ── helpers de fixture ──────────────────────────────────────────────────

def make_entry(pid, name, gid, special=False, prices=None, lows=None,
               number="BT1-073", rarity="Rare", cat=27):
    clean = _clean_card_name(name)
    return {"pid": pid, "name": name, "name_norm": _norm_name(clean),
            "base_norm": _norm_name(_base_name(clean)), "number": number,
            "rarity": rarity, "url": f"https://www.tcgplayer.com/product/{pid}",
            "prices": prices if prices is not None else {"Normal": 20.0},
            "lows": lows or {}, "cat": cat, "group_id": gid,
            "group_name": "G", "group_special": special}


def make_card(**kw):
    base = dict(
        name="Broly, The Rampaging Horror - BT1-073", en_name="Broly, The Rampaging Horror",
        code="BT1-073", code_norm="BT1-73", edition="Galactic Battle",
        section="dbsmasters", product_url="https://mypcards.com/dbsmasters/produto/1/x",
        myp_lowest_en_nm=100.0, en_nm_sellers=3, tcg_usd=40.0, tcg_brl=200.0,
        tcg_url="https://www.tcgplayer.com/product/1", rarity="Special Rare",
        subtype="Foil", join_via="codigo/grupo-edição/nome-exato (cat-seção)",
        margin_pct=1.0, margin_brl=100.0,
        last_updated="2026-08-03 00:00",
    )
    base.update(kw)
    return DbzCardData(**base)


# ── código de carta / nomes ─────────────────────────────────────────────

def test_normalize_code():
    assert normalize_code("FB01-001") == "FB1-1"
    assert normalize_code("BT31-055") == "BT31-55"
    assert normalize_code("P-001") == "P-1"
    assert normalize_code("FB02-135b") == "FB2-135B"   # sufixo preservado
    assert normalize_code("bt1-073") == "BT1-73"
    assert normalize_code("nada aqui") is None
    assert normalize_code(None) is None


def test_clean_card_name():
    # inconsistência REAL do tcgcsv FW (sonda v3): com e sem cauda de código
    assert _clean_card_name("Son Goku - FB01-139 (Alternate Art)") == \
        "Son Goku (Alternate Art)"
    assert _clean_card_name("Bardock - FB11-112") == "Bardock"
    assert _clean_card_name("Hit") == "Hit"
    assert _clean_card_name("Broly, The Rampaging Horror") == \
        "Broly, The Rampaging Horror"


def test_split_myp_title_fusion_dash():
    # h1 real do dbsfusion (sonda v2), com prefixo de pré-venda e \n interno
    d, c, e = split_myp_title("[Pré-Venda]\n         Bardock - FB11-112")
    assert (d, c, e) == ("Bardock - FB11-112", "FB11-112", "Bardock")


def test_split_myp_title_fusion_alt_art():
    d, c, e = split_myp_title("Son Goku - FB01-139 (Alternate Art)")
    assert c == "FB01-139"
    assert e == "Son Goku (Alternate Art)"


def test_split_myp_title_masters_sem_codigo():
    # h1 real do dbsmasters (sonda v2/v3): nome puro
    d, c, e = split_myp_title("Arack, Godly Power")
    assert c is None and e == "Arack, Godly Power" and d == e


def test_split_myp_title_bilingue_parens():
    # padrão bilíngue da plataforma (documentado no scanner Pokémon v5.8.7)
    d, c, e = split_myp_title(
        "Trunks, o Guerreiro (FB02-070)Trunks, Warrior of the Future")
    assert c == "FB02-070"
    assert e == "Trunks, Warrior of the Future"   # o EN, nunca o PT duplicado


# ── título de edição (concat real da sonda) ─────────────────────────────

def test_split_edition_title_casos_reais():
    cases = [
        ("Supreme RivalryB1307/05/2021", "supreme-rivalry",
         ("Supreme Rivalry", "B13")),
        ("Story Booster 01ST0121/08/2026", "story-booster-01",
         ("Story Booster 01", "ST01")),
        ("CollectorsSelection Vol.2CS-02", "collectorsselection-vol2",
         ("CollectorsSelection Vol.2", "CS-02")),
        ("Promotion CardsP01/01/2020", "promotion-cards",
         ("Promotion Cards", "P")),
        ("Mighty HeroesEX130/11/-0001", "mighty-heroes",   # data negativa real
         ("Mighty Heroes", "EX1")),
        ("Ultra Limit Release Event CardsFB04-PR08/11/2024",
         "ultra-limit-release-event-cards",
         ("Ultra Limit Release Event Cards", "FB04-PR")),
        ("Dual EvolutionFB09", "dual-evolution", ("Dual Evolution", "FB09")),
        ("5th Anniversary Set5TH", "5th-anniversary-set",
         ("5th Anniversary Set", "5TH")),
        ("ZENKAI Series Fighter's AmbitionBT19",
         "zenkai-series-fighters-ambition",
         ("ZENKAI Series Fighter's Ambition", "BT19")),
    ]
    for raw, slug, want in cases:
        assert split_edition_title(raw, slug) == want, (raw, slug)


# ── resolução edição → grupo tcgcsv ─────────────────────────────────────

def test_canonical_set_code_aliases():
    # classes de alias observadas nos DOIS catálogos (MYP ↔ tcgcsv)
    assert canonical_set_code("DBS-B09") == "BT9" == canonical_set_code("BT9")
    assert canonical_set_code("DBS-BE15") == "EX15" == canonical_set_code("EX15")
    assert canonical_set_code("DBS-EB06") == "EX6" == canonical_set_code("EX6")
    assert canonical_set_code("MB-01") == "MB1" == canonical_set_code("MB01")
    assert canonical_set_code("FB03_AR") == "FB3AR" == canonical_set_code("FB03AR")
    assert canonical_set_code("DBSFW_TPR") == "TPR" == canonical_set_code("TPR")
    assert canonical_set_code("DBS-B23PRE") == "BT23PRE" == canonical_set_code("B23PRE")
    assert canonical_set_code("5TH") == "5TH"
    assert canonical_set_code(None) is None


def test_resolve_edition_group_cascata():
    groups = [
        {"groupId": 1, "name": "Supreme Rivalry", "abbreviation": "DBS-B13"},
        {"groupId": 2, "name": "Expansion Deck Box Set 22: Ultimate Deck 2023",
         "abbreviation": "DBS-BE22"},
        # abbreviation DUPLICADA real do tcgcsv (FB03_AR em 2 grupos):
        {"groupId": 3, "name": "New Adventure Alternate Art Reprints",
         "abbreviation": "FB03_AR"},
        {"groupId": 4, "name": "Raging Roar Alternate Art Reprints",
         "abbreviation": "FB03_AR"},
        {"groupId": 5, "name": "Rivals Clash", "abbreviation": "FB06"},
        {"groupId": 6, "name": "Expansion Deck Box Set 15: Battle Enhanced",
         "abbreviation": "DBS-BE15"},
    ]
    # 1) nome exato
    assert resolve_edition_group("Supreme Rivalry", "B13", groups) == 1
    assert resolve_edition_group("Rivals Clash", "FB06", groups) == 5
    # 2) código canônico único (nome MYP ≠ nome tcgcsv)
    assert resolve_edition_group("Universe 15 Battle Enhanced", "EX15", groups) == 6
    # 3) contenção de nome única
    assert resolve_edition_group("Ultimate Deck 2023", "UD23", groups) == 2
    # abbr duplicada → código ambíguo → resolve pelo nome exato
    assert resolve_edition_group("Raging Roar Alternate Art Reprints",
                                 "FB03AR", groups) == 4
    # sem match → None (nunca chuta)
    assert resolve_edition_group("Inexistente Total", "ZZZ99", groups) is None
    assert resolve_edition_group("X", None, []) is None


# ── campo "Código" do produto ───────────────────────────────────────────

def test_parse_myp_product_code():
    # formatos REAIS da sonda v3
    assert parse_myp_product_code("dbsm_bt1-073_spr") == ("BT1-073", "spr")
    assert parse_myp_product_code("dbsf_fb01_fb01-139p1") == ("FB01-139", "p1")
    assert parse_myp_product_code("dbsm_bt1-073") == ("BT1-073", None)
    assert parse_myp_product_code("pokemon_swsh_123") == (None, None)
    assert parse_myp_product_code("") == (None, None)
    assert parse_myp_product_code(None) == (None, None)


def test_extract_myp_product_code_html():
    html = """<html><body><h1>Broly, The Rampaging Horror</h1>
    <table><tr><td>Energy</td><td>6</td></tr>
    <tr><td>Código</td><td>dbsm_bt1-073_spr</td></tr>
    <tr><td>Edição</td><td>Galactic Battle (BT1)</td></tr></table>
    </body></html>"""
    soup = BeautifulSoup(html, "lxml")
    assert extract_myp_product_code(soup) == "dbsm_bt1-073_spr"
    # sem o campo → None
    soup2 = BeautifulSoup("<html><body><h1>X</h1></body></html>", "lxml")
    assert extract_myp_product_code(soup2) is None


# ── threshold / subtipo / guardas ───────────────────────────────────────

def test_parse_threshold_percent_inteiro():
    assert parse_threshold(30) == 0.30      # convenção MYP
    assert parse_threshold(45) == 0.45
    assert parse_threshold(0.3) == 0.3      # fração aceita com warning


def test_pick_subtype_conservador():
    # menor market entre subtipos (nunca superestima a margem)
    assert pick_subtype({"Normal": 2.0, "Foil": 10.0}) == ("Normal", 2.0)
    assert pick_subtype({"Normal": 2.0, "Foil": 10.0}, prefer_foil=True) == \
        ("Foil", 10.0)
    assert pick_subtype({"Foil": 10.0}) == ("Foil", 10.0)
    assert pick_subtype({}) is None
    assert pick_subtype({"Normal": None}) is None   # sem market → nunca inventa


def test_ref_volatile():
    # caso real da frota: market arrastado (Bulma SB01-057, dbs_scanner)
    assert ref_volatile(506.0, 2000.0) is True
    assert ref_volatile(2000.0, 506.0) is True
    assert ref_volatile(10.0, 12.0) is False
    assert ref_volatile(10.0, None) is False
    assert ref_volatile(None, 10.0) is False


# ── join determinístico (variant-aware) ─────────────────────────────────

def test_join_spr_nunca_casa_base():
    base = make_entry(1, "Broly, The Rampaging Horror", 10)
    spr = make_entry(2, "Broly, The Rampaging Horror (SPR)", 10)
    primary = {"by_code": {"BT1-73": [base, spr]}, "by_group": {10: [base, spr]}}
    sec = {"by_code": {}, "by_group": {}}
    # MYP vendendo o SPR (sufixo _spr do campo Código) → casa SÓ o (SPR)
    mt, how = find_reference("BT1-73", "Broly, The Rampaging Horror",
                             primary, sec, group_id=10, variant_token="spr")
    assert mt is spr and "variante" in how
    # MYP vendendo o base → nome exato casa o base
    mt, how = find_reference("BT1-73", "Broly, The Rampaging Horror",
                             primary, sec, group_id=10)
    assert mt is base and "nome-exato" in how


def test_join_variante_sem_produto_fica_fora():
    base = make_entry(1, "Broly, The Rampaging Horror", 10)
    primary = {"by_code": {"BT1-73": [base]}, "by_group": {10: [base]}}
    sec = {"by_code": {}, "by_group": {}}
    mt, how = find_reference("BT1-73", "Broly, The Rampaging Horror",
                             primary, sec, group_id=10, variant_token="spr")
    assert mt is None and "sem produto qualificado" in how


def test_join_base_nunca_pega_ref_de_variante():
    spr = make_entry(2, "Broly, The Rampaging Horror (SPR)", 10)
    primary = {"by_code": {"BT1-73": [spr]}, "by_group": {10: [spr]}}
    sec = {"by_code": {}, "by_group": {}}
    mt, how = find_reference("BT1-73", "Broly, The Rampaging Horror",
                             primary, sec, group_id=10)
    assert mt is None and "variante qualificada" in how


def test_join_fw_alt_art_por_nome():
    b = make_entry(4, "Son Goku - FB01-139", 20, number="FB01-139", cat=80)
    alt = make_entry(3, "Son Goku - FB01-139 (Alternate Art)", 20,
                     number="FB01-139", cat=80)
    primary = {"by_code": {"FB1-139": [b, alt]}, "by_group": {20: [b, alt]}}
    sec = {"by_code": {}, "by_group": {}}
    mt, _ = find_reference("FB1-139", "Son Goku (Alternate Art)", primary, sec,
                           group_id=20)
    assert mt is alt
    mt, _ = find_reference("FB1-139", "Son Goku", primary, sec, group_id=20)
    assert mt is b


def test_join_escopo_grupo_principal_vs_prerelease():
    # BT31-001 existe no set principal E no grupo Release Event (sonda v2) —
    # sem grupo resolvido, o escopo "grupo-principal" evita a versão carimbada
    main_ = make_entry(1, "SS Gogeta, Invincible Power", 100)
    pr = make_entry(2, "SS Gogeta, Invincible Power", 200, special=True)
    primary = {"by_code": {"BT31-3": [main_, pr]},
               "by_group": {100: [main_], 200: [pr]}}
    sec = {"by_code": {}, "by_group": {}}
    mt, how = find_reference("BT31-3", "SS Gogeta, Invincible Power",
                             primary, sec, group_id=None)
    assert mt is main_ and "grupo-principal" in how
    # com grupo resolvido (edição Release Event do MYP) → escopo do grupo
    mt, how = find_reference("BT31-3", "SS Gogeta, Invincible Power",
                             primary, sec, group_id=200)
    assert mt is pr and "grupo-edição" in how


def test_join_masters_sem_codigo_grupo_nome():
    base = make_entry(1, "Arack, Godly Power", 10)
    primary = {"by_code": {}, "by_group": {10: [base]}}
    sec = {"by_code": {}, "by_group": {}}
    mt, how = find_reference(None, "Arack, Godly Power", primary, sec, group_id=10)
    assert mt is base and how == "grupo+nome-exato"
    mt, how = find_reference(None, "Arack, Godly Power", primary, sec, group_id=None)
    assert mt is None and "não resolvido" in how


def test_join_categoria_secundaria():
    # FB01 existe nas DUAS categorias (grupo legado na 27) — a categoria da
    # seção tem preferência; a irmã só entra se a primária não tem o código
    e80 = make_entry(1, "Son Goku - FB01-001", 20, number="FB01-001", cat=80)
    e27 = make_entry(2, "Son Goku - FB01-001", 30, number="FB01-001", cat=27)
    primary = {"by_code": {"FB1-1": [e80]}, "by_group": {20: [e80]}}
    secondary = {"by_code": {"FB1-1": [e27]}, "by_group": {30: [e27]}}
    mt, how = find_reference("FB1-1", "Son Goku", primary, secondary)
    assert mt is e80 and "cat-seção" in how
    mt, how = find_reference("FB1-1", "Son Goku",
                             {"by_code": {}, "by_group": {}}, secondary)
    assert mt is e27 and "cat-irmã" in how


# ── parser de sellers herdado (NM/EN) na página DBZ ─────────────────────

PRODUCT_HTML = """<html><body>
<h1>Broly, The Rampaging Horror</h1>
<table><tr><td>Código</td><td>dbsm_bt1-073_spr</td></tr></table>
<table class="table-striped table-bordered">
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td class="estoque-lista-nomeenfoil">Foil</td><td>R$ 100,00</td></tr>
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">SP - Pouco jogada</td>
    <td>R$ 80,00</td></tr>
<tr><td><span class="flag-icon" title="Português"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td>R$ 60,00</td></tr>
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td class="estoque-lista-nomeenfoil">Jumbo</td><td>R$ 55,00</td></tr>
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td>R$ 120,00</td></tr>
</table></body></html>"""


def _scraper():
    return MYPDbzScraper(delay=0, min_en_sellers=2, threshold=0.30,
                         min_price=50.0, tcg_cache_dir=None)


def test_parse_seller_table_herdado_en_nm():
    sc = _scraper()
    soup = BeautifulSoup(PRODUCT_HTML, "lxml")
    table = soup.select("table.table-striped.table-bordered")[0]
    st = sc._parse_seller_table(table)
    # EN NM = R$100 e R$120 (SP fora, PT fora, Jumbo fora)
    assert sorted(st["en_prices"]) == [100.0, 120.0]
    assert st["en"] == 2
    assert st["jumbo"] == 1


def test_scrape_dbz_product_end_to_end():
    sc = _scraper()
    soup = BeautifulSoup(PRODUCT_HTML, "lxml")
    sc._get = lambda url, save_debug=False: soup     # offline: sem rede
    spr = make_entry(2, "Broly, The Rampaging Horror (SPR)", 10,
                     prices={"Foil": 40.0}, lows={"Foil": 38.0})
    base = make_entry(1, "Broly, The Rampaging Horror", 10,
                      prices={"Normal": 1.0})
    sc.indices = {27: {"by_code": {"BT1-73": [base, spr]},
                       "by_group": {10: [base, spr]},
                       "groups": []},
                  80: {"by_code": {}, "by_group": {}, "groups": []}}
    ed = {"title": "Galactic Battle", "section": "dbsmasters"}
    card = sc.scrape_dbz_product("https://mypcards.com/dbsmasters/produto/1/x",
                                 ed, 27, 10, fx_usd_brl=5.0)
    assert isinstance(card, DbzCardData), card
    # o _spr casou o produto (SPR), nunca o base de US$1
    assert card.tcg_usd == 40.0 and "variante" in card.join_via
    assert card.myp_lowest_en_nm == 100.0 and card.en_nm_sellers == 2
    assert abs(card.margin_pct - 1.0) < 1e-9        # (200−100)/100
    assert card.code == "BT1-073" and "SPR" in card.name.upper()
    # gate de lixo: MYP 100 vs 50% da ref (100) — igual NÃO é lixo (< estrito)
    assert card.flag_lixo is False
    assert card.single_en_seller_risk is False       # 2 sellers = min default


def test_scrape_dbz_product_sem_ref_vai_pra_semref():
    sc = _scraper()
    soup = BeautifulSoup(PRODUCT_HTML.replace("dbsm_bt1-073_spr",
                                              "dbsm_zz9-999"), "lxml")
    sc._get = lambda url, save_debug=False: soup
    sc.indices = {27: {"by_code": {}, "by_group": {}, "groups": []},
                  80: {"by_code": {}, "by_group": {}, "groups": []}}
    ed = {"title": "Galactic Battle", "section": "dbsmasters"}
    row = sc.scrape_dbz_product("https://mypcards.com/dbsmasters/produto/1/x",
                                ed, 27, None, fx_usd_brl=5.0)
    assert isinstance(row, SemRefRow)
    assert row.myp_lowest_en_nm == 100.0
    assert "fora do índice" in row.motivo


def test_scrape_dbz_product_abaixo_do_piso_e_sem_en():
    sc = _scraper()
    html_baixo = PRODUCT_HTML.replace("R$ 100,00", "R$ 10,00").replace(
        "R$ 120,00", "R$ 12,00")
    sc._get = lambda url, save_debug=False: BeautifulSoup(html_baixo, "lxml")
    sc.indices = {27: {"by_code": {}, "by_group": {}, "groups": []},
                  80: {"by_code": {}, "by_group": {}, "groups": []}}
    ed = {"title": "Galactic Battle", "section": "dbsmasters"}
    assert sc.scrape_dbz_product("u", ed, 27, None, 5.0) is None
    assert sc._stats["skipped_low_price"] == 1


# ── XLSX + entrega (myp_dbz_summary) ────────────────────────────────────

def _build_xlsx(tmpdir: Path) -> Path:
    clean = make_card()                                   # 100% margem, limpo
    lixo = make_card(name="Golden Frieza - BT2-001", en_name="Golden Frieza",
                     code="BT2-001", myp_lowest_en_nm=60.0, tcg_usd=40.0,
                     tcg_brl=200.0, margin_pct=(200 - 60) / 60,
                     margin_brl=140.0, flag_lixo=True)
    volatil = make_card(name="Bulma - SB01-057", en_name="Bulma",
                        code="SB01-057", tcg_low_usd=2000.0, tcg_usd=506.0,
                        tcg_brl=2530.0, myp_lowest_en_nm=1000.0,
                        margin_pct=1.53, margin_brl=1530.0,
                        flag_ref_volatil=True)
    abaixo = make_card(name="Vegeta - BT3-002", en_name="Vegeta",
                       code="BT3-002", margin_pct=0.10, margin_brl=20.0)
    trunc = make_card(name="Zamasu - BT4-050", en_name="Zamasu",
                      code="BT4-050", margin_pct=0.55,
                      en_truncation_risk=True)
    semref = [SemRefRow(name="Beerus - XD1-01", en_name="Beerus",
                        code="XD1-01", edition="Universe 6 Assailants",
                        section="dbsmasters",
                        product_url="https://mypcards.com/dbsmasters/produto/9/x",
                        myp_lowest_en_nm=80.0, en_nm_sellers=2,
                        motivo="número fora do índice tcgcsv")]
    out = tmpdir / "dbz_test.xlsx"
    generate_dbz_xlsx([clean, lixo, volatil, abaixo, trunc], semref,
                      str(out), threshold=0.30, fx=5.0,
                      stats={"products_scanned": 5, "dbz_ref_ok": 5},
                      editions_scanned=4, min_price=50.0)
    return out


def test_xlsx_estrutura_e_threshold():
    with tempfile.TemporaryDirectory() as td:
        out = _build_xlsx(Path(td))
        wb = load_workbook(out, data_only=True)
        assert set(wb.sheetnames) == {"All EN Cards", "Sem Ref TCG", "Summary"}
        summary = {str(r[0]): r[1] for r in wb["Summary"].iter_rows(values_only=True)}
        assert summary["Margin Threshold"] == "30%"
        assert summary["Total EN Cards"] == 5
        assert summary["Sem Ref TCG"] == 1
        # Deals ≥30%: clean, lixo, volatil, trunc (0.55) — abaixo (0.10) não
        assert summary["Deals Found"] == 4
        headers = [c.value for c in wb["All EN Cards"][1]]
        for h in ("Card Name", "Código", "TCG US$", "TCG Source", "Margin %",
                  "URL", "TCG URL", "⚠️ Possível Lixo", "⚠️ Ref Volátil"):
            assert h in headers, h
        wb.close()


def test_summary_markdown_buckets_e_links():
    with tempfile.TemporaryDirectory() as td:
        out = _build_xlsx(Path(td))
        md_path = Path(td) / "dbz.md"
        rc = myp_dbz_summary.build_markdown(str(out), str(md_path), run_id="123")
        assert rc == 0
        md = md_path.read_text(encoding="utf-8")
        # buckets canônicos
        assert "## 🟢 Top 50 deals limpos" in md
        assert "## 🚨 REVISAR" in md
        assert "## ⚠️ Sem referência TCG" in md
        assert "## 🚨 EN truncation risk" in md
        # 2 links em linha de deal (contrato de entrega do repo)
        clean_line = next(l for l in md.splitlines()
                          if "Broly, The Rampaging Horror" in l and "**" in l)
        assert "[oferta](" in clean_line and "[TCG](" in clean_line
        # flag por linha no REVISAR
        assert "possível lixo" in md
        assert "ref volátil" in md
        # linha sem ref traz o motivo + link de BUSCA TCG (nunca URL inventada)
        semref_line = next(l for l in md.splitlines() if "Beerus" in l)
        assert "número fora do índice" in semref_line
        assert "[oferta](" in semref_line
        # link de BUSCA com o prefixo canônico ANCORADO no início do link
        # markdown (nunca substring solta de domínio)
        assert "[TCG](https://www.tcgplayer.com/search/" in semref_line
        # margem do deal limpo em negrito, percentual
        assert "**100.0%**" in md
        # cobertura honesta
        assert "Cobertura de referência real" in md
        # deal 10% NÃO aparece como deal
        assert not any("Vegeta" in l and "**" in l for l in md.splitlines())


def test_summary_threshold_vem_do_xlsx():
    # piso do summary = threshold REAL do scan (lição v5.14.1 do Pokémon)
    with tempfile.TemporaryDirectory() as td:
        card45 = make_card(margin_pct=0.45, margin_brl=45.0)
        out = Path(td) / "t50.xlsx"
        generate_dbz_xlsx([card45], [], str(out), threshold=0.50, fx=5.0)
        md_path = Path(td) / "t50.md"
        myp_dbz_summary.build_markdown(str(out), str(md_path))
        md = md_path.read_text(encoding="utf-8")
        assert "Deals (≥50%):** 0" in md


def test_carta_label_e_links_helpers():
    assert myp_dbz_summary.carta_label("Broly, The Rampaging Horror",
                                       "BT1-073") == \
        "Broly, The Rampaging Horror BT1-073"
    # não duplica quando o código já está no nome
    assert myp_dbz_summary.carta_label("Bardock - FB11-112", "FB11-112") == \
        "Bardock - FB11-112"
    links = myp_dbz_summary.delivery_links("https://m/x", "https://t/y")
    assert links == "[oferta](https://m/x) · [TCG](https://t/y)"
    # sem URL de produto TCG → busca gerada (nunca URL de produto inventada)
    links = myp_dbz_summary.delivery_links("https://m/x", None, "Beerus", "XD1-01")
    assert "[oferta](https://m/x)" in links and "search" in links


def test_search_url():
    u = tcg_dbz_search_url("Son Goku - FB01-139 (Alternate Art)", "FB01-139")
    # valida o HOST parseado, não substring (CodeQL: URL substring
    # sanitization incompleta — "tcgplayer.com" poderia estar em qualquer
    # posição de uma URL maliciosa)
    from urllib.parse import urlparse
    assert u is not None
    assert urlparse(u).netloc == "www.tcgplayer.com"
    assert "Son+Goku" in u
    assert tcg_dbz_search_url("", None) is None


# ── runner standalone (mesmo padrão do test_v5_8_offline.py) ────────────

def main() -> int:
    tests = [(n, f) for n, f in sorted(globals().items())
             if n.startswith("test_") and callable(f)]
    failed = 0
    for name, fn in tests:
        try:
            fn()
            print(f"  ✓ {name}")
        except AssertionError as e:
            failed += 1
            print(f"  ✗ {name}: {e}")
    print("═" * 50)
    print(f"{len(tests) - failed}/{len(tests)} passaram")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
