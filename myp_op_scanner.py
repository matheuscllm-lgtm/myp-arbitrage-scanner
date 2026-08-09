#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║   MYP Cards Arbitrage Scanner — ONE PIECE (paralelo ao Pokémon)      ║
║   mypcards.com (BR, R$) vs TCGplayer (tcgcsv.com) · v1.0             ║
╚══════════════════════════════════════════════════════════════════════╝

Scanner PARALELO ao fluxo Pokémon deste repo — mesmo precedente do
`myp_dbz_scanner.py` (jogo paralelo = script separado, nunca um modo --game
embutido). NÃO toca o scanner Pokémon: reusa por import a infra de site que é
da PLATAFORMA MYP (sessão cloudscraper, retry, parser de seller-table NM/EN,
paginação marketplace) e os helpers puros já provados do scanner DBZ.

O MYP tem UMA seção One Piece (provada pela sonda `probe-myp-onepiece`,
PR #98, run 31300834735, 2026-08-09), que espelha a categoria 68 do tcgcsv
(One Piece Card Game — catálogo INGLÊS do TCGplayer):

    /onepiece   One Piece Card Game (65 edições: OP01→OP17, ST01→ST-36,
                EB, PRB, LT, SD, promos)  ↔  tcgcsv cat 68 (84 groups)

Fatos de estrutura provados pela sonda (não re-descobrir):
  • Título de edição vem CONCATENADO com código de set + data
    ("Carrying On His WillOP1307/11/2025") — mesmo padrão do DBZ; o
    `split_edition_title` (slug como âncora) é reusado por import.
  • h1 de produto tem 3 formatos: "Sabo (OP13-004)" (código completo),
    "[Pré-Venda] Edward.Newgate (001) (Alternate Art)" (número CURTO
    "(001)" + qualificador de variante) e "Monkey.D.Dragon" (nome puro).
  • Campo "Código" da página = `one_<edição>_<código>[marcador]`
    ("one_op17_op17-001", "one_st-35_p-105", "one_op17_op17-020p1").
    ⚠️ O marcador `p1` de variante é NÃO-CONFIÁVEL — a sonda pegou o par
    Edward.Newgate OP17-001 com os códigos INVERTIDOS (o produto base com
    `p1` e o Alternate Art sem). O sinal de variante do lado MYP é o
    QUALIFICADOR DO H1 ("(Alternate Art)"), nunca o marcador.
  • Vocabulário de variante DIVERGE entre plataformas: MYP escreve
    "(Alternate Art)"; o tcgcsv cat 68 escreve "(Parallel)" (colisões
    reais por Number no OP01: base × "(Parallel)" × "(Box Topper)").
    A equivalência {alternate art ≡ parallel} é aplicada na normalização
    de nome — QUALQUER outro qualificador exige match exato.
  • O tcgcsv cat 68 duplica o número da carta entre grupos (starter decks
    REIMPRIMEM cartas OPxx com o número original, ex. Sabo OP13-004 no
    grupo ST-35) → o escopo grupo-edição do join resolve pro produto (e
    preço) da versão CERTA.
  • Página de produto usa os MESMOS seletores da seção Pokémon (flag-icon
    de idioma, célula NM, container marketplace) — parser herdado. Idiomas
    observados: "Inglês"/"Português"; JP ficaria fora pelo filtro EN
    herdado (lição do op_scanner do card-trader: idioma é o risco nº 1
    em One Piece).

Fluxo: menor oferta EN NM ao vivo no MYP  vs  market price TCGplayer (USD →
BRL com câmbio ao vivo). Margem BRUTA base compra. Entrega via
`myp_op_summary.py` (verbatim — contrato de entrega do repo).

Invariantes (frota + convenções DESTE repo):
  • Margem BRUTA: (TCG_BRL − MYP_BRL) / MYP_BRL — sem nenhuma taxa embutida.
  • `--threshold` em PERCENT INTEIRO (30 = 30%) — convenção MYP/Liga/eBay.
  • Piso de relevância R$50 (singles): `--min-price 50`.
  • Só Near Mint (célula dedicada, token exato "NM") e só EN (flag-icon
    "Inglês") — parser herdado do scanner Pokémon.
  • NUNCA inventar preço: carta sem referência tcgcsv fica FORA da margem e
    vai pra aba "Sem Ref TCG" com o motivo (sem fallback `.estat-tcg` —
    mesma decisão v1 do DBZ: referência real ou nada). Sem câmbio real o
    run FALHA ALTO.
  • NUNCA recomendar compra — buckets são classificação técnica.

Join MYP ↔ TCGplayer (DETERMINÍSTICO, nunca fuzzy):
  1. resolução edição→grupo tcgcsv (nome exato → código de set canônico
     único → contenção de nome única; canônico próprio do OP, SEM as
     classes de alias do DBZ — aqui EB = Extra Booster, nunca ≡ EX);
  2. carta por CÓDIGO (campo "Código" da página → fallback h1) com escopo
     em camadas: grupo resolvido → grupos principais (não Release Event/
     Anniversary/Demo Deck — duplicam números) → todos; desambiguação por
     nome com a equivalência {alternate art ≡ parallel} e a regra dura:
     nome QUALIFICADO nunca cai no produto base (e vice-versa);
  3. sem código: grupo resolvido + nome EXATO único dentro do grupo.
  Ambíguo/sem match → linha "Sem Ref TCG" com motivo (honestidade >
  cobertura; contagem explícita, nada some em silêncio).

Uso:
    python myp_op_scanner.py --list-editions
    python myp_op_scanner.py --editions "Royal Blood" --threshold 30 \
        --min-price 50 --delay 1.5 -o results/op_<data>.xlsx --resume
    python myp_op_scanner.py -o results/op_full.xlsx   # catálogo inteiro

Autor: Matheus Chillemi / Claude
Data: 2026-08-09 (v1.0)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote_plus

# Infra da PLATAFORMA MYP reusada por import (sessão CF, retry, parser de
# seller-table NM/EN, paginação marketplace, câmbio).
from myp_arbitrage_scanner import (
    BASE_URL,
    MAX_PAGES_PER_EDITION,
    MAX_SELLER_PAGES,
    MAX_EDITION_PAGES,
    OVERSIZED_TITLE_RE,
    MYPScraper,
    fetch_usd_brl,
    log,
)
# Helpers PUROS já provados no scanner DBZ (semântica idêntica nas duas
# plataformas — travados em teste lá; backlog da frota: extrair core comum).
from myp_dbz_scanner import (
    CARD_CODE_RE,
    _CARD_TOKEN_RE,
    _alnum,
    _base_name,
    _clean_card_name,
    _norm_name,
    _tcgcsv_json,
    normalize_code,
    parse_threshold,
    pick_subtype,
    ref_volatile,
    split_edition_title,
    split_myp_title,
)

OP_VERSION = "v1.0"

# Seção One Piece do MYP ↔ categoria tcgcsv (sonda 2026-08-09, PR #98).
OP_SECTION = "onepiece"
TCGCSV_OP_CAT = 68
SECTION_LABELS = {OP_SECTION: "One Piece"}

TCGCSV_OP_BASE = "https://tcgcsv.com/tcgplayer"

MARGIN_THRESHOLD_OP = 0.30   # 30% margem BRUTA mínima (fração interna)
MIN_PRICE_BRL_OP = 50.0      # piso canônico de singles da frota
# Guardas da frota (mesmos valores do dbs/op scanner do CardTrader):
JUNK_RATIO = 0.5             # MYP < 50% da ref = possível lixo/scam → REVISAR
VOLATILE_REF_RATIO = 2.0     # market vs menor anúncio TCG >2× → REVISAR

OP_CHECKPOINT_VERSION = 1

# Campo "Código" da página de produto One Piece (sonda 2026-08-09):
#   "one_op17_op17-001"    (edição + código da carta)
#   "one_st-35_p-105"      (⚠️ o token de edição TAMBÉM tem hífen — por isso
#                           o parse pega o ÚLTIMO token com cara de carta)
#   "one_op17_op17-020p1"  (marcador de variante colado — NÃO-CONFIÁVEL, ver
#                           cabeçalho; parseado só pra auditoria)
_OP_PROD_CODE_FIELD_RE = re.compile(r"\bone_[a-z0-9_-]{3,60}\b", re.IGNORECASE)

# Número CURTO entre parênteses no nome ("Edward.Newgate (001)" no MYP;
# "Trafalgar Law (002)" no tcgcsv) — disambiguador de personagem repetido,
# removido na normalização dos DOIS lados (simétrico).
_SHORT_NUM_PAREN_RE = re.compile(r"\(\s*\d{1,4}\s*\)")

# Grupo tcgcsv "especial" (nunca preferido sem escopo explícito): versões
# carimbadas/derivadas que DUPLICAM número e nome do set principal.
_SPECIAL_GROUP_RE = re.compile(
    r"release event|anniversary|demo deck|revision pack|promo|pre-?release|"
    r"judge|winner|reprint",
    re.IGNORECASE)

_TCG_OP_SEARCH_BASE = "https://www.tcgplayer.com/search/all/product?q="


# ══════════════════════════════════════════════════════════════════════
# Helpers puros (testáveis offline)
# ══════════════════════════════════════════════════════════════════════

def _clean_card_name_op(s: Optional[str]) -> str:
    """Nome comparável One Piece: remove código completo (herdado) E o
    número curto "(001)" — presente ora no MYP ("Shanks (020)"), ora no
    tcgcsv ("Trafalgar Law (002)"), nunca garantido nos dois lados."""
    t = _clean_card_name(s)
    t = _SHORT_NUM_PAREN_RE.sub(" ", t)
    t = re.sub(r"\(\s*\)", " ", t)
    t = re.sub(r"\s*[-–—]\s*$", "", t)
    return " ".join(t.split())


def _norm_name_op(s: Optional[str]) -> str:
    """Normalização de nome pro join OP: casefold + colapso + a equivalência
    de vocabulário {alternate art ≡ parallel} (MYP escreve "(Alternate
    Art)", o tcgcsv cat 68 escreve "(Parallel)" — provado na sonda).
    Aplicada NOS DOIS lados → simétrica; qualquer outro qualificador
    ("(Box Topper)", "(Manga)", "(Winner)"…) segue exigindo match exato."""
    return _norm_name(s).replace("(alternate art)", "(parallel)")


def parse_myp_op_product_code(code_str: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    """'one_op17_op17-001' → ('OP17-001', None) · 'one_st-35_p-105' →
    ('P-105', None) · 'one_op17_op17-020p1' → ('OP17-020', 'p1') ·
    sem código de carta → (None, None).

    O código da carta é o ÚLTIMO token com formato de carta (o token de
    edição pode ter hífen e "cara de carta", ex. "st-35" — por isso último,
    não primeiro). ⚠️ O marcador de variante retornado é SÓ auditoria: a
    sonda provou par com marcador invertido (Edward.Newgate OP17-001) —
    quem decide variante no join é o qualificador do h1."""
    s = str(code_str or "").strip().lower()
    if not s.startswith("one_"):
        return None, None
    toks = s.split("_")[1:]
    card = None
    marker_parts: list[str] = []
    last_idx = -1
    for i, tok in enumerate(toks):
        m = _CARD_TOKEN_RE.match(tok)
        if m and "-" in tok:
            card = m.group(1).upper()
            marker_parts = [m.group(2)] if m.group(2) else []
            last_idx = i
    if card is None:
        return None, None
    for tok in toks[last_idx + 1:]:
        if tok:
            marker_parts.append(tok)
    return card, ("_".join(marker_parts) or None)


def extract_myp_op_product_code(soup) -> Optional[str]:
    """Localiza o campo 'Código' na página de produto OP. Primário: célula
    cujo texto é exatamente 'Código' (a linha-pai traz o valor). Fallback:
    1º padrão one_ no texto da página."""
    try:
        for el in soup.find_all(["td", "th", "dt", "span", "div", "li"]):
            if el.get_text(strip=True) == "Código":
                parent = el.parent
                if parent is not None:
                    m = _OP_PROD_CODE_FIELD_RE.search(
                        parent.get_text(" ", strip=True))
                    if m:
                        return m.group(0)
        m = _OP_PROD_CODE_FIELD_RE.search(soup.get_text(" ", strip=True))
        return m.group(0) if m else None
    except Exception:  # noqa: BLE001 — extração é best-effort
        return None


def canonical_set_code_op(code: Optional[str]) -> Optional[str]:
    """Código de set → forma canônica pra equivalência MYP ↔ tcgcsv cat 68.

    Remove separadores/espaços e zeros à esquerda do número:
    "ST-35"≡"ST35" · "OP-04"≡"OP04"→"OP4" · "EB-03"≡"EB03"→"EB3" ·
    "OP17 RE"→"OP17RE" · "PRB-01"→"PRB1". SEM as classes de alias do DBZ
    ({B≡BT}/{BE≡EX}) — em One Piece EB é Extra Booster, nunca ≡ EX."""
    c = re.sub(r"[^A-Za-z0-9]", "", str(code or "").upper())
    if not c:
        return None
    m = re.match(r"^([A-Z]+)(\d+)([A-Z]*)$", c)
    if not m:
        return c
    alpha, num, tail = m.groups()
    return f"{alpha}{int(num)}{tail}"


def resolve_edition_group_op(clean_title: str, myp_code: Optional[str],
                             groups: list[dict]) -> Optional[int]:
    """Edição MYP → groupId tcgcsv cat 68. Cascata unique-match-only (mesma
    filosofia do DBZ; nunca chutar em ambíguo):
      1. nome EXATO (alnum) — "Romance Dawn" ↔ "Romance Dawn";
      2. código canônico igual e ÚNICO — "ST12" ↔ "ST-12" (cobre os casos
         em que o MYP encurta o nome, ex. "One Piece Film Edition" vs
         "Starter Deck 5: One Piece Film Edition");
      3. contenção de nome ÚNICA — "Memorial Collection" ⊂ "Extra Booster:
         Memorial Collection".
    Sem match → None (cartas da edição saem em Sem Ref, nunca margem)."""
    if not groups:
        return None
    title_a = _alnum(clean_title)
    if title_a:
        exact = [g for g in groups if _alnum(g.get("name")) == title_a]
        if len(exact) == 1:
            return exact[0]["groupId"]
    code_c = canonical_set_code_op(myp_code)
    if code_c:
        by_code = [g for g in groups
                   if canonical_set_code_op(g.get("abbreviation")) == code_c]
        if len(by_code) == 1:
            return by_code[0]["groupId"]
    if title_a and len(title_a) >= 6:
        contains = [g for g in groups if title_a in _alnum(g.get("name"))]
        if len(contains) == 1:
            return contains[0]["groupId"]
    return None


def _name_cascade_op(cands: list[dict], en_name: str) -> tuple[Optional[dict], str]:
    """Desambiguação por nome dentro de um escopo NÃO-vazio de candidatos.

    Variant-aware pela via do NOME (One Piece não tem token de variante
    confiável — ver cabeçalho): a normalização já converteu "(Alternate
    Art)" → "(parallel)", então o nome qualificado do MYP casa o produto
    "(Parallel)" do tcgcsv por igualdade exata. Regras, na ordem:
      1. nome EXATO normalizado único;
      2. nome QUALIFICADO sem produto qualificado no índice → SEM match
         (nunca cair no produto base — precificaria a variante errada,
         classe de bug nº 3 da frota);
      3. nome-base: candidato NÃO-qualificado único;
      4. candidato único não-qualificado.
    (match, rótulo) ou (None, motivo)."""
    clean = _clean_card_name_op(en_name)
    name_n = _norm_name_op(clean)
    base_n = _norm_name_op(_base_name(clean))
    exact = [c for c in cands if c["name_norm"] == name_n]
    if len(exact) == 1:
        return exact[0], "nome-exato"
    if len(exact) > 1:
        return None, f"nome exato ambíguo ({len(exact)})"
    if name_n != base_n:
        return None, "variante sem produto qualificado no índice"
    unqualified = [c for c in cands
                   if c["name_norm"] == c["base_norm"] and c["base_norm"] == base_n]
    if len(unqualified) == 1:
        return unqualified[0], "nome-base"
    if len(cands) == 1:
        only = cands[0]
        if only["name_norm"] != only["base_norm"]:
            return None, "índice só tem variante qualificada (base ausente)"
        return only, "unico"
    return None, f"ambíguo ({len(cands)} produtos tcgcsv)"


def find_reference_op(code_norm: Optional[str], en_name: str, index: dict,
                      group_id: Optional[int] = None) -> tuple[Optional[dict], str]:
    """Join determinístico carta MYP → produto tcgcsv cat 68 (NUNCA fuzzy).

    COM código: candidatos pelo código com escopo em camadas — (a) grupo
    resolvido da edição (crítico: starter decks reimprimem números OPxx
    com produto/preço PRÓPRIOS), (b) grupos principais (não Release Event/
    Anniversary/Demo Deck), (c) todos; o primeiro escopo NÃO-vazio decide.
    SEM código: grupo resolvido + nome exato único.
    Retorna (entry|None, rótulo/motivo)."""
    if code_norm:
        cands = (index.get("by_code") or {}).get(code_norm) or []
        if not cands:
            return None, "número fora do índice tcgcsv"
        scopes = []
        if group_id is not None:
            scopes.append(("grupo-edição",
                           [c for c in cands if c["group_id"] == group_id]))
        scopes.append(("grupo-principal",
                       [c for c in cands if not c["group_special"]]))
        scopes.append(("global", cands))
        for scope_label, scoped in scopes:
            if not scoped:
                continue
            match, how = _name_cascade_op(scoped, en_name)
            if match is not None:
                return match, f"codigo/{scope_label}/{how}"
            return None, f"código: {how} [{scope_label}]"
        return None, "código sem escopo utilizável"
    if group_id is None:
        return None, "sem código e grupo tcgcsv não resolvido"
    cands = (index.get("by_group") or {}).get(group_id) or []
    if not cands:
        return None, "grupo tcgcsv resolvido mas vazio no índice"
    match, how = _name_cascade_op(cands, en_name)
    if match is not None:
        return match, f"grupo+{how}"
    return None, f"grupo: {how}"


def tcg_op_search_url(name: str, code: Optional[str] = None) -> Optional[str]:
    """URL de BUSCA no TCGplayer (fallback de link pra linhas sem produto
    casado — mesmo papel do tcg_search_url do fluxo Pokémon)."""
    base = _base_name(_clean_card_name_op(name))
    if not base and not code:
        return None
    q = " ".join(p for p in (base, code or "") if p)
    return _TCG_OP_SEARCH_BASE + quote_plus(q)


# ══════════════════════════════════════════════════════════════════════
# Índice de referência tcgcsv (cat 68)
# ══════════════════════════════════════════════════════════════════════

def build_op_index(session, cache_dir: Optional[Path],
                   cache_hours: float = 20.0, cat: int = TCGCSV_OP_CAT) -> dict:
    """Índice da categoria One Piece do tcgcsv:
        {"groups": [{groupId, name, abbreviation}],
         "by_code": {code_norm: [entry]},
         "by_group": {group_id: [entry]}}
    entry = {pid, name, name_norm, base_norm, number, rarity, url, prices,
             lows, cat, group_id, group_name, group_special}.
    name_norm/base_norm usam a normalização OP (número curto removido +
    equivalência alternate-art/parallel). Produtos SEM Number (selados/
    acessórios — 9 no OP01, provados na sonda) ficam fora."""
    out = {"groups": [], "by_code": {}, "by_group": {}}
    groups = _tcgcsv_json(session, f"{TCGCSV_OP_BASE}/{cat}/groups",
                          cache_dir / f"groups_{cat}.json" if cache_dir else None,
                          cache_hours)
    glist = (groups or {}).get("results", [])
    if not glist:
        log.warning(f"tcgcsv categoria {cat}: 0 groups — índice vazio.")
        return out
    out["groups"] = [{"groupId": g.get("groupId"), "name": g.get("name", ""),
                      "abbreviation": g.get("abbreviation", "")} for g in glist]
    log.info(f"  [tcgcsv] categoria {cat}: {len(glist)} groups…")
    by_pid: dict[int, dict] = {}
    for g in glist:
        gid = g["groupId"]
        gname = g.get("name", "")
        special = bool(_SPECIAL_GROUP_RE.search(gname))
        prods = _tcgcsv_json(session, f"{TCGCSV_OP_BASE}/{cat}/{gid}/products",
                             cache_dir / f"products_{cat}_{gid}.json" if cache_dir else None,
                             cache_hours)
        prices = _tcgcsv_json(session, f"{TCGCSV_OP_BASE}/{cat}/{gid}/prices",
                              cache_dir / f"prices_{cat}_{gid}.json" if cache_dir else None,
                              cache_hours)
        for p in (prods or {}).get("results", []):
            ext = {e.get("name"): e.get("value")
                   for e in (p.get("extendedData") or [])}
            number = ext.get("Number") or ""
            code_norm = normalize_code(number)
            if not code_norm:
                continue  # selado/acessório — nunca vira referência de single
            clean = _clean_card_name_op(p.get("name"))
            entry = {
                "pid": int(p["productId"]),
                "name": p.get("name", ""),
                "name_norm": _norm_name_op(clean),
                "base_norm": _norm_name_op(_base_name(clean)),
                "number": number,
                "rarity": ext.get("Rarity") or "",
                "url": p.get("url", ""),
                "prices": {},
                "lows": {},
                "cat": cat,
                "group_id": gid,
                "group_name": gname,
                "group_special": special,
            }
            by_pid[entry["pid"]] = entry
            out["by_code"].setdefault(code_norm, []).append(entry)
            out["by_group"].setdefault(gid, []).append(entry)
        for r in (prices or {}).get("results", []):
            pid = int(r.get("productId") or 0)
            entry = by_pid.get(pid)
            if entry is None:
                continue
            sub = r.get("subTypeName") or "?"
            if r.get("marketPrice") is not None:
                entry["prices"][sub] = float(r["marketPrice"])
            if r.get("lowPrice") is not None:
                entry["lows"][sub] = float(r["lowPrice"])
    n = len(by_pid)
    log.info(f"  [tcgcsv] categoria {cat}: {n} singles indexados "
             f"({len(out['by_code'])} códigos)")
    return out


# ══════════════════════════════════════════════════════════════════════
# Dados
# ══════════════════════════════════════════════════════════════════════

@dataclass
class OpCardData:
    name: str = ""                 # nome display ("Sabo - OP13-004")
    en_name: str = ""              # nome comparável (sem código/cauda)
    code: str = ""                 # código cru (OP13-004; "" se ausente)
    code_norm: str = ""            # chave de join sem zeros ("OP13-004" → "OP13-4")
    edition: str = ""              # título limpo da edição
    section: str = OP_SECTION
    product_url: str = ""
    myp_lowest_en_nm: Optional[float] = None
    en_nm_sellers: int = 0
    en_truncation_risk: bool = False
    single_en_seller_risk: bool = False
    tcg_usd: Optional[float] = None       # market TCGplayer (USD)
    tcg_low_usd: Optional[float] = None   # menor anúncio atual (USD)
    tcg_brl: Optional[float] = None
    tcg_source: str = "tcgcsv"            # única fonte real do OP v1
    tcg_product_name: str = ""
    tcg_url: str = ""
    rarity: str = ""                      # Rarity do tcgcsv (não a do MYP)
    subtype: str = ""                     # Normal | Foil (subtipo usado)
    join_via: str = ""
    margin_pct: Optional[float] = None    # FRAÇÃO (0.42 = 42%)
    margin_brl: Optional[float] = None
    flag_lixo: bool = False               # MYP < 50% da ref → REVISAR
    flag_ref_volatil: bool = False        # market vs low >2× → REVISAR
    last_updated: str = ""


@dataclass
class OpSemRefRow:
    """Oferta EN NM viva ≥ piso SEM referência tcgcsv — nada some em
    silêncio (mesmo papel da aba Sem Ref do DBZ)."""
    name: str = ""
    en_name: str = ""
    code: str = ""
    edition: str = ""
    section: str = OP_SECTION
    product_url: str = ""
    myp_lowest_en_nm: Optional[float] = None
    en_nm_sellers: int = 0
    motivo: str = ""


# ══════════════════════════════════════════════════════════════════════
# Scraper
# ══════════════════════════════════════════════════════════════════════

class MYPOpScraper(MYPScraper):
    """Scraper da seção One Piece do MYP.

    Herda do MYPScraper a infra de PLATAFORMA (sessão cloudscraper firefox,
    `_get` com retry, `_parse_seller_table` NM/EN, `_max_seller_page`,
    `_parse_brl`). O pipeline de preço Pokémon (pokemontcg.io/estat) NÃO é
    usado: a referência OP é o índice tcgcsv cat 68 carregado no início."""

    def __init__(self, delay: float = 1.5, min_en_sellers: int = 2,
                 threshold: float = MARGIN_THRESHOLD_OP,
                 min_price: float = MIN_PRICE_BRL_OP,
                 tcg_cache_dir: Optional[Path] = None, cache_hours: float = 20.0):
        super().__init__(delay=delay, min_en_sellers=min_en_sellers,
                         threshold=threshold, min_price=min_price,
                         tcg_source="tcgcsv")
        self.tcg_cache_dir = tcg_cache_dir
        self.cache_hours = cache_hours
        self.op_cards: list[OpCardData] = []
        self.semref: list[OpSemRefRow] = []
        self.op_index: dict = {}
        self._stats.update({
            "op_sem_codigo": 0,
            "op_sem_ref": 0,
            "op_ref_ok": 0,
            "op_flag_lixo": 0,
            "op_flag_ref_volatil": 0,
            "op_edicoes_sem_grupo": 0,
        })

    # ── catálogo de edições ──
    def get_section_editions(self, section: str = OP_SECTION) -> list[dict]:
        """Edições da seção (/{section}/edicoes?page=N) — mesma cascata de
        seletores do fluxo Pokémon/DBZ; separa título limpo + código de set
        do card concatenado ("Romance DawnOP0102/12/2022")."""
        editions: list[dict] = []
        page = 1
        while page <= MAX_EDITION_PAGES:
            url = f"{BASE_URL}/{section}/edicoes?page={page}"
            log.info(f"  [{section}] edições página {page}…")
            soup = self._get(url)
            if not soup:
                break
            links = soup.select("a.edicao-link")
            if not links:
                for c in soup.select('[class*="edicao"]'):
                    a = c.select_one(f'a[href*="/{section}/"]')
                    if a and a not in links:
                        links.append(a)
            if not links:
                exclude = ("edicoes", "produto", "outros", "selados",
                           "acessorios", "#")
                for a in soup.select("a[href]"):
                    href = a.get("href", "")
                    if (re.match(rf"^/{section}/[a-z0-9][\w-]+$", href)
                            and not any(x in href for x in exclude)):
                        links.append(a)
            if not links:
                break
            found_on_page = 0
            seen = {e["url"] for e in editions}
            for a in links:
                href = a.get("href", "")
                if not href:
                    continue
                full = f"{BASE_URL}{href}" if href.startswith("/") else href
                if full in seen:
                    continue
                raw_title = a.get_text(strip=True)[:120]
                slug = href.rstrip("/").split("/")[-1]
                clean_title, myp_code = split_edition_title(raw_title, slug)
                if not clean_title or len(clean_title) < 2:
                    continue
                editions.append({"title": clean_title, "myp_code": myp_code,
                                 "raw_title": raw_title, "url": full,
                                 "href": href, "slug": slug,
                                 "section": section})
                seen.add(full)
                found_on_page += 1
            if found_on_page == 0:
                break
            page += 1
        log.info(f"  [{section}] {len(editions)} edições")
        return editions

    # ── produtos de uma edição ──
    def get_section_products(self, edition_url: str,
                             section: str = OP_SECTION) -> list[str]:
        """URLs de produto de uma edição (paginada) — espelho do fluxo DBZ
        com o seletor da seção."""
        product_urls: list[str] = []
        seen: set[str] = set()
        page = 1
        prev_first: Optional[str] = None
        while page <= MAX_PAGES_PER_EDITION:
            soup = self._get(f"{edition_url}?page={page}")
            if not soup:
                break
            links = soup.select(f'a[href*="/{section}/produto/"]')
            current_first = None
            for a in links:
                href = a.get("href", "")
                if href:
                    current_first = (f"{BASE_URL}{href}"
                                     if href.startswith("/") else href)
                    break
            if page > 1 and prev_first is not None and current_first == prev_first:
                log.warning(f"  🚨 Loop de paginação em {edition_url} "
                            f"(página {page} == {page - 1}) — parando.")
                break
            prev_first = current_first
            new_count = 0
            for a in links:
                href = a.get("href", "")
                full = f"{BASE_URL}{href}" if href.startswith("/") else href
                if full not in seen:
                    seen.add(full)
                    product_urls.append(full)
                    new_count += 1
            if new_count == 0:
                break
            page += 1
        return product_urls

    # ── página de produto ──
    def scrape_op_product(self, url: str, edition: dict,
                          group_id: Optional[int], fx_usd_brl: float):
        """Uma página de produto OP → OpCardData | OpSemRefRow | None.

        None = fora do funil (sem oferta EN NM, abaixo do piso, oversized,
        fetch falhou) — contado nos stats herdados."""
        soup = self._get(url)
        if not soup:
            return None

        h1 = soup.select_one("h1")
        h1txt = h1.get_text(" ", strip=True) if h1 else ""
        if not h1txt:
            title_tag = soup.find("title")
            if title_tag and title_tag.text:
                h1txt = title_tag.text.split("|")[0].strip()
        if h1txt and OVERSIZED_TITLE_RE.search(h1txt):
            self._stats["skipped_jumbo"] += 1
            log.info(f"  ⏭️ Oversized: {h1txt[:60]}")
            return None

        display, h1_code, en_name = split_myp_title(h1txt)
        # campo "Código" da página (one_st-35_op13-004): fonte PRIMÁRIA do
        # código (o h1 nem sempre traz — "Monkey.D.Dragon"). O marcador de
        # variante é IGNORADO no join (não-confiável — ver cabeçalho); quem
        # marca variante é o qualificador do h1 ("(Alternate Art)").
        field_code, _marker = parse_myp_op_product_code(
            extract_myp_op_product_code(soup))
        code_raw = field_code or h1_code
        code_norm = normalize_code(code_raw)
        if (field_code and h1_code
                and normalize_code(field_code) != normalize_code(h1_code)):
            log.warning(f"  ⚠️ Código do campo ({field_code}) ≠ código do h1 "
                        f"({h1_code}) em {url} — usando o campo.")
        if code_raw and code_raw.lower() not in display.lower():
            display = f"{en_name} - {code_raw}" if en_name else code_raw

        # referência ANTES dos sellers: barata (índice local) e alimenta o
        # cost-gate da paginação (só pagina quem ainda pode virar deal).
        entry, join_label = find_reference_op(code_norm, en_name,
                                              self.op_index, group_id=group_id)
        ref_chosen = None
        if entry is not None:
            prefer_foil = "foil" in _norm_name(display)
            ref_chosen = pick_subtype(entry["prices"], prefer_foil=prefer_foil)

        # ── sellers EN NM (parser herdado) + gate de truncation/paginação ──
        en_prices: list[float] = []
        en_sellers = 0
        per_table = []
        seller_tables = soup.select("table.table-striped.table-bordered") or [soup]
        for table in seller_tables:
            st = self._parse_seller_table(table)
            per_table.append(st)
            en_prices.extend(st["en_prices"])
            en_sellers += st["en"]
            if st["jumbo"]:
                self._stats["jumbo_rows_filtered"] += st["jumbo"]

        TABLE_CAP = 15
        lowest_seen = min(en_prices) if en_prices else None
        gate = False
        if lowest_seen is not None:
            for ts in per_table:
                if (ts["rows"] >= TABLE_CAP and ts["en"] == 0
                        and 0 < ts["max_price"] < lowest_seen):
                    gate = True
                    break

        truncation_risk = False
        max_seller_page = self._max_seller_page(soup)
        can_be_deal = bool(ref_chosen and ref_chosen[1] * fx_usd_brl >= self.min_price)
        if gate and max_seller_page >= 2 and not can_be_deal:
            self._stats["pagination_skipped_low_tcg"] += 1
            if ref_chosen is None:
                truncation_risk = True
        elif gate and max_seller_page >= 2:
            pages = min(max_seller_page, MAX_SELLER_PAGES)
            log.info(f"  📄 Paginando marketplace de {display or url} "
                     f"(2..{pages} de {max_seller_page})")
            for pg in range(2, pages + 1):
                page_soup = self._get(f"{url}?estoque-outros-page={pg}")
                if page_soup is None:
                    truncation_risk = True
                    self._stats["seller_page_fetch_failures"] += 1
                    break
                mkt = page_soup.select_one("#lista-anuncio-demais-vendedores")
                if mkt is None:
                    if pg < pages:
                        truncation_risk = True
                        self._stats["seller_page_empty_early"] += 1
                    break
                pst = self._parse_seller_table(mkt)
                en_prices.extend(pst["en_prices"])
                en_sellers += pst["en"]
                if pst["jumbo"]:
                    self._stats["jumbo_rows_filtered"] += pst["jumbo"]
                self._stats["seller_pages_followed"] += 1
            if max_seller_page > MAX_SELLER_PAGES:
                truncation_risk = True

        if not en_prices:
            self._stats["skipped_no_en_sellers"] += 1
            return None
        lowest_en = min(en_prices)
        if lowest_en < self.min_price:
            self._stats["skipped_low_price"] += 1
            return None

        if truncation_risk:
            self._stats["en_truncation_risks"] += 1

        # ── sem referência → linha honesta na aba própria ──
        if entry is None or ref_chosen is None:
            if not code_norm and "sem código" in join_label:
                self._stats["op_sem_codigo"] += 1
            motivo = (join_label if entry is None
                      else "produto tcgcsv sem market price")
            self._stats["op_sem_ref"] += 1
            return OpSemRefRow(
                name=display, en_name=en_name, code=code_raw or "",
                edition=edition["title"], section=edition["section"],
                product_url=url, myp_lowest_en_nm=lowest_en,
                en_nm_sellers=en_sellers, motivo=motivo,
            )

        subtype, tcg_usd = ref_chosen
        tcg_low = (entry.get("lows") or {}).get(subtype)
        tcg_brl = tcg_usd * fx_usd_brl
        card = OpCardData(
            name=display, en_name=en_name, code=code_raw or "",
            code_norm=code_norm or "", edition=edition["title"],
            section=edition["section"], product_url=url,
            myp_lowest_en_nm=lowest_en, en_nm_sellers=en_sellers,
            en_truncation_risk=truncation_risk,
            single_en_seller_risk=en_sellers < self.min_en_sellers,
            tcg_usd=tcg_usd, tcg_low_usd=tcg_low, tcg_brl=tcg_brl,
            tcg_product_name=entry["name"], tcg_url=entry["url"],
            rarity=entry["rarity"], subtype=subtype, join_via=join_label,
            margin_brl=tcg_brl - lowest_en,
            margin_pct=(tcg_brl - lowest_en) / lowest_en,
            flag_lixo=lowest_en < JUNK_RATIO * tcg_brl,
            flag_ref_volatil=ref_volatile(tcg_usd, tcg_low),
            last_updated=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
        self._stats["op_ref_ok"] += 1
        if card.flag_lixo:
            self._stats["op_flag_lixo"] += 1
        if card.flag_ref_volatil:
            self._stats["op_flag_ref_volatil"] += 1
        if card.single_en_seller_risk:
            self._stats["single_en_seller_risks"] += 1
        return card

    # ── checkpoint/resume (espelho do v5.11.4 do Pokémon) ──
    def _save_op_checkpoint(self, path: str, done: set) -> None:
        try:
            payload = {
                "version": OP_CHECKPOINT_VERSION,
                "cards": [asdict(c) for c in self.op_cards],
                "semref": [asdict(s) for s in self.semref],
                "done_editions": sorted(done),
                "stats": self._stats,
            }
            tmp = f"{path}.tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(payload, f)
            os.replace(tmp, path)
        except Exception as e:  # noqa: BLE001 — best-effort, nunca derruba o scan
            log.warning(f"  ⚠️ Falha ao salvar checkpoint {path}: {e!r}")

    def _load_op_checkpoint(self, path: str) -> set:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:  # noqa: BLE001
            log.warning(f"  ⚠️ Checkpoint {path} ilegível ({e!r}) — do zero.")
            return set()
        if data.get("version") != OP_CHECKPOINT_VERSION:
            log.warning(f"  ⚠️ Checkpoint {path} de versão antiga — ignorando.")
            return set()
        cfields = OpCardData.__dataclass_fields__
        sfields = OpSemRefRow.__dataclass_fields__
        self.op_cards = [OpCardData(**{k: v for k, v in c.items() if k in cfields})
                         for c in data.get("cards", [])]
        self.semref = [OpSemRefRow(**{k: v for k, v in s.items() if k in sfields})
                       for s in data.get("semref", [])]
        st = data.get("stats")
        if isinstance(st, dict):
            self._stats.update(st)
        return set(data.get("done_editions", []))

    # ── scan principal ──
    def scan_op(self, fx_usd_brl: float, edition_filter: Optional[list[str]] = None,
                max_editions: int = 0, max_products: int = 0,
                resume: bool = False, checkpoint_path: Optional[str] = None
                ) -> tuple[list[OpCardData], list[OpSemRefRow]]:
        log.info("═" * 60)
        log.info("  MYP OP Arbitrage Scanner (One Piece — paralelo ao Pokémon)")
        log.info(f"  Threshold: {self.margin_threshold * 100:.0f}% | EN only | NM only")
        log.info(f"  Piso: R${self.min_price:.0f} | Câmbio: US$1 = R${fx_usd_brl:.4f}")
        log.info("═" * 60)

        # índice de referência da categoria 68 (bulk, 1×)
        if not self.op_index:
            self.op_index = build_op_index(self.session, self.tcg_cache_dir,
                                           self.cache_hours)
        if not (self.op_index.get("by_code") or {}):
            raise RuntimeError(
                "Índice tcgcsv vazio na categoria 68 — sem referência de "
                "preço não há scan honesto (nunca inventar preço).")

        editions = self.get_section_editions(OP_SECTION)

        if edition_filter:
            fl = [f.lower().strip() for f in edition_filter]
            matched = []
            for ed in editions:
                hay = f"{ed['title']} {ed.get('myp_code') or ''}".lower()
                for f in fl:
                    if f in hay:
                        matched.append(ed)
                        log.info(f"  ✅ Matched: '{ed['title']}' "
                                 f"[{ed.get('myp_code') or '—'}] (filtro: '{f}')")
                        break
            editions = matched
            if not editions:
                log.warning("Nenhuma edição casou o filtro!")
                return [], []
        if max_editions:
            editions = editions[:max_editions]

        done: set = set()
        if resume and checkpoint_path and os.path.exists(checkpoint_path):
            done = self._load_op_checkpoint(checkpoint_path)
            log.info(f"  ⏯️ Resume: {len(done)} edições feitas, "
                     f"{len(self.op_cards)} cards restaurados.")

        for i, ed in enumerate(editions):
            if ed["url"] in done:
                log.info(f"[{i + 1}/{len(editions)}] ⏭️ (resume) {ed['title']}")
                continue
            group_id = resolve_edition_group_op(
                ed["title"], ed.get("myp_code"),
                self.op_index.get("groups") or [])
            if group_id is None:
                self._stats["op_edicoes_sem_grupo"] += 1
                log.warning(f"  ⚠️ Edição sem grupo tcgcsv resolvido: "
                            f"'{ed['title']}' [{ed.get('myp_code') or '—'}] — "
                            f"cartas sem código sairão em Sem Ref.")
            log.info(f"\n[{i + 1}/{len(editions)}] 📦 [{ed['section']}] "
                     f"{ed['title']} [{ed.get('myp_code') or '—'}]"
                     f"{f' → group {group_id}' if group_id else ''}")
            product_urls = self.get_section_products(ed["url"], ed["section"])
            if max_products:
                product_urls = product_urls[:max_products]
            log.info(f"  → {len(product_urls)} produtos")
            for j, purl in enumerate(product_urls):
                self._stats["products_scanned"] += 1
                if (j + 1) % 10 == 0:
                    log.info(f"  {j + 1}/{len(product_urls)}…")
                result = self.scrape_op_product(purl, ed, group_id, fx_usd_brl)
                if result is None:
                    continue
                self._stats["en_found"] += 1
                if isinstance(result, OpSemRefRow):
                    self.semref.append(result)
                    continue
                self.op_cards.append(result)
                if result.margin_pct is not None and \
                        result.margin_pct >= self.margin_threshold:
                    log.info(f"  🔥 DEAL: {result.name} | "
                             f"MYP R${result.myp_lowest_en_nm:,.2f} | "
                             f"TCG R${result.tcg_brl:,.2f} | "
                             f"{result.margin_pct * 100:.1f}%")
            if checkpoint_path:
                done.add(ed["url"])
                self._save_op_checkpoint(checkpoint_path, done)

        deals = [c for c in self.op_cards
                 if c.margin_pct and c.margin_pct >= self.margin_threshold]
        log.info("\n" + "═" * 60)
        log.info(f"  Produtos escaneados: {self._stats['products_scanned']}")
        log.info(f"  Cards EN NM ≥ piso: {len(self.op_cards) + len(self.semref)}")
        log.info(f"  Com referência tcgcsv: {self._stats['op_ref_ok']} | "
                 f"SEM referência: {self._stats['op_sem_ref']} "
                 f"(sem código: {self._stats['op_sem_codigo']})")
        log.info(f"  Edições sem grupo tcgcsv: {self._stats['op_edicoes_sem_grupo']}")
        log.info(f"  🔥 Deals (≥{self.margin_threshold * 100:.0f}%): {len(deals)}")
        log.info(f"  Flags: possível lixo={self._stats['op_flag_lixo']} | "
                 f"ref volátil={self._stats['op_flag_ref_volatil']} | "
                 f"truncation={self._stats['en_truncation_risks']}")
        log.info("═" * 60)

        if checkpoint_path and os.path.exists(checkpoint_path):
            try:
                os.remove(checkpoint_path)
            except OSError:
                pass
        return self.op_cards, self.semref


# ══════════════════════════════════════════════════════════════════════
# XLSX
# ══════════════════════════════════════════════════════════════════════

OP_HEADERS = [
    "Card Name", "EN Name", "Código", "Edition", "Seção", "Rarity",
    "MYP EN NM (R$)", "TCG US$", "TCG Low US$", "TCG Player (R$)",
    "TCG Source", "Subtipo", "Margin %", "Diff (R$)", "NM Sellers",
    "⚠️ EN Trunc", "⚠️ 1 Seller", "⚠️ Possível Lixo", "⚠️ Ref Volátil",
    "Join", "URL", "Updated", "TCG URL",
]
OP_SEMREF_HEADERS = [
    "Card Name", "EN Name", "Código", "Edition", "Seção",
    "MYP EN NM (R$)", "NM Sellers", "Motivo", "URL",
]


def generate_op_xlsx(cards: list[OpCardData], semref: list[OpSemRefRow],
                     output_path: str, threshold: float, fx: float,
                     stats: Optional[dict] = None,
                     editions_scanned: int = 0,
                     min_price: float = MIN_PRICE_BRL_OP) -> None:
    """XLSX de trabalho (insumo do myp_op_summary.py — a ENTREGA é o
    markdown do summary, nunca este arquivo)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1A5276")   # azul One Piece
    normal = Font(name="Arial", size=10)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    def write_headers(ws, headers, widths):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        ws.freeze_panes = "A2"

    # ── All EN Cards ──
    ws = wb.active
    ws.title = "All EN Cards"
    write_headers(ws, OP_HEADERS,
                  [34, 30, 11, 28, 12, 16, 14, 11, 11, 14, 10, 10, 10, 12,
                   10, 10, 10, 12, 12, 26, 55, 16, 55])
    margin_col = OP_HEADERS.index("Margin %") + 1
    for row, c in enumerate(
            sorted(cards, key=lambda x: -(x.margin_pct or -999)), start=2):
        vals = [
            c.name, c.en_name, c.code, c.edition,
            SECTION_LABELS.get(c.section, c.section), c.rarity,
            c.myp_lowest_en_nm, c.tcg_usd, c.tcg_low_usd, c.tcg_brl,
            c.tcg_source, c.subtype, c.margin_pct, c.margin_brl,
            c.en_nm_sellers,
            "⚠️ MAYBE" if c.en_truncation_risk else "",
            "⚠️ 1 SELLER" if c.single_en_seller_risk else "",
            "⚠️ LIXO?" if c.flag_lixo else "",
            "⚠️ VOLÁTIL" if c.flag_ref_volatil else "",
            c.join_via, c.product_url, c.last_updated, c.tcg_url,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = normal
            if col == margin_col and v is not None:
                cell.number_format = "0.00%"
                if v >= 0.50:
                    cell.fill = green_fill
                elif v >= threshold:
                    cell.fill = yellow_fill

    # ── Sem Ref TCG ──
    ws2 = wb.create_sheet("Sem Ref TCG")
    write_headers(ws2, OP_SEMREF_HEADERS, [34, 30, 11, 28, 12, 14, 10, 46, 55])
    for row, s in enumerate(semref, start=2):
        vals = [s.name, s.en_name, s.code, s.edition,
                SECTION_LABELS.get(s.section, s.section),
                s.myp_lowest_en_nm, s.en_nm_sellers, s.motivo, s.product_url]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=row, column=col, value=v).font = normal

    # ── Summary ──
    ws3 = wb.create_sheet("Summary")
    deals = [c for c in cards if c.margin_pct and c.margin_pct >= threshold]
    pairs = [
        ("Scanner", f"myp_op_scanner {OP_VERSION} (One Piece)"),
        ("Scan Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Margin Threshold", f"{threshold * 100:.0f}%"),
        ("Min Price (BRL)", min_price),
        ("FX USD→BRL", fx),
        ("Editions Scanned", editions_scanned),
        ("Total EN Cards", len(cards)),
        ("Sem Ref TCG", len(semref)),
        ("Deals Found", len(deals)),
    ]
    for k, v in list((stats or {}).items()):
        if k.startswith("op_") or k in (
                "products_scanned", "en_truncation_risks",
                "single_en_seller_risks", "skipped_no_en_sellers",
                "skipped_low_price", "skipped_jumbo", "http_retries"):
            pairs.append((f"stat:{k}", v))
    for row, (k, v) in enumerate(pairs, start=1):
        ws3.cell(row=row, column=1, value=k).font = Font(bold=True, name="Arial",
                                                         size=10)
        ws3.cell(row=row, column=2, value=v).font = normal
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 40

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info(f"  💾 XLSX: {out}")


# ══════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════

def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Scanner MYP → One Piece (seção onepiece vs TCGplayer "
                    "via tcgcsv cat 68). Paralelo ao scanner Pokémon.")
    ap.add_argument("--editions", nargs="*", default=None,
                    help="substrings de título/código de edição MYP (default: "
                         "TODAS as 65 edições da seção)")
    ap.add_argument("--threshold", type=float, default=30.0,
                    help="margem BRUTA mínima em PERCENT INTEIRO (30 = 30%%; "
                         "convenção MYP — CardTrader usa fração)")
    ap.add_argument("--min-price", type=float, default=MIN_PRICE_BRL_OP,
                    help="piso de relevância em R$ (default 50 — canônico)")
    ap.add_argument("--delay", type=float, default=1.5,
                    help="delay entre requests (s)")
    ap.add_argument("--min-en-sellers", type=int, default=2,
                    help="abaixo disso a linha ganha flag '1 SELLER' → REVISAR")
    ap.add_argument("--max-editions", type=int, default=0)
    ap.add_argument("--max-products", type=int, default=0)
    ap.add_argument("--fx", type=float, default=None,
                    help="câmbio USD→BRL manual (default: frankfurter/er-api "
                         "ao vivo; sem fonte real o run FALHA ALTO)")
    ap.add_argument("--cache-hours", type=float, default=20.0,
                    help="TTL do cache tcgcsv em disco (0 = sem cache)")
    ap.add_argument("--tcg-cache-dir", default="results/op_cache",
                    help="pasta do cache tcgcsv (dentro de results/, "
                         "gitignored; NUNCA compartilhada com o dbz_cache)")
    ap.add_argument("--resume", action="store_true",
                    help="retoma do checkpoint <output>.resume.json")
    ap.add_argument("--list-editions", action="store_true",
                    help="lista as edições OP do MYP e sai (sem scan)")
    ap.add_argument("-o", "--output", default=None,
                    help="XLSX de saída (default results/myp_op_<ts>.xlsx)")
    args = ap.parse_args(argv)

    threshold = parse_threshold(args.threshold)
    scraper = MYPOpScraper(
        delay=args.delay, min_en_sellers=args.min_en_sellers,
        threshold=threshold, min_price=args.min_price,
        tcg_cache_dir=Path(args.tcg_cache_dir) if args.tcg_cache_dir else None,
        cache_hours=args.cache_hours,
    )

    if args.list_editions:
        for ed in scraper.get_section_editions(OP_SECTION):
            print(f"[{OP_SECTION}] {ed['title']}  [{ed.get('myp_code') or '—'}]"
                  f"  →  {ed['url']}")
        return 0

    # câmbio: manual OU ao vivo; sem fonte real = falha alta (a referência
    # OP é 100% USD — sem FX não existe margem honesta possível).
    fx = args.fx or fetch_usd_brl(scraper.session)
    if not fx or fx <= 0:
        log.error("Sem câmbio USD→BRL real (frankfurter/er-api falharam e "
                  "--fx não foi passado). Nunca inventar câmbio — abortando.")
        return 2

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or f"results/myp_op_{stamp}.xlsx"
    checkpoint = f"{output}.resume.json"

    cards, semref = scraper.scan_op(
        fx_usd_brl=fx, edition_filter=args.editions,
        max_editions=args.max_editions, max_products=args.max_products,
        resume=args.resume, checkpoint_path=checkpoint,
    )
    generate_op_xlsx(cards, semref, output, threshold, fx,
                     stats=scraper._stats,
                     editions_scanned=len({c.edition for c in cards}
                                          | {s.edition for s in semref}),
                     min_price=args.min_price)
    print(f"\nOK: {output}")
    print(f"Entrega: python myp_op_summary.py {output} -o "
          f"results/op-{datetime.now().strftime('%Y-%m-%d')}.md")
    return 0


if __name__ == "__main__":
    sys.exit(main())
