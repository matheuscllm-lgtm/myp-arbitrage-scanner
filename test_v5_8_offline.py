"""
Offline regression test for v5.8 fix (Jirachi PR-SM_SM161 case).

Sem rede / sem CloudFlare. Exercita a lógica determinística com fixtures
sintéticas: TCG suspect filter + language-by-condition detection + XLSX
end-to-end surfacing.

Run: python test_v5_8_offline.py
Exit 0 = todos asserts passam, 1 = regressão.
"""
import os
import sys
import tempfile
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from myp_arbitrage_scanner import (
    CardData,
    MYPScraper,
    _clean_secret,
    generate_xlsx,
    tcg_search_url,
    tcg_direct_url,
    myp_edition_to_ptcg_setcode,
    TCG_SUSPECT_RATIO_THRESHOLD,
    OVERSIZED_TITLE_RE,
    OVERSIZED_FOIL_RE,
)

# NOTE (v5.8.8, 2026-05-29): PT_CONDITION_MARKERS / EN_CONDITION_MARKERS foram
# REMOVIDOS do scanner no commit a4d2111 (a detecção de idioma migrou de
# "substring na linha" pra leitura da célula dedicada td.estoque-lista-
# qualidadenome). O test_v5_8_offline.py importava esses símbolos e estava
# QUEBRADO em ImportError desde então (pré-existente, não introduzido aqui).
# Os dois testes que dependiam deles (language markers disjoint + H1 language
# logic) foram removidos por testarem lógica que não existe mais.


def make_jirachi():
    """Reproduz o caso real: TCG declarado 75x última venda."""
    return CardData(
        name="Jirachi PR-SM SM161",
        edition="Sol & Lua Promos",
        rarity="Holo Rara",
        product_url="https://mypcards.com/pokemon/sol-lua-promos/jirachi-sm161",
        myp_lowest_en_nm=99.99,
        tcg_player_price=1499.00,
        myp_last_sale_brl=19.99,
        tcg_suspect=True,
        margin_pct=14.0,
        margin_brl=1399.01,
        en_nm_sellers=1,
        single_en_seller_risk=True,  # MYP-LOW-a 2026-05-30: coerência com scanner real (en_nm_sellers=1 < min default 2)
        last_updated="2026-05-16 19:30",
    )


def make_clean_deal():
    """Deal legítimo: ratio TCG/last_sale dentro do normal (1.1x)."""
    return CardData(
        name="Charizard ex Holo",
        edition="Surging Sparks",
        rarity="Special Illustration",
        product_url="https://mypcards.com/pokemon/surging-sparks/charizard-ex",
        myp_lowest_en_nm=80.00,
        tcg_player_price=200.00,
        myp_last_sale_brl=180.00,
        tcg_suspect=False,
        margin_pct=1.5,
        margin_brl=120.00,
        en_nm_sellers=4,
        last_updated="2026-05-16 19:30",
    )


def make_borderline_deal():
    """Ratio 9.5x — abaixo do threshold 10x, NÃO deve ser suspect."""
    return CardData(
        name="Mew V Alt Art",
        edition="Surging Sparks",
        rarity="Ultra Rara",
        product_url="https://mypcards.com/pokemon/x/mew-v",
        myp_lowest_en_nm=50.00,
        tcg_player_price=950.00,
        myp_last_sale_brl=100.00,
        tcg_suspect=False,
        margin_pct=18.0,
        margin_brl=900.00,
        en_nm_sellers=2,
        last_updated="2026-05-16 19:30",
    )


def find_card_in_sheet(ws, name_substring):
    """Procura linha onde col A contém substring."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and name_substring.lower() in str(row[0]).lower():
            return row
    return None


def test_xlsx_end_to_end():
    """H2 + surfacing: Jirachi NÃO em 🔥 Deals, ESTÁ em 🚨 TCG Suspect."""
    jirachi = make_jirachi()
    clean = make_clean_deal()
    borderline = make_borderline_deal()
    cards = [jirachi, clean, borderline]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = f.name
    generate_xlsx(cards, out, threshold=0.25)

    wb = load_workbook(out)
    print(f"  Sheets: {wb.sheetnames}")

    # Assert 1: 🚨 TCG Suspect existe
    assert "🚨 TCG Suspect" in wb.sheetnames, "Sheet '🚨 TCG Suspect' ausente"

    # Assert 2: Jirachi NÃO em Deals
    ws_deals = wb["🔥 Deals"]
    jirachi_row = find_card_in_sheet(ws_deals, "Jirachi")
    assert jirachi_row is None, f"BUG: Jirachi vazou para 🔥 Deals: {jirachi_row}"

    # Assert 3: Jirachi ESTÁ em Suspect
    ws_susp = wb["🚨 TCG Suspect"]
    jirachi_susp = find_card_in_sheet(ws_susp, "Jirachi")
    assert jirachi_susp is not None, "BUG: Jirachi não apareceu em 🚨 TCG Suspect"
    # v5.11.1: coluna "TCG US$" inserida em idx 5 → last_sale agora idx 6.
    print(f"  Jirachi row em Suspect: name={jirachi_susp[0]!r}, last_sale={jirachi_susp[6]}")

    # Assert 4: clean deal está em Deals
    clean_row = find_card_in_sheet(ws_deals, "Charizard")
    # margin 1.5 (150%) >> threshold 0.25 → entra
    # (margin_pct é Decimal 1.5 = 150% no current code)
    # Borderline 18 (1800%) também entra
    # Actually let me check — threshold 0.25 = 25%. clean margin_pct=1.5 means 150%.
    assert clean_row is not None, "Clean deal sumiu de 🔥 Deals"
    # v5.11.1: "TCG US$" em idx 5 → margin agora idx 7.
    print(f"  Clean deal em Deals: name={clean_row[0]!r}, margin={clean_row[7]}")

    # Assert 5: borderline (ratio 9.5x, abaixo do 10x) NÃO é suspect
    border_susp = find_card_in_sheet(ws_susp, "Mew")
    assert border_susp is None, f"FALSO POSITIVO: borderline 9.5x foi marcado como suspect: {border_susp}"
    border_deal = find_card_in_sheet(ws_deals, "Mew")
    assert border_deal is not None, "Borderline sumiu de Deals (não deveria — não é suspect)"

    # Assert 6: Summary tem TCG Suspects count
    ws_sum = wb["Summary"]
    summary_text = []
    for row in ws_sum.iter_rows(values_only=True):
        summary_text.append([str(c) if c else "" for c in row])
    flat = "\n".join("|".join(r) for r in summary_text)
    assert "TCG Suspects" in flat, "Summary não menciona TCG Suspects"
    assert "Deals Found (clean)" in flat, "Summary não menciona 'Deals Found (clean)'"
    print(f"  Summary contém TCG Suspects ✓")

    # Cleanup
    Path(out).unlink()
    return True


def test_tcg_search_url():
    """v5.8.8: URL de busca TCGplayer remove o sufixo (NNN/MMM) e codifica."""
    u = tcg_search_url("Rayquaza VMAX (111/203)")
    assert u is not None
    assert "tcgplayer.com" in u, f"domínio errado: {u}"
    assert "Rayquaza" in u and "111" not in u, f"sufixo não removido: {u}"
    # apóstrofo e espaço codificados
    u2 = tcg_search_url("Team Aqua's Kyogre (003/95)")
    assert "Kyogre" in u2 and "%27" in u2, f"encoding falhou: {u2}"
    # promo sem (NNN/MMM) standard
    u3 = tcg_search_url("Pikachu (PR-SM SM229)")
    assert u3 is not None and "Pikachu" in u3 and "SM229" not in u3, f"promo: {u3}"
    # nome vazio → None (não gera link de busca vazio)
    assert tcg_search_url("") is None
    assert tcg_search_url(None) is None
    print(f"  tcg_search_url: 6 casos OK ✓")
    return True


def test_price_cell_hyperlinks():
    """v5.8.8: célula MYP EN NM linka pro produto MYP, TCG Player pra busca
    TCGplayer. Valor (número) preservado, hyperlink + fonte azul/sublinhada.
    Verifica nas 5 sheets de cards (não na Summary).

    v5.8.9 (2026-05-29): TCG link agora pode ser DIRETO (pokemontcg.io
    redirect) quando MYP edition mapeada + collector# in-range. Quando não,
    cai no fallback de busca (tcgplayer.com/search). Asserção genérica
    "tcgplayer.com in target" cobre ambos (search) E (product redirect
    final) — domínio pokemontcg.io é checado no test_tcg_direct_url.
    """
    clean = make_clean_deal()       # entra em Deals + All + Top50
    border = make_borderline_deal() # entra em Deals + All + Top50
    jirachi = make_jirachi()        # suspect → TCG Suspect sheet
    trunc = CardData(
        name="Psyduck (053/198)",
        edition="Scarlet & Violet",   # NÃO mapeado → fallback search
        product_url="https://mypcards.com/pokemon/produto/99999/psyduck",
        myp_lowest_en_nm=300.0,
        tcg_player_price=400.0,
        myp_last_sale_brl=380.0,
        margin_pct=0.30,
        margin_brl=100.0,
        en_nm_sellers=3,
        en_truncation_risk=True,     # → Validate Manually
        last_updated="2026-05-29 12:00",
    )
    # v5.8.9: card cuja edition É mapeada → hyperlink DIRETO pokemontcg.io.
    # Tef = Temporal Forces → sv5. Salamence ex (187/159) é oversized →
    # fallback DEVE acontecer (oversized_collector_risk=True).
    direct = CardData(
        name="Iron Hands ex (070/162)",
        edition="Temporal Forces",
        product_url="https://mypcards.com/pokemon/produto/12345/iron-hands-ex",
        myp_lowest_en_nm=120.0,
        tcg_player_price=240.0,
        myp_last_sale_brl=220.0,
        margin_pct=1.00,
        margin_brl=120.0,
        en_nm_sellers=5,
        last_updated="2026-05-29 12:00",
    )
    cards = [clean, border, jirachi, trunc, direct]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = f.name
    generate_xlsx(cards, out, threshold=0.25)
    wb = load_workbook(out)

    card_sheets = [
        "🔥 Deals", "All EN Cards", "🏆 Top 50 Margin",
        "🚨 Validate Manually", "🚨 TCG Suspect",
    ]
    checked = 0
    for sname in card_sheets:
        assert sname in wb.sheetnames, f"sheet ausente: {sname}"
        ws = wb[sname]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        myp_col = hdr.index("MYP EN NM (R$)") + 1
        tcg_col = hdr.index("TCG Player (R$)") + 1
        rows_seen = 0
        for r in range(2, ws.max_row + 1):
            myp_cell = ws.cell(row=r, column=myp_col)
            tcg_cell = ws.cell(row=r, column=tcg_col)
            if myp_cell.value is None and tcg_cell.value is None:
                continue
            rows_seen += 1
            # MYP price cell: número preservado + hyperlink mypcards
            if myp_cell.value is not None:
                assert isinstance(myp_cell.value, (int, float)), \
                    f"{sname} r{r}: valor MYP não numérico: {myp_cell.value!r}"
                assert myp_cell.hyperlink is not None, \
                    f"{sname} r{r}: MYP price sem hyperlink"
                assert "mypcards.com" in myp_cell.hyperlink.target, \
                    f"{sname} r{r}: hyperlink MYP errado: {myp_cell.hyperlink.target}"
                assert myp_cell.font.underline == "single", \
                    f"{sname} r{r}: MYP price sem underline"
            # TCG price cell: número preservado + hyperlink TCG.
            # v5.8.9: aceita ambos os esquemas — direct (pokemontcg.io
            # redirect → tcgplayer.com/product/<id>) ou fallback search
            # (tcgplayer.com/search). Assertions específicas por caso
            # estão no bloco abaixo.
            if tcg_cell.value is not None:
                assert isinstance(tcg_cell.value, (int, float)), \
                    f"{sname} r{r}: valor TCG não numérico: {tcg_cell.value!r}"
                assert tcg_cell.hyperlink is not None, \
                    f"{sname} r{r}: TCG price sem hyperlink"
                target = tcg_cell.hyperlink.target
                ok = ("tcgplayer.com" in target
                      or "prices.pokemontcg.io/tcgplayer/" in target)
                assert ok, f"{sname} r{r}: hyperlink TCG errado: {target}"
        assert rows_seen >= 1, f"{sname}: nenhuma row de card pra checar"
        checked += rows_seen

    # Card Name e URL NÃO devem ser quebrados: URL ainda string http, Card Name
    # ainda o (NNN/MMM) intacto.
    ws_all = wb["All EN Cards"]
    hdr = [c.value for c in next(ws_all.iter_rows(min_row=1, max_row=1))]
    name_col = hdr.index("Card Name") + 1
    url_col = hdr.index("URL") + 1
    # Localiza o Psyduck (053/198): tem o token (NNN/MMM) que NÃO pode quebrar.
    psyduck_name = None
    for r in range(2, ws_all.max_row + 1):
        nv = ws_all.cell(row=r, column=name_col).value
        uv = ws_all.cell(row=r, column=url_col).value
        # URL não pode ter sido quebrada em nenhuma row
        assert uv is None or str(uv).startswith("http"), f"URL quebrada r{r}: {uv!r}"
        if nv and "Psyduck" in str(nv):
            psyduck_name = nv
    assert psyduck_name == "Psyduck (053/198)", \
        f"Card Name (NNN/MMM) quebrado/alterado: {psyduck_name!r}"

    # v5.8.9: Iron Hands ex (Temporal Forces, 070/162, in-range) → DIRECT link
    # `prices.pokemontcg.io/tcgplayer/sv5-70`. Psyduck (Scarlet & Violet — não
    # mapeado) → fallback `tcgplayer.com/search`.
    direct_link = None
    psyduck_link = None
    for r in range(2, ws_all.max_row + 1):
        nv = ws_all.cell(row=r, column=name_col).value
        tcg_cell = ws_all.cell(row=r, column=hdr.index("TCG Player (R$)") + 1)
        if not nv or not tcg_cell.hyperlink:
            continue
        if "Iron Hands" in str(nv):
            direct_link = tcg_cell.hyperlink.target
        elif "Psyduck" in str(nv):
            psyduck_link = tcg_cell.hyperlink.target
    assert direct_link is not None, "Iron Hands sumiu de All EN Cards"
    assert direct_link == "https://prices.pokemontcg.io/tcgplayer/sv5-70", \
        f"Iron Hands não pegou direct link: {direct_link!r}"
    assert psyduck_link is not None, "Psyduck sumiu de All EN Cards"
    assert "tcgplayer.com/search" in psyduck_link, \
        f"Psyduck (Scarlet & Violet, unmapped) não caiu no fallback: {psyduck_link!r}"

    Path(out).unlink()
    print(f"  Hyperlinks de preço OK em {len(card_sheets)} sheets ({checked} células checadas) ✓")
    print(f"  Direct link (sv5-70) e fallback (search) confirmados ✓")
    return True


def test_tcg_url_column():
    """v5.11.2: coluna "TCG URL" texto plano (última coluna do XLSX).

    O scanner integrado lê o XLSX dict-by-name (hyperlink de célula não
    sobrevive) — a coluna expõe o MESMO link do hyperlink da célula
    "TCG Player (R$)": direct (pokemontcg.io redirect) quando a edição é
    mapeada + collector# in-range, senão fallback de busca por nome.
    """
    clean = make_clean_deal()  # sem (NNN/MMM) no nome → fallback search
    direct = CardData(
        name="Iron Hands ex (070/162)",
        edition="Temporal Forces",      # mapeada → sv5 → direct link
        product_url="https://mypcards.com/pokemon/produto/12345/iron-hands-ex",
        myp_lowest_en_nm=120.0,
        tcg_player_price=240.0,
        myp_last_sale_brl=220.0,
        margin_pct=1.00,
        margin_brl=120.0,
        en_nm_sellers=5,
        last_updated="2026-06-10 12:00",
    )
    cards = [clean, direct]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = f.name
    generate_xlsx(cards, out, threshold=0.25)
    wb = load_workbook(out)

    ws = wb["All EN Cards"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # v5.14: +1 coluna "TCG Source" (real/fallback) → 18 colunas. "TCG URL"
    # segue sendo a última (o source foi inserido no meio, após "TCG US$").
    assert len(hdr) == 18, f"esperava 18 colunas, veio {len(hdr)}: {hdr}"
    assert hdr[-1] == "TCG URL", f"última coluna deveria ser 'TCG URL': {hdr[-1]!r}"
    assert "TCG Source" in hdr, f"coluna 'TCG Source' (v5.14) ausente: {hdr}"
    url_col = hdr.index("TCG URL") + 1
    tcg_price_col = hdr.index("TCG Player (R$)") + 1
    name_col = hdr.index("Card Name") + 1

    rows_checked = 0
    for r in range(2, ws.max_row + 1):
        nv = ws.cell(row=r, column=name_col).value
        if not nv:
            continue
        plain = ws.cell(row=r, column=url_col).value
        price_cell = ws.cell(row=r, column=tcg_price_col)
        assert plain and str(plain).startswith("http"), \
            f"r{r} ({nv}): TCG URL vazia/inválida: {plain!r}"
        # texto plano == hyperlink da célula de preço (mesma fonte)
        if price_cell.hyperlink:
            assert plain == price_cell.hyperlink.target, \
                f"r{r} ({nv}): TCG URL {plain!r} != hyperlink {price_cell.hyperlink.target!r}"
        if "Iron Hands" in str(nv):
            assert plain == "https://prices.pokemontcg.io/tcgplayer/sv5-70", \
                f"direct esperado: {plain!r}"
        elif "Charizard" in str(nv):
            assert "tcgplayer.com/search" in plain, f"fallback search esperado: {plain!r}"
        rows_checked += 1
    assert rows_checked == 2, f"esperava 2 rows, vi {rows_checked}"

    Path(out).unlink()
    print(f"  Coluna 'TCG URL' (última, 18 cols) OK: direct + fallback ✓")
    return True


def test_myp_edition_to_setcode():
    """v5.8.9: mapeamento MYP edition → pokemontcg.io setcode.

    Cobre forma bilingual concat (MYP cola PT+EN) — substrings EN do mapa
    casam mesmo dentro da concat. Cobre longest-substring win (evita "Mega
    Evolution" overriding "Ascended Heroes").
    """
    # Direct match
    assert myp_edition_to_ptcg_setcode("Temporal Forces") == "sv5"
    assert myp_edition_to_ptcg_setcode("Stellar Crown") == "sv7"
    # Bilingual concat (MYP common form)
    assert myp_edition_to_ptcg_setcode(
        "Espada e Escudo 7: Céus em EvoluçãoSword & Shield 7: Evolving Skies"
    ) == "swsh7"
    assert myp_edition_to_ptcg_setcode(
        "Escarlate e Violeta: Amigos de JornadaSV09: Journey Together"
    ) == "sv9"
    # Longest-substring win: "Ascended Heroes" (mapped to me2pt5) beats
    # "Mega Evolution" (mapped to me1) when both substrings present.
    assert myp_edition_to_ptcg_setcode("ME: Ascended Heroes") == "me2pt5"
    # Case-insensitive
    assert myp_edition_to_ptcg_setcode("STELLAR CROWN") == "sv7"
    # v5.16: eras antigas agora MAPEADAS (substring EN única no título MYP).
    assert myp_edition_to_ptcg_setcode("Black & White 9: Plasma Freeze") == "bw9"
    assert myp_edition_to_ptcg_setcode("Sun & Moon 9: Team Up") == "sm9"
    assert myp_edition_to_ptcg_setcode("XY 7: Ancient Origins") == "xy7"
    assert myp_edition_to_ptcg_setcode("EX 3: Dragon") == "ex3"
    # v5.16: DP2 vem como 'Diamond & PEARLS 2' (typo MYP) → casa pelo nome do set.
    assert myp_edition_to_ptcg_setcode(
        "Diamante & Pérola 2: Tesouros MisteriososDiamond & Pearls 2: Mysterious Treasures"
    ) == "dp2"
    # Unmapped: nomes-BASE/promo/SV base seguem None (substrings distintivas do
    # v5.16 NÃO casam o título-base sozinho — sem over-match).
    assert myp_edition_to_ptcg_setcode("Diamond & Pearl") is None       # base DP, não numerado
    assert myp_edition_to_ptcg_setcode("Scarlet & Violet") is None
    assert myp_edition_to_ptcg_setcode("Sun & Moon Promos") is None
    assert myp_edition_to_ptcg_setcode("Black & White") is None         # base BW, não numerado
    # Edge cases
    assert myp_edition_to_ptcg_setcode("") is None
    assert myp_edition_to_ptcg_setcode(None) is None
    print(f"  Mapeamento edition→setcode: 16 casos OK ✓")
    return True


def test_tcg_direct_url():
    """v5.8.9: monta URL DIRETA via redirect pokemontcg.io.

    Forma: prices.pokemontcg.io/tcgplayer/{setcode}-{num}. Strip leading
    zeros (verificado contra Link TCG do CT handoff: base6-13 não base6-013).
    None em todos os casos onde caller deve cair no fallback.
    """
    # In-range, edition mapeada → direct
    u = tcg_direct_url("Iron Hands ex (070/162)", "Temporal Forces")
    assert u == "https://prices.pokemontcg.io/tcgplayer/sv5-70", f"got {u!r}"
    # Leading zeros stripped (base6-13 style)
    u = tcg_direct_url("Some Card (013/110)", "Evolving Skies")
    assert u == "https://prices.pokemontcg.io/tcgplayer/swsh7-13", f"got {u!r}"
    # Bilingual concat edition
    u = tcg_direct_url(
        "Regidrago V (184/195)",
        "Espada e Escudo 12: Tempestade PrateadaSword & Shield 12: Silver Tempest",
    )
    assert u == "https://prices.pokemontcg.io/tcgplayer/swsh12-184", f"got {u!r}"
    # Oversized (numerator > set_size) → None, caller fallback
    # (Mega Feraligatr ex 274/217 — confirmado 404 em pokemontcg.io 2026-05-29)
    u = tcg_direct_url("Mega Feraligatr ex (274/217)", "ME: Ascended Heroes",
                       oversized_collector_risk=True)
    assert u is None, f"oversized deve cair no fallback: {u!r}"
    # Edition não mapeada → None
    u = tcg_direct_url("Charizard (004/102)", "Base Set")
    assert u is None
    u = tcg_direct_url("Random (053/198)", "Scarlet & Violet")
    assert u is None  # SV base NÃO está no mapa
    # Sem (NNN/MMM) parseável → None (promo PR-SM_SM161 style)
    u = tcg_direct_url("Jirachi PR-SM SM161", "Sol & Lua Promos")
    assert u is None
    # Edge: nome vazio
    u = tcg_direct_url("", "Temporal Forces")
    assert u is None
    print(f"  tcg_direct_url: 8 casos OK ✓")
    return True


def test_threshold_constant():
    """Threshold 10x é o que documentação especifica."""
    assert TCG_SUSPECT_RATIO_THRESHOLD == 10.0, f"Threshold mudou: {TCG_SUSPECT_RATIO_THRESHOLD}"
    return True


def test_jirachi_ratio_math():
    """Sanity: ratio 1499/19.99 ≈ 75 deve disparar threshold 10."""
    ratio = 1499.00 / 19.99
    # v5.14.4: assert espelha o operador da produção (`>=`, boundary inclusivo).
    assert ratio >= TCG_SUSPECT_RATIO_THRESHOLD, f"Caso Jirachi não dispara: {ratio:.1f}x"
    print(f"  Jirachi ratio = {ratio:.1f}x ✓")
    return True


def test_tcg_suspect_boundary_exactly_10x():
    """v5.14.4: boundary INCLUSIVO via scrape_product real. Um ratio EXATAMENTE
    10x (declarado `.estat-tcg` R$1000 / última venda R$100) DEVE disparar
    tcg_suspect — com `>` ele escapava e virava deal "limpo" com margem possível-
    mente falsa (FP, erro caro). SEM preço real (`_fetch_ptcg_usd → None`), pra o
    clear-on-real não desfazer o suspect."""
    from myp_arbitrage_scanner import MYPScraper

    html = (
        '<html><body><h1>Gengar (100/191)</h1>'
        '<span class="estat-tcg">TCG Player: R$ 1.000,00</span>'
        '<span class="estatistica-ultimo">Última venda: R$ 100,00</span>'
        '<table class="table-striped table-bordered"><tbody>'
        + _seller_row("Inglês", "NM - Quase nova", "120,00")
        + _seller_row("Inglês", "NM - Quase nova", "130,00")
        + '</tbody></table></body></html>'
    )
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc._fetch_ptcg_usd = lambda cid: None     # sem preço real → suspect sobrevive
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/gengar",
                             "Surging Sparks")
    assert card is not None
    ratio = card.tcg_player_price / card.myp_last_sale_brl
    assert abs(ratio - 10.0) < 1e-9, f"ratio deveria ser 10.0x, é {ratio}"
    assert card.tcg_suspect is True, \
        "BUG: ratio EXATAMENTE 10x escapou do filtro suspect (boundary não-inclusivo)"
    assert sc._stats["tcg_suspects"] == 1, sc._stats["tcg_suspects"]
    print("  tcg_suspect boundary: ratio exatamente 10x → suspect (inclusivo `>=`) ✓")
    return True


# ── v5.9 (2026-06-03): marketplace pagination fix (truncation root cause) ──
def _seller_row(lang_title, qual, price_brl, foil="Normal"):
    """Uma <tr> de seller no formato MYP (flag-icon + células dedicadas)."""
    return (
        f'<tr>'
        f'<td><span class="flag-icon" title="{lang_title}"></span></td>'
        f'<td class="estoque-lista-qualidadenome">{qual}</td>'
        f'<td class="estoque-lista-nomeenfoil">{foil}</td>'
        f'<td>R$ {price_brl}</td>'
        f'</tr>'
    )


def _marketplace_container(rows_html, pagination_html=""):
    """Container #lista-anuncio-demais-vendedores (tabela marketplace + paginação)."""
    return (
        f'<div id="lista-anuncio-demais-vendedores">'
        f'<table class="table-striped table-bordered"><tbody>{rows_html}</tbody></table>'
        f'{pagination_html}'
        f'</div>'
    )


def _make_psyduck_page1():
    """Página 1: lojistas tem EN-NM mais caro (R$498); marketplace cheia de
    PT/JP baratos (0 EN visível) com paginação até a página 3 — réplica fiel
    do caso Psyduck 226/217 documentado no HANDOFF-TRUNCATION-2026-06-03 §2."""
    # Tabela lojistas: 2 EN-NM (498, 520) → lowest EN page-1 = 498 (inflado).
    lojistas = (
        '<table class="table-striped table-bordered"><tbody>'
        + _seller_row("Inglês", "NM - Quase nova", "498,00")
        + _seller_row("Inglês", "NM - Quase nova", "520,00")
        + '</tbody></table>'
    )
    # Marketplace page 1: 16 rows PT/JP NM, todas R$180–245 (< 498), 0 EN.
    mkt_rows = ""
    for i in range(16):
        price = 180 + i * 4  # 180..240
        lang = "Português" if i % 2 == 0 else "Japonês"
        mkt_rows += _seller_row(lang, "NM - Quase nova", f"{price},00")
    pagination = (
        '<ul class="pagination">'
        '<a href="?estoque-outros-page=2">2</a>'
        '<a href="?estoque-outros-page=3">3</a>'
        '</ul>'
    )
    marketplace = _marketplace_container(mkt_rows, pagination)
    return (
        '<html><body>'
        '<h1>Psyduck (053/198)</h1>'
        '<span class="estat-tcg">TCG Player: R$ 557,40</span>'
        f'{lojistas}{marketplace}'
        '</body></html>'
    )


def _make_psyduck_page2():
    """Página 2 marketplace: 3 EN-NM, a mais barata R$398 (o TRUE lowest)."""
    rows = (
        _seller_row("Inglês", "NM - Quase nova", "398,00")
        + _seller_row("Inglês", "NM - Quase nova", "410,00")
        + _seller_row("Português", "NM - Quase nova", "405,00")
    )
    return f'<html><body>{_marketplace_container(rows)}</body></html>'


def _make_psyduck_page3():
    """Página 3 marketplace: EN-NM mais caros (450, 650) — não muda o min."""
    rows = (
        _seller_row("Inglês", "NM - Quase nova", "450,00")
        + _seller_row("Inglês", "NM - Quase nova", "650,00")
    )
    return f'<html><body>{_marketplace_container(rows)}</body></html>'


def test_marketplace_pagination():
    """v5.9: scrape_product segue ?estoque-outros-page=N e acha o EN-NM oculto.

    Caso Psyduck: page 1 reportaria R$498,70 (tabela lojistas) porque a
    marketplace está cheia de PT/JP e os EN-NM caem na página 2. Com o fix,
    lowest EN-NM = R$398 (página 2), margem vs TCG R$557,40 = +40%, e
    truncation_risk é RESOLVIDO (False) porque a paginação foi seguida com
    sucesso. Sem rede: monkeypatch de _get serve fixtures por URL.
    """
    from myp_arbitrage_scanner import MYPScraper

    pages = {
        "https://mypcards.com/pokemon/produto/310463/psyduck": _make_psyduck_page1(),
        "https://mypcards.com/pokemon/produto/310463/psyduck?estoque-outros-page=2": _make_psyduck_page2(),
        "https://mypcards.com/pokemon/produto/310463/psyduck?estoque-outros-page=3": _make_psyduck_page3(),
    }
    fetched = []

    sc = MYPScraper(delay=0.0)

    def fake_get(url, save_debug=False):
        fetched.append(url)
        html = pages.get(url)
        return BeautifulSoup(html, "lxml") if html is not None else None

    sc._get = fake_get
    base = "https://mypcards.com/pokemon/produto/310463/psyduck"
    card = sc.scrape_product(base, "Scarlet & Violet")

    assert card is not None, "BUG: Psyduck retornou None (não deveria ser skipado)"
    # O CORAÇÃO DO FIX: lowest EN-NM = 398, não 498.
    assert card.myp_lowest_en_nm == 398.0, \
        f"BUG truncation: lowest EN-NM={card.myp_lowest_en_nm} (esperado 398.0 da pág 2)"
    # Margem vira deal real ≥25% (+40%).
    assert abs(card.margin_pct - 0.4005) < 0.01, \
        f"margem={card.margin_pct} (esperado ~0.40 / +40%)"
    # Paginação seguida com sucesso → risco resolvido.
    assert card.en_truncation_risk is False, \
        "truncation_risk deveria ser resolvido (paginação seguiu sem falha)"
    # Páginas 2 e 3 foram realmente buscadas.
    assert sc._stats["seller_pages_followed"] == 2, \
        f"esperado 2 páginas seguidas, got {sc._stats['seller_pages_followed']}"
    assert sc._stats["seller_page_fetch_failures"] == 0
    # EN sellers = lojistas(2) + pág2(2 EN) + pág3(2 EN) = 6.
    assert card.en_nm_sellers == 6, f"en_nm_sellers={card.en_nm_sellers} (esperado 6)"
    print(f"  Psyduck lowest EN-NM = R${card.myp_lowest_en_nm} "
          f"(margem +{card.margin_pct*100:.0f}%), {sc._stats['seller_pages_followed']} págs seguidas ✓")
    return True


def test_pagination_gate_skips_untruncated():
    """v5.9: produto SEM sinal de truncation NÃO pagina (controle de custo).

    Página 1 com EN-NM visível e barato → gate não dispara → nenhum fetch de
    ?estoque-outros-page mesmo que a paginação exista no HTML."""
    from myp_arbitrage_scanner import MYPScraper

    # Marketplace com EN-NM barato visível na página 1 (R$120) + paginação.
    mkt_rows = (
        _seller_row("Inglês", "NM - Quase nova", "120,00")
        + _seller_row("Português", "NM - Quase nova", "130,00")
    )
    pagination = '<ul class="pagination"><a href="?estoque-outros-page=2">2</a></ul>'
    html = (
        '<html><body><h1>Cheap Card (010/100)</h1>'
        '<span class="estat-tcg">TCG Player: R$ 300,00</span>'
        + _marketplace_container(mkt_rows, pagination)
        + '</body></html>'
    )
    fetched = []
    sc = MYPScraper(delay=0.0)

    def fake_get(url, save_debug=False):
        fetched.append(url)
        return BeautifulSoup(html, "lxml")

    sc._get = fake_get
    card = sc.scrape_product("https://mypcards.com/pokemon/produto/1/cheap", "x")
    assert card is not None
    assert card.myp_lowest_en_nm == 120.0
    # NENHUMA página extra buscada — só a página 1.
    assert sc._stats["seller_pages_followed"] == 0, \
        f"gate falhou: paginou produto não-truncado ({sc._stats['seller_pages_followed']} págs)"
    assert all("estoque-outros-page" not in u for u in fetched), \
        f"buscou página extra indevidamente: {fetched}"
    print(f"  Produto não-truncado: 0 páginas extras (truncation gate OK) ✓")
    return True


def test_pagination_cost_gate_low_tcg():
    """v5.9.1: produto TRUNCADO mas com TCG < min_price NÃO pagina (cost gate).

    Mesmo padrão de truncation do Psyduck (marketplace pág 1 cheia de PT/JP
    baratos, 0 EN, paginação disponível) MAS com TCG R$50 < min_price R$80.
    Como o card nunca pode virar deal, paginar pra resolver truncation é puro
    desperdício — o cost gate pula a paginação e conta pagination_skipped_low_tcg.
    Contraste com test_marketplace_pagination (TCG R$557 ⟹ ainda pagina)."""
    from myp_arbitrage_scanner import MYPScraper

    # Lojistas: 2 EN-NM (90, 95) ⟹ lowest EN visível pág-1 = 90 (≥ min_price).
    lojistas = (
        '<table class="table-striped table-bordered"><tbody>'
        + _seller_row("Inglês", "NM - Quase nova", "90,00")
        + _seller_row("Inglês", "NM - Quase nova", "95,00")
        + '</tbody></table>'
    )
    # Marketplace pág 1: 16 PT/JP baratos (40..70 < 90, 0 EN) ⟹ truncation gate
    # DISPARA. O cost gate é que deve barrar a paginação (TCG baixo).
    mkt_rows = ""
    for i in range(16):
        price = 40 + i * 2
        lang = "Português" if i % 2 == 0 else "Japonês"
        mkt_rows += _seller_row(lang, "NM - Quase nova", f"{price},00")
    pagination = '<ul class="pagination"><a href="?estoque-outros-page=2">2</a></ul>'
    html = (
        '<html><body><h1>Cheap Truncated (010/100)</h1>'
        '<span class="estat-tcg">TCG Player: R$ 50,00</span>'
        + lojistas
        + _marketplace_container(mkt_rows, pagination)
        + '</body></html>'
    )
    fetched = []
    # min_price pinado em 80 (não depende do default global, que virou R$50
    # no #20): card com TCG R$50 fica abaixo do piso ⟹ cost gate deve pular.
    sc = MYPScraper(delay=0.0, min_price=80.0)

    def fake_get(url, save_debug=False):
        fetched.append(url)
        return BeautifulSoup(html, "lxml")

    sc._get = fake_get
    card = sc.scrape_product("https://mypcards.com/pokemon/produto/2/cheaptrunc", "x")

    assert card is not None, "card não deveria ser None (EN visível R$90 ≥ min_price)"
    # O cost gate pulou a paginação APESAR do sinal de truncation.
    assert sc._stats["seller_pages_followed"] == 0, \
        f"cost gate falhou: paginou card TCG<min ({sc._stats['seller_pages_followed']} págs)"
    assert sc._stats["pagination_skipped_low_tcg"] == 1, \
        f"pagination_skipped_low_tcg={sc._stats['pagination_skipped_low_tcg']} (esperado 1)"
    assert all("estoque-outros-page" not in u for u in fetched), \
        f"buscou página extra indevidamente: {fetched}"
    print(f"  Truncado + TCG R$50 < R$80: 0 págs, gate contou 1 (cost gate v5.9.1) ✓")
    return True


def test_a3_real_price_rescues_pagination():
    """v5.11.5 (A3): card truncado com `.estat-tcg` DECLARADO baixo (< min_price)
    MAS preço REAL alto → a trava de custo consulta o real ANTES e PAGINA (não
    pula), achando o EN-NM barato da página 2. Sem o fix (decidindo pelo
    declarado), pularia e perderia o deal. Espelha test_pagination_cost_gate_low_tcg
    mas com o real resgatando."""
    from myp_arbitrage_scanner import MYPScraper

    base = "https://mypcards.com/pokemon/produto/9/blackbolt"
    # Lojistas: 2 EN-NM (90, 95) → lowest_en pág-1 = 90 (≥ min 80).
    lojistas = (
        '<table class="table-striped table-bordered"><tbody>'
        + _seller_row("Inglês", "NM - Quase nova", "90,00")
        + _seller_row("Inglês", "NM - Quase nova", "95,00")
        + '</tbody></table>'
    )
    # Marketplace pág-1: 16 PT/JP baratos (0 EN) < 90 → truncation gate dispara.
    mkt1 = "".join(
        _seller_row("Português" if i % 2 == 0 else "Japonês", "NM - Quase nova", f"{40 + i*2},00")
        for i in range(16)
    )
    pagination = '<ul class="pagination"><a href="?estoque-outros-page=2">2</a></ul>'
    page1 = (
        '<html><body><h1>Cheap Truncated (010/100)</h1>'
        '<span class="estat-tcg">TCG Player: R$ 50,00</span>'   # DECLARADO baixo (<80)
        + lojistas + _marketplace_container(mkt1, pagination) + '</body></html>'
    )
    # Página 2: um EN-NM barato (R$85) escondido — só achável paginando.
    page2 = (
        '<html><body>'
        + _marketplace_container(_seller_row("Inglês", "NM - Quase nova", "85,00"))
        + '</body></html>'
    )
    pages = {base: page1, f"{base}?estoque-outros-page=2": page2}

    sc = MYPScraper(delay=0.0, min_price=80.0)
    sc.fx_usd_brl = 5.0
    sc._fetch_ptcg_usd = lambda cid: 100.0   # real US$100 → R$500 (≥ min)
    sc._get = lambda url, save_debug=False: BeautifulSoup(pages[url], "lxml") if url in pages else None

    card = sc.scrape_product(base, "SV: Black Bolt")   # edição mapeada → setcode resolve

    assert card is not None
    # PAGINOU (não pulou) porque o real resgatou a decisão.
    assert sc._stats["pagination_skipped_low_tcg"] == 0, "A3: NÃO devia pular (real é alto)"
    assert sc._stats["seller_pages_followed"] == 1, \
        f"A3: devia paginar pág 2, got {sc._stats['seller_pages_followed']}"
    # Achou o EN-NM barato da página 2 (R$85 < 90 visível).
    assert card.myp_lowest_en_nm == 85.0, f"lowest_en={card.myp_lowest_en_nm} (esperado 85 da pág 2)"
    assert card.tcg_source == "pokemontcg.io"
    print("  A3: declarado R$50<80 mas real R$500 → paginou, achou EN R$85 (deal salvo) ✓")
    return True


def _real_price_page(card_h1, estat_tcg_brl, en_prices):
    """Página simples: h1 com (NNN/MMM), .estat-tcg declarado, N sellers EN-NM."""
    rows = "".join(_seller_row("Inglês", "NM - Quase nova", p) for p in en_prices)
    return (
        f'<html><body><h1>{card_h1}</h1>'
        f'<span class="estat-tcg">TCG Player: R$ {estat_tcg_brl}</span>'
        f'<table class="table-striped table-bordered"><tbody>{rows}</tbody></table>'
        f'</body></html>'
    )


def test_real_tcg_overrides_estat():
    """v5.11: preço real do pokemontcg.io SOBREPÕE o `.estat-tcg` inflado do MYP.

    Réplica do Darumaka 097/086 (Black Bolt): MYP declara R$2.867 (.estat-tcg
    mapeia a carta errada), mas o TCGplayer real é US$13,42. Com câmbio 5,0 o
    TCG vira R$67,10 → o 'deal' fake de +2289% morre (margem negativa)."""
    from myp_arbitrage_scanner import MYPScraper

    html = _real_price_page("Darumaka (097/086)", "2.867,75", ["120,00", "130,00"])
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0                       # câmbio fixo (sem rede)
    sc._fetch_ptcg_usd = lambda cid: 13.42    # mock pokemontcg.io (sem rede)
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/darumaka",
                             "SV: Black Bolt")
    assert card is not None
    assert card.tcg_source == "pokemontcg.io", f"source={card.tcg_source}"
    assert abs(card.tcg_real_usd - 13.42) < 0.001, card.tcg_real_usd
    assert abs(card.tcg_player_price - 67.10) < 0.01, card.tcg_player_price
    assert abs(card.myp_declared_tcg_brl - 2867.75) < 0.01, card.myp_declared_tcg_brl
    # O deal fake morreu: TCG real R$67,10 < EN-NM R$120 ⟹ margem negativa.
    assert card.margin_pct is not None and card.margin_pct < 0, card.margin_pct
    assert sc._stats["tcg_from_real"] == 1
    print("  Darumaka: .estat-tcg R$2867 → real US$13,42×5=R$67,10, deal fake morto ✓")
    return True


def test_fallback_to_estat_when_no_coverage():
    """v5.11: sem cobertura no pokemontcg.io → mantém o `.estat-tcg` do MYP."""
    from myp_arbitrage_scanner import MYPScraper

    html = _real_price_page("Mega Gengar ex (269/217)", "437,95", ["300,00", "310,00"])
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc._fetch_ptcg_usd = lambda cid: None     # me2pt5-269 sem preço (caso real)
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/gengar",
                             "Ascended Heroes")
    assert card is not None
    assert card.tcg_source == "myp_estat", f"source={card.tcg_source}"
    assert abs(card.tcg_player_price - 437.95) < 0.01, card.tcg_player_price
    assert card.tcg_real_usd is None
    assert sc._stats["tcg_from_myp_fallback"] == 1
    print("  Mega Gengar: sem cobertura pokemontcg.io → fallback .estat-tcg R$437,95 ✓")
    return True


def test_no_fx_keeps_estat():
    """v5.11: sem câmbio (fx None) → real-price desativado, usa `.estat-tcg`.
    Garante que o caminho v5.11 é INERTE quando scan() não rodou (testes
    offline / FX indisponível)."""
    from myp_arbitrage_scanner import MYPScraper

    called = []
    html = _real_price_page("Pikachu ex (179/086)", "200,00", ["100,00", "110,00"])
    sc = MYPScraper(delay=0.0, min_price=50.0)
    # fx_usd_brl fica None (default). _fetch_ptcg_usd NÃO deve ser chamado.
    sc._fetch_ptcg_usd = lambda cid: called.append(cid) or 9.99
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/pika",
                             "SV: Black Bolt")
    assert card is not None
    assert card.tcg_source == "myp_estat", f"source={card.tcg_source}"
    assert abs(card.tcg_player_price - 200.0) < 0.01, card.tcg_player_price
    assert called == [], f"pokemontcg.io chamado sem câmbio: {called}"
    print("  Sem câmbio: real-price inerte, usa .estat-tcg R$200 (sem chamar API) ✓")
    return True


def _no_estat_page(card_h1, en_prices):
    """Página SEM o `.estat-tcg` do MYP (card que o MYP não declara preço TCG)."""
    rows = "".join(_seller_row("Inglês", "NM - Quase nova", p) for p in en_prices)
    return (
        f'<html><body><h1>{card_h1}</h1>'
        f'<table class="table-striped table-bordered"><tbody>{rows}</tbody></table>'
        f'</body></html>'
    )


def test_real_price_clears_suspect():
    """v5.11.3 (A1, resgatado do PR #25): card com `.estat-tcg` inflado vira
    suspect, mas o preço REAL sobrepõe → a flag suspect é LIMPA (senão o card
    sumiria da sheet 🔥 Deals mesmo com margem real legítima)."""
    from myp_arbitrage_scanner import MYPScraper

    # Darumaka: declarado R$2.867 vs última venda R$32 → ratio 89 > 10 = suspect.
    html = (
        '<html><body><h1>Darumaka (097/086)</h1>'
        '<span class="estat-tcg">TCG Player: R$ 2.867,75</span>'
        '<span class="estatistica-ultimo">Última venda: R$ 32,00</span>'
        '<table class="table-striped table-bordered"><tbody>'
        + _seller_row("Inglês", "NM - Quase nova", "120,00")
        + _seller_row("Inglês", "NM - Quase nova", "130,00")
        + '</tbody></table></body></html>'
    )
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc._fetch_ptcg_usd = lambda cid: 13.42   # real US$13,42 → R$67,10
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/daru",
                             "SV: Black Bolt")
    assert card is not None
    assert card.tcg_source == "pokemontcg.io", card.tcg_source
    assert card.tcg_suspect is False, "A1: suspect deveria ter sido limpo"
    assert sc._stats["tcg_suspects"] == 0, sc._stats["tcg_suspects"]
    assert abs(card.tcg_player_price - 67.10) < 0.01, card.tcg_player_price
    print("  A1: .estat-tcg inflado → suspect setado e LIMPO após preço real ✓")
    return True


def test_prices_card_without_estat_tcg():
    """v5.11.3 (A2): card SEM `.estat-tcg` não é mais skipado prematuramente —
    o preço REAL do pokemontcg.io o precifica."""
    from myp_arbitrage_scanner import MYPScraper

    html = _no_estat_page("Foobar ex (170/086)", ["100,00", "110,00"])
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc._fetch_ptcg_usd = lambda cid: 40.0    # real US$40 → R$200
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/foo",
                             "SV: Black Bolt")
    assert card is not None, "A2: não deveria skipar — preço real disponível"
    assert card.myp_declared_tcg_brl is None, card.myp_declared_tcg_brl
    assert card.tcg_source == "pokemontcg.io", card.tcg_source
    assert abs(card.tcg_player_price - 200.0) < 0.01, card.tcg_player_price
    assert card.margin_pct is not None and abs(card.margin_pct - 1.0) < 0.01, card.margin_pct
    print("  A2: card sem .estat-tcg precificado via fonte real (R$200, +100%) ✓")
    return True


def test_skip_when_no_tcg_at_all():
    """v5.11.3 (A2): sem `.estat-tcg` E sem cobertura no pokemontcg.io → skip
    (skipped_no_tcg_price), pois não há preço TCG nenhum pra comparar."""
    from myp_arbitrage_scanner import MYPScraper

    html = _no_estat_page("Nada ex (171/086)", ["100,00", "110,00"])
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc._fetch_ptcg_usd = lambda cid: None    # sem cobertura
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")

    card = sc.scrape_product("https://mypcards.com/pokemon/produto/9/nada",
                             "SV: Black Bolt")
    assert card is None, "deveria skipar: sem TCG declarado nem real"
    assert sc._stats["skipped_no_tcg_price"] == 1, sc._stats["skipped_no_tcg_price"]
    print("  A2: sem .estat-tcg e sem cobertura real → skip correto ✓")
    return True


def test_parse_brl_formats():
    """v5.8.10: _parse_brl — BR canonical, US-decimal leak, milhares, edge cases.

    Regressão do bug v5.8.2 ('30.00' lido como 3000.0). É a função mais
    propensa a erro do parser e não tinha cobertura direta.
    """
    f = MYPScraper._parse_brl
    cases = [
        ("R$ 1.900,00", 1900.0),      # BR canonical
        ("R$ 30,00", 30.0),           # BR só vírgula
        ("R$ 30.00", 30.0),           # US-decimal leak (bug v5.8.2)
        ("R$ 1,500.00", 1500.0),      # US milhares + decimal
        ("R$ 30.000", 30000.0),       # BR milhares (sufixo 3 dígitos)
        ("R$ 1.500.000", 1500000.0),  # BR milhares (multi-ponto)
        ("1234.56", 1234.56),         # US decimal sem prefixo
        ("R$ 0,00", None),            # zero → None (guard val > 0)
        ("", None),
        ("   ", None),
        (None, None),                 # guard v5.8.4 (não-str)
        ("sem preço", None),
    ]
    for text, expected in cases:
        got = f(text)
        assert got == expected, f"_parse_brl({text!r}) = {got}, esperado {expected}"
    print(f"  _parse_brl: {len(cases)} casos OK ✓")
    return True


def test_last_brl():
    """v5.8.10: _last_brl extrai o ÚLTIMO R$ (caso multi-preço '.estat-tcg')."""
    f = MYPScraper._last_brl
    assert f("Last R$ 26,00 | Avg R$ 30,00") == 30.0, "deve pegar o último valor"
    assert f("R$ 99,90") == 99.9
    assert f("sem preço aqui") is None
    assert f("") is None
    assert f(None) is None
    print(f"  _last_brl: extração do último valor OK ✓")
    return True


def test_oversized_regex():
    """v5.8.10: filtros Jumbo/oversized — title (2ª camada) + foil (1ª camada)."""
    assert OVERSIZED_TITLE_RE.search("Pikachu Jumbo (SWSH039)")
    assert OVERSIZED_TITLE_RE.search("Charizard Oversized Promo")
    assert OVERSIZED_TITLE_RE.search("Mewtwo Box Topper")
    assert not OVERSIZED_TITLE_RE.search("Charizard ex 234/091"), "standard não deve casar"
    assert OVERSIZED_FOIL_RE.search("Jumbo")
    assert not OVERSIZED_FOIL_RE.search("Holo")
    print(f"  oversized/jumbo regex OK ✓")
    return True


def test_delivery_table_format():
    """v5.11.1: tabela de ENTREGA do myp_summary.py no formato aprovado pelo
    operador (links clicáveis MYP + TCG, igual scanner COMC).

    Exercita: helpers de coluna (carta_label/delivery_links/fmt_usd) + round-trip
    XLSX → markdown (myp_summary.main) garantindo header + links no output.
    """
    import myp_summary as S

    # ── Helpers de coluna ──
    # Carta: nome + número numa coluna só, sem duplicar.
    assert S.carta_label("Pikachu (173/165)") == "Pikachu 173/165", \
        f"carta_label não juntou nome+número: {S.carta_label('Pikachu (173/165)')!r}"
    assert S.carta_label("Charizard ex") == "Charizard ex", \
        "carta_label não deve alterar nome sem número"
    base, num = S.split_card_name("Iron Hands ex (070/162)")
    assert base == "Iron Hands ex" and num == "070/162", f"split errado: {base!r}/{num!r}"

    # fmt_usd
    assert S.fmt_usd(13.42) == "US$13.42", f"fmt_usd errado: {S.fmt_usd(13.42)!r}"
    assert S.fmt_usd(None) == "—"

    # Links: oferta (MYP) + TCG direto (Temporal Forces → sv5-70).
    links = S.delivery_links(
        "https://mypcards.com/pokemon/produto/123/iron-hands",
        "Iron Hands ex (070/162)", "Temporal Forces", oversized=False,
    )
    assert "[oferta](https://mypcards.com/pokemon/produto/123/iron-hands)" in links, \
        f"link de oferta MYP ausente: {links!r}"
    assert "[TCG](https://prices.pokemontcg.io/tcgplayer/sv5-70)" in links, \
        f"link TCG direto ausente: {links!r}"
    assert " · " in links, f"separador de links ausente: {links!r}"

    # Sem URL MYP mas com nome → só o link TCG (busca por nome).
    only_tcg = S.delivery_links(None, "X", "edição inexistente")
    assert only_tcg.startswith("[TCG]") and "[oferta]" not in only_tcg, \
        f"sem MYP url deveria sobrar só TCG: {only_tcg!r}"
    # Sem nada → '—'.
    assert S.delivery_links(None, "", "") == "—"

    # ── Round-trip XLSX → markdown ──
    clean = make_clean_deal()
    clean.name = "Charizard ex (125/191)"   # in-range (não supranumerário)
    clean.tcg_real_usd = 36.50      # v5.11.1: USD real exposto no XLSX
    clean.tcg_source = "pokemontcg.io"  # v5.14.3: consistência — quem tem USD real
                                        # tem source real (no scanner os dois andam
                                        # juntos, L1398); senão o deal cai no balde
                                        # fallback (preço não-confiável), não no limpo.
    cards = [clean, make_borderline_deal()]

    import tempfile as _tf, os as _os
    with _tf.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_out = f.name
    generate_xlsx(cards, xlsx_out, threshold=0.25)

    # Confere que a coluna "TCG US$" entrou no XLSX e foi populada.
    wb = load_workbook(xlsx_out)
    ws = wb["All EN Cards"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "TCG US$" in hdr, f"coluna TCG US$ ausente do XLSX: {hdr}"
    usd_col = hdr.index("TCG US$") + 1
    name_col = hdr.index("Card Name") + 1
    usd_seen = None
    for r in range(2, ws.max_row + 1):
        if "Charizard" in str(ws.cell(row=r, column=name_col).value or ""):
            usd_seen = ws.cell(row=r, column=usd_col).value
    wb.close()  # Windows: solta o handle antes do unlink
    assert usd_seen == 36.50, f"TCG US$ não persistiu no XLSX: {usd_seen!r}"

    md_out = xlsx_out.replace(".xlsx", ".md")
    rc = S.build_markdown(xlsx_out, md_out, scan_type="daily", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    md = Path(md_out).read_text(encoding="utf-8")

    # Header no formato aprovado.
    assert "| # | Margem % | MYP R$ | TCG US$ | Dif | Carta | Set | Raridade | Cond | Qtd | Links |" in md, \
        "header da tabela de entrega não bate o formato aprovado"
    # Carta composta + link de oferta clicável + Cond NM + USD.
    assert "Charizard ex 125/191" in md, "Carta composta ausente do markdown"
    assert "[oferta](https://mypcards.com/pokemon/surging-sparks/charizard-ex)" in md, \
        "link de oferta MYP ausente do markdown"
    assert "US$36.50" in md, "TCG US$ ausente do markdown"
    assert "| NM |" in md, "coluna Cond=NM ausente do markdown"

    _os.unlink(xlsx_out)
    _os.unlink(md_out)
    print("  delivery table format (header + Carta + links + USD + Cond) OK ✓")
    return True


def test_checkpoint_save_load():
    """v5.11.4: _save_checkpoint → _load_checkpoint preserva cards, stats e o
    set de edições feitas (round-trip)."""
    import os as _os, tempfile
    from myp_arbitrage_scanner import MYPScraper, CardData

    sc = MYPScraper(delay=0.0)
    c = CardData(); c.name = "Pikachu (058/102)"; c.myp_lowest_en_nm = 100.0
    c.tcg_player_price = 180.0; c.margin_pct = 0.8; c.tcg_source = "pokemontcg.io"
    sc.cards = [c]
    sc._stats["en_found"] = 1
    fd, path = tempfile.mkstemp(suffix=".resume.json"); _os.close(fd)
    try:
        sc._save_checkpoint(path, {"u1", "u2"})
        sc2 = MYPScraper(delay=0.0)
        done = sc2._load_checkpoint(path)
        assert done == {"u1", "u2"}, done
        assert len(sc2.cards) == 1 and sc2.cards[0].name == "Pikachu (058/102)", sc2.cards
        assert abs(sc2.cards[0].tcg_player_price - 180.0) < 0.01
        assert sc2.cards[0].tcg_source == "pokemontcg.io"
        assert sc2._stats["en_found"] == 1, sc2._stats["en_found"]
    finally:
        _os.path.exists(path) and _os.unlink(path)
    print("  checkpoint save/load round-trip (cards + stats + done) ✓")
    return True


def test_scan_resume_skips_done_editions():
    """v5.11.4: scan(resume=True) pula edições já no checkpoint, escaneia só as
    que faltam, e remove o checkpoint ao concluir."""
    import os as _os, tempfile, json
    import myp_arbitrage_scanner as M
    from myp_arbitrage_scanner import MYPScraper, CardData, CHECKPOINT_VERSION

    # checkpoint pré-existente: edição u1 já feita, 1 card restaurado.
    pre = CardData(); pre.name = "Done (1/100)"; pre.myp_lowest_en_nm = 100.0
    pre.tcg_player_price = 200.0; pre.margin_pct = 1.0
    fd, ckpt = tempfile.mkstemp(suffix=".resume.json"); _os.close(fd)
    with open(ckpt, "w", encoding="utf-8") as f:
        json.dump({"version": CHECKPOINT_VERSION,
                   "cards": [{"name": "Done (1/100)", "myp_lowest_en_nm": 100.0,
                              "tcg_player_price": 200.0, "margin_pct": 1.0}],
                   "done_editions": ["u1"], "stats": {}}, f)

    sc = MYPScraper(delay=0.0)
    # evita rede: câmbio mockado + catálogo/produtos/scrape mockados.
    M.fetch_usd_brl = lambda session: 5.0
    sc.get_all_editions = lambda: [{"title": "E1", "url": "u1"},
                                   {"title": "E2", "url": "u2"}]
    sc.get_edition_products = lambda url: [f"{url}/p1"]
    scraped = []

    def fake_scrape(purl, title):
        scraped.append(purl)
        c = CardData(); c.name = f"{title} (2/100)"; c.myp_lowest_en_nm = 100.0
        c.tcg_player_price = 150.0; c.margin_pct = 0.5
        return c
    sc.scrape_product = fake_scrape

    cards = sc.scan(resume=True, checkpoint_path=ckpt)

    assert scraped == ["u2/p1"], f"deveria escanear só u2 (E1 já feita): {scraped}"
    names = sorted(c.name for c in cards)
    assert names == ["Done (1/100)", "E2 (2/100)"], names
    assert not _os.path.exists(ckpt), "checkpoint deveria ser removido ao concluir"
    print("  resume: pula edição feita (u1), escaneia só u2, limpa checkpoint ✓")
    return True


def test_prefill_ptcg_set_batch():
    """v5.12: _prefill_ptcg_set popula o cache do set inteiro num request, e
    _real_tcg_brl passa a usar o cache SEM chamar _fetch_ptcg_usd (round-trip
    por-card eliminado). Cobre min(market|mid), strip de zero à esquerda na
    chave e skip de número não-numérico (TG/GG)."""
    from myp_arbitrage_scanner import MYPScraper

    batch = {"data": [
        {"id": "sv5-70", "number": "070",
         "tcgplayer": {"prices": {"holofoil": {"market": 13.42, "mid": 20.0}}}},
        {"id": "sv5-1", "number": "1",
         "tcgplayer": {"prices": {"normal": {"mid": 5.0}}}},          # só mid
        {"id": "sv5-GG01", "number": "GG01",
         "tcgplayer": {"prices": {"holofoil": {"market": 99.0}}}},    # não-numérico → skip
    ], "page": 1, "pageSize": 250, "count": 3, "totalCount": 3}
    calls = {"batch": 0, "single": 0}

    class _Resp:
        status_code = 200
        def raise_for_status(self):
            pass
        def json(self):
            return batch

    class _Sess:
        def get(self, url, **kw):
            calls["batch" if "q=set.id" in url else "single"] += 1
            return _Resp()

    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc.session = _Sess()
    sc._prefill_ptcg_set("sv5")

    # min(market|mid): market 13.42 vence mid 20; zero à esquerda normalizado (070→70)
    assert sc._ptcg_cache.get("sv5-70") == 13.42, sc._ptcg_cache
    assert sc._ptcg_cache.get("sv5-1") == 5.0, sc._ptcg_cache       # só mid disponível
    assert "sv5-GG01" not in sc._ptcg_cache, "número não-numérico não deve entrar no cache"
    assert calls["batch"] == 1, f"esperava 1 request batch, got {calls['batch']}"

    # idempotente: 2ª chamada do mesmo setcode não refaz request
    sc._prefill_ptcg_set("sv5")
    assert calls["batch"] == 1, "prefill repetido do mesmo set não deveria refazer request"

    # _real_tcg_brl agora resolve pelo cache, SEM round-trip por-card
    def _boom(cid):
        raise AssertionError(f"não devia chamar _fetch_ptcg_usd: {cid}")
    sc._fetch_ptcg_usd = _boom
    brl = sc._real_tcg_brl("Iron Hands ex (070/162)", "Temporal Forces")  # Temporal Forces → sv5
    assert abs(brl - 13.42 * 5.0) < 1e-6, brl
    assert calls["single"] == 0, "não devia haver fetch por-card após prefill"
    print("  v5.12 prefill: batch popula cache + _real_tcg_brl usa cache (0 round-trips por-card) ✓")
    return True


# ══════════════════════════════════════════════════════════════════════
# v5.15: fonte de preço tcgcsv.com (funciona no CI)
# ══════════════════════════════════════════════════════════════════════
def _tcgcsv_session(groups=None, products=None, prices=None, log_calls=None):
    """Sessão fake que responde aos 3 endpoints tcgcsv com fixtures sintéticas.

    Schema fiel ao dump REAL (confirmado 2026-06-21): /groups → results c/
    {groupId,name,abbreviation}; /products → results c/ extendedData[Number];
    /prices → results c/ {productId, marketPrice, midPrice, subTypeName}."""
    bodies = {"groups": groups, "products": products, "prices": prices}

    class _Resp:
        def __init__(self, body, status=200):
            self._body = body
            self.status_code = status
        def json(self):
            return self._body

    class _Sess:
        def get(self, url, **kw):
            if log_calls is not None:
                log_calls.append(url)
            if url.endswith("/groups"):
                return _Resp(bodies["groups"]) if bodies["groups"] is not None else _Resp(None, 500)
            if url.endswith("/products"):
                return _Resp(bodies["products"]) if bodies["products"] is not None else _Resp(None, 500)
            if url.endswith("/prices"):
                return _Resp(bodies["prices"]) if bodies["prices"] is not None else _Resp(None, 500)
            return _Resp(None, 404)
    return _Sess()


def _tcgcsv_fixtures():
    """Fixture mínima de 1 set (Stellar Crown → sv7, abbr SCR, groupId 23537)."""
    groups = {"results": [
        {"groupId": 23537, "name": "SV07: Stellar Crown", "abbreviation": "SCR"},
        {"groupId": 99999, "name": "Outro Set Qualquer", "abbreviation": "ZZZ"},
    ]}
    products = {"results": [
        {"productId": 1001, "name": "Venusaur ex",
         "extendedData": [{"name": "Number", "value": "001/142"},
                          {"name": "Rarity", "value": "Double Rare"}]},
        {"productId": 1002, "name": "Ledyba",
         "extendedData": [{"name": "Number", "value": "070/142"}]},
        # número não-numérico (promo/GG) → não casa o cid {setcode}-{num}
        {"productId": 1003, "name": "Promo Card",
         "extendedData": [{"name": "Number", "value": "GG01/GG10"}]},
    ]}
    prices = {"results": [
        # productId 1001: 2 subtypes → _min_tcg_usd pega o menor market (12.0)
        {"productId": 1001, "lowPrice": 10.0, "midPrice": 20.0,
         "highPrice": 99.0, "marketPrice": 18.0, "subTypeName": "Normal"},
        {"productId": 1001, "lowPrice": 8.0, "midPrice": 15.0,
         "highPrice": 50.0, "marketPrice": 12.0, "subTypeName": "Holofoil"},
        # productId 1002: só mid (market None) → cai no mid 5.0
        {"productId": 1002, "lowPrice": 4.0, "midPrice": 5.0,
         "highPrice": 9.0, "marketPrice": None, "subTypeName": "Normal"},
        {"productId": 1003, "lowPrice": 1.0, "midPrice": 2.0,
         "highPrice": 3.0, "marketPrice": 1.5, "subTypeName": "Normal"},
    ]}
    return groups, products, prices


def test_tcgcsv_prefill_parses_schema():
    """v5.15: _prefill_tcgcsv_set parseia o schema REAL do tcgcsv (groups+products
    +prices), junta por productId, aplica _min_tcg_usd (menor market/mid entre
    subtypes — IDÊNTICO ao pokemontcg.io) e popula o MESMO _ptcg_cache."""
    from myp_arbitrage_scanner import MYPScraper

    groups, products, prices = _tcgcsv_fixtures()
    calls = []
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc.session = _tcgcsv_session(groups, products, prices, log_calls=calls)

    ok = sc._prefill_tcgcsv_set("sv7", "SV07: Stellar Crown")
    assert ok is True, "prefill devia ter preenchido ≥1 preço"
    # 1001: min(market 18, market 12) = 12.0 (menor subtype vence); cid sv7-1
    assert sc._ptcg_cache.get("sv7-1") == 12.0, sc._ptcg_cache
    # 1002: market None → mid 5.0; zero à esquerda normalizado (070→70)
    assert sc._ptcg_cache.get("sv7-70") == 5.0, sc._ptcg_cache
    # 1003: número não-numérico (GG01) → fora do cache
    assert not any(k.startswith("sv7-GG") for k in sc._ptcg_cache), sc._ptcg_cache
    # ambos os cids são marcados como provenientes do tcgcsv
    assert "sv7-1" in sc._tcgcsv_cids and "sv7-70" in sc._tcgcsv_cids
    # 3 requests: groups + products + prices (1 set)
    assert sum("/groups" in u for u in calls) == 1
    assert sum("/products" in u for u in calls) == 1
    assert sum("/prices" in u for u in calls) == 1
    assert sc._stats["tcgcsv_prefill_sets"] == 1
    print("  v5.15 tcgcsv prefill: parse schema + min(subtype) + cache compartilhado ✓")
    return True


def test_tcgcsv_groupid_resolution():
    """v5.15: resolve_tcgcsv_group_id casa por abreviação (primário) e por nome
    (fallback). Set sem correspondência → None (→ fallback honesto, sem chute)."""
    from myp_arbitrage_scanner import resolve_tcgcsv_group_id
    groups = [
        {"groupId": 23537, "name": "SV07: Stellar Crown", "abbreviation": "SCR"},
        {"groupId": 24541, "name": "ME: Ascended Heroes", "abbreviation": "ASC"},
    ]
    # por abreviação (mapa conhecido sv7→SCR)
    assert resolve_tcgcsv_group_id("sv7", "SV07: Stellar Crown", groups) == 23537
    assert resolve_tcgcsv_group_id("me2pt5", "ME: Ascended Heroes", groups) == 24541
    # setcode fora do mapa de abreviação → fallback por nome (substring MYP)
    groups2 = [{"groupId": 555, "name": "SV07: Stellar Crown", "abbreviation": "XXX"}]
    assert resolve_tcgcsv_group_id("sv7", "Stellar Crown", groups2) == 555
    # sem correspondência nenhuma → None (NUNCA chuta um groupId)
    assert resolve_tcgcsv_group_id("sv99", "Set Inexistente", groups) is None
    # AMBÍGUO: a substring casa >1 group e nenhuma abbr exata → None (não chuta o
    # primeiro). Guard anti "preço de promo rotulado real" (review v5.15).
    groups_ambig = [
        {"groupId": 901, "name": "SV07: Stellar Crown", "abbreviation": "XXX"},
        {"groupId": 902, "name": "Stellar Crown Promo", "abbreviation": "YYY"},
    ]
    assert resolve_tcgcsv_group_id("sv7", "Stellar Crown", groups_ambig) is None
    print("  v5.15 groupId: abbr primário + nome fallback ÚNICO + None honesto ✓")
    return True


def test_setcode_abbr_table_is_self_consistent():
    """v5.16: GUARD anti-bug-set-errado (estrutural, sem rede).

    Toda abreviação em PTCG_SETCODE_TO_TCGCSV_ABBR tem que ser ÚNICA dentro da
    própria tabela (duas setcodes distintas apontando pra mesma abbr injetaria
    o preço do mesmo group em sets diferentes — exceto subset legítimo, que
    não temos aqui). Também checa: nenhuma chave/valor vazio. Não depende de
    rede; é a primeira linha de defesa do mapa."""
    from myp_arbitrage_scanner import PTCG_SETCODE_TO_TCGCSV_ABBR as T
    from collections import Counter
    assert all(k and v for k, v in T.items()), "chave/valor vazio na tabela"
    c = Counter(v.upper() for v in T.values())
    dups = {a: n for a, n in c.items() if n > 1}
    assert not dups, f"abbr repetida na tabela (set-wrong risk): {dups}"
    print(f"  v5.16 tabela setcode→abbr: {len(T)} entradas, abbr únicas ✓")
    return True


def test_setcode_abbr_resolves_1to1_against_groups_fixture():
    """v5.16: cada abbr da tabela resolve para EXATAMENTE 1 group num snapshot
    REAL e COMPLETO do /groups do tcgcsv (fixture de metadados groupId/name/
    abbreviation, sem preços). Pega: (a) abbr que sumiu/renomeou no tcgcsv
    (0 matches), (b) abbr que colide com outro set (>1 match) → BLOCKER de
    preço de set errado.

    O fixture (`test_tcgcsv_groups_fixture.json`) é o dump INTEIRO do /groups ao
    vivo (217 groups, 2026-06-22) — TODOS os groups da categoria Pokémon, não só
    os usados pela tabela. Isso é o que torna o guard NÃO-circular: o dump
    contém abbreviations colisoras reais (ex.: `RR` = Rising Rivals + Team Rocket
    Returns = 2 groups; `CL`, `BKP`, `BLW`, `GEN`, `LTR` = 2 cada; `PR`/`POP` =
    promos com dezenas). Se alguém adicionar à tabela uma abbr ambígua dessas,
    `by_abbr[abbr] > 1` e o teste FALHA. As 106 entradas atuais são todas únicas
    no dump real (confirmado ao vivo)."""
    import json as _json
    from pathlib import Path
    from collections import Counter
    from myp_arbitrage_scanner import (
        PTCG_SETCODE_TO_TCGCSV_ABBR as T, resolve_tcgcsv_group_id,
        MYP_EDITION_SUBSTR_TO_PTCG as S,
    )
    fx_path = Path(__file__).parent / "test_tcgcsv_groups_fixture.json"
    groups = _json.loads(fx_path.read_text(encoding="utf-8"))
    # sanity: é o snapshot COMPLETO (não um subset montado p/ a tabela), e é
    # só metadados (sem preços) — senão o guard volta a ser circular.
    assert len(groups) >= 200, (
        f"fixture deve ser o dump COMPLETO do /groups (217), tem {len(groups)} "
        "— um subset reintroduz a circularidade que este teste corrige")
    assert all(set(g.keys()) == {"groupId", "name", "abbreviation"} for g in groups), \
        "fixture deve conter SÓ metadados (groupId/name/abbreviation), sem preços"

    by_abbr = Counter(str(g.get("abbreviation") or "").upper() for g in groups)
    # o dump real PRECISA conter abbrs colisoras, senão o guard seria inócuo
    assert any(n > 1 for n in by_abbr.values()), (
        "dump real deveria ter abbreviations colisoras (ex. RR/CL/PR); "
        "sua ausência indica fixture truncado/sanitizado demais")

    # 1) cada abbr da tabela aparece exatamente 1× no dump COMPLETO (1-a-1):
    #    0 = abbr sumiu/renomeou no tcgcsv; >1 = colide com outro set (set-wrong).
    bad = {}
    for setcode, abbr in T.items():
        n = by_abbr.get(abbr.upper(), 0)
        if n != 1:
            bad[setcode] = (abbr, n)
    assert not bad, f"abbr não-única no /groups (set-wrong/missing): {bad}"

    # 2) resolve_tcgcsv_group_id retorna um groupId p/ cada setcode da tabela,
    #    via o caminho de ABREVIAÇÃO (caminho primário), usando uma edição que
    #    de fato mapeia pra esse setcode (longest-substring), garantindo que o
    #    par (substring MYP → setcode → abbr → group) é coerente ponta-a-ponta.
    substr_for = {}
    for substr, sc in S.items():
        substr_for.setdefault(sc, substr)
    for setcode, abbr in T.items():
        ed = substr_for.get(setcode, "")
        gid = resolve_tcgcsv_group_id(setcode, ed, groups)
        assert gid is not None, f"setcode {setcode} ({abbr}) não resolveu groupId"
    print(f"  v5.16 {len(T)} abbrs resolvem 1-a-1 no snapshot /groups real ✓")
    return True


def test_tcgcsv_no_match_falls_back_honestly():
    """v5.15: set sem groupId no tcgcsv → _prefill_tcgcsv_set retorna False e NÃO
    popula cache (NUNCA preço inventado). Honestidade dura."""
    from myp_arbitrage_scanner import MYPScraper
    groups = {"results": [{"groupId": 1, "name": "Algum Set", "abbreviation": "ZZZ"}]}
    sc = MYPScraper(delay=0.0, min_price=50.0)
    sc.fx_usd_brl = 5.0
    sc.session = _tcgcsv_session(groups, {"results": []}, {"results": []})
    # sv7 não casa nenhum group da fixture → sem groupId
    ok = sc._prefill_tcgcsv_set("sv7", "SV07: Stellar Crown")
    assert ok is False
    assert sc._ptcg_cache == {}, "cache devia ficar vazio (sem preço inventado)"
    assert sc._tcgcsv_cids == set()
    print("  v5.15 tcgcsv sem match → fallback honesto (cache vazio) ✓")
    return True


def test_tcgcsv_end_to_end_real_source_label():
    """v5.15: ponta-a-ponta — prefill tcgcsv + scrape_product → card sai com
    tcg_source='tcgcsv' (REAL), preço real sobrepõe o `.estat-tcg`, margem real."""
    from myp_arbitrage_scanner import MYPScraper

    groups, products, prices = _tcgcsv_fixtures()
    sc = MYPScraper(delay=0.0, min_price=50.0, tcg_source="tcgcsv")
    sc.fx_usd_brl = 5.0
    sc.session = _tcgcsv_session(groups, products, prices)
    # pré-carrega o set (como faz o loop de edições)
    sc._prefill_tcgcsv_set("sv7", "SV07: Stellar Crown")
    # NÃO deve haver round-trip pokemontcg.io em modo tcgcsv
    sc._fetch_ptcg_usd = lambda cid: (_ for _ in ()).throw(
        AssertionError("modo tcgcsv não deve chamar pokemontcg.io"))

    # Ledyba 070/142: tcgcsv real só-mid US$5 → R$25? não — usamos Venusaur 001
    # (US$12 → R$60). .estat-tcg declara R$300 (fake) → o real sobrepõe.
    # EN-NM lowest = 50 (≥ min_price 50, passa o filtro). Margem real = (60-50)/50
    # = 0.20.
    html = _real_price_page("Venusaur ex (001/142)", "300,00", ["50,00", "55,00"])
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")
    card = sc.scrape_product("https://mypcards.com/pokemon/produto/1/venusaur",
                             "SV07: Stellar Crown")
    assert card is not None
    assert card.tcg_source == "tcgcsv", f"source={card.tcg_source}"
    assert abs(card.tcg_real_usd - 12.0) < 1e-6, card.tcg_real_usd
    assert abs(card.tcg_player_price - 60.0) < 1e-6, card.tcg_player_price
    # o preço real (R$60) sobrepôs o `.estat-tcg` declarado (R$300, fake)
    assert abs(card.myp_declared_tcg_brl - 300.0) < 1e-6, card.myp_declared_tcg_brl
    assert abs(card.margin_pct - 0.20) < 1e-6, card.margin_pct
    assert sc._stats["tcg_from_tcgcsv"] == 1, sc._stats["tcg_from_tcgcsv"]
    assert sc._stats["tcg_from_real"] == 1
    print("  v5.15 e2e: card sai tcg_source='tcgcsv' (REAL), preço real na margem ✓")
    return True


def _tcgcsv_deal(name, myp, tcg_brl, usd, rarity="Double Rare"):
    """v5.15: deal com preço REAL via tcgcsv (tcg_source='tcgcsv')."""
    return CardData(
        name=name, edition="Stellar Crown", rarity=rarity,
        product_url=f"https://myp/{name}", myp_lowest_en_nm=myp,
        tcg_player_price=tcg_brl, tcg_real_usd=usd, tcg_source="tcgcsv",
        myp_last_sale_brl=myp, margin_pct=(tcg_brl - myp) / myp,
        margin_brl=tcg_brl - myp, en_nm_sellers=3, last_updated="2026-06-20",
    )


def test_tcgcsv_recognized_as_real_in_summary():
    """v5.15: o gate de honestidade do myp_summary (_is_real) reconhece
    'real (tcgcsv)' como preço REAL — senão deals reais do CI cairiam no balde
    'validar manualmente'. Regra dura: tcgcsv é fonte real verificável."""
    from myp_summary import build_markdown

    deal = _tcgcsv_deal("Venusaur ex (001/142)", 50.0, 80.0, 16.0)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([deal], xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="daily", run_id="", repo="x/y")
    assert rc == 0
    text = Path(md).read_text(encoding="utf-8")
    sec = _section_of(text, "Venusaur ex")
    assert sec is not None and "limpos" in sec.lower(), \
        f"deal tcgcsv (REAL) devia estar no balde limpo, está em: {sec!r}\n{text[:900]}"
    # NÃO pode aparecer no balde fallback
    assert "FALLBACK `.estat-tcg`" not in text or "Venusaur ex" not in \
        text.split("FALLBACK `.estat-tcg`")[-1], \
        "deal tcgcsv (REAL) não pode estar no balde fallback"
    # cobertura conta 1/1 real
    assert "1/1 cartas EN com preço REAL" in text, \
        f"cobertura devia contar 1/1 real:\n{text[:600]}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  v5.15 summary: 'real (tcgcsv)' reconhecido como REAL (balde limpo) ✓")
    return True


def test_fallback_attribution():
    """v5.13 (Iteração #2): _attribute_fallback classifica POR QUE o card caiu no
    fallback `.estat-tcg`, na MESMA cascata de _real_tcg_brl. Cada motivo cai no
    seu balde — base de medição pro ataque a falso-positivo (cobertura)."""
    from myp_arbitrage_scanner import MYPScraper

    sc = MYPScraper(delay=0.0, min_price=50.0)

    # 1) sem câmbio → no_fx (condição global precede checagem de cobertura)
    sc.fx_usd_brl = None
    sc._attribute_fallback("Iron Hands ex (070/162)", "Temporal Forces")
    assert sc._stats["fallback_no_fx"] == 1, sc._stats

    sc.fx_usd_brl = 5.0  # câmbio presente daqui pra frente

    # 2) edição fora do mapa → unmapped_set (mesmo com nº colecionador válido)
    sc._attribute_fallback("Charizard (004/102)", "Base Set 1999")
    assert sc._stats["fallback_unmapped_set"] == 1, sc._stats

    # 3) edição mapeada mas nome sem token (NNN/MMM) → no_collector_num
    sc._attribute_fallback("Iron Hands ex", "Temporal Forces")  # → sv5, sem (N/M)
    assert sc._stats["fallback_no_collector_num"] == 1, sc._stats

    # 4) edição mapeada + nº presente → no_coverage (cid existe, pokemontcg.io 404)
    sc._attribute_fallback("Iron Hands ex (070/162)", "Temporal Forces")
    assert sc._stats["fallback_no_coverage"] == 1, sc._stats

    # baldes somam o total de fallbacks atribuídos (invariante do relatório)
    total = (sc._stats["fallback_no_fx"] + sc._stats["fallback_unmapped_set"]
             + sc._stats["fallback_no_collector_num"] + sc._stats["fallback_no_coverage"])
    assert total == 4, total
    print("  v5.13 fallback attribution: no_fx/unmapped_set/no_collector_num/no_coverage ✓")
    return True


def test_rarity_mislabel_gate():
    """2026-06-19: supranumerário + raridade 'Comum' = MYP provavelmente errou a
    RARIDADE (review/não-confiar no label), NÃO carta falsa. Supranumerário com
    raridade real (Rara/Hiper/etc.) = carta normal, sem flag. Match EXATO em 'Comum'."""
    from myp_summary import is_rarity_mislabel
    # Comum supranumerário -> flag (raridade provavelmente mal-rotulada)
    assert is_rarity_mislabel("Zoroark ex do N (286/217)", "Comum")
    assert is_rarity_mislabel("Mega Feraligatr ex (274/217)", "Comum")
    # Supranumerário com raridade REAL -> NÃO flaga (carta real aprovada em scans)
    assert not is_rarity_mislabel("Salamence ex (187/159)", "Rara")
    assert not is_rarity_mislabel("Bellibolt ex da Kissera (188/159)", "Rara Hiper")
    # In-range 'Comum' -> NÃO flaga (não é supranumerário)
    assert not is_rarity_mislabel("Pikachu (058/078)", "Comum")
    # EXATO em 'Comum': 'Incomum' NÃO casa (lição NM-only: nunca substring)
    assert not is_rarity_mislabel("Froakie (088/086)", "Incomum")
    # rarity ausente -> não casa
    assert not is_rarity_mislabel("Psyduck (226/217)", None)
    print("  rarity-mislabel gate: Comum-supr=flag, não-Comum-supr=real, exato ✓")
    return True


def test_tcg_source_column_explicit():
    """v5.14: coluna "TCG Source" declara EXPLICITAMENTE real vs fallback por card.

    Elimina a degradação silenciosa: antes, real-vs-fallback era inferido pela
    presença de "TCG US$". Agora cada row diz `real (pokemontcg.io)` ou
    `fallback (.estat-tcg)`. Honestidade do sinal = regra dura do projeto."""
    real = CardData(
        name="Iron Crown ex (081/162)", edition="Temporal Forces", rarity="Ultra Rara",
        product_url="https://mypcards.com/x/iron-crown", myp_lowest_en_nm=100.0,
        tcg_player_price=200.0, tcg_real_usd=37.0, tcg_source="pokemontcg.io",
        myp_last_sale_brl=190.0, margin_pct=1.0, margin_brl=100.0, en_nm_sellers=3,
    )
    fb = CardData(
        name="Pikachu Comum", edition="Set Antigo", rarity="Comum",
        product_url="https://mypcards.com/x/pika", myp_lowest_en_nm=60.0,
        tcg_player_price=150.0, tcg_real_usd=None, tcg_source="myp_estat",
        myp_last_sale_brl=140.0, margin_pct=1.5, margin_brl=90.0, en_nm_sellers=2,
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = f.name
    generate_xlsx([real, fb], out, threshold=0.25)
    wb = load_workbook(out)
    ws = wb["All EN Cards"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "TCG Source" in hdr, f"coluna 'TCG Source' ausente: {hdr}"
    src_col = hdr.index("TCG Source") + 1
    name_col = hdr.index("Card Name") + 1
    by_name = {}
    for r in range(2, ws.max_row + 1):
        nv = ws.cell(row=r, column=name_col).value
        if nv:
            by_name[str(nv)] = ws.cell(row=r, column=src_col).value
    assert "pokemontcg" in str(by_name["Iron Crown ex (081/162)"]).lower(), by_name
    assert "fallback" in str(by_name["Pikachu Comum"]).lower(), by_name
    assert "estat" in str(by_name["Pikachu Comum"]).lower(), by_name
    Path(out).unlink()
    print("  TCG Source col: real=pokemontcg.io, fallback=.estat-tcg explícito ✓")
    return True


def test_tcg_source_roundtrip_aggregate():
    """v5.14: o aggregate/round-trip PRESERVA a fonte (real/fallback). Sem isso,
    o enrich não saberia o que é real ao reler o XLSX. Inclui o fallback de
    inferência p/ XLSX antigos (sem a coluna 'TCG Source')."""
    from myp_aggregate import card_from_row, load_chunk_cards

    # Com a coluna nova: respeita o rótulo escrito.
    hdr = ["Card Name", "Edition", "Rarity", "MYP EN NM (R$)", "TCG Player (R$)",
           "TCG US$", "TCG Source", "MYP Last Sale (R$)", "Margin %", "Diff (R$)",
           "NM Sellers", "⚠️ EN Trunc", "⚠️ TCG Suspect", "⚠️ Single Seller",
           "⚠️ COLLECTOR#", "URL", "Updated", "TCG URL"]
    real_row = ("Charizard ex", "Surging Sparks", "SIR", 80.0, 200.0, 37.0,
                "real (pokemontcg.io)", 180.0, 1.5, 120.0, 4, "", "", "", "",
                "http://myp/x", "2026-06-20", "http://tcg/x")
    # v5.15.1 REGRESSÃO DURA: a fonte REAL do CI desde a v5.15 é o tcgcsv. O
    # rótulo "real (tcgcsv)" NÃO contém "pokemontcg" — o parser antigo o tratava
    # como fallback, zerando os 537 reais do CI no consolidado. Esta row garante
    # que tcgcsv sobrevive ao round-trip como REAL (token interno `tcgcsv`).
    real_csv_row = ("Garchomp ex", "Surging Sparks", "Double Rare", 70.0, 95.0, 18.4,
                    "real (tcgcsv)", 90.0, 0.35, 25.0, 5, "", "", "", "",
                    "http://myp/z", "2026-06-22", "http://tcg/z")
    fb_row = ("Pikachu", "Velho", "Comum", 60.0, 150.0, None,
              "fallback (.estat-tcg)", 140.0, 1.5, 90.0, 2, "", "", "", "",
              "http://myp/y", "2026-06-20", "http://tcg/y")
    cr = card_from_row(hdr, real_row)
    cc = card_from_row(hdr, real_csv_row)
    cf = card_from_row(hdr, fb_row)
    assert cr.tcg_source == "pokemontcg.io", cr.tcg_source
    assert cc.tcg_source == "tcgcsv", cc.tcg_source  # v5.15.1: tcgcsv preservado
    assert cf.tcg_source == "myp_estat", cf.tcg_source

    # Round-trip completo: generate_xlsx -> load_chunk_cards preserva a fonte.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = f.name
    generate_xlsx([cr, cc, cf], out, threshold=0.25)
    reloaded = {c.name: c.tcg_source for c in load_chunk_cards(Path(out))}
    assert reloaded["Charizard ex"] == "pokemontcg.io", reloaded
    assert reloaded["Garchomp ex"] == "tcgcsv", reloaded  # v5.15.1: NÃO vira fallback
    assert reloaded["Pikachu"] == "myp_estat", reloaded
    # Contagem real-vs-fallback preservada no round-trip (2 real, 1 fallback).
    real_ct = sum(1 for s in reloaded.values() if s in ("pokemontcg.io", "tcgcsv"))
    assert real_ct == 2, f"esperava 2 reais preservados, obteve {real_ct}: {reloaded}"
    Path(out).unlink()

    # XLSX ANTIGO (sem 'TCG Source'): infere pela presença de "TCG US$" — NÃO
    # mascara fallback como real (USD ausente => fallback).
    old_hdr = [h for h in hdr if h != "TCG Source"]
    old_real = tuple(v for h, v in zip(hdr, real_row) if h != "TCG Source")
    old_fb = tuple(v for h, v in zip(hdr, fb_row) if h != "TCG Source")
    assert card_from_row(old_hdr, old_real).tcg_source == "pokemontcg.io"
    assert card_from_row(old_hdr, old_fb).tcg_source == "myp_estat"
    print("  TCG Source round-trip: preserva real/fallback + infere XLSX antigo ✓")
    return True


def _make_mini_chunk(path, cards_spec):
    """Escreve um mini-XLSX de chunk (sheet 'All EN Cards') via generate_xlsx, a
    partir de uma lista (name, src, usd) — src em {'tcgcsv','pokemontcg.io','myp_estat'}."""
    cards = []
    for i, (name, src, usd) in enumerate(cards_spec):
        cards.append(CardData(
            name=name, edition="Surging Sparks", rarity="Double Rare",
            product_url=f"https://myp/{name}{i}",
            myp_lowest_en_nm=70.0, tcg_player_price=95.0,
            tcg_real_usd=usd, tcg_source=src,
            myp_last_sale_brl=90.0, margin_pct=0.35, margin_brl=25.0,
            en_nm_sellers=5, last_updated="2026-06-22",
        ))
    generate_xlsx(cards, str(path), threshold=0.30)


def test_aggregate_multichunk_preserves_real_counts():
    """v5.15.1 REGRESSÃO DURA (bug do run 27926311953): a agregação de MÚLTIPLOS
    chunks deve PRESERVAR as contagens real-vs-fallback. O bug: chunks gravavam
    'real (tcgcsv)' mas o consolidado saía 100% 'fallback (.estat-tcg)' porque o
    parser do aggregate não reconhecia o rótulo tcgcsv. Aqui montamos 2 mini-chunks
    com mix real(tcgcsv)/real(pokemontcg)/fallback, rodamos o aggregate.main() de
    verdade (o caminho do workflow), e assertamos que o consolidado preserva os
    reais — não os rebaixa pra fallback."""
    import myp_aggregate

    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        c0 = d / "myp_chunk_0.xlsx"
        c1 = d / "myp_chunk_1.xlsx"
        out = d / "consolidated.xlsx"
        # chunk 0: 2 tcgcsv real + 1 fallback ; chunk 1: 1 pokemontcg real + 1 fallback
        _make_mini_chunk(c0, [
            ("Garchomp ex", "tcgcsv", 18.4),
            ("Garbodor", "tcgcsv", 18.6),
            ("Entei", "myp_estat", None),
        ])
        _make_mini_chunk(c1, [
            ("Charizard ex", "pokemontcg.io", 37.0),
            ("Raichu", "myp_estat", None),
        ])

        # Roda o aggregate EXATAMENTE como o workflow (via main()).
        argv = ["myp_aggregate.py", str(c0), str(c1), "--output", str(out),
                "--threshold", "0.30"]
        old_argv = sys.argv[:]
        sys.argv = argv
        try:
            rc = myp_aggregate.main()
        finally:
            sys.argv = old_argv
        assert rc == 0, f"aggregate.main() retornou {rc}"

        # Lê a coluna 'TCG Source' do consolidado e conta real vs fallback.
        wb = load_workbook(out, read_only=True, data_only=True)
        ws = wb["All EN Cards"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = list(rows[0])
        si = headers.index("TCG Source")
        labels = [r[si] for r in rows[1:]]
        real_ct = sum(1 for s in labels if s and "real" in str(s).lower())
        fb_ct = sum(1 for s in labels if s and "fallback" in str(s).lower())
        tcgcsv_ct = sum(1 for s in labels if s and "tcgcsv" in str(s).lower())

        # Input total = 3 real (2 tcgcsv + 1 pokemontcg) + 2 fallback.
        assert len(labels) == 5, f"esperava 5 cards consolidados, obteve {len(labels)}"
        assert real_ct == 3, f"esperava 3 reais preservados, obteve {real_ct}: {labels}"
        assert fb_ct == 2, f"esperava 2 fallback, obteve {fb_ct}: {labels}"
        assert tcgcsv_ct == 2, f"esperava 2 tcgcsv preservados, obteve {tcgcsv_ct}: {labels}"
    print("  aggregate multi-chunk: preserva 3 real (2 tcgcsv) + 2 fallback ✓")
    return True


def test_summary_real_coverage_signal():
    """v5.14: o resumo de entrega TORNA VISÍVEL a cobertura de preço real. Se
    todos os deals limpos forem fallback, marca '🛑 ZERO preço real' (degradação
    do CI). Honestidade do sinal na ENTREGA, não só no XLSX."""
    from myp_summary import build_markdown

    # XLSX com 2 deals limpos, ambos FALLBACK (sem preço real) → deve gritar.
    fb1 = CardData(
        name="Charizard ex Holo", edition="Surging Sparks", rarity="SIR",
        product_url="https://myp/c1", myp_lowest_en_nm=80.0, tcg_player_price=200.0,
        tcg_real_usd=None, tcg_source="myp_estat", myp_last_sale_brl=190.0,
        margin_pct=1.5, margin_brl=120.0, en_nm_sellers=3, last_updated="2026-06-20",
    )
    fb2 = CardData(
        name="Pikachu ex Holo", edition="Surging Sparks", rarity="Ultra Rara",
        product_url="https://myp/c2", myp_lowest_en_nm=60.0, tcg_player_price=150.0,
        tcg_real_usd=None, tcg_source="myp_estat", myp_last_sale_brl=140.0,
        margin_pct=1.5, margin_brl=90.0, en_nm_sellers=2, last_updated="2026-06-20",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([fb1, fb2], xlsx, threshold=0.25)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    assert "Cobertura de preço TCG real" in text, "linha de cobertura ausente"
    assert "ZERO preço real" in text, f"esperava alerta de zero real:\n{text[:600]}"
    # v5.14.1: cobertura é medida sobre o UNIVERSO de cartas EN (2 cartas), não
    # sobre o subconjunto de deals. Ambas são fallback → 0/2 cartas EN reais.
    assert "0/2 cartas EN" in text, f"cobertura deve ser sobre o universo (0/2):\n{text[:600]}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary real-coverage: '🛑 ZERO preço real' quando tudo é fallback ✓")
    return True


def test_summary_coverage_real_universe_with_zero_deals():
    """v5.14.1: o BUG corrigido. Quando o catálogo EN tem preço 100% REAL mas
    NENHUMA carta bate o threshold (0 deals limpos ≥30%), o resumo NÃO pode
    gritar '🛑 ZERO' nem '0/0' — isso fazia o operador (médico, não-programador)
    achar que a key falhou. A cobertura deve refletir TODAS as cartas EN
    (`TCG Source`), não o balde de deals.

    Caso: 3 cartas EN com preço REAL (pokemontcg.io), margens baixas (sem deal
    ≥30%). Esperado: '✅ 3/3 cartas EN com preço REAL' + nota de 0 deals limpos.
    """
    from myp_summary import build_markdown

    def _real_lowmargin(name, myp, usd, ed="Surging Sparks"):
        return CardData(
            name=name, edition=ed, rarity="Ultra Rara",
            product_url=f"https://myp/{name}", myp_lowest_en_nm=myp,
            tcg_player_price=myp * 1.05,   # margem ~5% → NÃO é deal ≥30%
            tcg_real_usd=usd, tcg_source="pokemontcg.io",
            myp_last_sale_brl=myp, margin_pct=0.05, margin_brl=myp * 0.05,
            en_nm_sellers=3, last_updated="2026-06-20",
        )

    cards = [
        _real_lowmargin("Charizard ex", 80.0, 16.0),
        _real_lowmargin("Pikachu ex", 60.0, 12.0),
        _real_lowmargin("Mewtwo ex", 90.0, 18.0),
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx(cards, xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    # NÃO pode falsear ZERO nem 0/0 quando o universo é 100% real.
    assert "ZERO preço real" not in text, \
        f"BUG: gritou ZERO com universo 100% real:\n{text[:700]}"
    assert "0/0" not in text, f"BUG: imprimiu '0/0' enganoso:\n{text[:700]}"
    assert "3/3 cartas EN com preço REAL" in text, \
        f"cobertura do universo (3/3 real) ausente:\n{text[:700]}"
    # E deixa claro que não houve deal ≥threshold (sem confundir com cobertura).
    assert "nenhum deal limpo" in text, \
        f"esclarecimento de 0 deals ausente:\n{text[:700]}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary coverage: universo 100% real + 0 deals → '✅ 3/3', não ZERO ✓")
    return True


def test_summary_coverage_mixed_real_fallback():
    """v5.14.1: cobertura parcial sobre o universo. Mix real/fallback nas cartas
    EN → '⚠️ N/M cartas EN com preço REAL' contando TODAS, não só os deals."""
    from myp_summary import build_markdown

    real = CardData(
        name="Charizard ex", edition="Surging Sparks", rarity="SIR",
        product_url="https://myp/r", myp_lowest_en_nm=80.0, tcg_player_price=200.0,
        tcg_real_usd=36.0, tcg_source="pokemontcg.io", myp_last_sale_brl=180.0,
        margin_pct=1.5, margin_brl=120.0, en_nm_sellers=3, last_updated="2026-06-20",
    )
    fb = CardData(
        name="Pikachu ex", edition="Surging Sparks", rarity="Ultra Rara",
        product_url="https://myp/f", myp_lowest_en_nm=60.0, tcg_player_price=150.0,
        tcg_real_usd=None, tcg_source="myp_estat", myp_last_sale_brl=140.0,
        margin_pct=1.5, margin_brl=90.0, en_nm_sellers=2, last_updated="2026-06-20",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([real, fb], xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    assert "1/2 cartas EN com preço REAL" in text, \
        f"cobertura mista (1/2) ausente:\n{text[:700]}"
    assert "ZERO preço real" not in text, "não é zero, é parcial"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary coverage: mix real/fallback → '⚠️ 1/2 cartas EN' ✓")
    return True


def test_summary_deal_floor_matches_real_threshold():
    """v5.14.1 (fix): o piso de "deal limpo" no summary casa o threshold REAL do
    scan (lido do XLSX), não o 0.25 hardcoded legado. Uma carta de 27% sob um
    scan de threshold 30% NÃO pode ser contada/impressa como 'deal limpo (≥30%)'
    — isso era uma afirmação FALSA (carta sub-threshold rotulada como ≥30%).

    Caso: 1 carta EN, margem 27%, preço REAL, scan threshold=0.30. Esperado:
    'nenhum deal limpo' (não vaza a banda 25–30%), e cobertura segue '✅ 1/1'.
    """
    from myp_summary import build_markdown

    card = CardData(
        name="Snorlax ex", edition="Surging Sparks", rarity="Ultra Rara",
        product_url="https://myp/s", myp_lowest_en_nm=100.0,
        tcg_player_price=127.0,            # margem 27% → sub-threshold de 30%
        tcg_real_usd=25.4, tcg_source="pokemontcg.io", myp_last_sale_brl=100.0,
        margin_pct=0.27, margin_brl=27.0, en_nm_sellers=3, last_updated="2026-06-20",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([card], xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    # A carta de 27% NÃO é um deal limpo ≥30%.
    assert "nenhum deal limpo" in text, \
        f"BUG: carta de 27% vazou como deal limpo sob threshold 30%:\n{text[:700]}"
    assert "deals limpos (≥30%)" not in text, \
        f"BUG: imprimiu contagem de 'deals limpos (≥30%)' com carta sub-threshold:\n{text[:700]}"
    # Cobertura (universo) não é afetada: a carta tem preço real.
    assert "1/1 cartas EN com preço REAL" in text, \
        f"cobertura do universo (1/1 real) ausente:\n{text[:700]}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary deal-floor: 27% sob threshold 30% → 'nenhum deal limpo', não ≥30% ✓")
    return True


def _section_of(text: str, needle: str):
    """Retorna o último header '## …' antes da 1ª ocorrência de `needle` (ou None)."""
    section = None
    for ln in text.splitlines():
        if ln.startswith("## "):
            section = ln.strip()
        if needle in ln:
            return section
    return None


def _fallback_deal(name, myp, estat_brl, rarity="Ultra Rara", last_sale=None):
    """Deal com preço FALLBACK (.estat-tcg): sem USD real, source=myp_estat."""
    return CardData(
        name=name, edition="Surging Sparks", rarity=rarity,
        product_url=f"https://myp/{name}", myp_lowest_en_nm=myp,
        tcg_player_price=estat_brl, tcg_real_usd=None, tcg_source="myp_estat",
        myp_last_sale_brl=last_sale, margin_pct=(estat_brl - myp) / myp,
        margin_brl=estat_brl - myp, en_nm_sellers=3, last_updated="2026-06-20",
    )


def _real_deal(name, myp, tcg_brl, usd, rarity="Ultra Rara"):
    """Deal com preço REAL (pokemontcg.io)."""
    return CardData(
        name=name, edition="Surging Sparks", rarity=rarity,
        product_url=f"https://myp/{name}", myp_lowest_en_nm=myp,
        tcg_player_price=tcg_brl, tcg_real_usd=usd, tcg_source="pokemontcg.io",
        myp_last_sale_brl=myp, margin_pct=(tcg_brl - myp) / myp,
        margin_brl=tcg_brl - myp, en_nm_sellers=3, last_updated="2026-06-20",
    )


def test_summary_fallback_deal_not_in_clean():
    """v5.14.3 (BLOCKER reproduzido + corrigido): um deal com preço FALLBACK
    inflado, SEM última venda (gate de suspect pulado), raridade não-Comum e em
    range NÃO pode entrar no balde 'Top 50 deals limpos' — a margem é ilusória
    (caso Darumaka R$2867 vs R$60 → 4678%). Deve ir pro balde FALLBACK dedicado."""
    from myp_summary import build_markdown

    darumaka = _fallback_deal("Darumaka", 60.0, 2867.0, rarity="Ultra Rara",
                              last_sale=None)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([darumaka], xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    sec = _section_of(text, "Darumaka")
    assert sec is not None and "FALLBACK" in sec, \
        f"Darumaka (fallback) devia estar no balde FALLBACK, está em: {sec!r}\n{text[:900]}"
    assert "limpos" not in (sec or "").lower(), \
        f"BUG: fallback inflado vazou pro balde limpo: {sec!r}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  fallback-not-clean: Darumaka inflado sem last-sale → balde FALLBACK ✓")
    return True


def test_summary_real_deal_stays_clean():
    """v5.14.3: um deal com preço REAL (pokemontcg.io) continua no balde limpo e
    NÃO aparece no balde fallback."""
    from myp_summary import build_markdown

    real = _real_deal("Pikachu ex", 50.0, 120.0, 24.0)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([real], xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0
    text = Path(md).read_text(encoding="utf-8")
    sec = _section_of(text, "Pikachu ex")
    assert sec is not None and "limpos" in sec.lower(), \
        f"deal real devia estar no balde limpo, está em: {sec!r}"
    assert "FALLBACK `.estat-tcg`" not in text or "Pikachu ex" not in \
        text.split("FALLBACK `.estat-tcg`")[-1], "deal real não pode estar no balde fallback"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  real-stays-clean: deal real → balde limpo (não fallback) ✓")
    return True


def test_summary_ci_all_fallback_zero_clean():
    """v5.14.3: cenário CI (runners não alcançam pokemontcg.io → 100% fallback):
    ZERO deals limpos + balde fallback populado. Não engana o operador."""
    from myp_summary import build_markdown

    cards = [_fallback_deal("Charizard ex", 80.0, 900.0),
             _fallback_deal("Mewtwo ex", 60.0, 700.0)]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx(cards, xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0
    text = Path(md).read_text(encoding="utf-8")
    # Bloco limpo presente mas vazio:
    clean_block = text.split("## ⚠️")[0]
    assert "Nenhum deal limpo nesta run" in clean_block, \
        f"CI all-fallback devia dar 0 deals limpos:\n{clean_block[:600]}"
    # Os 2 fallbacks listados no balde dedicado:
    assert "FALLBACK `.estat-tcg`" in text, "balde fallback ausente"
    assert _section_of(text, "Charizard ex") and "FALLBACK" in _section_of(text, "Charizard ex")
    assert _section_of(text, "Mewtwo ex") and "FALLBACK" in _section_of(text, "Mewtwo ex")
    Path(xlsx).unlink(); Path(md).unlink()
    print("  ci-all-fallback: 0 limpos + 2 no balde fallback ✓")
    return True


def test_summary_mix_real_and_fallback_deals():
    """v5.14.3: mix — 1 deal real + 1 fallback (ambos ≥threshold). Real → limpo;
    fallback → balde fallback. Cada um no seu lugar."""
    from myp_summary import build_markdown

    cards = [_real_deal("Iron Hands ex", 50.0, 150.0, 30.0),
             _fallback_deal("Roaring Moon", 70.0, 1200.0)]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx(cards, xlsx, threshold=0.30)
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0
    text = Path(md).read_text(encoding="utf-8")
    assert "limpos" in (_section_of(text, "Iron Hands ex") or "").lower(), \
        "deal real devia estar no balde limpo"
    assert "FALLBACK" in (_section_of(text, "Roaring Moon") or ""), \
        "deal fallback devia estar no balde fallback"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  mix: real→limpo, fallback→fallback ✓")
    return True


def test_summary_fallback_gate_old_xlsx():
    """v5.14.3 + XLSX antigo (sem coluna 'TCG Source'): o gate de fallback usa a
    inferência por 'TCG US$'. Deal antigo real (com USD) fica limpo; deal antigo
    fallback (sem USD) vai pro balde fallback. Não regride leitura pré-v5.14."""
    from myp_summary import build_markdown

    cards = [_real_deal("Iron Hands ex", 50.0, 150.0, 30.0),
             _fallback_deal("Roaring Moon", 70.0, 1200.0)]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx(cards, xlsx, threshold=0.30)
    _drop_all_en_column(xlsx, "TCG Source")   # XLSX "antigo"
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0
    text = Path(md).read_text(encoding="utf-8")
    assert "limpos" in (_section_of(text, "Iron Hands ex") or "").lower(), \
        "deal real (com USD) devia ficar limpo mesmo em XLSX antigo"
    assert "FALLBACK" in (_section_of(text, "Roaring Moon") or ""), \
        "deal fallback (sem USD) devia ir pro balde fallback em XLSX antigo"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  fallback-gate (XLSX antigo): real→limpo, sem-USD→fallback ✓")
    return True


def _drop_all_en_column(xlsx_path: str, col_name: str):
    """Remove uma coluna da aba 'All EN Cards' (simula XLSX antigo sem ela)."""
    wb = load_workbook(xlsx_path)
    ws = wb["All EN Cards"]
    hdr = [c.value for c in ws[1]]
    if col_name in hdr:
        ws.delete_cols(hdr.index(col_name) + 1, 1)
    wb.save(xlsx_path)
    wb.close()


def _blank_all_en_cells(xlsx_path: str, card_name: str, cols):
    """Apaga (None) células `cols` da linha de `card_name` na aba 'All EN Cards'."""
    wb = load_workbook(xlsx_path)
    ws = wb["All EN Cards"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(hdr)}
    for row in ws.iter_rows(min_row=2):
        if row[idx["Card Name"] - 1].value == card_name:
            for col in cols:
                if col in idx:
                    ws.cell(row=row[0].row, column=idx[col]).value = None
    wb.save(xlsx_path)
    wb.close()


def test_summary_coverage_old_xlsx_inference():
    """v5.14.1 (gap fechado): cobertura medida no `build_markdown` sobre um XLSX
    ANTIGO (sem a coluna 'TCG Source'). `_is_real` deve INFERIR real/fallback pela
    presença de 'TCG US$' (que só o preço real preenche). Exercita o ramo de
    inferência DENTRO do summary (o round-trip aggregate cobria outro caminho)."""
    from myp_summary import build_markdown

    real = CardData(
        name="Charizard ex", edition="Surging Sparks", rarity="SIR",
        product_url="https://myp/r", myp_lowest_en_nm=80.0, tcg_player_price=200.0,
        tcg_real_usd=36.0, tcg_source="pokemontcg.io", myp_last_sale_brl=180.0,
        margin_pct=1.5, margin_brl=120.0, en_nm_sellers=3, last_updated="2026-06-20",
    )
    fb = CardData(
        name="Pikachu ex", edition="Surging Sparks", rarity="Ultra Rara",
        product_url="https://myp/f", myp_lowest_en_nm=60.0, tcg_player_price=150.0,
        tcg_real_usd=None, tcg_source="myp_estat", myp_last_sale_brl=140.0,
        margin_pct=1.5, margin_brl=90.0, en_nm_sellers=2, last_updated="2026-06-20",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([real, fb], xlsx, threshold=0.30)
    _drop_all_en_column(xlsx, "TCG Source")   # → XLSX "antigo"
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    # Inferido só por 'TCG US$': real tem USD, fallback não → 1/2.
    assert "1/2 cartas EN com preço REAL" in text, \
        f"inferência de XLSX antigo (1/2) ausente:\n{text[:700]}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary coverage: XLSX antigo s/ 'TCG Source' infere por 'TCG US$' → 1/2 ✓")
    return True


def test_summary_coverage_excludes_unpriced_card():
    """v5.14.1 (gap fechado): uma carta SEM nenhum preço TCG (sem USD, sem R$ e
    sem 'TCG Source') NÃO entra no denominador de cobertura — não há o que ser
    real/fallback nela. Mix: 1 real + 1 sem-preço → '1/1 cartas EN', não 1/2."""
    from myp_summary import build_markdown

    real = CardData(
        name="Charizard ex", edition="Surging Sparks", rarity="SIR",
        product_url="https://myp/r", myp_lowest_en_nm=80.0, tcg_player_price=200.0,
        tcg_real_usd=36.0, tcg_source="pokemontcg.io", myp_last_sale_brl=180.0,
        margin_pct=1.5, margin_brl=120.0, en_nm_sellers=3, last_updated="2026-06-20",
    )
    noprice = CardData(
        name="Pikachu ex", edition="Surging Sparks", rarity="Ultra Rara",
        product_url="https://myp/f", myp_lowest_en_nm=60.0, tcg_player_price=150.0,
        tcg_real_usd=None, tcg_source="myp_estat", myp_last_sale_brl=140.0,
        margin_pct=1.5, margin_brl=90.0, en_nm_sellers=2, last_updated="2026-06-20",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([real, noprice], xlsx, threshold=0.30)
    # Tira TODO preço TCG da 2ª carta → fica fora do universo de cobertura.
    _blank_all_en_cells(xlsx, "Pikachu ex",
                        ["TCG US$", "TCG Player (R$)", "TCG Source"])
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    assert "1/1 cartas EN com preço REAL" in text, \
        f"carta sem preço deve sair do denominador (1/1, não 1/2):\n{text[:700]}"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary coverage: carta sem preço TCG fora do denominador → 1/1 ✓")
    return True


def test_summary_coverage_no_price_at_all_branch():
    """v5.14.1 (gap fechado): universo SEM nenhum preço TCG (total_priced == 0)
    → ramo '⚠️ Sem preço TCG', nunca um falso ZERO nem '0/0'."""
    from myp_summary import build_markdown

    card = CardData(
        name="Charizard ex", edition="Surging Sparks", rarity="SIR",
        product_url="https://myp/r", myp_lowest_en_nm=80.0, tcg_player_price=200.0,
        tcg_real_usd=36.0, tcg_source="pokemontcg.io", myp_last_sale_brl=180.0,
        margin_pct=1.5, margin_brl=120.0, en_nm_sellers=3, last_updated="2026-06-20",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx = f.name
    generate_xlsx([card], xlsx, threshold=0.30)
    _blank_all_en_cells(xlsx, "Charizard ex",
                        ["TCG US$", "TCG Player (R$)", "TCG Source"])
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as f:
        md = f.name

    rc = build_markdown(xlsx, md, scan_type="weekly", run_id="", repo="x/y")
    assert rc == 0, f"build_markdown retornou {rc}"
    text = Path(md).read_text(encoding="utf-8")
    assert "Sem preço TCG" in text, f"ramo 'Sem preço TCG' ausente:\n{text[:700]}"
    assert "ZERO preço real" not in text, "sem preço ≠ ZERO real (não pode confundir)"
    Path(xlsx).unlink(); Path(md).unlink()
    print("  summary coverage: universo sem preço → '⚠️ Sem preço TCG', não ZERO ✓")
    return True


# ─── Regressão: segredo com BOM/zero-width → header latin-1-encodável ──────────
# Bug real (scanner irmão CardTrader, GH Actions): o secret POKEMONTCG_API_KEY
# tinha um BOM (U+FEFF) na frente. Headers HTTP são codificados em latin-1 pelo
# `requests`, então "\ufeff..." virava
# `UnicodeEncodeError: 'latin-1' codec can't encode '\ufeff'` em TODA chamada de
# pricing → mass pricing failure → scan "verde" mas vazio. O MYP está hoje
# insulado no CI (usa --tcg-source tcgcsv, sem header pokemontcg), mas em modo
# pokemontcg/auto o mesmo header X-Api-Key é montado — _clean_secret trava isso.
# NB: usamos escapes "\ufeff"/"\u200b" de propósito (nada de invisível literal
# no fonte do teste).
_BOM = "\ufeff"
_ZWSP = "\u200b"


def test_clean_secret_strips_bom_and_zero_width():
    assert _clean_secret(_BOM + "abc123") == "abc123"
    assert _clean_secret(_ZWSP + "abc123") == "abc123"
    assert _clean_secret(_BOM + "  abc123  \n") == "abc123"
    assert _clean_secret("  abc123\n") == "abc123"
    # vazio / só-invisível → None (caller trata como 'sem key')
    assert _clean_secret(None) is None
    assert _clean_secret("") is None
    assert _clean_secret("   ") is None
    assert _clean_secret(_BOM) is None
    assert _clean_secret(_ZWSP + _BOM) is None
    # valor limpo passa intacto
    assert _clean_secret("eyJhbGciOi.token.sig") == "eyJhbGciOi.token.sig"
    print("  _clean_secret: BOM/ZWSP/whitespace removidos, vazio→None ✓")
    return True


def test_ptcg_api_key_bom_header_is_latin1_encodable():
    """Com POKEMONTCG_API_KEY prefixada de BOM, o scanner sanitiza no read e o
    header X-Api-Key resultante codifica em latin-1 (o que o requests faz). Antes
    do _clean_secret isto estouraria UnicodeEncodeError em toda chamada."""
    prev = os.environ.get("POKEMONTCG_API_KEY")
    try:
        os.environ["POKEMONTCG_API_KEY"] = _BOM + "secretkey"
        sc = MYPScraper(delay=0.0)
        assert sc.ptcg_api_key == "secretkey", repr(sc.ptcg_api_key)
        header = {"X-Api-Key": sc.ptcg_api_key}["X-Api-Key"]
        # O ponto da regressão: NÃO pode levantar UnicodeEncodeError.
        header.encode("latin-1")
    finally:
        if prev is None:
            os.environ.pop("POKEMONTCG_API_KEY", None)
        else:
            os.environ["POKEMONTCG_API_KEY"] = prev
    print("  X-Api-Key com BOM: sanitizado → latin-1-encodável ✓")
    return True


def test_ptcg_api_key_bom_only_yields_no_key():
    """Key que era só BOM/zero-width vira None → sem header X-Api-Key (fallback
    sem-key, válido na pokemontcg.io). Não vira header inválido."""
    prev = os.environ.get("POKEMONTCG_API_KEY")
    try:
        os.environ["POKEMONTCG_API_KEY"] = _ZWSP + _BOM
        sc = MYPScraper(delay=0.0)
        assert sc.ptcg_api_key is None, repr(sc.ptcg_api_key)
        headers = {"X-Api-Key": sc.ptcg_api_key} if sc.ptcg_api_key else {}
        assert "X-Api-Key" not in headers
    finally:
        if prev is None:
            os.environ.pop("POKEMONTCG_API_KEY", None)
        else:
            os.environ["POKEMONTCG_API_KEY"] = prev
    print("  X-Api-Key só-BOM: vira None → sem header ✓")
    return True


def main():
    tests = [
        ("threshold constant", test_threshold_constant),
        ("secret BOM/zero-width sanitization", test_clean_secret_strips_bom_and_zero_width),
        ("X-Api-Key c/ BOM → latin-1-encodável", test_ptcg_api_key_bom_header_is_latin1_encodable),
        ("X-Api-Key só-BOM → sem header", test_ptcg_api_key_bom_only_yields_no_key),
        ("rarity-mislabel gate (2026-06-19)", test_rarity_mislabel_gate),
        ("Jirachi ratio math", test_jirachi_ratio_math),
        ("v5.14.4 tcg_suspect boundary exatamente 10x", test_tcg_suspect_boundary_exactly_10x),
        ("parse_brl BR/US formats (v5.8.10)", test_parse_brl_formats),
        ("_last_brl extraction (v5.8.10)", test_last_brl),
        ("oversized/jumbo regex (v5.8.10)", test_oversized_regex),
        ("XLSX end-to-end", test_xlsx_end_to_end),
        ("tcg_search_url (v5.8.8)", test_tcg_search_url),
        ("price cell hyperlinks (v5.8.8/v5.8.9)", test_price_cell_hyperlinks),
        ("coluna TCG URL texto plano (v5.11.2)", test_tcg_url_column),
        ("myp_edition_to_ptcg_setcode (v5.8.9)", test_myp_edition_to_setcode),
        ("tcg_direct_url (v5.8.9)", test_tcg_direct_url),
        ("marketplace pagination (v5.9)", test_marketplace_pagination),
        ("pagination truncation gate (v5.9)", test_pagination_gate_skips_untruncated),
        ("pagination cost gate TCG<min (v5.9.1)", test_pagination_cost_gate_low_tcg),
        ("A3 real price rescues pagination (v5.11.5)", test_a3_real_price_rescues_pagination),
        ("real TCG overrides .estat-tcg (v5.11)", test_real_tcg_overrides_estat),
        ("fallback to .estat-tcg sem cobertura (v5.11)", test_fallback_to_estat_when_no_coverage),
        ("sem câmbio mantém .estat-tcg (v5.11)", test_no_fx_keeps_estat),
        ("A1 preço real limpa suspect (v5.11.3)", test_real_price_clears_suspect),
        ("A2 precifica sem .estat-tcg (v5.11.3)", test_prices_card_without_estat_tcg),
        ("A2 skip sem TCG nenhum (v5.11.3)", test_skip_when_no_tcg_at_all),
        ("delivery table format (v5.11.1)", test_delivery_table_format),
        ("checkpoint save/load (v5.11.4)", test_checkpoint_save_load),
        ("scan resume skips done editions (v5.11.4)", test_scan_resume_skips_done_editions),
        ("v5.12 prefill batch pokemontcg.io por set", test_prefill_ptcg_set_batch),
        ("v5.15 tcgcsv prefill parseia schema real", test_tcgcsv_prefill_parses_schema),
        ("v5.15 tcgcsv resolução de groupId", test_tcgcsv_groupid_resolution),
        ("v5.16 tabela setcode→abbr auto-consistente", test_setcode_abbr_table_is_self_consistent),
        ("v5.16 abbrs resolvem 1-a-1 no /groups (fixture)", test_setcode_abbr_resolves_1to1_against_groups_fixture),
        ("v5.15 tcgcsv sem match → fallback honesto", test_tcgcsv_no_match_falls_back_honestly),
        ("v5.15 tcgcsv e2e source='tcgcsv' (REAL)", test_tcgcsv_end_to_end_real_source_label),
        ("v5.15 tcgcsv reconhecido como REAL no summary", test_tcgcsv_recognized_as_real_in_summary),
        ("v5.13 atribuição de cobertura do fallback", test_fallback_attribution),
        ("v5.14 coluna TCG Source explícita", test_tcg_source_column_explicit),
        ("v5.14 TCG Source round-trip aggregate", test_tcg_source_roundtrip_aggregate),
        ("v5.15.1 aggregate multi-chunk preserva contagens real", test_aggregate_multichunk_preserves_real_counts),
        ("v5.14 sinal de cobertura real no resumo", test_summary_real_coverage_signal),
        ("v5.14.1 cobertura sobre universo EN c/ 0 deals", test_summary_coverage_real_universe_with_zero_deals),
        ("v5.14.1 cobertura mista real/fallback (universo)", test_summary_coverage_mixed_real_fallback),
        ("v5.14.1 piso de deal casa threshold real (não 0.25)", test_summary_deal_floor_matches_real_threshold),
        ("v5.14.2 cobertura infere XLSX antigo s/ TCG Source", test_summary_coverage_old_xlsx_inference),
        ("v5.14.2 cobertura exclui carta sem preço TCG", test_summary_coverage_excludes_unpriced_card),
        ("v5.14.2 cobertura ramo 'sem preço TCG' (≠ZERO)", test_summary_coverage_no_price_at_all_branch),
        ("v5.14.3 fallback inflado NÃO entra em deals limpos (Darumaka)", test_summary_fallback_deal_not_in_clean),
        ("v5.14.3 deal real continua limpo", test_summary_real_deal_stays_clean),
        ("v5.14.3 CI all-fallback → 0 limpos + balde fallback", test_summary_ci_all_fallback_zero_clean),
        ("v5.14.3 mix real/fallback → cada um no seu balde", test_summary_mix_real_and_fallback_deals),
        ("v5.14.3 gate fallback em XLSX antigo (infere por USD)", test_summary_fallback_gate_old_xlsx),
    ]
    failed = 0
    for name, fn in tests:
        print(f"\n[TEST] {name}")
        try:
            fn()
            print(f"  ✓ PASS")
        except AssertionError as e:
            print(f"  ✗ FAIL: {e}")
            failed += 1
        except Exception as e:
            print(f"  ✗ ERROR: {type(e).__name__}: {e}")
            failed += 1

    print(f"\n{'═' * 50}")
    if failed:
        print(f"❌ {failed}/{len(tests)} testes falharam")
        sys.exit(1)
    print(f"✅ Todos os {len(tests)} testes passaram — fixes v5.8 + v5.9 OK")
    sys.exit(0)


if __name__ == "__main__":
    main()
