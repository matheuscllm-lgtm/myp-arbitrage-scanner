"""Suíte offline do scanner MYP → ONE PIECE (myp_op_scanner + summary).

Sem rede / sem Cloudflare / sem credencial — fixtures sintéticas calcadas nos
fatos PROVADOS pela sonda probe-myp-onepiece (PR #98, run 31300834735,
2026-08-09):
  - título de edição concatenado ("Carrying On His WillOP1307/11/2025");
  - h1 em 3 formatos: "Sabo (OP13-004)" (código completo), "[Pré-Venda]
    Edward.Newgate (001) (Alternate Art)" (número CURTO + qualificador),
    "Monkey.D.Dragon" (nome puro);
  - campo "Código" = one_<edição>_<código>[marcador] com token de edição
    que TAMBÉM tem hífen ("one_st-35_p-105") e marcador p1 NÃO-CONFIÁVEL
    (par Edward.Newgate com códigos invertidos — provado);
  - tcgcsv cat 68 com colisão base × "(Parallel)" × "(Box Topper)" no
    mesmo Number (OP01) e vocabulário divergente do MYP ("(Alternate
    Art)" ↔ "(Parallel)");
  - starter decks REIMPRIMEM números OPxx (Sabo OP13-004 no grupo ST-35).

Run: python test_myp_op_offline.py   (ou via pytest, coletado na raiz)
"""
import sys
import tempfile
from pathlib import Path

if (sys.stdout.encoding or "").lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from myp_dbz_scanner import normalize_code, split_edition_title, split_myp_title
from myp_op_scanner import (
    MYPOpScraper,
    OpCardData,
    OpSemRefRow,
    _base_name,
    _clean_card_name_op,
    _norm_name_op,
    canonical_set_code_op,
    extract_myp_op_product_code,
    find_reference_op,
    generate_op_xlsx,
    parse_myp_op_product_code,
    resolve_edition_group_op,
    tcg_op_search_url,
)
import myp_op_summary


# ── helpers de fixture ──────────────────────────────────────────────────

def make_entry(pid, name, gid, special=False, prices=None, lows=None,
               number="OP01-025", rarity="Super Rare"):
    clean = _clean_card_name_op(name)
    return {"pid": pid, "name": name, "name_norm": _norm_name_op(clean),
            "base_norm": _norm_name_op(_base_name(clean)), "number": number,
            "rarity": rarity, "url": f"https://www.tcgplayer.com/product/{pid}",
            "prices": prices if prices is not None else {"Normal": 20.0},
            "lows": lows or {}, "cat": 68, "group_id": gid,
            "group_name": "G", "group_special": special}


def make_card(**kw):
    base = dict(
        name="Roronoa Zoro - OP01-025", en_name="Roronoa Zoro (025)",
        code="OP01-025", code_norm="OP1-25", edition="Romance Dawn",
        section="onepiece", product_url="https://mypcards.com/onepiece/produto/1/x",
        myp_lowest_en_nm=100.0, en_nm_sellers=3, tcg_usd=40.0, tcg_brl=200.0,
        tcg_url="https://www.tcgplayer.com/product/1", rarity="Super Rare",
        subtype="Normal", join_via="codigo/grupo-edição/nome-exato",
        margin_pct=1.0, margin_brl=100.0,
        last_updated="2026-08-09 00:00",
    )
    base.update(kw)
    return OpCardData(**base)


# ── nomes / normalização OP ─────────────────────────────────────────────

def test_clean_card_name_op_numero_curto():
    # número CURTO "(001)" (h1 real do OP17 e nomes do tcgcsv OP01) sai
    assert _clean_card_name_op("Edward.Newgate (001) (Alternate Art)") == \
        "Edward.Newgate (Alternate Art)"
    assert _clean_card_name_op("Shanks (020)") == "Shanks"
    assert _clean_card_name_op("Trafalgar Law (002) (Parallel)") == \
        "Trafalgar Law (Parallel)"
    # código completo (formato "Sabo (OP13-004)") também sai, sem "( )" órfão
    assert _clean_card_name_op("Sabo (OP13-004)") == "Sabo"
    assert _clean_card_name_op('Eustass"Captain"Kid (OP10-099)') == \
        'Eustass"Captain"Kid'
    assert _clean_card_name_op("Monkey.D.Dragon") == "Monkey.D.Dragon"
    assert _clean_card_name_op("Cavendish (Box Topper)") == \
        "Cavendish (Box Topper)"


def test_norm_name_op_equivalencia_alternate_art_parallel():
    # vocabulário divergente provado: MYP "(Alternate Art)" ↔ tcgcsv
    # "(Parallel)" — convergem na normalização; outros qualificadores não
    assert _norm_name_op("Edward.Newgate (Alternate Art)") == \
        _norm_name_op("Edward.Newgate (Parallel)")
    assert _norm_name_op("Cavendish (Box Topper)") != \
        _norm_name_op("Cavendish (Parallel)")
    assert _norm_name_op("Nami  (Manga)") == "nami (manga)"


def test_split_myp_title_formatos_op():
    # h1 reais da sonda (função reusada do DBZ — fixtures OP)
    d, c, e = split_myp_title("Sabo (OP13-004)")
    assert c == "OP13-004" and e == "Sabo"
    d, c, e = split_myp_title('Eustass"Captain"Kid (OP10-099)')
    assert c == "OP10-099" and e == 'Eustass"Captain"Kid'
    d, c, e = split_myp_title("[Pré-Venda] Edward.Newgate (001) (Alternate Art)")
    assert c is None                      # número curto NÃO é código completo
    assert e == "Edward.Newgate (001) (Alternate Art)"
    d, c, e = split_myp_title("Monkey.D.Dragon")
    assert c is None and e == "Monkey.D.Dragon"


# ── título de edição (concat real da sonda) ─────────────────────────────

def test_split_edition_title_casos_reais_op():
    cases = [
        ("Carrying On His WillOP1307/11/2025", "carrying-on-his-will",
         ("Carrying On His Will", "OP13")),
        ("Starter Deck 35: RED/BLACK SaboST-3531/07/2026",
         "starter-deck-35-redblack-sabo",
         ("Starter Deck 35: RED/BLACK Sabo", "ST-35")),
        ("Starter Deck EX: Gear 5ST2114/03/2025", "starter-deck-ex-gear-5",
         ("Starter Deck EX: Gear 5", "ST21")),   # dígito antes do código
        ("Premium Booster -The Best-PRB-0108/11/2024",
         "premium-booster-the-best",
         ("Premium Booster -The Best-", "PRB-01")),
        ("Kingdoms of IntrigueOP-0422/09/2023", "kingdoms-of-intrigue",
         ("Kingdoms of Intrigue", "OP-04")),
        ("Emperors in the New World: 2nd Anniversary Tournament CardsOP09ANN06/12/2024",
         "emperors-in-the-new-world-2nd-anniversary-tournament-cards",
         ("Emperors in the New World: 2nd Anniversary Tournament Cards",
          "OP09ANN")),
        ("Promotion CardsPR30/09/2022", "promotion-cards",
         ("Promotion Cards", "PR")),
        ("Gift Collection 2023GC01", "gift-collection-2023",   # sem data
         ("Gift Collection 2023", "GC01")),
        ("STARTER DECK -Zoro and Sanji-ST1215/03/2023",
         "starter-deck-zoro-and-sanji",
         ("STARTER DECK -Zoro and Sanji-", "ST12")),
    ]
    for raw, slug, want in cases:
        assert split_edition_title(raw, slug) == want, (raw, slug)


# ── resolução edição → grupo tcgcsv cat 68 ──────────────────────────────

def test_canonical_set_code_op_sem_aliases_dbz():
    assert canonical_set_code_op("ST-35") == "ST35" == canonical_set_code_op("ST35")
    assert canonical_set_code_op("OP-04") == "OP4" == canonical_set_code_op("OP04")
    assert canonical_set_code_op("EB-03") == "EB3" == canonical_set_code_op("EB03")
    assert canonical_set_code_op("PRB-01") == "PRB1"
    assert canonical_set_code_op("OP17 RE") == "OP17RE" == \
        canonical_set_code_op("OP17RE")
    assert canonical_set_code_op("OP09 ANN") == "OP09ANN".replace("09", "9")
    # ⚠️ NUNCA herdar os aliases do DBZ: em One Piece EB = Extra Booster
    assert canonical_set_code_op("EB-01") == "EB1" != "EX1"
    # abbr composta real do tcgcsv (OP15) não casa o padrão → fica crua
    assert canonical_set_code_op("OP15-EB04") == "OP15EB04"
    assert canonical_set_code_op("PR") == "PR"
    assert canonical_set_code_op(None) is None


def test_resolve_edition_group_op_cascata():
    # nomes/abbrs REAIS da cat 68 (sonda 2026-08-09)
    groups = [
        {"groupId": 3188, "name": "Romance Dawn", "abbreviation": "OP01"},
        {"groupId": 24753, "name": "Starter Deck 35: RED/BLACK Sabo",
         "abbreviation": "ST-35"},
        {"groupId": 23834, "name": "Extra Booster: Anime 25th Collection",
         "abbreviation": "EB-02"},
        {"groupId": 24637, "name": "Adventure on Kami's Island",
         "abbreviation": "OP15-EB04"},
        {"groupId": 24775, "name": "The World's Strongest Warriors Release Event Cards",
         "abbreviation": "OP17 RE"},
        {"groupId": 9999, "name": "Starter Deck 12: Zoro & Sanji",
         "abbreviation": "ST-12"},
        {"groupId": 8888, "name": "Extra Booster: Memorial Collection",
         "abbreviation": "EB-01"},
    ]
    # 1) nome exato
    assert resolve_edition_group_op("Romance Dawn", "OP01", groups) == 3188
    assert resolve_edition_group_op("Starter Deck 35: RED/BLACK Sabo",
                                    "ST-35", groups) == 24753
    # nome exato mesmo com abbr composta que não casa por código (OP15)
    assert resolve_edition_group_op("Adventure on Kami's Island",
                                    "OP15", groups) == 24637
    # 2) código canônico único (nome MYP ≠ nome tcgcsv)
    assert resolve_edition_group_op("STARTER DECK -Zoro and Sanji-",
                                    "ST12", groups) == 9999
    # 3) contenção de nome única
    assert resolve_edition_group_op("Memorial Collection", "EB01",
                                    groups) == 8888
    # sem match → None (nunca chuta)
    assert resolve_edition_group_op("Inexistente Total", "ZZ99", groups) is None
    assert resolve_edition_group_op("X", None, []) is None


# ── campo "Código" do produto ───────────────────────────────────────────

def test_parse_myp_op_product_code():
    # formatos REAIS da sonda — o token de edição pode ter hífen (st-35),
    # por isso o código da carta é o ÚLTIMO token com formato de carta
    assert parse_myp_op_product_code("one_op17_op17-001") == ("OP17-001", None)
    assert parse_myp_op_product_code("one_st-35_p-105") == ("P-105", None)
    assert parse_myp_op_product_code("one_st-35_st35-001") == ("ST35-001", None)
    assert parse_myp_op_product_code("one_st-35_op13-004") == ("OP13-004", None)
    assert parse_myp_op_product_code("one_op17_op17-020p1") == ("OP17-020", "p1")
    assert parse_myp_op_product_code("dbsm_bt1-073_spr") == (None, None)
    assert parse_myp_op_product_code("") == (None, None)
    assert parse_myp_op_product_code(None) == (None, None)


def test_extract_myp_op_product_code_html():
    html = """<html><body><h1>Sabo (OP13-004)</h1>
    <table><tr><td>Cor</td><td>Vermelho</td></tr>
    <tr><td>Código</td><td>one_st-35_op13-004</td></tr>
    <tr><td>Edição</td><td>Starter Deck 35: RED/BLACK Sabo (ST-35)</td></tr></table>
    </body></html>"""
    soup = BeautifulSoup(html, "lxml")
    assert extract_myp_op_product_code(soup) == "one_st-35_op13-004"
    soup2 = BeautifulSoup("<html><body><h1>X</h1></body></html>", "lxml")
    assert extract_myp_op_product_code(soup2) is None


# ── join determinístico (variant-aware pelo NOME) ───────────────────────

def test_join_alternate_art_casa_parallel():
    # colisão REAL do tcgcsv (OP01): base × "(Parallel)" no mesmo Number.
    # O MYP anuncia "(Alternate Art)" → equivale a "(Parallel)" no join.
    base = make_entry(1, "Roronoa Zoro (025)", 3188)
    par = make_entry(2, "Roronoa Zoro (025) (Parallel)", 3188)
    idx = {"by_code": {"OP1-25": [base, par]}, "by_group": {3188: [base, par]}}
    mt, how = find_reference_op("OP1-25", "Roronoa Zoro (025) (Alternate Art)",
                                idx, group_id=3188)
    assert mt is par and "nome-exato" in how
    # MYP vendendo o base → casa o base
    mt, how = find_reference_op("OP1-25", "Roronoa Zoro (025)", idx,
                                group_id=3188)
    assert mt is base and "nome-exato" in how


def test_join_variante_sem_produto_qualificado_fica_fora():
    # índice só tem o base → o "(Alternate Art)" do MYP NUNCA cai no base
    # (precificaria a variante errada — classe de bug nº 3 da frota)
    base = make_entry(1, "Roronoa Zoro (025)", 3188)
    idx = {"by_code": {"OP1-25": [base]}, "by_group": {3188: [base]}}
    mt, how = find_reference_op("OP1-25", "Roronoa Zoro (025) (Alternate Art)",
                                idx, group_id=3188)
    assert mt is None and "variante sem produto qualificado" in how


def test_join_base_nunca_pega_ref_de_variante():
    par = make_entry(2, "Roronoa Zoro (025) (Parallel)", 3188)
    idx = {"by_code": {"OP1-25": [par]}, "by_group": {3188: [par]}}
    mt, how = find_reference_op("OP1-25", "Roronoa Zoro (025)", idx,
                                group_id=3188)
    assert mt is None and "variante qualificada" in how


def test_join_box_topper_nao_colide_com_base():
    # colisão real OP01-008: 'Cavendish' × 'Cavendish (Box Topper)'
    base = make_entry(1, "Cavendish", 3188, number="OP01-008")
    bt = make_entry(2, "Cavendish (Box Topper)", 3188, number="OP01-008")
    idx = {"by_code": {"OP1-8": [base, bt]}, "by_group": {3188: [base, bt]}}
    mt, _ = find_reference_op("OP1-8", "Cavendish", idx, group_id=3188)
    assert mt is base
    mt, _ = find_reference_op("OP1-8", "Cavendish (Box Topper)", idx,
                              group_id=3188)
    assert mt is bt


def test_join_reprint_starter_deck_escopo_grupo_edicao():
    # Sabo OP13-004 existe no booster OP13 E reimpresso no ST-35 (provado):
    # com o grupo da edição resolvido, casa o produto (e preço) da versão
    # CERTA; sem grupo, 2 candidatos não-especiais homônimos → recusa
    op13 = make_entry(1, "Sabo (004)", 24303, number="OP13-004")
    st35 = make_entry(2, "Sabo (004)", 24753, number="OP13-004")
    idx = {"by_code": {"OP13-4": [op13, st35]},
           "by_group": {24303: [op13], 24753: [st35]}}
    mt, how = find_reference_op("OP13-4", "Sabo", idx, group_id=24753)
    assert mt is st35 and "grupo-edição" in how
    mt, how = find_reference_op("OP13-4", "Sabo", idx, group_id=24303)
    assert mt is op13 and "grupo-edição" in how
    mt, how = find_reference_op("OP13-4", "Sabo", idx, group_id=None)
    assert mt is None and "ambíguo" in how


def test_join_escopo_grupo_principal_vs_release_event():
    # número duplicado entre set principal e Release Event (grupo especial):
    # sem grupo resolvido, o escopo "grupo-principal" evita o carimbado
    main_ = make_entry(1, "Monkey.D.Luffy (003)", 100, number="OP15-003")
    re_ = make_entry(2, "Monkey.D.Luffy (003)", 200, number="OP15-003",
                     special=True)
    idx = {"by_code": {"OP15-3": [main_, re_]},
           "by_group": {100: [main_], 200: [re_]}}
    mt, how = find_reference_op("OP15-3", "Monkey.D.Luffy", idx, group_id=None)
    assert mt is main_ and "grupo-principal" in how
    # com grupo resolvido (edição Release Event do MYP) → escopo do grupo
    mt, how = find_reference_op("OP15-3", "Monkey.D.Luffy", idx, group_id=200)
    assert mt is re_ and "grupo-edição" in how


def test_join_sem_codigo_grupo_nome():
    base = make_entry(1, "Monkey.D.Dragon", 24753, number="OP13-017")
    idx = {"by_code": {}, "by_group": {24753: [base]}}
    mt, how = find_reference_op(None, "Monkey.D.Dragon", idx, group_id=24753)
    assert mt is base and how == "grupo+nome-exato"
    mt, how = find_reference_op(None, "Monkey.D.Dragon", idx, group_id=None)
    assert mt is None and "não resolvido" in how


def test_join_numero_fora_do_indice():
    idx = {"by_code": {}, "by_group": {}}
    mt, how = find_reference_op("ZZ9-999", "Nada", idx, group_id=None)
    assert mt is None and "fora do índice" in how


# ── parser de sellers herdado (NM/EN) na página OP ──────────────────────

PRODUCT_HTML = """<html><body>
<h1>Roronoa Zoro (OP01-025)</h1>
<table><tr><td>Código</td><td>one_op01_op01-025</td></tr></table>
<table class="table-striped table-bordered">
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td class="estoque-lista-nomeenfoil">Foil</td><td>R$ 100,00</td></tr>
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">SP - Pouco jogada</td>
    <td>R$ 80,00</td></tr>
<tr><td><span class="flag-icon" title="Português"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td>R$ 60,00</td></tr>
<tr><td><span class="flag-icon" title="Inglês"></span></td>
    <td class="estoque-lista-qualidadenome">NM - Quase nova</td>
    <td>R$ 120,00</td></tr>
</table></body></html>"""


def _scraper():
    return MYPOpScraper(delay=0, min_en_sellers=2, threshold=0.30,
                        min_price=50.0, tcg_cache_dir=None)


def test_parse_seller_table_herdado_en_nm():
    sc = _scraper()
    soup = BeautifulSoup(PRODUCT_HTML, "lxml")
    table = soup.select("table.table-striped.table-bordered")[0]
    st = sc._parse_seller_table(table)
    # EN NM = R$100 e R$120 (SP fora, PT fora)
    assert sorted(st["en_prices"]) == [100.0, 120.0]
    assert st["en"] == 2


def test_scrape_op_product_end_to_end():
    sc = _scraper()
    soup = BeautifulSoup(PRODUCT_HTML, "lxml")
    sc._get = lambda url, save_debug=False: soup     # offline: sem rede
    base = make_entry(1, "Roronoa Zoro (025)", 3188,
                      prices={"Normal": 40.0}, lows={"Normal": 38.0})
    par = make_entry(2, "Roronoa Zoro (025) (Parallel)", 3188,
                     prices={"Normal": 400.0})
    sc.op_index = {"by_code": {"OP1-25": [base, par]},
                   "by_group": {3188: [base, par]}, "groups": []}
    ed = {"title": "Romance Dawn", "section": "onepiece"}
    card = sc.scrape_op_product("https://mypcards.com/onepiece/produto/1/x",
                                ed, 3188, fx_usd_brl=5.0)
    assert isinstance(card, OpCardData), card
    # h1 sem qualificador → casa o BASE (nunca o Parallel de US$400)
    assert card.tcg_usd == 40.0 and "nome-exato" in card.join_via
    assert card.myp_lowest_en_nm == 100.0 and card.en_nm_sellers == 2
    assert abs(card.margin_pct - 1.0) < 1e-9        # (200−100)/100
    assert card.code == "OP01-025"
    assert card.flag_lixo is False                   # 100 = 50% de 200: < estrito
    assert card.single_en_seller_risk is False


def test_scrape_op_product_alternate_art_casa_parallel():
    html = PRODUCT_HTML.replace(
        "<h1>Roronoa Zoro (OP01-025)</h1>",
        "<h1>Roronoa Zoro (025) (Alternate Art)</h1>")
    sc = _scraper()
    sc._get = lambda url, save_debug=False: BeautifulSoup(html, "lxml")
    base = make_entry(1, "Roronoa Zoro (025)", 3188, prices={"Normal": 1.0})
    par = make_entry(2, "Roronoa Zoro (025) (Parallel)", 3188,
                     prices={"Normal": 400.0}, lows={"Normal": 380.0})
    sc.op_index = {"by_code": {"OP1-25": [base, par]},
                   "by_group": {3188: [base, par]}, "groups": []}
    ed = {"title": "Romance Dawn", "section": "onepiece"}
    card = sc.scrape_op_product("u", ed, 3188, fx_usd_brl=5.0)
    assert isinstance(card, OpCardData), card
    # o "(Alternate Art)" casou o produto (Parallel), nunca o base de US$1
    assert card.tcg_usd == 400.0
    assert "Parallel" in card.tcg_product_name


def test_scrape_op_product_sem_ref_vai_pra_semref():
    sc = _scraper()
    soup = BeautifulSoup(PRODUCT_HTML.replace("one_op01_op01-025",
                                              "one_op01_zz9-999"), "lxml")
    sc._get = lambda url, save_debug=False: soup
    sc.op_index = {"by_code": {}, "by_group": {}, "groups": []}
    ed = {"title": "Romance Dawn", "section": "onepiece"}
    row = sc.scrape_op_product("https://mypcards.com/onepiece/produto/1/x",
                               ed, None, fx_usd_brl=5.0)
    assert isinstance(row, OpSemRefRow)
    assert row.myp_lowest_en_nm == 100.0
    assert "fora do índice" in row.motivo


def test_scrape_op_product_abaixo_do_piso():
    sc = _scraper()
    html_baixo = PRODUCT_HTML.replace("R$ 100,00", "R$ 10,00").replace(
        "R$ 120,00", "R$ 12,00")
    sc._get = lambda url, save_debug=False: BeautifulSoup(html_baixo, "lxml")
    sc.op_index = {"by_code": {}, "by_group": {}, "groups": []}
    ed = {"title": "Romance Dawn", "section": "onepiece"}
    assert sc.scrape_op_product("u", ed, None, 5.0) is None
    assert sc._stats["skipped_low_price"] == 1


# ── XLSX + entrega (myp_op_summary) ─────────────────────────────────────

def _build_xlsx(tmpdir: Path) -> Path:
    clean = make_card()                                   # 100% margem, limpo
    lixo = make_card(name="Shanks - OP01-120", en_name="Shanks",
                     code="OP01-120", myp_lowest_en_nm=60.0, tcg_usd=40.0,
                     tcg_brl=200.0, margin_pct=(200 - 60) / 60,
                     margin_brl=140.0, flag_lixo=True)
    volatil = make_card(name="Nami - OP01-016", en_name="Nami",
                        code="OP01-016", tcg_low_usd=2000.0, tcg_usd=506.0,
                        tcg_brl=2530.0, myp_lowest_en_nm=1000.0,
                        margin_pct=1.53, margin_brl=1530.0,
                        flag_ref_volatil=True)
    abaixo = make_card(name="Usopp - OP01-002", en_name="Usopp",
                       code="OP01-002", margin_pct=0.10, margin_brl=20.0)
    trunc = make_card(name="Sanji - OP01-013", en_name="Sanji",
                      code="OP01-013", margin_pct=0.55,
                      en_truncation_risk=True)
    semref = [OpSemRefRow(name="Buggy - OP09-999", en_name="Buggy",
                          code="OP09-999", edition="Emperors in the New World",
                          section="onepiece",
                          product_url="https://mypcards.com/onepiece/produto/9/x",
                          myp_lowest_en_nm=80.0, en_nm_sellers=2,
                          motivo="número fora do índice tcgcsv")]
    out = tmpdir / "op_test.xlsx"
    generate_op_xlsx([clean, lixo, volatil, abaixo, trunc], semref,
                     str(out), threshold=0.30, fx=5.0,
                     stats={"products_scanned": 5, "op_ref_ok": 5},
                     editions_scanned=2, min_price=50.0)
    return out


def test_xlsx_estrutura_e_threshold():
    with tempfile.TemporaryDirectory() as td:
        out = _build_xlsx(Path(td))
        wb = load_workbook(out, data_only=True)
        assert set(wb.sheetnames) == {"All EN Cards", "Sem Ref TCG", "Summary"}
        summary = {str(r[0]): r[1] for r in wb["Summary"].iter_rows(values_only=True)}
        assert summary["Margin Threshold"] == "30%"
        assert summary["Total EN Cards"] == 5
        assert summary["Sem Ref TCG"] == 1
        # Deals ≥30%: clean, lixo, volatil, trunc (0.55) — abaixo (0.10) não
        assert summary["Deals Found"] == 4
        headers = [c.value for c in wb["All EN Cards"][1]]
        for h in ("Card Name", "Código", "TCG US$", "TCG Source", "Margin %",
                  "URL", "TCG URL", "⚠️ Possível Lixo", "⚠️ Ref Volátil"):
            assert h in headers, h
        wb.close()


def test_summary_markdown_buckets_e_links():
    with tempfile.TemporaryDirectory() as td:
        out = _build_xlsx(Path(td))
        md_path = Path(td) / "op.md"
        rc = myp_op_summary.build_markdown(str(out), str(md_path), run_id="123")
        assert rc == 0
        md = md_path.read_text(encoding="utf-8")
        assert "# MYP Scan ONE PIECE" in md
        assert "## 🟢 Top 50 deals limpos" in md
        assert "## 🚨 REVISAR" in md
        assert "## ⚠️ Sem referência TCG" in md
        assert "## 🚨 EN truncation risk" in md
        # 2 links em linha de deal (contrato de entrega do repo)
        clean_line = next(l for l in md.splitlines()
                          if "Roronoa Zoro" in l and "**" in l)
        assert "[oferta](" in clean_line and "[TCG](" in clean_line
        # flag por linha no REVISAR
        assert "possível lixo" in md
        assert "ref volátil" in md
        # linha sem ref traz o motivo + link de BUSCA TCG (nunca URL inventada)
        semref_line = next(l for l in md.splitlines() if "Buggy" in l)
        assert "número fora do índice" in semref_line
        assert "[oferta](" in semref_line
        assert "[TCG](https://www.tcgplayer.com/search/" in semref_line
        assert "**100.0%**" in md
        assert "Cobertura de referência real" in md
        # referência declarada = cat 68 (One Piece)
        assert "cat 68" in md
        # deal 10% NÃO aparece como deal
        assert not any("Usopp" in l and "**" in l for l in md.splitlines())


def test_summary_threshold_vem_do_xlsx():
    with tempfile.TemporaryDirectory() as td:
        card45 = make_card(margin_pct=0.45, margin_brl=45.0)
        out = Path(td) / "t50.xlsx"
        generate_op_xlsx([card45], [], str(out), threshold=0.50, fx=5.0)
        md_path = Path(td) / "t50.md"
        myp_op_summary.build_markdown(str(out), str(md_path))
        md = md_path.read_text(encoding="utf-8")
        assert "Deals (≥50%):** 0" in md


def test_carta_label_e_links_helpers():
    assert myp_op_summary.carta_label("Roronoa Zoro", "OP01-025") == \
        "Roronoa Zoro OP01-025"
    assert myp_op_summary.carta_label("Sabo - OP13-004", "OP13-004") == \
        "Sabo - OP13-004"
    links = myp_op_summary.delivery_links("https://m/x", "https://t/y")
    assert links == "[oferta](https://m/x) · [TCG](https://t/y)"
    links = myp_op_summary.delivery_links("https://m/x", None, "Buggy", "OP09-999")
    assert "[oferta](https://m/x)" in links and "search" in links


def test_search_url():
    u = tcg_op_search_url("Roronoa Zoro (025) (Parallel)", "OP01-025")
    from urllib.parse import urlparse
    assert u is not None
    assert urlparse(u).netloc == "www.tcgplayer.com"
    assert "Roronoa+Zoro" in u
    assert tcg_op_search_url("", None) is None


def test_normalize_code_formatos_op():
    # helper reusado do DBZ — travar os formatos One Piece nele também
    assert normalize_code("OP01-001") == "OP1-1"
    assert normalize_code("OP13-004") == "OP13-4"
    assert normalize_code("ST35-001") == "ST35-1"
    assert normalize_code("P-105") == "P-105"
    assert normalize_code("EB01-006") == "EB1-6"


# ── runner standalone (mesmo padrão do test_v5_8_offline.py) ────────────

def main() -> int:
    tests = [(n, f) for n, f in sorted(globals().items())
             if n.startswith("test_") and callable(f)]
    failed = 0
    for name, fn in tests:
        try:
            fn()
            print(f"  ✓ {name}")
        except AssertionError as e:
            failed += 1
            print(f"  ✗ {name}: {e}")
    print("═" * 50)
    print(f"{len(tests) - failed}/{len(tests)} passaram")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
