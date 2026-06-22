"""Gera markdown summary a partir de XLSX consolidado do MYP scanner.

Output otimizado pra leitura rápida em GitHub UI (renderiza markdown nativo)
e em Obsidian (frontmatter + tags + tabela). Foca em:
  - Top 10 deals limpos (sem flag SIR/HR/SAR)
  - Top deals com flag SIR (sinaliza pra validação manual)
  - Stats do scan (editions, EN cards, deals encontrados)

Uso:
    python myp_summary.py path/to/myp_arbitrage.xlsx \\
        --output results/daily-2026-05-15.md \\
        --type daily
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

if sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from openpyxl import load_workbook

# v5.11.1 (2026-06-09): reaproveita os helpers de URL do scanner pra montar o
# link DIRETO do TCGplayer (via redirect pokemontcg.io) na tabela de ENTREGA.
# Mesma lógica do XLSX (generate_xlsx) — direct link quando a edição é mapeada
# e o collector# está in-range; senão cai pra busca-por-nome. Import tolerante:
# se o módulo do scanner não estiver importável (ex.: teste isolado), o link
# TCG some mas o resto da tabela segue.
try:
    from myp_arbitrage_scanner import tcg_direct_url, tcg_search_url
except Exception:  # pragma: no cover - fallback defensivo
    def tcg_direct_url(name, edition, oversized_collector_risk=False):
        return None

    def tcg_search_url(name):
        return None


def fmt_usd(v) -> str:
    """Formata valor USD pra display (US$1,234.56). '—' se ausente."""
    if v is None:
        return "—"
    try:
        return f"US${float(v):,.2f}"
    except (ValueError, TypeError):
        return "—"


def split_card_name(name: str | None) -> tuple[str, str]:
    """Separa 'Pikachu (173/165)' → ('Pikachu', '173/165').

    Retorna (nome_base, numero). Numero vazio se o nome não embute (NNN/MMM).
    A coluna `Carta` da entrega junta como 'Pikachu 173/165' (sem duplicar o
    número quando já está no nome)."""
    if not name:
        return ("", "")
    import re
    m = re.search(r"\((\d+/\d+)\)\s*$", name)
    if not m:
        return (name.strip(), "")
    base = name[: m.start()].strip()
    return (base, m.group(1))


def carta_label(name: str | None) -> str:
    """Coluna `Carta` = nome + número numa string só ('Pikachu 173/165').

    Se o nome não embute número, retorna só o nome (sem duplicar)."""
    base, num = split_card_name(name)
    if num and num not in base:
        return f"{base} {num}"
    return base or (name or "").strip()


def delivery_links(myp_url: str | None, name: str | None, edition: str | None,
                   oversized: bool = False, tcg_url: str | None = None) -> str:
    """Coluna `Links`: '[oferta](myp_url) · [TCG](tcg_url)'.

    - oferta → página do produto no MYP (validação do preço/seller).
    - TCG → produto/busca TCGplayer (workflow manual de conferir preço NM).
    Emite só os links que existirem; '—' se nenhum.

    v5.11.4: aceita `tcg_url` explícito (coluna `TCG URL` do XLSX, plain-text
    desde v5.11.2) e o prefere sobre o recompute via import do scanner — assim a
    entrega usa o MESMO link que o XLSX já carrega, e funciona mesmo quando
    `myp_arbitrage_scanner` não é importável (teste isolado / env sem o módulo)."""
    parts = []
    if myp_url:
        parts.append(f"[oferta]({myp_url})")
    tcg = tcg_url or (
        tcg_direct_url(name or "", edition or "", oversized_collector_risk=oversized)
        or tcg_search_url(name or "")
    )
    if tcg:
        parts.append(f"[TCG]({tcg})")
    return " · ".join(parts) if parts else "—"


def is_supranumerary(name: str | None) -> bool:
    """Detecta se card_num > set_total no nome '(N/M)'."""
    import re
    if not name:
        return False
    m = re.search(r"\((\d+)/(\d+)\)", name)
    if not m:
        return False
    try:
        return int(m.group(1)) > int(m.group(2))
    except (ValueError, TypeError):
        return False


def is_rarity_mislabel(name: str | None, rarity: str | None) -> bool:
    """Supranumerário ('(N/M)' com N>M) E raridade EXATAMENTE 'Comum'.

    Operador 2026-06-19: uma carta supranumerária costuma ser REAL — a MYP é que
    às vezes erra a RARIDADE, marcando uma SIR/SAR/ex real como 'Comum'. Esse sinal
    significa "não confie no rótulo 'Comum', valide", NÃO "deal falso". Uma carta
    supranumerária com raridade real (Rara/Hiper/etc.) é normal e NÃO é flagada.
    Match EXATO em 'Comum' (lição NM-only: nunca substring).
    """
    return is_supranumerary(name) and (str(rarity or "").strip() == "Comum")


def fmt_brl(v) -> str:
    """Formata valor BRL pra display."""
    if v is None:
        return "—"
    try:
        return f"R${float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "—"


def fmt_pct(v) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v) * 100:.1f}%"
    except (ValueError, TypeError):
        return "—"


def build_markdown(xlsx: str, output: str, scan_type: str,
                   run_id: str = "",
                   repo: str = "matheuscllm-lgtm/myp-arbitrage-scanner") -> int:
    """Gera o markdown de ENTREGA a partir do XLSX consolidado.

    Extraído do antigo main() (v5.11.1) pra ser testável sem argv/subprocess.
    `scan_type` ∈ {"daily","weekly"}."""
    # nomes legados usados no corpo (mantidos pra diff mínimo)
    class _A:  # namespace leve, evita reescrever args.* abaixo
        pass
    args = _A()
    args.xlsx = xlsx
    args.output = output
    args.type = scan_type
    args.run_id = run_id
    args.repo = repo

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: XLSX não encontrado: {xlsx_path}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Stats do Summary sheet
    summary_data = {}
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                summary_data[str(row[0]).strip().rstrip(":")] = row[1]

    # All EN Cards (fonte mais rica que só Deals)
    all_cards = []
    if "All EN Cards" in wb.sheetnames:
        ws = wb["All EN Cards"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = list(rows[0])
            for r in rows[1:]:
                rec = dict(zip(headers, r))
                if rec.get("Card Name"):
                    all_cards.append(rec)

    # v5.11.1: solta o handle do XLSX assim que a extração termina (Windows
    # segura o arquivo em read_only até fechar — relevante p/ testes/uso in-process).
    wb.close()

    # v5.14.1 (fix): o piso de "deal" no summary tem de casar o threshold REAL do
    # scan (lido do XLSX `Margin Threshold`), NÃO um 0.25 hardcoded legado (o
    # default era 25% até v5.10; hoje é 30%). Sem isso, cards na banda 25–30%
    # vazavam para o balde de "deals limpos" (Top-50) e — pior, com o novo
    # `deals_clarif` desta versão — eram impressos como "deals limpos (≥30%)",
    # uma afirmação falsa (carta de 27% rotulada como ≥30%). XLSX antigo sem o
    # campo cai no default 0.25 (preserva comportamento histórico).
    _thr_raw = summary_data.get("Margin Threshold", "25%")
    try:
        deal_floor = float(str(_thr_raw).strip().rstrip("%")) / 100.0
    except (TypeError, ValueError):
        deal_floor = 0.25
    deals = [c for c in all_cards if c.get("Margin %") and c["Margin %"] >= deal_floor]
    deals_sorted = sorted(deals, key=lambda c: c.get("Margin %") or 0, reverse=True)

    # v5.8 (2026-05-16): Top 15 "limpos" exclui agora 3 buckets: supranumerários,
    # truncation-risk E tcg_suspect (Jirachi PR-SM_SM161 caso). Antes Jirachi
    # aparecia como #1 em latest-weekly.md mesmo com TCG inflado 75x.
    def _is_suspect(c) -> bool:
        return bool(c.get("⚠️ TCG Suspect"))

    def _is_rarity_mislabel(c) -> bool:
        return is_rarity_mislabel(c.get("Card Name"), c.get("Rarity"))

    # v5.14.3: "o preço é de verdade?" — fonte canônica = coluna "TCG Source"
    # (real `pokemontcg.io`/`tcgcsv` vs fallback `.estat-tcg`); XLSX antigo
    # (pré-v5.14, sem a coluna) infere pela presença de "TCG US$" (que só o preço
    # real preenche). Definido AQUI (e não mais abaixo, na cobertura) porque agora
    # gateia também `deals_clean`. Definição ÚNICA — não duplicar.
    # v5.15: tcgcsv.com é a SEGUNDA fonte de preço REAL (a que funciona no CI). O
    # rótulo `real (tcgcsv)` NÃO contém "pokemontcg" → precisa ser reconhecido
    # explicitamente como real, senão o gate de honestidade o trataria como
    # fallback e jogaria deals reais do CI pro balde "validar manualmente".
    def _is_real(c) -> bool:
        src = c.get("TCG Source")
        if src is not None and str(src).strip() != "":
            s = str(src).lower()
            return "pokemontcg" in s or "tcgcsv" in s
        return c.get("TCG US$") not in (None, "", "—")

    # v5.14.3 (fix BLOCKER de honestidade): um "deal limpo" precisa de preço REAL.
    # Um preço FALLBACK (.estat-tcg) tem margem NÃO-confiável por definição (o
    # próprio v5.11 nasceu pra não confiar nele). Um fallback inflado SEM última
    # venda (o gate de `tcg_suspect` é pulado quando não há última venda) vazava
    # pro Top-50 como compra "limpa" com margem ilusória (caso Darumaka: .estat-tcg
    # R$2867 vs MYP R$60 → 4678%). Agora fallback sai do balde limpo e vai pro
    # balde dedicado `deals_fallback` ("validar manualmente"). Regra dura do
    # CLAUDE.md: "Nunca trate fallback como real".
    deals_clean = [
        c for c in deals_sorted
        if not _is_rarity_mislabel(c) and not _is_suspect(c) and _is_real(c)
    ]
    deals_fallback = [
        c for c in deals_sorted
        if not _is_rarity_mislabel(c) and not _is_suspect(c) and not _is_real(c)
    ]
    deals_supranum = [c for c in deals_sorted if _is_rarity_mislabel(c)]
    deals_suspect = [c for c in deals_sorted if _is_suspect(c)]

    truncations = [c for c in all_cards if c.get("⚠️ EN Trunc")]

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    scan_type_label = "Daily Quick" if args.type == "daily" else "Weekly Full"

    # ── Build markdown ──
    lines = []
    lines.append(f"---")
    lines.append(f"tags: [tcg, scanner, myp, arbitrage, scan-{args.type}]")
    lines.append(f"date: {today}")
    lines.append(f"type: {args.type}")
    lines.append(f"source: GH Actions run {args.run_id}" if args.run_id else "source: local scan")
    lines.append(f"---")
    lines.append("")
    lines.append(f"# MYP Scan {scan_type_label} — {today}")
    lines.append("")

    # Stats line
    total = summary_data.get("Total EN Cards", len(all_cards))
    deals_n = summary_data.get("Deals Found (clean)", summary_data.get("Deals Found", len(deals)))
    threshold = summary_data.get("Margin Threshold", "25%")
    lines.append(f"**Cards EN escaneados:** {total} | **Deals (≥{threshold}):** {deals_n} | "
                 f"**Limpos (preço real):** {len(deals_clean)} | "
                 f"**Fallback:** {len(deals_fallback)} | "
                 f"**🚨 TCG suspects:** {len(deals_suspect)} | "
                 f"**Truncation:** {len(truncations)}")
    lines.append("")

    # ── Sinal de honestidade: cobertura de preço REAL (v5.14, corrigido v5.14.1) ──
    # Quantas cartas EN têm preço TCGplayer REAL (pokemontcg.io) vs FALLBACK
    # (.estat-tcg, margem NÃO-confiável). Torna VISÍVEL a degradação silenciosa
    # do CI (runners não alcançam a pokemontcg.io → tudo fallback). A fonte é a
    # coluna "TCG Source" (canônica); fallback p/ presença de "TCG US$" em XLSX
    # antigos (pré-v5.14, sem a coluna).
    #
    # v5.14.1: a cobertura é medida sobre o UNIVERSO de cartas EN (aba
    # `All EN Cards`), NÃO sobre o subconjunto de deals ≥threshold. O bug anterior
    # contava só `deals_clean`: quando 0 cartas batiam o threshold (mas o catálogo
    # inteiro tinha preço real), o resumo imprimia "✅ 0/0" ou, pior, "🛑 ZERO"
    # falso — fazendo o operador achar que a key falhou quando estava tudo certo.
    # Cobertura ("o preço usado é de verdade?") e deals ("a margem bate 30%?") são
    # dois números distintos; agora cada um vem do seu universo correto.
    # (`_is_real` é definido acima, junto da partição de deals — definição única.)

    # Universo = cartas EN que de fato têm ALGUM preço TCG (real ou fallback). Uma
    # carta sem nenhum preço TCG não entra no denominador de cobertura (não há o
    # que ser "real" ou "fallback"); seria ruído contar como se a key tivesse
    # falhado nela.
    priced = [c for c in all_cards
              if _is_real(c) or c.get("TCG Player (R$)") not in (None, "", "—")
              or c.get("TCG US$") not in (None, "", "—")]
    real_n = sum(1 for c in priced if _is_real(c))
    total_priced = len(priced)
    fb_n = total_priced - real_n

    # v5.14.3: `deals_clean` já é 100% preço real por construção (o fallback foi
    # pro balde `deals_fallback`), então não há mais "X de N com preço real" a
    # reportar. O esclarecimento agora diz quantos deals limpos (preço real) há e,
    # se houver, quantos ficaram só com preço fallback (em balde à parte).
    deals_clarif = (f" — {len(deals_clean)} deal(s) limpo(s) ≥{threshold} (preço real)"
                    if deals_clean else
                    f" — nenhum deal limpo ≥{threshold} nesta run")
    if deals_fallback:
        deals_clarif += (f"; {len(deals_fallback)} deal(s) só com preço fallback "
                         f"`.estat-tcg` (margem NÃO-confiável — em balde à parte)")
    deals_clarif += "."

    if total_priced == 0:
        cov_note = ("⚠️ **Sem preço TCG** — nenhuma carta EN com preço TCGplayer "
                    "(real ou fallback). Nada a reportar de cobertura.")
    elif real_n == 0:
        cov_note = (f"🛑 **ZERO preço real** — 0/{total_priced} cartas EN com preço "
                    f"REAL (pokemontcg.io); todas em fallback `.estat-tcg` (margens "
                    f"NÃO-confiáveis). Provável run em runner do GitHub (que não "
                    f"alcança a pokemontcg.io): enriqueça LOCAL com `myp_enrich.py` "
                    f"antes de operar.{deals_clarif}")
    elif fb_n:
        cov_note = (f"⚠️ **{real_n}/{total_priced} cartas EN com preço REAL** "
                    f"(pokemontcg.io); {fb_n} em fallback `.estat-tcg` (margem "
                    f"NÃO-confiável — validar manual ou enriquecer com "
                    f"`myp_enrich.py`).{deals_clarif}")
    else:
        cov_note = (f"✅ **{real_n}/{total_priced} cartas EN com preço REAL** "
                    f"(pokemontcg.io).{deals_clarif}")
    lines.append(f"**Cobertura de preço TCG real:** {cov_note}")
    lines.append("")

    if args.run_id:
        lines.append(f"**Artifact XLSX:** [`myp-{args.type}-consolidated-{args.run_id}`](https://github.com/{args.repo}/actions/runs/{args.run_id})")
        lines.append("")

    # ── Top 50 deals limpos — FORMATO DE ENTREGA (links clicáveis) ──
    # v5.11.1 (2026-06-09): formato aprovado pelo operador (espelha a entrega do
    # scanner COMC). Colunas:
    #   # | Margem % | MYP R$ | TCG US$ | Dif | Carta | Set | Raridade | Cond | Qtd | Links
    # - Carta = nome + número numa coluna só ('Pikachu 173/165'), sem duplicar.
    # - TCG US$ = preço REAL do TCGplayer em USD (pokemontcg.io). '—' onde só
    #   houve fallback .estat-tcg (sem USD real).
    # - Dif = lucro bruto em R$ (Diff (R$) = TCG R$ − MYP R$); margem segue BRUTA.
    # - Cond = NM (invariante NM-only).
    # - Qtd = nº de ofertas EN-NM (NM Sellers) — quantos lotes o operador pode
    #   comprar; o scanner não captura estoque por seller, então é a contagem.
    # - Links = [oferta](MYP) · [TCG](TCGplayer) — clicáveis; TCG p/ validação NM.
    lines.append("## 🟢 Top 50 deals limpos (sem flag SIR/HR/SAR)")
    lines.append("")
    if not deals_clean:
        lines.append("> Nenhum deal limpo nesta run.")
    else:
        lines.append("| # | Margem % | MYP R$ | TCG US$ | Dif | Carta | Set | Raridade | Cond | Qtd | Links |")
        lines.append("|---|---:|---:|---:|---:|---|---|---|---|---:|---|")
        for i, c in enumerate(deals_clean[:50], 1):
            name = c.get("Card Name")
            carta = carta_label(name)
            ed = (c.get("Edition") or "").strip()
            rarity = (c.get("Rarity") or "").strip() or "—"
            myp = fmt_brl(c.get("MYP EN NM (R$)"))
            tcg_usd = fmt_usd(c.get("TCG US$"))
            margin = fmt_pct(c.get("Margin %"))
            diff = fmt_brl(c.get("Diff (R$)"))
            qty = c.get("NM Sellers") or 0
            links = delivery_links(
                c.get("URL"), name, ed,
                oversized=bool(c.get("⚠️ COLLECTOR#")),
                tcg_url=c.get("TCG URL"),
            )
            lines.append(
                f"| {i} | **{margin}** | {myp} | {tcg_usd} | {diff} | "
                f"{carta} | {ed} | {rarity} | NM | {qty} | {links} |"
            )
    lines.append("")

    # ── Deals com flag SIR (alto risco) ──
    lines.append("## ⚠️ Deals com raridade suspeita (supranumerário + 'Comum' — validar)")
    lines.append("")
    lines.append("> `card_num > set_total` **com raridade 'Comum'** = a MYP provavelmente errou a "
                 "RARIDADE (a carta costuma ser uma SIR/SAR/ex real marcada 'Comum'). **A carta/deal "
                 "pode ser REAL** — o que NÃO se deve confiar é no rótulo 'Comum' (afeta valor/tier). "
                 "Valide a raridade e o preço (Link TCG) antes de operar; não trate como falso "
                 "automaticamente. Supranumerário com raridade real (Rara/Hiper/etc.) fica no bucket limpo.")
    lines.append("")
    if not deals_supranum:
        lines.append("> Nenhum deal com flag supranumerário nesta run.")
    else:
        # v5.11.4: Carta (nome+número via carta_label) + Links (oferta MYP · TCG)
        # no formato canônico — esses deals são a MAIOR PARTE da entrega do
        # operador e antes saíam sem links clicáveis.
        lines.append("| # | Carta | Edição | MYP R$ | TCG R$ | Margem (suspeita) | Links |")
        lines.append("|---|---|---|---:|---:|---:|---|")
        for i, c in enumerate(deals_supranum[:50], 1):
            name = c.get("Card Name")
            carta = carta_label(name)
            ed = (c.get("Edition") or "")[:30]
            myp = fmt_brl(c.get("MYP EN NM (R$)"))
            tcg = fmt_brl(c.get("TCG Player (R$)"))
            margin = fmt_pct(c.get("Margin %"))
            links = delivery_links(
                c.get("URL"), name, (c.get("Edition") or "").strip(),
                oversized=bool(c.get("⚠️ COLLECTOR#")),
                tcg_url=c.get("TCG URL"),
            )
            lines.append(f"| {i} | {carta} | {ed} | {myp} | {tcg} | {margin} | {links} |")
    lines.append("")

    # ── TCG Suspect (v5.8 — MYP infla .estat-tcg) ──
    if deals_suspect:
        lines.append("## 🚨 TCG Suspect (campo .estat-tcg inflado pelo MYP)")
        lines.append("")
        lines.append("> Cards onde TCG declarado pelo MYP é ≥10x a última venda real "
                     "do próprio MYP. Provável bug do `.estat-tcg`. Caso Jirachi "
                     "PR-SM_SM161: MYP declarava R$1499 vs última venda R$19,99 (75x). "
                     "Margens absurdas aqui são quase certamente artefato. **Já excluídos "
                     "do Top 15 limpos**, listados aqui pra auditoria.")
        lines.append("")
        # v5.11.4: Carta + Links também no balde de suspeitos (auditoria do
        # operador precisa do link de oferta MYP + TCG pra validar manualmente).
        lines.append("| # | Carta | Edição | MYP R$ | TCG decl R$ | Última venda R$ | Margem (fake) | Links |")
        lines.append("|---|---|---|---:|---:|---:|---:|---|")
        for i, c in enumerate(deals_suspect[:50], 1):
            name = c.get("Card Name")
            carta = carta_label(name)
            ed = (c.get("Edition") or "")[:30]
            myp = fmt_brl(c.get("MYP EN NM (R$)"))
            tcg = fmt_brl(c.get("TCG Player (R$)"))
            last = fmt_brl(c.get("MYP Last Sale (R$)"))
            margin = fmt_pct(c.get("Margin %"))
            links = delivery_links(
                c.get("URL"), name, (c.get("Edition") or "").strip(),
                oversized=bool(c.get("⚠️ COLLECTOR#")),
                tcg_url=c.get("TCG URL"),
            )
            lines.append(f"| {i} | {carta} | {ed} | {myp} | {tcg} | {last} | {margin} | {links} |")
        lines.append("")

    # ── Deals com preço FALLBACK (v5.14.3 — margem NÃO-confiável) ──
    # Deals ≥threshold cujo preço TCG veio do FALLBACK `.estat-tcg` (não do preço
    # real pokemontcg.io). A margem aqui pode ser ILUSÓRIA — o `.estat-tcg` às
    # vezes mapeia a carta errada e infla o "preço TCG" (caso Darumaka). Saem do
    # balde limpo de propósito; ficam aqui pra validação manual.
    if deals_fallback:
        lines.append("## ⚠️ Deals com preço FALLBACK `.estat-tcg` (margem NÃO-confiável — validar)")
        lines.append("")
        lines.append("> O preço TCG destes deals é uma **estimativa do próprio MYP** "
                     "(`.estat-tcg`), **não** o preço real do TCGplayer (pokemontcg.io). "
                     "Essa estimativa às vezes aponta pra carta errada e **infla a margem** "
                     "— ou seja, a margem abaixo **pode ser ilusória**. Por isso estes deals "
                     "**NÃO** entram no balde 'limpos'. **Antes de operar:** confira o preço "
                     "NM no Link TCG, ou rode `myp_enrich.py` local pra obter o preço real. "
                     "(Coberturas de CI saem 100% aqui — os runners não alcançam a "
                     "pokemontcg.io.)")
        lines.append("")
        lines.append("| # | Margem (estimada) | MYP R$ | TCG est. R$ | Dif (est.) | Carta | Set | Raridade | Cond | Qtd | Links |")
        lines.append("|---|---:|---:|---:|---:|---|---|---|---|---:|---|")
        for i, c in enumerate(deals_fallback[:50], 1):
            name = c.get("Card Name")
            carta = carta_label(name)
            ed = (c.get("Edition") or "").strip()
            rarity = (c.get("Rarity") or "").strip() or "—"
            myp = fmt_brl(c.get("MYP EN NM (R$)"))
            tcg = fmt_brl(c.get("TCG Player (R$)"))
            margin = fmt_pct(c.get("Margin %"))
            diff = fmt_brl(c.get("Diff (R$)"))
            qty = c.get("NM Sellers") or 0
            links = delivery_links(
                c.get("URL"), name, ed,
                oversized=bool(c.get("⚠️ COLLECTOR#")),
                tcg_url=c.get("TCG URL"),
            )
            lines.append(
                f"| {i} | {margin} | {myp} | {tcg} | {diff} | "
                f"{carta} | {ed} | {rarity} | NM | {qty} | {links} |"
            )
        lines.append("")

    # ── Truncation risks ──
    if truncations:
        lines.append("## 🚨 EN truncation risk (preço pode estar superestimado)")
        lines.append("")
        lines.append("> Seller table do MYP bateu cap com zero EN visível — listing real "
                     "pode ser mais barato. Validar via página direta.")
        lines.append("")
        lines.append("| Carta | Edição | MYP R$ reportado | TCG R$ |")
        lines.append("|---|---|---:|---:|")
        for c in truncations[:50]:
            name = (c.get("Card Name") or "")[:55]
            ed = (c.get("Edition") or "")[:30]
            myp = fmt_brl(c.get("MYP EN NM (R$)"))
            tcg = fmt_brl(c.get("TCG Player (R$)"))
            lines.append(f"| {name} | {ed} | {myp} | {tcg} |")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append(f"*Gerado em {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} via "
                 f"`myp_summary.py` (single source: XLSX consolidado).*")

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"OK markdown summary: {out_path}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Gera markdown summary do scan MYP")
    parser.add_argument("xlsx", help="Caminho do XLSX consolidado")
    parser.add_argument("-o", "--output", required=True, help="Caminho do .md de saída")
    parser.add_argument("--type", choices=["daily", "weekly"], required=True,
                        help="Tipo do scan (afeta título + tags)")
    parser.add_argument("--run-id", default="",
                        help="GH Actions run ID (link no markdown)")
    parser.add_argument("--repo", default="matheuscllm-lgtm/myp-arbitrage-scanner",
                        help="Repo GH pra link do artifact")
    args = parser.parse_args()
    return build_markdown(
        xlsx=args.xlsx, output=args.output, scan_type=args.type,
        run_id=args.run_id, repo=args.repo,
    )


if __name__ == "__main__":
    sys.exit(main())
