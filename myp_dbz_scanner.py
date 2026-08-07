#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║   MYP Cards Arbitrage Scanner — DRAGON BALL (paralelo ao Pokémon)    ║
║   mypcards.com (BR, R$) vs TCGplayer (tcgcsv.com) · v1.1             ║
╚══════════════════════════════════════════════════════════════════════╝

Scanner PARALELO ao fluxo Pokémon deste repo — precedente da frota:
`dbs_scanner.py`/`op_scanner.py` no card-trader-scanner (jogo paralelo =
script separado; o modo `--game` embutido foi rejeitado lá, PR #56 fechado
como superado). Este script NÃO toca o scanner Pokémon: reusa por import a
infra de site que é da PLATAFORMA MYP (sessão cloudscraper, retry, parser de
seller-table NM/EN, paginação marketplace) e implementa o resto próprio.

O MYP tem DUAS seções Dragon Ball (provadas pela sonda `probe-myp-dragonball`
no PR #95, runs 30780981509/30781111250, 2026-08-03), que espelham 1:1 as
duas categorias do tcgcsv:

    /dbsfusion   Dragon Ball Super CG: Fusion World (35 edições)  ↔  cat 80
    /dbsmasters  Dragon Ball Super CG "Masters"     (88 edições)  ↔  cat 27

Fatos de estrutura provados pela sonda (não re-descobrir):
  • O título de edição vem CONCATENADO com o código de set do MYP e a data
    ("Supreme RivalryB1307/05/2021") — `split_edition_title` separa usando o
    slug da URL como âncora.
  • h1 de produto dbsfusion = "Nome - FB11-112" (código com hífen no fim,
    às vezes com prefixo "[Pré-Venda]"); h1 de produto dbsmasters = só o
    nome ("Arack, Godly Power") — SEM código. Por isso o join tem DOIS
    caminhos: por código (quando o título traz) e por grupo+nome exato.
  • Página de produto usa os MESMOS seletores da seção Pokémon (`.estat-tcg`,
    `.estatistica-ultimo`, `table.table-striped.table-bordered`) — o parser
    de sellers NM/EN é herdado do MYPScraper.

Fluxo: menor oferta EN NM ao vivo no MYP  vs  market price TCGplayer (USD →
BRL com câmbio ao vivo). Margem BRUTA base compra. Entrega via
`myp_dbz_summary.py` (verbatim — contrato de entrega do repo).

Invariantes (frota + convenções DESTE repo):
  • Margem BRUTA: (TCG_BRL − MYP_BRL) / MYP_BRL — sem nenhuma taxa embutida.
  • `--threshold` em PERCENT INTEIRO (30 = 30%) — convenção MYP/Liga/eBay;
    valor < 1.0 é tratado como fração com warning (mesmo guard do Pokémon).
  • Piso de relevância R$50 (singles): `--min-price 50`.
  • Só Near Mint (célula dedicada, token exato "NM") e só EN (flag-icon
    "Inglês"/"English") — parser herdado do scanner Pokémon.
  • NUNCA inventar preço: carta sem referência tcgcsv fica FORA da margem e
    vai pra aba "Sem Ref TCG" com o motivo (não há fallback `.estat-tcg`
    aqui — decisão v1, espelha o dbs_scanner do CardTrader: referência real
    ou nada). Sem câmbio real o run FALHA ALTO.
  • NUNCA recomendar compra — buckets são classificação técnica.

Join MYP ↔ TCGplayer (DETERMINÍSTICO, nunca fuzzy):
  1. resolução edição→grupo tcgcsv (nome exato → código de set equivalente
     único → contenção de nome única; classes de alias {B≡BT}, {BE≡EB≡EX}
     cobrem a divergência de nomenclatura entre MYP e TCGplayer);
  2. carta por CÓDIGO ("FB11-112", normalizado de zeros) com escopo em
     camadas: grupo resolvido → grupos principais (não Pre-Release/Alt
     Art/Promo) → todos; colisão resolve por nome EN EXATO normalizado;
  3. sem código no título (caso dbsmasters): grupo resolvido + nome EXATO
     único dentro do grupo.
  Ambíguo/sem match → linha "Sem Ref TCG" com motivo (honestidade >
  cobertura; contagem explícita, nada some em silêncio).

Uso:
    python myp_dbz_scanner.py --list-editions
    python myp_dbz_scanner.py --editions "Awakened Pulse" --threshold 30 \
        --min-price 50 --delay 1.5 -o results/dbz_<data>.xlsx --resume
    python myp_dbz_scanner.py -o results/dbz_full.xlsx   # catálogo inteiro

Autor: Matheus Chillemi / Claude
Data: 2026-08-07 (v1.1)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote_plus

# Infra da PLATAFORMA MYP reusada por import (sessão CF, retry, parser de
# seller-table NM/EN, paginação marketplace, câmbio). O que é Pokémon-only
# (pokemontcg.io, mapas de set, TG/GG, supranumerário) NÃO é usado aqui.
from myp_arbitrage_scanner import (
    BASE_URL,
    MAX_PAGES_PER_EDITION,
    MAX_SELLER_PAGES,
    MAX_EDITION_PAGES,
    OVERSIZED_TITLE_RE,
    MYPScraper,
    fetch_usd_brl,
    log,
)

DBZ_VERSION = "v1.1"

# Seções Dragon Ball do MYP ↔ categoria tcgcsv correspondente (sonda 2026-08-03).
GAME_SECTIONS: tuple[tuple[str, int], ...] = (
    ("dbsfusion", 80),    # Fusion World
    ("dbsmasters", 27),   # DBS Masters (jogo clássico)
)
SECTION_LABELS = {"dbsfusion": "Fusion World", "dbsmasters": "Masters"}

TCGCSV_DBZ_BASE = "https://tcgcsv.com/tcgplayer"
# tcgcsv exige User-Agent identificando a aplicação (sem ele = 401).
TCGCSV_DBZ_USER_AGENT = (
    "myp-arbitrage-scanner-dbz/1.0 "
    "(+github.com/matheuscllm-lgtm/myp-arbitrage-scanner)"
)

MARGIN_THRESHOLD_DBZ = 0.30   # 30% margem BRUTA mínima (fração interna)
MIN_PRICE_BRL_DBZ = 50.0      # piso canônico de singles da frota
# Guardas da frota (mesmos valores do dbs_scanner do CardTrader):
JUNK_RATIO = 0.5              # MYP < 50% da ref = possível lixo/scam → REVISAR
VOLATILE_REF_RATIO = 2.0      # market vs menor anúncio TCG divergindo >2× → REVISAR

DBZ_CHECKPOINT_VERSION = 1

# Código de carta DBZ: prefixo alfa (+dígitos de set) + hífen + número
# (+ sufixo de letra). Ex.: FB01-001, FS06-15, BT31-055, P-001, EX23-01.
CARD_CODE_RE = re.compile(r"\b([A-Za-z]{1,8})(\d{0,3})-(\d{1,4})([A-Za-z]?)\b")
# Código entre parênteses ("Nome PT (FB01-001)Nome EN" — padrão bilíngue).
_TITLE_CODE_RE = re.compile(r"\(\s*([A-Za-z]{1,8}\d{0,3}-\d{1,4}[A-Za-z]?)\s*\)")
# Prefixo de status no h1 do MYP (ex. "[Pré-Venda] Bardock - FB11-112").
_H1_TAG_RE = re.compile(r"^\[[^\]]{1,24}\]\s*")
# Data concatenada no fim do título de edição ("…BT2805/09/2025"; o site
# emite até "30/11/-0001" pra edição sem data).
_EDITION_DATE_RE = re.compile(r"(\d{2}/\d{2}/-?\d{1,5})$")
# Grupo tcgcsv "especial" (nunca preferido sem escopo explícito): versões
# carimbadas/reimpressões que DUPLICAM número e nome do set principal.
_SPECIAL_GROUP_RE = re.compile(
    r"pre-?release|release event|alternate art|promo|judge|tournament|reprint",
    re.IGNORECASE)

_TCG_DBZ_SEARCH_BASE = "https://www.tcgplayer.com/search/all/product?q="


# ══════════════════════════════════════════════════════════════════════
# Helpers puros (testáveis offline)
# ══════════════════════════════════════════════════════════════════════

def normalize_code(raw: Optional[str]) -> Optional[str]:
    """Normaliza um código de carta pra chave de join.

    "FB01-001" → "FB1-1" · "BT31-055" → "BT31-55" · "P-001" → "P-1" ·
    "FB02-135b" → "FB2-135B" (sufixo de alt-art PRESERVADO — alt art nunca
    colapsa na carta base; lição do op_scanner da frota). Sem código
    parseável → None (nunca chuta)."""
    if not raw:
        return None
    m = CARD_CODE_RE.search(str(raw))
    if not m:
        return None
    alpha, prefix_digits, num, suffix = m.groups()
    pd = prefix_digits.lstrip("0")
    tail = num.lstrip("0") or "0"
    return f"{alpha.upper()}{pd}-{tail}{suffix.upper()}"


def _norm_name(s: Optional[str]) -> str:
    """Normalização de nome pra join: casefold + colapso de espaços."""
    return " ".join(str(s or "").casefold().split())


def _clean_card_name(s: Optional[str]) -> str:
    """Nome comparável: remove o código de carta em QUALQUER posição e os
    hífens órfãos que sobram. O TCGplayer FW é inconsistente ('Hit' vs
    'Son Goku - FB01-001' vs 'Son Goku - FB01-139 (Alternate Art)' — provado
    na sonda v3); o MYP idem. Todos convergem pra 'Nome (Qualificador)'."""
    t = CARD_CODE_RE.sub(" ", str(s or ""))
    t = re.sub(r"\(\s*\)", " ", t)             # "Marker (E-60)" → sem "( )" órfão
    t = re.sub(r"\s*[-–—]\s+(?=\()", " ", t)   # "Goku - (Alt)" → "Goku (Alt)"
    t = re.sub(r"\s*[-–—]\s*$", "", t)
    return " ".join(t.split())


def _strip_code_tail(s: str) -> str:
    """Compat: nome sem código (delegado ao _clean_card_name)."""
    return _clean_card_name(s)


def _base_name(s: Optional[str]) -> str:
    """Nome sem o qualificador '(...)' final ("Son Goku (Alternate Art)" →
    "Son Goku") — 2º passo da desambiguação por nome."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).strip()


# Campo "Código" da página de produto MYP (provado na sonda v3, 2026-08-03):
#   dbsmasters → "dbsm_bt1-073_spr"      (código da carta + variante _spr)
#   dbsfusion  → "dbsf_fb01_fb01-139p1"  (set + código da carta + marcador p1)
# É a ÚNICA fonte do código no dbsmasters (o h1 de lá não traz código).
_PROD_CODE_FIELD_RE = re.compile(r"\bdbs[fm]_[a-z0-9_-]{3,60}\b", re.IGNORECASE)
_CARD_TOKEN_RE = re.compile(r"^([A-Za-z]{1,8}\d{0,3}-\d{1,4})([A-Za-z0-9]*)$")


def parse_myp_product_code(code_str: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    """'dbsm_bt1-073_spr' → ('BT1-073', 'spr') · 'dbsf_fb01_fb01-139p1' →
    ('FB01-139', 'p1') · sem código de carta → (None, None).

    O sufixo de variante é DETERMINANTE no join: um produto '_spr' só pode
    casar o produto '(SPR)' do TCGplayer — casar com o base infla/deflaciona
    a margem com o preço da versão errada (classe de bug nº 3 da frota)."""
    s = str(code_str or "").strip().lower()
    if not s.startswith(("dbsm_", "dbsf_")):
        return None, None
    card = None
    variant_parts: list[str] = []
    for tok in s.split("_")[1:]:
        if card is None:
            m = _CARD_TOKEN_RE.match(tok)
            if m and "-" in tok:
                card = m.group(1).upper()
                if m.group(2):
                    variant_parts.append(m.group(2))
            continue
        if tok:
            variant_parts.append(tok)
    return card, ("_".join(variant_parts) or None)


def extract_myp_product_code(soup) -> Optional[str]:
    """Localiza o campo 'Código' na página de produto. Primário: célula cujo
    texto é exatamente 'Código' (a linha-pai traz o valor). Fallback: 1º
    padrão dbsm_/dbsf_ no texto da página."""
    try:
        for el in soup.find_all(["td", "th", "dt", "span", "div", "li"]):
            if el.get_text(strip=True) == "Código":
                parent = el.parent
                if parent is not None:
                    m = _PROD_CODE_FIELD_RE.search(
                        parent.get_text(" ", strip=True))
                    if m:
                        return m.group(0)
        m = _PROD_CODE_FIELD_RE.search(soup.get_text(" ", strip=True))
        return m.group(0) if m else None
    except Exception:  # noqa: BLE001 — extração é best-effort
        return None


def _alnum(s: Optional[str]) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").casefold())


def split_myp_title(h1_text: str) -> tuple[str, Optional[str], str]:
    """Título (h1) do produto MYP → (nome_display, código_cru, nome_EN_limpo).

    Formatos reais (sonda 2026-08-03):
      dbsfusion:  "[Pré-Venda] Bardock - FB11-112"  → código no FIM, hífen;
      dbsmasters: "Arack, Godly Power"              → SEM código (join por
                  grupo+nome cuida desse caso).
    Também aceita o formato "(CODIGO)" por robustez. Whitespace interno é
    normalizado (o h1 real traz \\n dentro do texto)."""
    txt = " ".join((h1_text or "").split())
    if not txt:
        return "", None, ""
    while _H1_TAG_RE.match(txt):
        txt = _H1_TAG_RE.sub("", txt, count=1)
    # formato "(CODIGO)" = padrão bilíngue da plataforma ("Nome PT (COD)Nome
    # EN", já documentado no scanner Pokémon v5.8.7): o nome EN vem DEPOIS.
    m = _TITLE_CODE_RE.search(txt)
    if m:
        code = m.group(1)
        before = _clean_card_name(txt[: m.start()])
        after = _clean_card_name(txt[m.end():])
        if before and after.startswith("("):
            # "Energy Marker (E-128) (Gold)": o que segue o código é
            # QUALIFICADOR do nome que o precede — um nome EN bilíngue
            # nunca começa com parêntese
            en_name = f"{before} {after}"
        else:
            en_name = after or before
        display = f"{en_name} - {code}" if en_name else code
        return display, code, en_name
    # formato dbsfusion real: "Bardock - FB11-112 [(Alternate Art)]"
    matches = list(CARD_CODE_RE.finditer(txt))
    if not matches:
        return txt, None, txt
    code = matches[-1].group(0)
    en_name = _clean_card_name(txt)
    display = f"{en_name} - {code}" if en_name else code
    return display, code, en_name


def split_edition_title(raw_title: str, slug: str) -> tuple[str, Optional[str]]:
    """Título de edição do MYP → (título limpo, código de set do MYP).

    O card de edição concatena título + código + data sem separador
    ("Supreme RivalryB1307/05/2021", slug "supreme-rivalry"). O corte é
    ancorado no SLUG (alnum do prefixo == alnum do slug) — nunca em
    heurística de maiúscula, que quebraria em "5th Anniversary Set5TH"."""
    txt = " ".join((raw_title or "").split())
    txt = _EDITION_DATE_RE.sub("", txt).strip()
    slug_a = _alnum(slug)
    if slug_a:
        for cut in range(len(txt), 0, -1):
            if _alnum(txt[:cut]) == slug_a:
                code = txt[cut:].strip(" -–—") or None
                return txt[:cut].strip(), code
    # fallback defensivo: sufixo maiúsculo/dígito após o último minúsculo
    m = re.search(r"([A-Z0-9][A-Z0-9_-]{0,11})$", txt)
    if m and not txt[: m.start()].strip() == "":
        return txt[: m.start()].strip(), m.group(1)
    return txt, None


def canonical_set_code(code: Optional[str]) -> Optional[str]:
    """Código de set → forma canônica pra equivalência MYP ↔ tcgcsv.

    Remove separadores e os prefixos "DBS"/"DBSFW" do tcgcsv, tira zeros à
    esquerda do número e aplica as classes de alias observadas nos dois
    catálogos ({B≡BT} boosters antigos; {BE≡EB≡EX} expansion decks):
    "DBS-B09"≡"BT9" · "DBS-BE15"≡"EX15" · "MB-01"≡"MB1" · "FB03_AR"≡"FB3AR".
    """
    c = re.sub(r"[^A-Za-z0-9]", "", str(code or "").upper())
    if not c:
        return None
    for pref in ("DBSFW", "DBS"):
        if c.startswith(pref) and len(c) > len(pref):
            c = c[len(pref):]
            break
    m = re.match(r"^([A-Z]+)(\d+)([A-Z]*)$", c)
    if not m:
        return c
    alpha, num, tail = m.groups()
    alpha = {"B": "BT", "EB": "EX", "BE": "EX"}.get(alpha, alpha)
    return f"{alpha}{int(num)}{tail}"


def resolve_edition_group(clean_title: str, myp_code: Optional[str],
                          groups: list[dict]) -> Optional[int]:
    """Edição MYP → groupId tcgcsv. Cascata unique-match-only (lição da
    frota: nunca chutar em ambíguo; alias deduzido já saiu alucinado):
      1. nome EXATO (alnum) — "Rivals Clash" ↔ "Rivals Clash";
      2. código canônico igual e ÚNICO — "B13" ↔ "DBS-B13" (o tcgcsv tem
         abbr duplicada real, ex. FB03_AR em 2 grupos → ambíguo → passo 3);
      3. contenção de nome ÚNICA — "Ultimate Deck 2023" ⊂ "Expansion Deck
         Box Set 22: Ultimate Deck 2023".
    Sem match → None (cartas da edição saem em Sem Ref, nunca margem)."""
    if not groups:
        return None
    title_a = _alnum(clean_title)
    if title_a:
        exact = [g for g in groups if _alnum(g.get("name")) == title_a]
        if len(exact) == 1:
            return exact[0]["groupId"]
    code_c = canonical_set_code(myp_code)
    if code_c:
        by_code = [g for g in groups
                   if canonical_set_code(g.get("abbreviation")) == code_c]
        if len(by_code) == 1:
            return by_code[0]["groupId"]
    if title_a and len(title_a) >= 6:
        contains = [g for g in groups if title_a in _alnum(g.get("name"))]
        if len(contains) == 1:
            return contains[0]["groupId"]
    return None


def parse_threshold(value: float) -> float:
    """Threshold CLI → fração interna. Convenção MYP: PERCENT INTEIRO
    (30 = 30%). Valor < 1.0 é aceito como fração com warning (mesmo guard
    do scanner Pokémon — trap inverso ao CardTrader)."""
    v = float(value)
    if v < 1.0:
        log.warning(
            f"⚠️ --threshold {v} < 1.0 — interpretando como FRAÇÃO "
            f"({v * 100:.0f}%). A convenção deste repo é percent inteiro "
            f"(ex.: --threshold 30).")
        return v
    return v / 100.0


def pick_subtype(prices: dict, prefer_foil: bool = False) -> Optional[tuple[str, float]]:
    """Escolhe o subtipo de preço TCGplayer (Normal/Foil) da referência.

    - 1 subtipo com market → ele;
    - dica de foil no anúncio MYP → Foil/Holofoil se existir;
    - senão o MENOR market entre os subtipos — conservador: nunca superestima
      a margem (mesma convenção do scanner Pokémon deste repo, que pega o
      menor preço entre variantes).
    Sem nenhum market price → None (nunca inventa)."""
    avail = {k: v for k, v in (prices or {}).items() if v is not None and v > 0}
    if not avail:
        return None
    if len(avail) == 1:
        return next(iter(avail.items()))
    if prefer_foil:
        for k in ("Foil", "Holofoil"):
            if k in avail:
                return k, avail[k]
    key = min(avail, key=lambda k: avail[k])
    return key, avail[key]


def ref_volatile(market_usd: Optional[float], low_usd: Optional[float]) -> bool:
    """Menor anúncio ATUAL do TCGplayer diverge do market em >2× (qualquer
    direção) = referência volátil/fina → sinal-only, rebaixa pra REVISAR
    (mesma guarda do dbs_scanner da frota)."""
    if not market_usd or not low_usd or market_usd <= 0 or low_usd <= 0:
        return False
    ratio = max(market_usd, low_usd) / min(market_usd, low_usd)
    return ratio > VOLATILE_REF_RATIO


def _name_cascade(cands: list[dict], en_name: str,
                  variant_token: Optional[str] = None) -> tuple[Optional[dict], str]:
    """Desambiguação por nome dentro de um escopo NÃO-vazio de candidatos.

    Variant-aware (sonda v3): o mesmo número tem produto base E variante
    especial no tcgcsv ('Broly…' vs 'Broly… (SPR)'; 'Son Goku - FB01-001' vs
    '… (Alternate Art)'). Regras, na ordem:
      1. COM variant_token (ex. '_spr' do campo Código do MYP): só candidato
         cujo nome contém o token E cujo nome-base bate — o exato do nome
         SEM token casaria o produto BASE (preço da versão errada, classe de
         bug nº 3 da frota). 0 candidatos com o token → SEM match.
      2. SEM token: nome EXATO (cobre o alt art do FW, cujo h1 já traz
         '(Alternate Art)') → base não-qualificado → candidato único (desde
         que NÃO seja variante qualificada — nunca precificar base com ref
         de variante).
    (match, rótulo) ou (None, motivo)."""
    clean = _clean_card_name(en_name)
    name_n = _norm_name(clean)
    base_n = _norm_name(_base_name(clean))
    if variant_token:
        tok = re.sub(r"[^a-z0-9]", "", variant_token.casefold())
        with_tok = [c for c in cands
                    if tok and tok in re.sub(r"[^a-z0-9]", "", c["name_norm"])
                    and (not base_n or c["base_norm"] == base_n)]
        if len(with_tok) == 1:
            return with_tok[0], f"nome+variante({variant_token})"
        if not with_tok:
            return None, f"variante '{variant_token}' sem produto qualificado"
        return None, f"variante '{variant_token}' ambígua ({len(with_tok)})"
    exact = [c for c in cands if c["name_norm"] == name_n]
    if len(exact) == 1:
        return exact[0], "nome-exato"
    unqualified = [c for c in cands
                   if c["name_norm"] == c["base_norm"] and c["base_norm"] == base_n]
    if len(unqualified) == 1:
        return unqualified[0], "nome-base"
    if len(cands) == 1:
        only = cands[0]
        if only["name_norm"] != only["base_norm"]:
            return None, "índice só tem variante qualificada (base ausente)"
        return only, "unico"
    return None, f"ambíguo ({len(cands)} produtos tcgcsv)"


def find_reference(code_norm: Optional[str], en_name: str,
                   primary: dict, secondary: dict,
                   group_id: Optional[int] = None,
                   variant_token: Optional[str] = None) -> tuple[Optional[dict], str]:
    """Join determinístico carta MYP → produto tcgcsv (NUNCA fuzzy).

    `primary`/`secondary` = índices por categoria ({"by_code", "by_group"}).
    COM código: candidatos pelo código com escopo em camadas — (a) grupo
    resolvido da edição, (b) grupos principais (não Pre-Release/Alt Art/
    Promo — esses DUPLICAM número e nome do set principal), (c) todos; o
    primeiro escopo NÃO-vazio decide (match único/nome exato, senão
    ambíguo → fora). SEM código: grupo resolvido + nome exato único.
    A ambiguidade na categoria primária encerra (não cai pra irmã — seria
    chute). Retorna (entry|None, rótulo/motivo)."""
    if code_norm:
        for idx, origem in ((primary, "cat-seção"), (secondary, "cat-irmã")):
            cands = (idx.get("by_code") or {}).get(code_norm) or []
            if not cands:
                continue
            scopes = []
            if group_id is not None:
                scopes.append(("grupo-edição",
                               [c for c in cands if c["group_id"] == group_id]))
            scopes.append(("grupo-principal",
                           [c for c in cands if not c["group_special"]]))
            scopes.append(("global", cands))
            for scope_label, scoped in scopes:
                if not scoped:
                    continue
                match, how = _name_cascade(scoped, en_name, variant_token)
                if match is not None:
                    return match, f"codigo/{scope_label}/{how} ({origem})"
                return None, f"código: {how} [{scope_label}]"
            return None, "código sem escopo utilizável"
        return None, "número fora do índice tcgcsv"
    # sem código NENHUM (nem no título nem no campo Código): grupo + nome
    if group_id is None:
        return None, "sem código e grupo tcgcsv não resolvido"
    cands = (primary.get("by_group") or {}).get(group_id) or []
    if not cands:
        return None, "grupo tcgcsv resolvido mas vazio no índice"
    match, how = _name_cascade(cands, en_name, variant_token)
    if match is not None:
        return match, f"grupo+{how}"
    return None, f"grupo: {how}"


def tcg_dbz_search_url(name: str, code: Optional[str] = None) -> Optional[str]:
    """URL de BUSCA no TCGplayer (fallback de link pra linhas sem produto
    casado — mesmo papel do tcg_search_url do fluxo Pokémon)."""
    base = _base_name(_strip_code_tail(name))
    if not base and not code:
        return None
    q = " ".join(p for p in (base, code or "") if p)
    return _TCG_DBZ_SEARCH_BASE + quote_plus(q)


# ══════════════════════════════════════════════════════════════════════
# Índice de referência tcgcsv (cats 80 + 27)
# ══════════════════════════════════════════════════════════════════════

def _tcgcsv_json(session, url: str, cache_file: Optional[Path],
                 cache_hours: float) -> Optional[dict]:
    """GET JSON do tcgcsv com cache em disco opcional (TTL em horas)."""
    if cache_file is not None and cache_hours > 0 and cache_file.exists():
        age_h = (time.time() - cache_file.stat().st_mtime) / 3600.0
        if age_h < cache_hours:
            try:
                return json.loads(cache_file.read_text(encoding="utf-8"))
            except (OSError, ValueError):
                pass  # cache corrompido → refetch
    try:
        r = session.get(url, headers={"User-Agent": TCGCSV_DBZ_USER_AGENT},
                        timeout=30)
        if r.status_code != 200:
            log.warning(f"tcgcsv {url} → HTTP {r.status_code}")
            return None
        data = r.json()
    except Exception as e:  # noqa: BLE001 — rede/JSON: reporta e segue
        log.warning(f"tcgcsv {url} falhou: {e!r}")
        return None
    if cache_file is not None:
        try:
            cache_file.parent.mkdir(parents=True, exist_ok=True)
            cache_file.write_text(json.dumps(data), encoding="utf-8")
        except OSError:
            pass
    return data


def build_category_index(session, cat: int, cache_dir: Optional[Path],
                         cache_hours: float = 20.0) -> dict:
    """Índice de UMA categoria tcgcsv:
        {"groups": [{groupId, name, abbreviation}],
         "by_code": {code_norm: [entry]},
         "by_group": {group_id: [entry]}}
    entry = {pid, name, name_norm, base_norm, number, rarity, url, prices,
             lows, cat, group_id, group_name, group_special}.
    Produtos SEM Number (selados/acessórios) ficam fora — regra provada no
    dbs_scanner da frota (todo produto sem número era selado)."""
    out = {"groups": [], "by_code": {}, "by_group": {}}
    groups = _tcgcsv_json(session, f"{TCGCSV_DBZ_BASE}/{cat}/groups",
                          cache_dir / f"groups_{cat}.json" if cache_dir else None,
                          cache_hours)
    glist = (groups or {}).get("results", [])
    if not glist:
        log.warning(f"tcgcsv categoria {cat}: 0 groups — índice vazio.")
        return out
    out["groups"] = [{"groupId": g.get("groupId"), "name": g.get("name", ""),
                      "abbreviation": g.get("abbreviation", "")} for g in glist]
    log.info(f"  [tcgcsv] categoria {cat}: {len(glist)} groups…")
    by_pid: dict[int, dict] = {}
    for g in glist:
        gid = g["groupId"]
        gname = g.get("name", "")
        special = bool(_SPECIAL_GROUP_RE.search(gname))
        prods = _tcgcsv_json(session, f"{TCGCSV_DBZ_BASE}/{cat}/{gid}/products",
                             cache_dir / f"products_{cat}_{gid}.json" if cache_dir else None,
                             cache_hours)
        prices = _tcgcsv_json(session, f"{TCGCSV_DBZ_BASE}/{cat}/{gid}/prices",
                              cache_dir / f"prices_{cat}_{gid}.json" if cache_dir else None,
                              cache_hours)
        for p in (prods or {}).get("results", []):
            ext = {e.get("name"): e.get("value")
                   for e in (p.get("extendedData") or [])}
            number = ext.get("Number") or ""
            code_norm = normalize_code(number)
            if not code_norm:
                continue  # selado/acessório — nunca vira referência de single
            clean = _strip_code_tail(p.get("name"))
            entry = {
                "pid": int(p["productId"]),
                "name": p.get("name", ""),
                "name_norm": _norm_name(clean),
                "base_norm": _norm_name(_base_name(clean)),
                "number": number,
                "rarity": ext.get("Rarity") or "",
                "url": p.get("url", ""),
                "prices": {},
                "lows": {},
                "cat": cat,
                "group_id": gid,
                "group_name": gname,
                "group_special": special,
            }
            by_pid[entry["pid"]] = entry
            out["by_code"].setdefault(code_norm, []).append(entry)
            out["by_group"].setdefault(gid, []).append(entry)
        for r in (prices or {}).get("results", []):
            pid = int(r.get("productId") or 0)
            entry = by_pid.get(pid)
            if entry is None:
                continue
            sub = r.get("subTypeName") or "?"
            if r.get("marketPrice") is not None:
                entry["prices"][sub] = float(r["marketPrice"])
            if r.get("lowPrice") is not None:
                entry["lows"][sub] = float(r["lowPrice"])
    n = len(by_pid)
    log.info(f"  [tcgcsv] categoria {cat}: {n} singles indexados "
             f"({len(out['by_code'])} códigos)")
    return out


# ══════════════════════════════════════════════════════════════════════
# Dados
# ══════════════════════════════════════════════════════════════════════

@dataclass
class DbzCardData:
    name: str = ""                 # nome display ("Bardock - FB11-112")
    en_name: str = ""              # nome comparável (sem código/cauda)
    code: str = ""                 # código cru do título (FB11-112; "" se ausente)
    code_norm: str = ""            # chave de join sem zeros ("FB01-001" → "FB1-1")
    edition: str = ""              # título limpo da edição
    section: str = ""              # dbsfusion | dbsmasters
    product_url: str = ""
    myp_lowest_en_nm: Optional[float] = None
    en_nm_sellers: int = 0
    en_truncation_risk: bool = False
    single_en_seller_risk: bool = False
    tcg_usd: Optional[float] = None       # market TCGplayer (USD)
    tcg_low_usd: Optional[float] = None   # menor anúncio atual (USD)
    tcg_brl: Optional[float] = None
    tcg_source: str = "tcgcsv"            # única fonte real do DBZ v1
    tcg_product_name: str = ""
    tcg_url: str = ""
    rarity: str = ""                      # Rarity do tcgcsv (não a do MYP)
    subtype: str = ""                     # Normal | Foil (subtipo usado)
    join_via: str = ""
    margin_pct: Optional[float] = None    # FRAÇÃO (0.42 = 42%)
    margin_brl: Optional[float] = None
    flag_lixo: bool = False               # MYP < 50% da ref → REVISAR
    flag_ref_volatil: bool = False        # market vs low >2× → REVISAR
    last_updated: str = ""


@dataclass
class SemRefRow:
    """Oferta EN NM viva ≥ piso SEM referência tcgcsv — nada some em silêncio
    (mesmo papel do sidecar `_semref.csv` do dbs_scanner)."""
    name: str = ""
    en_name: str = ""
    code: str = ""
    edition: str = ""
    section: str = ""
    product_url: str = ""
    myp_lowest_en_nm: Optional[float] = None
    en_nm_sellers: int = 0
    motivo: str = ""


# ══════════════════════════════════════════════════════════════════════
# Scraper
# ══════════════════════════════════════════════════════════════════════

class MYPDbzScraper(MYPScraper):
    """Scraper das seções Dragon Ball do MYP.

    Herda do MYPScraper a infra de PLATAFORMA (sessão cloudscraper firefox,
    `_get` com retry, `_parse_seller_table` NM/EN, `_max_seller_page`,
    `_parse_brl`). O pipeline de preço Pokémon (pokemontcg.io/estat) NÃO é
    usado: a referência DBZ é o índice tcgcsv carregado no início do scan."""

    def __init__(self, delay: float = 1.5, min_en_sellers: int = 2,
                 threshold: float = MARGIN_THRESHOLD_DBZ,
                 min_price: float = MIN_PRICE_BRL_DBZ,
                 sections: tuple[tuple[str, int], ...] = GAME_SECTIONS,
                 tcg_cache_dir: Optional[Path] = None, cache_hours: float = 20.0):
        super().__init__(delay=delay, min_en_sellers=min_en_sellers,
                         threshold=threshold, min_price=min_price,
                         tcg_source="tcgcsv")
        self.sections = sections
        self.tcg_cache_dir = tcg_cache_dir
        self.cache_hours = cache_hours
        self.dbz_cards: list[DbzCardData] = []
        self.semref: list[SemRefRow] = []
        self.indices: dict[int, dict] = {}   # cat → índice da categoria
        self._stats.update({
            "dbz_sem_codigo": 0,
            "dbz_sem_ref": 0,
            "dbz_ref_ok": 0,
            "dbz_flag_lixo": 0,
            "dbz_flag_ref_volatil": 0,
            "dbz_edicoes_sem_grupo": 0,
        })

    # ── catálogo de edições (por seção) ──
    def get_section_editions(self, section: str) -> list[dict]:
        """Edições de UMA seção (/{section}/edicoes?page=N) — mesma cascata
        de seletores do get_all_editions do Pokémon, parametrizada; separa
        título limpo + código de set do card concatenado da sonda."""
        editions: list[dict] = []
        page = 1
        while page <= MAX_EDITION_PAGES:
            url = f"{BASE_URL}/{section}/edicoes?page={page}"
            log.info(f"  [{section}] edições página {page}…")
            soup = self._get(url)
            if not soup:
                break
            links = soup.select("a.edicao-link")
            if not links:
                for c in soup.select('[class*="edicao"]'):
                    a = c.select_one(f'a[href*="/{section}/"]')
                    if a and a not in links:
                        links.append(a)
            if not links:
                exclude = ("edicoes", "produto", "outros", "selados",
                           "acessorios", "#")
                for a in soup.select("a[href]"):
                    href = a.get("href", "")
                    if (re.match(rf"^/{section}/[a-z0-9][\w-]+$", href)
                            and not any(x in href for x in exclude)):
                        links.append(a)
            if not links:
                break
            found_on_page = 0
            seen = {e["url"] for e in editions}
            for a in links:
                href = a.get("href", "")
                if not href:
                    continue
                full = f"{BASE_URL}{href}" if href.startswith("/") else href
                if full in seen:
                    continue
                raw_title = a.get_text(strip=True)[:120]
                slug = href.rstrip("/").split("/")[-1]
                clean_title, myp_code = split_edition_title(raw_title, slug)
                if not clean_title or len(clean_title) < 2:
                    continue
                editions.append({"title": clean_title, "myp_code": myp_code,
                                 "raw_title": raw_title, "url": full,
                                 "href": href, "slug": slug,
                                 "section": section})
                seen.add(full)
                found_on_page += 1
            if found_on_page == 0:
                break
            page += 1
        log.info(f"  [{section}] {len(editions)} edições")
        return editions

    # ── produtos de uma edição ──
    def get_section_products(self, edition_url: str, section: str) -> list[str]:
        """URLs de produto de uma edição (paginada) — espelho do
        get_edition_products do Pokémon com o seletor da seção."""
        product_urls: list[str] = []
        seen: set[str] = set()
        page = 1
        prev_first: Optional[str] = None
        while page <= MAX_PAGES_PER_EDITION:
            soup = self._get(f"{edition_url}?page={page}")
            if not soup:
                break
            links = soup.select(f'a[href*="/{section}/produto/"]')
            current_first = None
            for a in links:
                href = a.get("href", "")
                if href:
                    current_first = (f"{BASE_URL}{href}"
                                     if href.startswith("/") else href)
                    break
            if page > 1 and prev_first is not None and current_first == prev_first:
                log.warning(f"  🚨 Loop de paginação em {edition_url} "
                            f"(página {page} == {page - 1}) — parando.")
                break
            prev_first = current_first
            new_count = 0
            for a in links:
                href = a.get("href", "")
                full = f"{BASE_URL}{href}" if href.startswith("/") else href
                if full not in seen:
                    seen.add(full)
                    product_urls.append(full)
                    new_count += 1
            if new_count == 0:
                break
            page += 1
        return product_urls

    # ── página de produto ──
    def scrape_dbz_product(self, url: str, edition: dict, cat_primary: int,
                           group_id: Optional[int], fx_usd_brl: float):
        """Uma página de produto DBZ → DbzCardData | SemRefRow | None.

        None = fora do funil (sem oferta EN NM, abaixo do piso, oversized,
        fetch falhou) — contado nos stats herdados."""
        soup = self._get(url)
        if not soup:
            return None

        h1 = soup.select_one("h1")
        h1txt = h1.get_text(" ", strip=True) if h1 else ""
        if not h1txt:
            title_tag = soup.find("title")
            if title_tag and title_tag.text:
                h1txt = title_tag.text.split("|")[0].strip()
        if h1txt and OVERSIZED_TITLE_RE.search(h1txt):
            self._stats["skipped_jumbo"] += 1
            log.info(f"  ⏭️ Oversized: {h1txt[:60]}")
            return None

        display, h1_code, en_name = split_myp_title(h1txt)
        # campo "Código" da página (dbsm_bt1-073_spr / dbsf_fb01_fb01-139p1):
        # ÚNICA fonte de código no dbsmasters + traz o sufixo de variante.
        field_code, variant_token = parse_myp_product_code(
            extract_myp_product_code(soup))
        code_raw = field_code or h1_code
        code_norm = normalize_code(code_raw)
        if (field_code and h1_code
                and normalize_code(field_code) != normalize_code(h1_code)):
            log.warning(f"  ⚠️ Código do campo ({field_code}) ≠ código do h1 "
                        f"({h1_code}) em {url} — usando o campo.")
        # nome já qualificado ("(Alternate Art)" no h1 do FW) dispensa o
        # token de variante — o join por nome exato é mais forte.
        if variant_token and "(" in en_name:
            variant_token = None
        if code_raw and code_raw.lower() not in display.lower():
            display = f"{en_name} - {code_raw}" if en_name else code_raw
        if variant_token:
            display = f"{display} ({variant_token.upper()})"

        # referência ANTES dos sellers: barata (índice local) e alimenta o
        # cost-gate da paginação (só pagina quem ainda pode virar deal).
        primary = self.indices.get(cat_primary) or {}
        secondary_cat = next((c for _s, c in self.sections if c != cat_primary),
                             None)
        secondary = self.indices.get(secondary_cat) or {} if secondary_cat else {}
        entry, join_label = find_reference(code_norm, en_name, primary,
                                           secondary, group_id=group_id,
                                           variant_token=variant_token)
        ref_chosen = None
        if entry is not None:
            prefer_foil = "foil" in _norm_name(display)
            ref_chosen = pick_subtype(entry["prices"], prefer_foil=prefer_foil)

        # ── sellers EN NM (parser herdado) + gate de truncation/paginação ──
        en_prices: list[float] = []
        en_sellers = 0
        per_table = []
        seller_tables = soup.select("table.table-striped.table-bordered") or [soup]
        for table in seller_tables:
            st = self._parse_seller_table(table)
            per_table.append(st)
            en_prices.extend(st["en_prices"])
            en_sellers += st["en"]
            if st["jumbo"]:
                self._stats["jumbo_rows_filtered"] += st["jumbo"]

        TABLE_CAP = 15
        lowest_seen = min(en_prices) if en_prices else None
        gate = False
        if lowest_seen is not None:
            for ts in per_table:
                if (ts["rows"] >= TABLE_CAP and ts["en"] == 0
                        and 0 < ts["max_price"] < lowest_seen):
                    gate = True
                    break

        truncation_risk = False
        max_seller_page = self._max_seller_page(soup)
        # cost-gate com a REF real (sem ref → linha é validação manual de
        # qualquer jeito; não paginamos, mas registramos o risco).
        can_be_deal = bool(ref_chosen and ref_chosen[1] * fx_usd_brl >= self.min_price)
        if gate and max_seller_page >= 2 and not can_be_deal:
            self._stats["pagination_skipped_low_tcg"] += 1
            if ref_chosen is None:
                truncation_risk = True
        elif gate and max_seller_page >= 2:
            pages = min(max_seller_page, MAX_SELLER_PAGES)
            log.info(f"  📄 Paginando marketplace de {display or url} "
                     f"(2..{pages} de {max_seller_page})")
            for pg in range(2, pages + 1):
                page_soup = self._get(f"{url}?estoque-outros-page={pg}")
                if page_soup is None:
                    truncation_risk = True
                    self._stats["seller_page_fetch_failures"] += 1
                    break
                mkt = page_soup.select_one("#lista-anuncio-demais-vendedores")
                if mkt is None:
                    if pg < pages:
                        truncation_risk = True
                        self._stats["seller_page_empty_early"] += 1
                    break
                pst = self._parse_seller_table(mkt)
                en_prices.extend(pst["en_prices"])
                en_sellers += pst["en"]
                if pst["jumbo"]:
                    self._stats["jumbo_rows_filtered"] += pst["jumbo"]
                self._stats["seller_pages_followed"] += 1
            if max_seller_page > MAX_SELLER_PAGES:
                truncation_risk = True

        if not en_prices:
            self._stats["skipped_no_en_sellers"] += 1
            return None
        lowest_en = min(en_prices)
        if lowest_en < self.min_price:
            self._stats["skipped_low_price"] += 1
            return None

        if truncation_risk:
            self._stats["en_truncation_risks"] += 1

        # ── sem referência → linha honesta na aba própria ──
        if entry is None or ref_chosen is None:
            if not code_norm and "sem código" in join_label:
                self._stats["dbz_sem_codigo"] += 1
            motivo = (join_label if entry is None
                      else "produto tcgcsv sem market price")
            self._stats["dbz_sem_ref"] += 1
            return SemRefRow(
                name=display, en_name=en_name, code=code_raw or "",
                edition=edition["title"], section=edition["section"],
                product_url=url, myp_lowest_en_nm=lowest_en,
                en_nm_sellers=en_sellers, motivo=motivo,
            )

        subtype, tcg_usd = ref_chosen
        tcg_low = (entry.get("lows") or {}).get(subtype)
        tcg_brl = tcg_usd * fx_usd_brl
        card = DbzCardData(
            name=display, en_name=en_name, code=code_raw or "",
            code_norm=code_norm or "", edition=edition["title"],
            section=edition["section"], product_url=url,
            myp_lowest_en_nm=lowest_en, en_nm_sellers=en_sellers,
            en_truncation_risk=truncation_risk,
            single_en_seller_risk=en_sellers < self.min_en_sellers,
            tcg_usd=tcg_usd, tcg_low_usd=tcg_low, tcg_brl=tcg_brl,
            tcg_product_name=entry["name"], tcg_url=entry["url"],
            rarity=entry["rarity"], subtype=subtype, join_via=join_label,
            margin_brl=tcg_brl - lowest_en,
            margin_pct=(tcg_brl - lowest_en) / lowest_en,
            flag_lixo=lowest_en < JUNK_RATIO * tcg_brl,
            flag_ref_volatil=ref_volatile(tcg_usd, tcg_low),
            last_updated=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
        self._stats["dbz_ref_ok"] += 1
        if card.flag_lixo:
            self._stats["dbz_flag_lixo"] += 1
        if card.flag_ref_volatil:
            self._stats["dbz_flag_ref_volatil"] += 1
        if card.single_en_seller_risk:
            self._stats["single_en_seller_risks"] += 1
        return card

    # ── checkpoint/resume (espelho do v5.11.4 do Pokémon) ──
    def _save_dbz_checkpoint(self, path: str, done: set) -> None:
        try:
            payload = {
                "version": DBZ_CHECKPOINT_VERSION,
                "cards": [asdict(c) for c in self.dbz_cards],
                "semref": [asdict(s) for s in self.semref],
                "done_editions": sorted(done),
                "stats": self._stats,
            }
            tmp = f"{path}.tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(payload, f)
            os.replace(tmp, path)
        except Exception as e:  # noqa: BLE001 — best-effort, nunca derruba o scan
            log.warning(f"  ⚠️ Falha ao salvar checkpoint {path}: {e!r}")

    def _load_dbz_checkpoint(self, path: str) -> set:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:  # noqa: BLE001
            log.warning(f"  ⚠️ Checkpoint {path} ilegível ({e!r}) — do zero.")
            return set()
        if data.get("version") != DBZ_CHECKPOINT_VERSION:
            log.warning(f"  ⚠️ Checkpoint {path} de versão antiga — ignorando.")
            return set()
        cfields = DbzCardData.__dataclass_fields__
        sfields = SemRefRow.__dataclass_fields__
        self.dbz_cards = [DbzCardData(**{k: v for k, v in c.items() if k in cfields})
                          for c in data.get("cards", [])]
        self.semref = [SemRefRow(**{k: v for k, v in s.items() if k in sfields})
                       for s in data.get("semref", [])]
        st = data.get("stats")
        if isinstance(st, dict):
            self._stats.update(st)
        return set(data.get("done_editions", []))

    # ── scan principal ──
    def scan_dbz(self, fx_usd_brl: float, edition_filter: Optional[list[str]] = None,
                 sections_filter: Optional[list[str]] = None,
                 max_editions: int = 0, max_products: int = 0,
                 resume: bool = False, checkpoint_path: Optional[str] = None
                 ) -> tuple[list[DbzCardData], list[SemRefRow]]:
        log.info("═" * 60)
        log.info("  MYP DBZ Arbitrage Scanner (Dragon Ball — paralelo ao Pokémon)")
        log.info(f"  Threshold: {self.margin_threshold * 100:.0f}% | EN only | NM only")
        log.info(f"  Piso: R${self.min_price:.0f} | Câmbio: US$1 = R${fx_usd_brl:.4f}")
        log.info("═" * 60)

        sections = [s for s in self.sections
                    if not sections_filter or s[0] in sections_filter]
        if not sections:
            log.warning("Nenhuma seção selecionada — nada a fazer.")
            return [], []

        # índice de referência das categorias das seções escolhidas (bulk, 1×)
        for _sec, cat in sections:
            if cat not in self.indices:
                self.indices[cat] = build_category_index(
                    self.session, cat, self.tcg_cache_dir, self.cache_hours)
        if not any((idx.get("by_code") or {}) for idx in self.indices.values()):
            raise RuntimeError(
                "Índice tcgcsv vazio nas categorias 80/27 — sem referência "
                "de preço não há scan honesto (nunca inventar preço).")

        editions: list[dict] = []
        for section, _cat in sections:
            editions.extend(self.get_section_editions(section))

        if edition_filter:
            fl = [f.lower().strip() for f in edition_filter]
            matched = []
            for ed in editions:
                hay = f"{ed['title']} {ed.get('myp_code') or ''}".lower()
                for f in fl:
                    if f in hay:
                        matched.append(ed)
                        log.info(f"  ✅ Matched: '{ed['title']}' "
                                 f"[{ed.get('myp_code') or '—'}] (filtro: '{f}')")
                        break
            editions = matched
            if not editions:
                log.warning("Nenhuma edição casou o filtro!")
                return [], []
        if max_editions:
            editions = editions[:max_editions]

        done: set = set()
        if resume and checkpoint_path and os.path.exists(checkpoint_path):
            done = self._load_dbz_checkpoint(checkpoint_path)
            log.info(f"  ⏯️ Resume: {len(done)} edições feitas, "
                     f"{len(self.dbz_cards)} cards restaurados.")

        cat_by_section = dict(self.sections)
        for i, ed in enumerate(editions):
            if ed["url"] in done:
                log.info(f"[{i + 1}/{len(editions)}] ⏭️ (resume) {ed['title']}")
                continue
            cat = cat_by_section[ed["section"]]
            group_id = resolve_edition_group(
                ed["title"], ed.get("myp_code"),
                (self.indices.get(cat) or {}).get("groups") or [])
            if group_id is None:
                self._stats["dbz_edicoes_sem_grupo"] += 1
                log.warning(f"  ⚠️ Edição sem grupo tcgcsv resolvido: "
                            f"'{ed['title']}' [{ed.get('myp_code') or '—'}] — "
                            f"cartas sem código no título sairão em Sem Ref.")
            log.info(f"\n[{i + 1}/{len(editions)}] 📦 [{ed['section']}] "
                     f"{ed['title']} [{ed.get('myp_code') or '—'}]"
                     f"{f' → group {group_id}' if group_id else ''}")
            product_urls = self.get_section_products(ed["url"], ed["section"])
            if max_products:
                product_urls = product_urls[:max_products]
            log.info(f"  → {len(product_urls)} produtos")
            for j, purl in enumerate(product_urls):
                self._stats["products_scanned"] += 1
                if (j + 1) % 10 == 0:
                    log.info(f"  {j + 1}/{len(product_urls)}…")
                result = self.scrape_dbz_product(purl, ed, cat, group_id,
                                                 fx_usd_brl)
                if result is None:
                    continue
                self._stats["en_found"] += 1
                if isinstance(result, SemRefRow):
                    self.semref.append(result)
                    continue
                self.dbz_cards.append(result)
                if result.margin_pct is not None and \
                        result.margin_pct >= self.margin_threshold:
                    log.info(f"  🔥 DEAL: {result.name} | "
                             f"MYP R${result.myp_lowest_en_nm:,.2f} | "
                             f"TCG R${result.tcg_brl:,.2f} | "
                             f"{result.margin_pct * 100:.1f}%")
            if checkpoint_path:
                done.add(ed["url"])
                self._save_dbz_checkpoint(checkpoint_path, done)

        deals = [c for c in self.dbz_cards
                 if c.margin_pct and c.margin_pct >= self.margin_threshold]
        log.info("\n" + "═" * 60)
        log.info(f"  Produtos escaneados: {self._stats['products_scanned']}")
        log.info(f"  Cards EN NM ≥ piso: {len(self.dbz_cards) + len(self.semref)}")
        log.info(f"  Com referência tcgcsv: {self._stats['dbz_ref_ok']} | "
                 f"SEM referência: {self._stats['dbz_sem_ref']} "
                 f"(sem código no título: {self._stats['dbz_sem_codigo']})")
        log.info(f"  Edições sem grupo tcgcsv: {self._stats['dbz_edicoes_sem_grupo']}")
        log.info(f"  🔥 Deals (≥{self.margin_threshold * 100:.0f}%): {len(deals)}")
        log.info(f"  Flags: possível lixo={self._stats['dbz_flag_lixo']} | "
                 f"ref volátil={self._stats['dbz_flag_ref_volatil']} | "
                 f"truncation={self._stats['en_truncation_risks']}")
        log.info("═" * 60)

        if checkpoint_path and os.path.exists(checkpoint_path):
            try:
                os.remove(checkpoint_path)
            except OSError:
                pass
        return self.dbz_cards, self.semref


# ══════════════════════════════════════════════════════════════════════
# XLSX
# ══════════════════════════════════════════════════════════════════════

DBZ_HEADERS = [
    "Card Name", "EN Name", "Código", "Edition", "Seção", "Rarity",
    "MYP EN NM (R$)", "TCG US$", "TCG Low US$", "TCG Player (R$)",
    "TCG Source", "Subtipo", "Margin %", "Diff (R$)", "NM Sellers",
    "⚠️ EN Trunc", "⚠️ 1 Seller", "⚠️ Possível Lixo", "⚠️ Ref Volátil",
    "Join", "URL", "Updated", "TCG URL",
]
SEMREF_HEADERS = [
    "Card Name", "EN Name", "Código", "Edition", "Seção",
    "MYP EN NM (R$)", "NM Sellers", "Motivo", "URL",
]


def generate_dbz_xlsx(cards: list[DbzCardData], semref: list[SemRefRow],
                      output_path: str, threshold: float, fx: float,
                      stats: Optional[dict] = None,
                      editions_scanned: int = 0,
                      min_price: float = MIN_PRICE_BRL_DBZ) -> None:
    """XLSX de trabalho (insumo do myp_dbz_summary.py — a ENTREGA é o
    markdown do summary, nunca este arquivo)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="B03A2E")   # vermelho DBZ
    normal = Font(name="Arial", size=10)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    def write_headers(ws, headers, widths):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        ws.freeze_panes = "A2"

    # ── All EN Cards ──
    ws = wb.active
    ws.title = "All EN Cards"
    write_headers(ws, DBZ_HEADERS,
                  [34, 30, 11, 28, 12, 16, 14, 11, 11, 14, 10, 10, 10, 12,
                   10, 10, 10, 12, 12, 26, 55, 16, 55])
    margin_col = DBZ_HEADERS.index("Margin %") + 1
    for row, c in enumerate(
            sorted(cards, key=lambda x: -(x.margin_pct or -999)), start=2):
        vals = [
            c.name, c.en_name, c.code, c.edition,
            SECTION_LABELS.get(c.section, c.section), c.rarity,
            c.myp_lowest_en_nm, c.tcg_usd, c.tcg_low_usd, c.tcg_brl,
            c.tcg_source, c.subtype, c.margin_pct, c.margin_brl,
            c.en_nm_sellers,
            "⚠️ MAYBE" if c.en_truncation_risk else "",
            "⚠️ 1 SELLER" if c.single_en_seller_risk else "",
            "⚠️ LIXO?" if c.flag_lixo else "",
            "⚠️ VOLÁTIL" if c.flag_ref_volatil else "",
            c.join_via, c.product_url, c.last_updated, c.tcg_url,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = normal
            if col == margin_col and v is not None:
                cell.number_format = "0.00%"
                if v >= 0.50:
                    cell.fill = green_fill
                elif v >= threshold:
                    cell.fill = yellow_fill

    # ── Sem Ref TCG ──
    ws2 = wb.create_sheet("Sem Ref TCG")
    write_headers(ws2, SEMREF_HEADERS, [34, 30, 11, 28, 12, 14, 10, 46, 55])
    for row, s in enumerate(semref, start=2):
        vals = [s.name, s.en_name, s.code, s.edition,
                SECTION_LABELS.get(s.section, s.section),
                s.myp_lowest_en_nm, s.en_nm_sellers, s.motivo, s.product_url]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=row, column=col, value=v).font = normal

    # ── Summary ──
    ws3 = wb.create_sheet("Summary")
    deals = [c for c in cards if c.margin_pct and c.margin_pct >= threshold]
    pairs = [
        ("Scanner", f"myp_dbz_scanner {DBZ_VERSION} (Dragon Ball)"),
        ("Scan Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Margin Threshold", f"{threshold * 100:.0f}%"),
        ("Min Price (BRL)", min_price),
        ("FX USD→BRL", fx),
        ("Editions Scanned", editions_scanned),
        ("Total EN Cards", len(cards)),
        ("Sem Ref TCG", len(semref)),
        ("Deals Found", len(deals)),
    ]
    for k, v in list((stats or {}).items()):
        if k.startswith("dbz_") or k in (
                "products_scanned", "en_truncation_risks",
                "single_en_seller_risks", "skipped_no_en_sellers",
                "skipped_low_price", "skipped_jumbo", "http_retries"):
            pairs.append((f"stat:{k}", v))
    for row, (k, v) in enumerate(pairs, start=1):
        ws3.cell(row=row, column=1, value=k).font = Font(bold=True, name="Arial",
                                                         size=10)
        ws3.cell(row=row, column=2, value=v).font = normal
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 40

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info(f"  💾 XLSX: {out}")


# ══════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════

def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Scanner MYP → Dragon Ball (dbsfusion + dbsmasters vs "
                    "TCGplayer via tcgcsv). Paralelo ao scanner Pokémon.")
    ap.add_argument("--editions", nargs="*", default=None,
                    help="substrings de título/código de edição MYP (default: "
                         "TODAS as edições das seções escolhidas)")
    ap.add_argument("--sections", nargs="*", default=None,
                    choices=[s for s, _ in GAME_SECTIONS],
                    help="dbsfusion e/ou dbsmasters (default: ambas)")
    ap.add_argument("--threshold", type=float, default=30.0,
                    help="margem BRUTA mínima em PERCENT INTEIRO (30 = 30%%; "
                         "convenção MYP — CardTrader usa fração)")
    ap.add_argument("--min-price", type=float, default=MIN_PRICE_BRL_DBZ,
                    help="piso de relevância em R$ (default 50 — canônico)")
    ap.add_argument("--delay", type=float, default=1.5,
                    help="delay entre requests (s)")
    ap.add_argument("--min-en-sellers", type=int, default=2,
                    help="abaixo disso a linha ganha flag '1 SELLER' → REVISAR")
    ap.add_argument("--max-editions", type=int, default=0)
    ap.add_argument("--max-products", type=int, default=0)
    ap.add_argument("--fx", type=float, default=None,
                    help="câmbio USD→BRL manual (default: frankfurter/er-api "
                         "ao vivo; sem fonte real o run FALHA ALTO)")
    ap.add_argument("--cache-hours", type=float, default=20.0,
                    help="TTL do cache tcgcsv em disco (0 = sem cache)")
    ap.add_argument("--tcg-cache-dir", default="results/dbz_cache",
                    help="pasta do cache tcgcsv (dentro de results/, gitignored)")
    ap.add_argument("--resume", action="store_true",
                    help="retoma do checkpoint <output>.resume.json")
    ap.add_argument("--list-editions", action="store_true",
                    help="lista as edições DBZ do MYP e sai (sem scan)")
    ap.add_argument("-o", "--output", default=None,
                    help="XLSX de saída (default results/myp_dbz_<ts>.xlsx)")
    args = ap.parse_args(argv)

    threshold = parse_threshold(args.threshold)
    scraper = MYPDbzScraper(
        delay=args.delay, min_en_sellers=args.min_en_sellers,
        threshold=threshold, min_price=args.min_price,
        tcg_cache_dir=Path(args.tcg_cache_dir) if args.tcg_cache_dir else None,
        cache_hours=args.cache_hours,
    )

    if args.list_editions:
        sections = [s for s in GAME_SECTIONS
                    if not args.sections or s[0] in args.sections]
        for section, _cat in sections:
            for ed in scraper.get_section_editions(section):
                print(f"[{section}] {ed['title']}  [{ed.get('myp_code') or '—'}]"
                      f"  →  {ed['url']}")
        return 0

    # câmbio: manual OU ao vivo; sem fonte real = falha alta (a referência
    # DBZ é 100% USD — sem FX não existe margem honesta possível).
    fx = args.fx or fetch_usd_brl(scraper.session)
    if not fx or fx <= 0:
        log.error("Sem câmbio USD→BRL real (frankfurter/er-api falharam e "
                  "--fx não foi passado). Nunca inventar câmbio — abortando.")
        return 2

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or f"results/myp_dbz_{stamp}.xlsx"
    checkpoint = f"{output}.resume.json"

    cards, semref = scraper.scan_dbz(
        fx_usd_brl=fx, edition_filter=args.editions,
        sections_filter=args.sections, max_editions=args.max_editions,
        max_products=args.max_products, resume=args.resume,
        checkpoint_path=checkpoint,
    )
    generate_dbz_xlsx(cards, semref, output, threshold, fx,
                      stats=scraper._stats,
                      editions_scanned=len({c.edition for c in cards}
                                           | {s.edition for s in semref}),
                      min_price=args.min_price)
    print(f"\nOK: {output}")
    print(f"Entrega: python myp_dbz_summary.py {output} -o "
          f"results/dbz-{datetime.now().strftime('%Y-%m-%d')}.md")
    return 0


if __name__ == "__main__":
    sys.exit(main())
