"""Aggregate múltiplos chunks de myp_arbitrage_scanner.py em um único XLSX.

Usado pelo GH Actions matrix job: cada chunk gera um xlsx parcial, depois
este script merge tudo em um relatório consolidado idêntico ao que um run
single-thread produziria.

Reconstrói objetos CardData a partir das rows da sheet "All EN Cards" de
cada chunk, dedupe por product_url (chunks são interleaved então em teoria
não deveria haver duplicata, mas defensivo), e re-invoca generate_xlsx pra
preservar formatação exata.

Uso:
    python myp_aggregate.py chunk_*.xlsx -o myp_arbitrage_FULL.xlsx
"""
from __future__ import annotations

import argparse
import glob
import logging
import sys
from pathlib import Path

# UTF-8 stdout obrigatório no Windows (emoji nas sheets)
if sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from openpyxl import load_workbook

# Reusa CardData + generate_xlsx do scanner (single source of truth)
from myp_arbitrage_scanner import CardData, generate_xlsx

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger(__name__)


def card_from_row(headers: list[str], row: tuple) -> CardData | None:
    """Reconstrói CardData a partir de uma row de 'All EN Cards' sheet."""
    rec = dict(zip(headers, row))
    name = rec.get("Card Name")
    if not name:
        return None
    card = CardData()
    card.name = name
    card.edition = rec.get("Edition") or ""
    card.rarity = rec.get("Rarity") or ""
    card.myp_lowest_en_nm = rec.get("MYP EN NM (R$)")
    card.tcg_player_price = rec.get("TCG Player (R$)")
    # v5.11.1 (2026-06-09): preservar preço real em USD entre chunks pra a
    # tabela de ENTREGA (myp_summary.py). .get() → None em chunks antigos.
    card.tcg_real_usd = rec.get("TCG US$")
    # v5.14 (2026-06-20): preservar a FONTE do preço entre chunks/round-trips
    # (real pokemontcg.io/tcgcsv vs fallback .estat-tcg). É o sinal de
    # honestidade do output. Chunks antigos não têm a coluna "TCG Source" →
    # infere pela presença de "TCG US$" (real só era populado no caminho real),
    # mantendo o comportamento legado sem mascarar fallback como real.
    #
    # v5.15.1 (2026-06-22) BUG FIX: a célula "TCG Source" guarda o RÓTULO LEGÍVEL
    # que generate_xlsx escreve ("real (tcgcsv)", "real (pokemontcg.io)",
    # "fallback (.estat-tcg)") — NÃO o token interno (`tcgcsv`/`pokemontcg.io`/
    # `myp_estat`). O parser antigo só reconhecia a substring "pokemontcg", então
    # TODO chunk via tcgcsv (a fonte do CI desde v5.15) caía em "myp_estat" →
    # o consolidado mostrava 0 real / 100% fallback mesmo com 537 cards REAIS
    # gravados pelos chunks. Aqui reconstruímos o token interno do rótulo (inverso
    # de _REAL_SOURCES de generate_xlsx): tcgcsv e pokemontcg.io são REAIS e
    # preservados; só "fallback" (ou rótulo desconhecido) vira myp_estat.
    # Honestidade dura: real só permanece real; fallback continua fallback.
    _src = rec.get("TCG Source")
    if _src:
        _src_lc = str(_src).lower()
        if "tcgcsv" in _src_lc:
            card.tcg_source = "tcgcsv"
        elif "pokemontcg" in _src_lc:
            card.tcg_source = "pokemontcg.io"
        else:
            card.tcg_source = "myp_estat"
    else:
        # Chunk legado sem a coluna "TCG Source": real só era populado no caminho
        # real (com USD), então infere pela presença de "TCG US$".
        card.tcg_source = "pokemontcg.io" if card.tcg_real_usd not in (None, "", "—") else "myp_estat"
    # v5.8 (2026-05-16): preservar sanity-check fields entre chunks. Aggregate
    # estava strip-ando tcg_suspect → consolidated XLSX volta a mostrar Jirachi
    # como deal #1. .get() retorna None se chunk antigo não tem essas colunas.
    card.myp_last_sale_brl = rec.get("MYP Last Sale (R$)")
    card.tcg_suspect = bool(rec.get("⚠️ TCG Suspect"))
    card.margin_pct = rec.get("Margin %")
    # Diff é calculado em generate_xlsx, não precisamos preservar
    card.en_nm_sellers = rec.get("NM Sellers") or 0
    card.en_truncation_risk = bool(rec.get("⚠️ EN Trunc"))
    # v5.8.3 (2026-05-18): preservar flag de single-seller risk entre chunks
    card.single_en_seller_risk = bool(rec.get("⚠️ Single Seller"))
    # v5.8.5 (2026-05-19): preservar flag de oversized-collector-risk entre
    # chunks. Chunks antigos sem essa coluna retornam None → False.
    card.oversized_collector_risk = bool(rec.get("⚠️ COLLECTOR#"))
    card.product_url = rec.get("URL") or ""
    card.last_updated = rec.get("Updated") or ""
    # margin_brl: derived
    if card.tcg_player_price and card.myp_lowest_en_nm:
        card.margin_brl = card.tcg_player_price - card.myp_lowest_en_nm
    return card


def load_chunk_cards(xlsx_path: Path) -> list[CardData]:
    """Carrega todas as cards de um chunk XLSX."""
    cards: list[CardData] = []
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"  Falha abrindo {xlsx_path}: {e} — pulando")
        return cards

    sheet_name = "All EN Cards"
    if sheet_name not in wb.sheetnames:
        log.warning(f"  {xlsx_path} sem sheet '{sheet_name}' (sheets: {wb.sheetnames}) — pulando")
        return cards

    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return cards
        headers = list(rows[0])
        for r in rows[1:]:
            card = card_from_row(headers, r)
            if card:
                cards.append(card)
        return cards
    finally:
        # read_only=True mantém o handle do arquivo aberto até .close()
        # explícito (no Windows isso trava unlink/reescrita do XLSX). Fecha
        # sempre — sem isso o enrich não consegue reescrever sobre o input.
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Aggregate múltiplos chunks XLSX em um único relatório MYP",
    )
    parser.add_argument("inputs", nargs="+",
                        help="Caminhos dos XLSX de chunks (aceita glob: chunk_*.xlsx)")
    parser.add_argument("-o", "--output", required=True,
                        help="Caminho do XLSX consolidado")
    parser.add_argument("--threshold", type=float, default=0.25,
                        help="Margem mínima como FRAÇÃO (default 0.25 = 25%%) — só pra formatação das sheets")
    args = parser.parse_args()

    # Expand globs (Windows shell não expande)
    paths: list[Path] = []
    for p in args.inputs:
        matches = glob.glob(p)
        if matches:
            paths.extend(Path(m) for m in matches)
        else:
            # Não é glob ou não casou — tratar como path literal
            paths.append(Path(p))

    if not paths:
        log.error("Nenhum chunk XLSX encontrado.")
        return 1

    log.info(f"Aggregating {len(paths)} chunks:")
    for p in paths:
        log.info(f"  • {p}")

    all_cards: list[CardData] = []
    for p in paths:
        chunk_cards = load_chunk_cards(p)
        log.info(f"  {p.name}: {len(chunk_cards)} cards")
        all_cards.extend(chunk_cards)

    # Dedupe por product_url (defensivo — interleaved chunks NÃO deveriam
    # gerar duplicatas, mas se o operador rodar 2 chunks que se sobrepõem
    # acidentalmente, evita inflar o relatório)
    seen_urls: set[str] = set()
    deduped: list[CardData] = []
    duplicates = 0
    for c in all_cards:
        key = c.product_url or f"{c.name}|{c.edition}"
        if key in seen_urls:
            duplicates += 1
            continue
        seen_urls.add(key)
        deduped.append(c)

    if duplicates:
        log.warning(f"  {duplicates} duplicates removidas (mesma product_url ou name+edition)")

    log.info(f"Total: {len(deduped)} cards únicos consolidados")

    if not deduped:
        log.error("Zero cards após aggregação — não há nada pra gerar XLSX.")
        return 1

    generate_xlsx(deduped, args.output, args.threshold)
    log.info(f"OK consolidated XLSX: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
