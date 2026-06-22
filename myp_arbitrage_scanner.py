#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║         MYP Cards Arbitrage Scanner — Pokémon TCG Singles          ║
║                                                                      ║
║  Compara preços de singles (EN) no mypcards.com vs TCG Player.     ║
║  Gera planilha .xlsx com alertas de arbitragem (margem >= 30%).    ║
╚══════════════════════════════════════════════════════════════════════╝

Uso:
    python myp_arbitrage_scanner.py                          # Scan completo
    python myp_arbitrage_scanner.py --max-editions 5         # Apenas 5 edições
    python myp_arbitrage_scanner.py --threshold 40           # Margem mínima 40%
    python myp_arbitrage_scanner.py --delay 2.0              # 2s entre requests

Requisitos:
    pip install cloudscraper beautifulsoup4 openpyxl lxml

Autor: Matheus Chillemi / Claude
Data: 2026-04-15 (v5) | 2026-05-12 (v5.1 → v5.3) | 2026-05-14 (v5.4 → v5.6) | 2026-05-16 (v5.8) | 2026-05-19 (v5.8.4 → v5.8.6) | 2026-05-29 (v5.8.7 → v5.8.9) | 2026-06-01 (v5.8.10) | 2026-06-03 (v5.9) | 2026-06-06 (v5.10) | 2026-06-07 (v5.10.1 → v5.11) | 2026-06-09 (v5.11.1) | 2026-06-10 (v5.11.2 → v5.11.3) | 2026-06-16 (v5.11.4 → v5.11.6) | 2026-06-13 (v5.11.7, doc-only) | 2026-06-17 (v5.11.8 — loop: timing + bench) | 2026-06-17 (v5.12 — batch pokemontcg.io por set) | 2026-06-17 (v5.13 — Iteração #2: atribuição de cobertura do fallback) | 2026-06-20 (v5.14 — coluna "TCG Source" explícita + enrich off-runner p/ preço real) | 2026-06-20 (v5.14.1 — cobertura de preço real no summary medida sobre o universo de cartas EN) | 2026-06-21 (v5.14.3 — deal com preço FALLBACK sai do balde "limpos" → balde dedicado; fix BLOCKER de honestidade) | 2026-06-21 (v5.14.4 — tcg_suspect boundary inclusivo `>=` (pega exatamente-10x); regressão de precisão minerada do eval asi-evolve)
Versão: v5.16

Changelog v5.1 (2026-05-12 — auditoria C/H/M, mesma metodologia do CT scanner):
  - C1: --threshold < 1.0 auto-converte com warning (UX guard contra trap
    inverso ao CT scanner — MYP usa percent integer, CT usa fração)
  - H3: detecção heurística SIR/HR/SAR — warning quando rarity="Comum" mas
    TCG price alto (>R$200). Reduz falso positivo documentado em memória.
  - M1: HTTP retry com backoff (3 tentativas, 2s→4s) em transient errors
  - M4: debug_*.html agora salvo em subpasta .debug/ do script,
    não polui CWD
  - M5: novos stat counters (skipped_no_tcg, skipped_no_en_sellers,
    skipped_low_price) pra auditoria do funnel

Changelog v5.2 (2026-05-12):
  - Default --threshold de 35 → 25 (mais discovery, menos filtragem)
  - Nova sheet "🏆 Top 50 Margin" no xlsx: cards ordenados por margem
    decrescente sem filtro, pra inspeção visual chase-card

Changelog v5.3 (2026-05-12 — após caso Psyduck/bartsimpson):
  - T1: novo campo CardData.en_truncation_risk; parser itera por seller table
    individualmente (Tabela 0=lojistas/15-cap, Tabela 1=marketplace/20-cap).
    Heurística refinada: dispara só quando uma tabela está no cap (≥15 rows),
    com zero EN visível, E max_price visível < lowest_en reportado (= hidden
    listings podem ser EN mais baratos que o reportado). Evita false alarm
    quando max visível já está acima do lowest_en (hidden não pode quebrar).
  - H3 refinada: agora também exige card_num > set_total quando o sufixo
    (X/Y) é extraível do nome — evita falso alarm em commons in-set caros.
  - Nova sheet "🚨 Validate Manually" no xlsx: lista os cards com
    en_truncation_risk pra punch-list de validação manual.
  - Nova coluna "⚠️ EN Trunc" nas sheets de cards.
  - Novo stat counter en_truncation_risks no summary final.
  - Bug fix: pricing promocional. Rows com "R$ X (riscado) R$ Y" usavam X
    (preço antigo, mais caro) via re.search; agora re.findall + [-1] pega
    Y (preço ativo). Caso: MatchampTCG Psyduck "R$ 275,00 R$ 220,00" lido
    como R$275 quando deveria ser R$220.
"""

# v5.6.1 fix: requests é transitive dep do cloudscraper E é referenciado em
# `except (requests.RequestException, ...)` no _get retry loop (v5.4 C3 fix).
# Antes era importado APENAS no fallback ImportError do cloudscraper, causando
# NameError em qualquer setup que tenha cloudscraper (todos os production runs).
import os
import requests
try:
    import cloudscraper
    HAS_CLOUDSCRAPER = True
except ImportError:
    HAS_CLOUDSCRAPER = False

from bs4 import BeautifulSoup
import re
import time
import logging
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import Optional
from pathlib import Path
from urllib.parse import quote_plus

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════
BASE_URL = "https://mypcards.com"
MARGIN_THRESHOLD = 0.30          # v5.10: 30% margem BRUTA mínima para alerta (era 0.25). Política cross-scanner 2026-06-06: 30% margem bruta = só (preço_alvo − preço_BR)/preço_BR, SEM taxa embutida; operador calcula custos por fora.
MIN_PRICE_BRL = 50.0             # piso padrão cross-scanner: "carta valiosa" = > R$50 (ignora cartas baratas)
REQUEST_DELAY = 1.5              # segundos entre requests
MAX_PAGES_PER_EDITION = 30       # max páginas por edição
MAX_EDITION_PAGES = 50           # v5.4 H4: cap em get_all_editions (evita infinite loop)
# v5.9 (2026-06-03): a tabela marketplace (#lista-anuncio-demais-vendedores)
# pagina via ?estoque-outros-page=N e ordena por preço crescente across-idiomas.
# Quando a página 1 enche de PT/JP baratos, os EN-NM caem em páginas 2+. Cap de
# páginas a seguir por produto (evita loop infinito + limita custo de scan).
MAX_SELLER_PAGES = 10            # v5.9: max páginas da tabela marketplace por produto
MIN_EDITIONS_EXPECTED = 200      # v5.4 C2: catalog scrape sanity floor (~326 esperado, alarme em <200)
# v5.11.4 (2026-06-16): checkpoint/resume. O container da nuvem é reciclado na
# inatividade e mata o processo — mas o disco sobrevive. Salvando o progresso por
# edição (cards + edições já feitas) num sidecar `<output>.resume.json`, um
# `--resume` retoma de onde parou em vez de perder horas de scan.
CHECKPOINT_VERSION = 1
TIMEOUT = 20                     # timeout HTTP em segundos
HTTP_MAX_RETRIES = 3             # M1 fix: retries em transient errors
DEBUG_DIR = Path(__file__).resolve().parent / ".debug"   # M4 fix: subpasta dedicada
SUPRANUMERARY_PRICE_THRESHOLD = 200.0  # H3 fix: TCG R$ acima disso + rarity="Comum" = SIR/HR suspeito
# v5.8.3 (2026-05-18): cartas Jumbo (oversized ~25×35cm) têm mercado/preço
# distintos da versão standard. MYP agrupa standard + jumbo na MESMA página de
# produto; a variante é indicada por seller-row na coluna `.estoque-lista-nomeenfoil`
# ("Foil"). Detectamos e excluímos rows Jumbo da contagem EN. Caso M-Rayquaza-EX
# 098/98 XY 7 (produto 32737): h1 sem "Jumbo" mas 5 sellers com Jumbo no Foil col
# inflavam o min preço EN.
#
# v5.8.4 (2026-05-19): regex broader cobre também 'oversized', 'box topper',
# 'poster card'. MYP usa qualquer um desses pra produtos físicos não-standard.
# \b word-boundary em todos pra consistência (foil_re antes não tinha; rara
# colisão tipo "jumbocard" não documentada mas defensivo).
OVERSIZED_FOIL_RE = re.compile(
    r"\b(jumbo|oversized)\b", re.IGNORECASE,
)
# Filtro por título como segunda camada — caso MYP liste oversized como
# produto standalone (sem distinção via coluna foil).
OVERSIZED_TITLE_RE = re.compile(
    r"\b(jumbo|oversized|box\s?topper|poster\s?card)\b", re.IGNORECASE,
)
# Regex único pro padrão de preço BRL ('R$ 1.900,00'). Centralizado porque
# estava duplicado em 5 call-sites (TCG price, última venda, row price + 2 no
# revalidate_deals.py); drift no markup do MYP quebraria todos de uma vez.
PRICE_RE = re.compile(r'R\$\s*[\d.,]+')

# v5.8.7 (2026-05-29): o <h1> da página de produto MYP concatena, sem
# separador, o título PT "Nome (NNN/MMM)" seguido de uma cópia do nome EN.
# `h1.get_text(strip=True)` colapsa isso em strings como
# "Heatran-EX (109/116)Heatran-EX" ou "Kyogre da Equipe Aqua (003/95)Team
# Aqua's Kyogre". No XLSX 2026-05-27 isso atingia 275/1190 rows (23%).
# Dois problemas: (1) copy-paste sujo (operador cola o nome no MYP/TCGplayer
# pra buscar a carta); (2) o `merge_myp_ct.py` casa cartas via
# NUM_IN_NAME_RE = r"\(\s*(\d+)\s*/\s*(\d+)\s*\)\s*$" — ANCORADO no fim da
# string. O lixo após "(NNN/MMM)" fazia o regex falhar → essas 275 linhas
# eram silenciosamente descartadas do índice de cross-reference.
#
# Fix: quando o nome contém "(NNN/MMM)", trunca logo após o ")", preservando
# EXATAMENTE o formato "(NNN/MMM)" que o merge depende. Promos "(PR-...)",
# formatos "RCxx/RCyy", e nomes sem número ficam intocados (zero regressão).
NAME_NNN_MMM_RE = re.compile(r"^(.*?\(\s*\d+\s*/\s*\d+\s*\))")


def _clean_secret(value: Optional[str]) -> Optional[str]:
    """Sanitiza um segredo (API key) lido de env/secret do CI antes de usá-lo
    num header HTTP.

    Remove BOM (U+FEFF), zero-width space (U+200B) e espaços/quebras nas pontas.
    O `requests` codifica headers HTTP em latin-1; um BOM grudado no secret por
    engano (arquivo salvo como UTF-8-with-BOM, copy/paste do site) vira
    `UnicodeEncodeError: 'latin-1' codec can't encode '\\ufeff'` e derruba 100%
    das chamadas — foi exatamente o que abortou o scan do scanner irmão no
    GitHub Actions (mass pricing failure, scan "verde" mas vazio). `str.strip()`
    NÃO remove BOM (U+FEFF não é whitespace pra Python), então tratamos
    explicitamente. Retorna None se sobrar vazio — caller trata como 'sem key'
    (sem header X-Api-Key, que é fallback válido na pokemontcg.io).
    """
    if value is None:
        return None
    cleaned = value.replace("\ufeff", "").replace("\u200b", "").strip()
    return cleaned or None


def clean_card_name(raw: str) -> str:
    """Trunca o título do <h1> logo após '(NNN/MMM)' quando presente,
    removendo o nome EN duplicado que o MYP concatena sem separador.

    Mantém intacto o formato '(NNN/MMM)' (load-bearing pro merge) e não
    altera nomes que não casam o padrão (promos, RCxx, sem número)."""
    if not isinstance(raw, str):
        return ""
    raw = raw.strip()
    m = NAME_NNN_MMM_RE.match(raw)
    return m.group(1).strip() if m else raw


# v5.8.8 (2026-05-29): TCGplayer search-by-name URL para hyperlink na célula
# de preço "TCG Player (R$)". O scanner é HTML-scrape e a página de produto MYP
# NÃO embute tcg_productId nem link TCGplayer (probe 2026-05-29: 0 hits);
# o `mypcards.com/api/v1` que embutiria tcg_productId está 404 hoje (instável
# desde 2026-05-07). Logo, NÃO há link DIRETO de produto barato/estável — o
# fallback é busca por nome. Remove o sufixo "(NNN/MMM)"/"(PR-...)" pra busca
# mais limpa (o número não ajuda na busca TCGplayer e atrapalha o match).
_TCG_SEARCH_BASE = "https://www.tcgplayer.com/search/pokemon/product?productLineName=pokemon&q="

# v5.8.9 (2026-05-29): mapa MYP edition substring → pokemontcg.io setcode pra
# montar URL DIRETA de produto TCGplayer via redirect
# `https://prices.pokemontcg.io/tcgplayer/{setcode}-{num}` (pokemontcg.io
# resolve pra `tcgplayer.com/product/<id>` da carta exata, mesma mecânica que
# o CardTrader scanner usa no `Link TCG`). Custo zero (string build).
#
# Mapa derivado de:
#   - C:/Users/mathe/Scripts/merge_myp_ct.py CT_ALIAS_TO_MYP_SUBSTRING
#     (MYP edition substring ↔ CT 3-letter alias)
#   - CardTrader Scanner SET_ALIAS_TO_PTCG (CT alias ↔ pokemontcg.io setcode,
#     SV+SWSH+ME eras, verificado 2026-05-19 contra pokemontcg.io /sets)
#
# Chain MYP substring → CT alias → ptcg setcode. Sets-base validados via
# probe HTTP 2026-05-29 (all 25 setcodes returnam 200 no redirect base).
# Vintage / promo / pre-SWSH NÃO mapeados — fallback de busca-por-nome cobre.
#
# Substrings usam forma EN ("Silver Tempest", "Evolving Skies"). Funciona
# mesmo com o bilingual concat bug do MYP (ex "Sword & Shield 7: ...Evolving
# Skies"): a porção EN está sempre presente no concat.
MYP_EDITION_SUBSTR_TO_PTCG = {
    # Scarlet & Violet era (longest-substring match prevention pra "SV09":
    # listar substring específica em vez de "151"-ish ambiguidade)
    "Scarlet & Violet: Destined Rivals": "sv10",
    "SV09: Journey Together":            "sv9",
    "Prismatic Evolutions":              "sv8pt5",
    "Surging Sparks":                    "sv8",
    "Stellar Crown":                     "sv7",
    "Shrouded Fable":                    "sv6pt5",
    "Twilight Masquerade":               "sv6",
    "Temporal Forces":                   "sv5",
    "Paldean Fates":                     "sv4pt5",
    "Paradox Rift":                      "sv4",
    "Obsidian Flames":                   "sv3",
    # v5.8.10 (2026-05-30, MYP-M2 fix): chave era "Pokémon 151" mas MYP emite
    # 'Escarlate e Violeta: 151Scarlet & Violet—151' (sem "Pokémon"). Trocado
    # pra "151" curto — nenhum outro set no mapa contém "151" como substring
    # (auditado: Paldea Evolved, Paradox Rift, etc. — sem conflito). Longest-
    # substring match ainda protegeria caso futuro set tipo "151ish" surja.
    "151":                               "sv3pt5",
    "Paldea Evolved":                    "sv2",
    # Mega Evolution era (commit 2026-05-19): pokemontcg.io ainda tem
    # cobertura parcial pros ME sets (`me2pt5-274` 404 em 2026-05-29 mas
    # `me2pt5-10` 200). Quando oversized_collector_risk=True, write_card_row
    # CAI no fallback de busca (variant fora de range é o caso 404 típico).
    #
    # ⚠️ ACHADO 2026-06-17 (Iteração #2, medido no quick ao vivo): a era ME
    # INTEIRA está com **0% de preço TCGplayer** no pokemontcg.io — `me2pt5`,
    # `me3` (Perfect Order), `me4` (Chaos Rising) retornam 200 mas o objeto
    # `tcgplayer` traz só `url`, sem `prices` (vs SV = 100%, atualizado no dia).
    # CONSEQUÊNCIA: NÃO adicionar `Perfect Order`→me3 / `Chaos Rising`→me4
    # esperando reduzir fallback — não recupera preço nenhum, só troca o card de
    # `unmapped_set` pra `no_coverage`. Os "deals" ME (Ascended Heroes etc.) saem
    # do `.estat-tcg` e por isso caem nos baldes validar-manualmente. Só um preço
    # TCGplayer alternativo (scrape direto / outra fonte) destrava margem real
    # pra ME (item de backlog; o playbook docs/optimization-loop.md é local-only/
    # gitignored e pode não existir no clone — ver seção "Otimizar" no CLAUDE.md).
    "Mega Evolution: Phantasmal Flames": "me2",
    "Ascended Heroes":                   "me2pt5",
    "Mega Evolution":                    "me1",   # base ME (catch-all curto, longest-substr win pros específicos acima)
    # Sword & Shield era
    "Crown Zenith":                      "swsh12pt5",
    "Silver Tempest":                    "swsh12",
    "Lost Origin":                       "swsh11",
    "Astral Radiance":                   "swsh10",
    "Brilliant Stars":                   "swsh9",
    "Fusion Strike":                     "swsh8",
    "Evolving Skies":                    "swsh7",
    # Special sets (não-numerados, fora das eras principais)
    "Pokémon GO":                        "pgo",  # MYP-LOW-d 2026-05-30: special set SWSH-era, não numerado
    # v5.11 (2026-06-07): Black Bolt / White Flare ADICIONADOS. Probe ao vivo
    # confirmou cobertura pokemontcg.io estável base+oversized (zsv10pt5-1
    # Snivy, zsv10pt5-97 Darumaka IR, zsv10pt5-172 Zekrom ex, rsv10pt5-102
    # Lampent, rsv10pt5-168 Jellicent ex — todos 200 com preço TCGplayer real).
    # Motivava o switch p/ preço real: o `.estat-tcg` do MYP nesses sets base-086
    # mapeia a carta errada (Darumaka declarado R$2.867 vs real US$13,42).
    "Black Bolt":                        "zsv10pt5",
    "White Flare":                       "rsv10pt5",
    # ══════════════════════════════════════════════════════════════════
    # v5.16 (2026-06-22): EXPANSÃO de cobertura de set (eras SM/XY/SWSH-
    # antigo/BW/DP/EX/e-Card/Neo/Gym + alguns clássicos). Cada substring
    # abaixo foi VERIFICADA contra a lista REAL de edições do MYP (scrape ao
    # vivo de 362 edições, 2026-06-22) — match ÚNICO por título (zero
    # colisão/ambiguidade na simulação longest-substring-wins sobre TODOS os
    # títulos reais) — E cada setcode tem abbreviação tcgcsv resolvendo 1-a-1
    # contra /groups (ver PTCG_SETCODE_TO_TCGCSV_ABBR). Cobertura de edições
    # mapeadas: 32→112 / 362. Substrings escolhidas em EN (o MYP emite título
    # bilíngue PT+EN concatenado, ex.: 'Sol & Lua 7…Sun & Moon 7: Celestial
    # Storm' contém a substring EN). Onde o MYP usa nome distintivo do set
    # (DP2 vem como 'Diamond & PEARLS 2' — typo plural do MYP), a chave é o
    # nome do set ('Mysterious Treasures') que casa 1 título só. NÃO mapeado
    # de propósito (incerto/ambíguo, deixa em fallback honesto): EX7 Team
    # Rocket Returns + 'Team Rocket' base (substring colide c/ Returns/JP),
    # XY9 BREAKpoint + Platinum2 Rising Rivals (abbr tcgcsv ambígua), Call of
    # Legends + TCG Classic (abbr 'CL' ambígua), e TODOS os sets só-JP
    # (Eevee Heroes, VMAX Climax, etc. — sem print EN no TCGplayer/tcgcsv).
    # Sun & Moon era
    "Sun & Moon 2: Guardians Rising":            "sm2",
    "Sun & Moon 3: Burning Shadows":             "sm3",
    "Sun & Moon 3.5: Shining Legends":           "sm35",
    "Sun & Moon 4: Crimson Invasion":            "sm4",
    "Sun & Moon 5: Ultra Prism":                 "sm5",
    "Sun & Moon 6: Forbidden Light":             "sm6",
    "Sun & Moon 7: Celestial Storm":             "sm7",
    "Sun & Moon 7.5: Dragon Majesty":            "sm75",
    "Sun & Moon 8: Lost Thunder":                "sm8",
    "Sun & Moon 9: Team Up":                     "sm9",
    "Sun & Moon 10: Unbroken Bonds":             "sm10",
    "Sun & Moon 11: Unified Minds":              "sm11",
    "Sun & Moon 11.5: Hidden Fates":             "sm115",
    "Sun & Moon 12: Cosmic Eclipse":             "sm12",
    # XY era
    "XY 2: Flashfire":                           "xy2",
    "XY 3: Furious Fists":                       "xy3",
    "XY 4: Phantom Forces":                      "xy4",
    "XY 5: Primal Clash":                        "xy5",
    "XY 6: Roaring Skies":                       "xy6",
    "XY 7: Ancient Origins":                     "xy7",
    "XY 8: BREAKthrough":                        "xy8",
    "XY 10: Fates Collide":                      "xy10",
    "XY 11: Steam Siege":                        "xy11",
    "XY 12: Evolutions":                         "xy12",
    "XY: Double Crisis":                         "dc1",
    "XY: Kalos Starter Set":                     "xy0",
    # Sword & Shield (mais antigos que os já mapeados acima)
    "Sword & Shield 2: Rebel Clash":             "swsh2",
    "Sword & Shield 3: Darkness Ablaze":         "swsh3",
    "Sword & Shield 3.5: Champion's Path":       "swsh35",
    "Sword & Shield 4: Vivid Voltage":           "swsh4",
    "Sword & Shield 5: Battle Styles":           "swsh5",
    "Sword & Shield 6: Chilling Reign":          "swsh6",
    "Shining Fates":                             "swsh45",
    "Celebrations: Classic Collection":          "cel25c",
    # HeartGold & SoulSilver era
    "HeartGold & SoulSilver 2: Unleashed":       "hgss2",
    "HeartGold & SoulSilver 3: Undaunted":       "hgss3",
    "HeartGold & SoulSilver 4: Triumphant":      "hgss4",
    # Black & White era
    "Black & White 2: Emerging Powers":          "bw2",
    "Black & White 3: Noble Victories":          "bw3",
    "Black & White 4: Next Destinies":           "bw4",
    "Black & White 5: Dark Explorers":           "bw5",
    "Black & White 6: Dragons Exalted":          "bw6",
    "Black & White 7: Boundaries Crossed":       "bw7",
    "Black & White 8: Plasma Storm":             "bw8",
    "Black & White 9: Plasma Freeze":            "bw9",
    "Black & White 10: Plasma Blast":            "bw10",
    "Black & White: Dragon Vault":               "dv1",
    # Diamond & Pearl / Platinum era
    "Platinum 3: Supreme Victors":               "pl3",
    "Platinum 4: Arceus":                        "pl4",
    "Mysterious Treasures":                      "dp2",   # MYP titula 'Diamond & PEARLS 2' (typo) → casa pelo nome do set
    "Secret Wonders":                            "dp3",
    "Great Encounters":                          "dp4",
    "Majestic Dawn":                             "dp5",
    "Legends Awakened":                          "dp6",
    "Diamond & Pearl 7: Stormfront":             "dp7",
    # EX era
    "EX 1: Ruby & Sapphire":                     "ex1",
    "EX 2: Sandstorm":                           "ex2",
    "EX 3: Dragon":                              "ex3",
    "EX 4: Team Magma vs Team Aqua":             "ex4",
    "EX 5: Hidden Legends":                      "ex5",
    "EX 6: Fire Red & Leaf Green":               "ex6",
    "EX 8: Deoxys":                              "ex8",
    "EX 9: Emerald":                             "ex9",
    "EX 10: Unseen Forces":                      "ex10",
    "EX 11: Delta Species":                      "ex11",
    "EX 12: Legend Maker":                       "ex12",
    "EX 13: Holon Phantoms":                     "ex13",
    "EX 14: Crystal Guardians":                  "ex14",
    "EX 15: Dragon Frontiers":                   "ex15",
    "EX 16: Power Keepers":                      "ex16",
    # e-Card era
    "E-Card 1: Expedition Base Set":             "ecard1",
    "E-Card 2: Aquapolis":                       "ecard2",
    "E-Card 3: Skyridge":                        "ecard3",
    # WOTC clássicos
    "Neo Genesis":                               "neo1",
    "Neo Discovery":                             "neo2",
    "Neo Revelation":                            "neo3",
    "Neo Destiny":                               "neo4",
    "Gym Heroes":                                "gym1",
    "Gym Challenge":                             "gym2",
    "Legendary Collection":                      "base6",
}

# Regex (NNN/MMM) — captura numerator e denominator. Reutilizado de
# write_card_row L871. Definido aqui pra tcg_direct_url também.
_COLLECTOR_NUM_RE = re.compile(r"\((\d+)\s*/\s*(\d+)\)")

# ══════════════════════════════════════════════════════════════════════
# v5.11 (2026-06-07): PREÇO TCG REAL via pokemontcg.io + câmbio USD→BRL
# ──────────────────────────────────────────────────────────────────────
# Antes (≤v5.10.1) o "TCG R$" vinha do campo `.estat-tcg` da página MYP — um
# número que o MYP declara. Em sets base-086 (Black Bolt/White Flare) e parte de
# Destined Rivais esse campo mapeia a carta errada → preço furado (Darumaka
# 097/086: MYP declarava R$2.867 vs TCGplayer real US$13,42). A partir da v5.11
# o scanner busca o preço REAL do TCGplayer via pokemontcg.io (USD) e converte
# pra BRL com câmbio ao vivo, com FALLBACK pro `.estat-tcg` do MYP onde o
# pokemontcg.io não tem cobertura (ex.: alguns Mega — me2pt5-269 sem preço).
# ══════════════════════════════════════════════════════════════════════
PTCG_API_BASE = "https://api.pokemontcg.io/v2/cards/"

# ══════════════════════════════════════════════════════════════════════
# v5.15 (2026-06-21): PREÇO TCG REAL via tcgcsv.com (fonte que FUNCIONA no CI)
# ──────────────────────────────────────────────────────────────────────
# Achado empírico (sonda probe-price-sources.yml run 27918333945): os runners
# do GitHub Actions ALCANÇAM `tcgcsv.com` (HTTP 200, JSON real), enquanto
# `api.pokemontcg.io` é bloqueado pelo CF da API nos IPs de datacenter do
# GitHub/Azure → toda chamada cai no fallback `.estat-tcg` (margens infladas).
# tcgcsv.com é um dump diário grátis dos preços do TCGplayer; cross-check local
# 2026-06-21 confirmou que o preço tcgcsv concorda com o pokemontcg.io em
# 0–1,5% (sv7-1: 1,34 vs 1,36; sv7-2: 0,06 idêntico) — é o MESMO preço
# TCGplayer, só capturado por outra rota. BÔNUS: tcgcsv TEM preço pros sets ME
# (Ascended Heroes etc.) que o pokemontcg.io devolve SEM `prices` (a era ME
# inteira estava 0% de cobertura — ver MYP_EDITION_SUBSTR_TO_PTCG nota ME).
#
# Schema confirmado (categoria 3 = Pokémon):
#   GET /tcgplayer/3/groups            → {results:[{groupId, name, abbreviation}]}
#   GET /tcgplayer/3/<groupId>/products→ {results:[{productId, name,
#                                          extendedData:[{name:"Number",
#                                          value:"001/142"}, ...]}]}
#   GET /tcgplayer/3/<groupId>/prices  → {results:[{productId, lowPrice,
#                                          midPrice, marketPrice, subTypeName}]}
#     subTypeName ∈ {Normal, Holofoil, Reverse Holofoil}; 1 produto pode ter
#     várias linhas (1 por subtype). Header `User-Agent` é OBRIGATÓRIO (sem ele
#     a API devolve 401). O preço usado é o MESMO `_min_tcg_usd` (menor market/
#     mid entre subtypes) — resultado idêntico ao caminho pokemontcg.io.
#
# PONTE set→groupId: o scanner já mapeia edição MYP → setcode pokemontcg.io
# (MYP_EDITION_SUBSTR_TO_PTCG). Aqui ligamos setcode → tcgcsv groupId por
# abreviação/nome do group (resolvido 1× contra /groups, cacheado). O cache de
# preço (`_ptcg_cache`) segue keyed por `{setcode}-{num}` — os dois provedores
# preenchem a MESMA estrutura, então TODO o caminho de margem a jusante
# (`_real_tcg_brl`, override, margem) é reusado sem mudança.
# ══════════════════════════════════════════════════════════════════════
TCGCSV_BASE = "https://tcgcsv.com/tcgplayer/3"  # categoria 3 = Pokémon
TCGCSV_USER_AGENT = "myp-arbitrage-scanner/5.16 (+github.com/matheuscllm-lgtm/myp-arbitrage-scanner)"

# setcode pokemontcg.io → abreviação tcgcsv (confirmado 2026-06-21 contra
# /groups: sv7=SCR=Stellar Crown, me2pt5=ASC=Ascended Heroes, etc.). Quando o
# setcode NÃO está aqui, a ponte cai no match por nome contra /groups (substring
# do nome da edição), e se nem isso casar → sem groupId → fallback `.estat-tcg`
# honesto (NUNCA preço inventado).
PTCG_SETCODE_TO_TCGCSV_ABBR = {
    "sv10": "DRI",      # Destined Rivals
    "sv9": "JTG",       # Journey Together
    "sv8pt5": "PRE",    # Prismatic Evolutions
    "sv8": "SSP",       # Surging Sparks
    "sv7": "SCR",       # Stellar Crown
    "sv6pt5": "SFA",    # Shrouded Fable
    "sv6": "TWM",       # Twilight Masquerade
    "sv5": "TEF",       # Temporal Forces
    "sv4pt5": "PAF",    # Paldean Fates
    "sv4": "PAR",       # Paradox Rift
    "sv3": "OBF",       # Obsidian Flames
    "sv3pt5": "MEW",    # 151
    "sv2": "PAL",       # Paldea Evolved
    "me2pt5": "ASC",    # Ascended Heroes (pokemontcg.io SEM preço → tcgcsv resgata)
    "me2": "PFL",       # Phantasmal Flames
    "me1": "MEG",       # Mega Evolution (base)
    "swsh12pt5": "CRZ",   # Crown Zenith
    "swsh12": "SWSH12",   # Silver Tempest
    "swsh11": "SWSH11",   # Lost Origin
    "swsh10": "SWSH10",   # Astral Radiance
    "swsh9": "SWSH09",    # Brilliant Stars (tcgcsv usa zero-pad: SWSH09)
    "swsh8": "SWSH08",    # Fusion Strike
    "swsh7": "SWSH07",    # Evolving Skies
    "pgo": "PGO",         # Pokémon GO
    "zsv10pt5": "BLK",  # Black Bolt
    "rsv10pt5": "WHT",  # White Flare
    # ── v5.16 (2026-06-22): abbreviações VERIFICADAS 1-a-1 contra /groups ao
    # vivo (cada uma resolve p/ EXATAMENTE 1 group, nome coerente c/ o set).
    # ⚠️ A abbr tcgcsv NEM SEMPRE bate com o ptcgoCode do pokemontcg.io: SM/
    # SWSH usam o esquema próprio do tcgcsv (SM02, SWSH02, …) e alguns sets
    # usam código alternativo (sm7→CES, sm35→SHL, swsh35→CHP). Tudo abaixo foi
    # casado contra o NOME do group no dump real, não por suposição.
    # Sun & Moon
    "sm2": "SM02",      # SM - Guardians Rising
    "sm3": "SM03",      # SM - Burning Shadows
    "sm35": "SHL",      # Shining Legends
    "sm4": "SM04",      # SM - Crimson Invasion
    "sm5": "SM05",      # SM - Ultra Prism
    "sm6": "SM06",      # SM - Forbidden Light
    "sm7": "CES",       # SM - Celestial Storm
    "sm75": "DRM",      # Dragon Majesty
    "sm8": "SM8",       # SM - Lost Thunder (tcgcsv usa SM8, sem zero-pad)
    "sm9": "SM9",       # SM - Team Up
    "sm10": "SM10",     # SM - Unbroken Bonds
    "sm11": "SM11",     # SM - Unified Minds
    "sm115": "HIF",     # Hidden Fates
    "sm12": "SM12",     # SM - Cosmic Eclipse
    # XY
    "xy2": "FLF",       # XY - Flashfire
    "xy3": "FFI",       # XY - Furious Fists
    "xy4": "PHF",       # XY - Phantom Forces
    "xy5": "PRC",       # XY - Primal Clash
    "xy6": "ROS",       # XY - Roaring Skies
    "xy7": "AOR",       # XY - Ancient Origins
    "xy8": "BKT",       # XY - BREAKthrough
    "xy10": "FCO",      # XY - Fates Collide
    "xy11": "STS",      # XY - Steam Siege
    "xy12": "EVO",      # XY - Evolutions
    "dc1": "DCR",       # Double Crisis
    "xy0": "KSS",       # Kalos Starter Set
    # Sword & Shield (antigos)
    "swsh2": "SWSH02",  # Rebel Clash
    "swsh3": "SWSH03",  # Darkness Ablaze
    "swsh35": "CHP",    # Champion's Path
    "swsh4": "SWSH04",  # Vivid Voltage
    "swsh5": "SWSH05",  # Battle Styles (NÃO 'BST' — esse é EX Battle Stadium!)
    "swsh6": "SWSH06",  # Chilling Reign
    "swsh45": "SHF",    # Shining Fates
    "cel25c": "CCC",    # Celebrations: Classic Collection
    # HeartGold & SoulSilver
    "hgss2": "UL",      # Unleashed
    "hgss3": "UD",      # Undaunted
    "hgss4": "TM",      # Triumphant
    # Black & White
    "bw2": "EPO",       # Emerging Powers
    "bw3": "NVI",       # Noble Victories
    "bw4": "NXD",       # Next Destinies
    "bw5": "DEX",       # Dark Explorers
    "bw6": "DRX",       # Dragons Exalted
    "bw7": "BCR",       # Boundaries Crossed
    "bw8": "PLS",       # Plasma Storm
    "bw9": "PLF",       # Plasma Freeze
    "bw10": "PLB",      # Plasma Blast
    "dv1": "DRV",       # Dragon Vault
    # Diamond & Pearl / Platinum
    "pl3": "SV",        # Supreme Victors
    "pl4": "AR",        # Arceus
    "dp2": "MT",        # Mysterious Treasures
    "dp3": "SW",        # Secret Wonders
    "dp4": "GE",        # Great Encounters
    "dp5": "MD",        # Majestic Dawn
    "dp6": "LA",        # Legends Awakened
    "dp7": "SF",        # Stormfront
    # EX
    "ex1": "RS",        # Ruby & Sapphire
    "ex2": "SS",        # Sandstorm
    "ex3": "DR",        # Dragon
    "ex4": "MA",        # Team Magma vs Team Aqua
    "ex5": "HL",        # Hidden Legends
    "ex6": "RG",        # FireRed & LeafGreen
    "ex8": "DX",        # Deoxys
    "ex9": "EM",        # Emerald
    "ex10": "UF",       # Unseen Forces
    "ex11": "DS",       # Delta Species
    "ex12": "LM",       # Legend Maker
    "ex13": "HP",       # Holon Phantoms
    "ex14": "CG",       # Crystal Guardians
    "ex15": "DF",       # Dragon Frontiers
    "ex16": "PK",       # Power Keepers
    # e-Card
    "ecard1": "EX",     # Expedition Base Set
    "ecard2": "AQ",     # Aquapolis
    "ecard3": "SK",     # Skyridge
    # WOTC clássicos
    "neo1": "N1",       # Neo Genesis
    "neo2": "N2",       # Neo Discovery
    "neo3": "N3",       # Neo Revelation
    "neo4": "N4",       # Neo Destiny
    "gym1": "G1",       # Gym Heroes
    "gym2": "G2",       # Gym Challenge
    "base6": "LC",      # Legendary Collection
}

# Câmbio: frankfurter.app (ECB, grátis, sem key); fallback open.er-api.com.
_FX_SOURCES = (
    ("frankfurter", "https://api.frankfurter.app/latest?from=USD&to=BRL",
     lambda j: j.get("rates", {}).get("BRL")),
    ("er-api", "https://open.er-api.com/v6/latest/USD",
     lambda j: j.get("rates", {}).get("BRL")),
)


def fetch_usd_brl(session) -> Optional[float]:
    """Cotação USD→BRL ao vivo (v5.11). Tenta frankfurter, depois er-api.

    Retorna None se ambas falharem — o caller cai no `.estat-tcg` do MYP pra
    a run inteira (sem câmbio não dá pra converter o preço real em USD)."""
    for label, url, extract in _FX_SOURCES:
        try:
            r = session.get(url, timeout=15)
            if r.status_code != 200:
                continue
            rate = extract(r.json())
            if rate and float(rate) > 0:
                log.info(f"  💱 Câmbio USD→BRL: {float(rate):.4f} (fonte: {label})")
                return float(rate)
        except Exception as e:  # noqa: BLE001
            log.debug(f"FX {label} falhou: {e!r}")
    return None


def tcgcsv_fetch_groups(session) -> Optional[list]:
    """v5.15: baixa a lista de groups (sets) do tcgcsv (categoria 3 = Pokémon).

    Retorna a lista `results` (cada item: {groupId, name, abbreviation, ...}) ou
    None se a request falhar. Header User-Agent é OBRIGATÓRIO (sem ele = 401)."""
    try:
        r = session.get(f"{TCGCSV_BASE}/groups",
                         headers={"User-Agent": TCGCSV_USER_AGENT}, timeout=20)
        if r.status_code != 200:
            log.debug(f"tcgcsv /groups status {r.status_code}")
            return None
        results = (r.json() or {}).get("results")
        return results if isinstance(results, list) else None
    except Exception as e:  # noqa: BLE001
        log.debug(f"tcgcsv /groups falhou: {e!r}")
        return None


def resolve_tcgcsv_group_id(setcode: str, edition: str, groups: list) -> Optional[int]:
    """v5.15: setcode pokemontcg.io (+ nome da edição MYP) → tcgcsv groupId.

    Estratégia em cascata (todas contra o dump REAL de /groups, sem chute):
      1. Abreviação conhecida (PTCG_SETCODE_TO_TCGCSV_ABBR) casando o campo
         `abbreviation` do group — caminho primário, confirmado 2026-06-21.
      2. Fallback: match por NOME — a substring EN da edição MYP que está em
         MYP_EDITION_SUBSTR_TO_PTCG (a chave cujo valor é este setcode) aparece
         no `name` do group (ex.: "Stellar Crown" ∈ "SV07: Stellar Crown").
         ⚠️ Só aceita match por nome se ele for **ÚNICO**: se a substring casar
         >1 group (ex.: "Mega Evolution" ∈ ME01/MEP/MEE), é AMBÍGUO → None →
         fallback `.estat-tcg` honesto. NUNCA chuta o primeiro — injetar preço
         de promo/energy como "real" é a classe de bug que a v5.11 corrigiu no
         `.estat-tcg`. (O caminho 1 por abreviação é exato e cobre os sets
         mapeados; este fallback é a rede de segurança contra renomeação futura.)
    Retorna None se nada casar → caller mantém fallback `.estat-tcg` honesto.
    """
    if not groups:
        return None
    abbr = PTCG_SETCODE_TO_TCGCSV_ABBR.get(setcode)
    if abbr:
        for g in groups:
            if str(g.get("abbreviation") or "").upper() == abbr.upper():
                return g.get("groupId")
    # Fallback por nome: acha a(s) substring(s) MYP que mapeiam a este setcode e
    # casa contra o name do group — exigindo match ÚNICO (sem chute em ambíguo).
    substrs = [s.lower() for s, c in MYP_EDITION_SUBSTR_TO_PTCG.items() if c == setcode]
    matches = [g.get("groupId") for g in groups
               if any(sl in str(g.get("name") or "").lower() for sl in substrs)]
    uniq = list(dict.fromkeys(m for m in matches if m is not None))
    return uniq[0] if len(uniq) == 1 else None


def tcg_search_url(name: str) -> Optional[str]:
    """URL de busca TCGplayer pelo nome da carta (sem o sufixo (NNN/MMM)).

    Retorna None se o nome for vazio — evita gerar link de busca vazio."""
    if not name or not isinstance(name, str):
        return None
    # Remove o token (...) final, igual ao que clean_card_name preserva no
    # Card Name: aqui queremos o nome puro pra query de busca.
    base = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()
    if not base:
        return None
    return _TCG_SEARCH_BASE + quote_plus(base)


def myp_edition_to_ptcg_setcode(edition: str) -> Optional[str]:
    """MYP edition string → pokemontcg.io setcode (longest substring wins).

    Returns None pra edition não mapeada (vintage/promo/Black Bolt/etc).
    Case-insensitive. Tolera bilingual concat do MYP (ex "Sword & Shield 7:
    ...Evolving Skies" contém EN substring).
    """
    if not isinstance(edition, str) or not edition:
        return None
    el = edition.lower()
    hits = [(s, c) for s, c in MYP_EDITION_SUBSTR_TO_PTCG.items() if s.lower() in el]
    if not hits:
        return None
    # Longest substring wins — evita "151" matching "1518" etc.
    hits.sort(key=lambda t: -len(t[0]))
    return hits[0][1]


def tcg_direct_url(card_name: str, edition: str,
                   oversized_collector_risk: bool = False) -> Optional[str]:
    """URL DIRETA do produto TCGplayer via redirect pokemontcg.io.

    Forma: `https://prices.pokemontcg.io/tcgplayer/{setcode}-{num}`.
    pokemontcg.io devolve 302 → `tcgplayer.com/product/<id>` da carta exata.

    Retorna None (caller deve cair no `tcg_search_url`) quando:
      - edition não está em MYP_EDITION_SUBSTR_TO_PTCG;
      - card_name não tem o token (NNN/MMM) parseável;
      - oversized_collector_risk=True (numerator > set_size → variant SIR/HR
        que pokemontcg.io frequentemente NÃO tem indexada → dead link).
        Caso documentado: `me2pt5-274` (Mega Feraligatr ex 274/217) 404 em
        2026-05-29. Sets-base mapeados sempre resolvem.
    """
    if oversized_collector_risk:
        return None
    setcode = myp_edition_to_ptcg_setcode(edition)
    if not setcode:
        return None
    m = _COLLECTOR_NUM_RE.search(card_name or "")
    if not m:
        return None
    # pokemontcg.io usa o number sem leading zeros (ex sv9-187 não sv9-0187,
    # base6-13 não base6-013). Verificado contra Link TCG do CT handoff
    # 2026-05-19.
    num = m.group(1).lstrip("0") or "0"
    return f"https://prices.pokemontcg.io/tcgplayer/{setcode}-{num}"
# v5.8.3 (2026-05-18): Flareon VMAX (018/203) "Prize Pack Series" — observado
# 1 seller único (`gvrgyn`) listando como Inglês quando a carta não tem print
# EN nessa edição (mislabeling). Sem cross-check pokemontcg.io confiável, a
# heurística defensiva é tratar 1-seller-EN como single_seller_risk e mover
# pra Validate Manually. Não suprime — apenas escala visibilidade.
# v5.8.4 (2026-05-19): CLI override via --min-en-sellers. Default permanece 1
# (legacy v5.8.3 behavior). Card é flagged quando `en_sellers <
# MIN_EN_SELLERS_FOR_DEALS` (strict less-than). Threshold pode ser elevado
# pra cenários mais conservadores (ex.: --min-en-sellers 2 trata 1 OU 2
# sellers como risco).
MIN_EN_SELLERS_FOR_DEALS_DEFAULT = 2  # < default = flagged (legacy era ≤1)
# v5.8 H2 (2026-05-16): se TCG declarado >> última venda real, MYP infla o
# preço de referência. Caso Jirachi PR-SM_SM161: declarava R$1499 vs última
# venda real R$19,99 (75x). Threshold 10x captura inflação grosseira sem
# false-positive em cards com pouca liquidez (last sale antigo + alta).
TCG_SUSPECT_RATIO_THRESHOLD = 10.0
# v5.4 H1: idiomas EN reconhecidos. Tudo fora dessa lista que parecer um title de
# flag-icon (não vazio) é tratado como "unknown" e contado pra warn-once.
KNOWN_LANGUAGES = {
    "Inglês", "Português", "Japonês", "Italiano",
    "Espanhol", "Francês", "Alemão", "Coreano",
    "English", "Portuguese", "Japanese",
}
EN_LANGUAGES = {"Inglês", "English"}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
}

# Mapeamento de códigos de idioma no produto
LANG_MAP = {
    "eng": "EN", "por": "PT", "jpn": "JP", "ita": "IT",
    "esp": "ES", "fra": "FR", "deu": "DE", "kor": "KR",
    "chi": "CN", "tha": "TH",
}


@dataclass
class CardData:
    name: str = ""
    edition: str = ""
    edition_url: str = ""
    product_url: str = ""
    product_code: str = ""
    language: str = ""
    condition: str = "NM"
    rarity: str = ""
    myp_lowest_en_nm: Optional[float] = None   # menor preço EN NM no MYP
    tcg_player_price: Optional[float] = None    # preço TCG usado na margem (BRL)
    # v5.11 (2026-06-07): proveniência do preço TCG. "pokemontcg.io" = preço
    # real do TCGplayer (USD convertido p/ BRL); "myp_estat" = fallback no campo
    # `.estat-tcg` declarado pelo MYP (usado onde o pokemontcg.io não cobre).
    tcg_source: str = "myp_estat"
    tcg_real_usd: Optional[float] = None         # preço real em USD (se via pokemontcg.io)
    myp_declared_tcg_brl: Optional[float] = None # `.estat-tcg` cru do MYP (auditoria)
    # v5.8 H2 (2026-05-16): MYP às vezes reporta .estat-tcg inflado (caso
    # Jirachi PR-SM_SM161: MYP=R$1499 vs TCGPlayer real $26=R$132 = 11x off).
    # Capturar última venda real do MYP pra sanity check.
    myp_last_sale_brl: Optional[float] = None
    tcg_suspect: bool = False                    # True se TCG declarado >> última venda real
    margin_pct: Optional[float] = None
    margin_brl: Optional[float] = None
    en_nm_sellers: int = 0                      # qtd vendedores EN NM
    en_truncation_risk: bool = False            # 2026-05-12: alguma seller table está no cap (15/20) sem EN visível → EN possivelmente escondido
    # v5.8.3 (2026-05-18): único seller EN visível → risco de mislabeling de idioma
    # (caso Flareon VMAX 018/203 Prize Pack: 1 seller gvrgyn lista como EN mas
    # carta não tem print EN nessa edição). Sinaliza pra Validate Manually.
    single_en_seller_risk: bool = False
    # v5.8.5 (2026-05-19): collector# > set_size = variant (SIR/HR/promo extra/
    # special illustration rare). Frequente JP-only. Caso Darumaka 097/086, Mew
    # ex 232/091, Charizard ex 234/091. Parse de (NNN/MMM) no card.name; quando
    # numerator > denominator, flag oversized_collector_risk = True.
    oversized_collector_risk: bool = False
    last_updated: str = ""


# ══════════════════════════════════════════════════════════════════════
# SCRAPER
# ══════════════════════════════════════════════════════════════════════
class MYPScraper:
    def __init__(
        self,
        delay: float = REQUEST_DELAY,
        min_en_sellers: int = MIN_EN_SELLERS_FOR_DEALS_DEFAULT,
        threshold: float = MARGIN_THRESHOLD,
        min_price: float = MIN_PRICE_BRL,
        tcg_source: str = "auto",
    ):
        # v5.8.4 (2026-05-19): threshold configurable via CLI. Card é flagged
        # quando `en_sellers < min_en_sellers`. Default 2 reproduz v5.8.3
        # (que checava `en_sellers <= 1`).
        self.min_en_sellers = min_en_sellers
        # Config por-instância. Antes `threshold`/`min_price` eram globais
        # reatribuídas no __main__ — frágil (vazava estado entre instâncias no
        # mesmo processo, e era inconsistente com min_en_sellers). Defaults =
        # constantes do módulo, então MYPScraper() sem args = comportamento legado.
        self.margin_threshold = threshold
        self.min_price = min_price
        if HAS_CLOUDSCRAPER:
            # 2026-05-17: Cloudflare passou a bloquear o fingerprint chrome/windows
            # do cloudscraper (HTTP 403 cf-mitigated: challenge). Firefox/windows
            # ainda passa. Mantemos chrome no env var pra rollback fácil.
            browser_fp = os.environ.get("MYP_CLOUDSCRAPER_BROWSER", "firefox")
            self.session = cloudscraper.create_scraper(
                browser={"browser": browser_fp, "platform": "windows", "desktop": True},
            )
            log.info(f"Using cloudscraper (browser={browser_fp}, CloudFlare bypass enabled)")
            # Não sobrescreve User-Agent — o cloudscraper já configura UA coerente
            # com o TLS fingerprint do browser escolhido. Forçar UA Chrome num
            # fingerprint Firefox = mismatch detectado pelo CF (403).
            non_ua_headers = {k: v for k, v in HEADERS.items() if k.lower() != "user-agent"}
            self.session.headers.update(non_ua_headers)
        else:
            self.session = requests.Session()
            log.warning("cloudscraper not installed — may get 403 errors!")
            log.warning("Fix: pip install cloudscraper")
            self.session.headers.update(HEADERS)
        self.delay = delay
        self.cards: list[CardData] = []
        # v5.11 (2026-06-07): preço TCG real via pokemontcg.io + câmbio.
        # fx_usd_brl é buscado uma vez no início de scan(); fica None em testes
        # offline (scan() não roda) → real-price path inerte, usa `.estat-tcg`.
        self.fx_usd_brl: Optional[float] = None
        # _clean_secret: tira BOM/zero-width/whitespace antes de virar header
        # HTTP (X-Api-Key). Sem isso, um BOM no secret quebra o latin-1 encode do
        # requests e derruba TODA chamada à pokemontcg.io (mass pricing failure,
        # scan "verde" mas vazio — bug real que pegou o scanner irmão CardTrader).
        self.ptcg_api_key: Optional[str] = _clean_secret(os.environ.get("POKEMONTCG_API_KEY"))
        self._ptcg_cache: dict[str, Optional[float]] = {}  # cid → preço USD (ou None)
        self._prefilled_sets: set[str] = set()  # v5.12: setcodes já pré-carregados (batch)
        # v5.15: fonte do preço TCG real. "auto" = tcgcsv primeiro (funciona no
        # CI), pokemontcg.io como complemento por-card; "tcgcsv" = só tcgcsv;
        # "pokemontcg" = comportamento ≤v5.14 (só pokemontcg.io). O preço de
        # AMBOS popula o MESMO _ptcg_cache (keyed por {setcode}-{num}).
        self.tcg_source_mode: str = (tcg_source or "auto").lower()
        # cids cujo preço veio do tcgcsv (p/ rotular tcg_source corretamente —
        # "tcgcsv" vs "pokemontcg.io"; ambos são REAIS, rótulos distintos só p/
        # auditoria/proveniência). Setado por _prefill_tcgcsv_set.
        self._tcgcsv_cids: set[str] = set()
        self._tcgcsv_groups: Optional[list] = None   # cache de /groups (1× por run)
        self._tcgcsv_groups_fetched: bool = False
        self._stats = {
            "pages_fetched": 0, "products_scanned": 0, "en_found": 0,
            # M5 fix: counters por motivo de skip (auditoria do funnel)
            "skipped_no_tcg_price": 0,
            "skipped_no_en_sellers": 0,
            "skipped_low_price": 0,
            "supranumerary_warnings": 0,
            "http_retries": 0,
            # 2026-05-12: contador de risco de truncamento de EN-NM
            # (alguma seller table cheia sem EN visível — caso bartsimpson Psyduck)
            "en_truncation_risks": 0,
            # v5.4 H1: títulos de idioma fora de KNOWN_LANGUAGES (drop silencioso)
            "skipped_unknown_lang_titles": 0,
            # v5.8 H2 (2026-05-16): cards com TCG declarado >> última venda
            # real (Jirachi PR-SM_SM161 caso). Não filtra do funnel — fica
            # em All EN Cards, mas é excluído da sheet 🔥 Deals.
            "tcg_suspects": 0,
            # v5.8.3 (2026-05-18): produtos Jumbo (oversized) skipados pelo
            # filtro de TÍTULO (camada 2). `.estat-tcg` reflete preço da carta
            # standard, deal é fictício.
            "skipped_jumbo": 0,
            # v5.8.3 (2026-05-18): seller rows com foil="Jumbo" filtradas pela
            # camada 1 (caso M-Rayquaza-EX XY 7: standard + jumbo no mesmo
            # produto, MYP diferencia via coluna `.estoque-lista-nomeenfoil`).
            "jumbo_rows_filtered": 0,
            # v5.8.3 (2026-05-18): cards com apenas 1 seller EN visível —
            # risco de seller mislabeling (caso Flareon VMAX 018/203).
            "single_en_seller_risks": 0,
            # v5.8.5 (2026-05-19): collector# > set_size (variant fora do
            # numbered set, frequentemente JP-only). Caso Darumaka 097/086.
            "oversized_collector_risks": 0,
            # v5.9 (2026-06-03): paginação da tabela marketplace seguida quando
            # a página 1 sinaliza truncation (?estoque-outros-page=N). Conta
            # páginas extras lidas com sucesso e falhas de fetch dessas páginas.
            "seller_pages_followed": 0,
            "seller_page_fetch_failures": 0,
            # v5.10.1 (2026-06-07): cost gate — paginações puladas porque o card
            # tem TCG < min_price (não pode virar deal, não vale o request).
            "pagination_skipped_low_tcg": 0,
            # v5.11 (2026-06-07): proveniência do preço TCG na margem.
            "tcg_from_real": 0,         # preço real do TCGplayer (pokemontcg.io)
            "tcg_from_myp_fallback": 0, # fallback no `.estat-tcg` do MYP
            # v5.13 (Iteração #2): POR QUE o candidato caiu no fallback `.estat-tcg`
            # — atribuição de cobertura. Cobertura faltante é a RAIZ dos falso-
            # positivos: tcg_suspect/supranumerário só sobrevivem sem preço real.
            # A soma dos 4 baldes = tcg_from_myp_fallback. Num scan ao vivo aponta
            # o maior balde FIXÁVEL (ex.: unmapped_set → adicionar o setcode).
            "fallback_no_fx": 0,            # run sem câmbio (condição global, não cobertura)
            "fallback_unmapped_set": 0,     # edição fora de MYP_EDITION_SUBSTR_TO_PTCG
            "fallback_no_collector_num": 0, # nome sem token (NNN/MMM) parseável
            "fallback_no_coverage": 0,      # cid existe mas pokemontcg.io 404/sem preço/429
            # loop plumbing: medição de tempo (perf_counter, sempre ligada, ~0
            # overhead) — base do loop de otimização iterativo. `ptcg_calls`
            # conta só round-trips REAIS (cache-hit em _real_tcg_brl não passa
            # por _fetch_ptcg_usd, então não conta).
            "t_http_total": 0.0,      # tempo acumulado dentro de _get (s)
            "t_ptcg_total": 0.0,      # tempo acumulado em _fetch_ptcg_usd (s)
            "ptcg_calls": 0,          # nº de chamadas reais à pokemontcg.io (por-card)
            "ptcg_prefill_calls": 0,  # v5.12: requests de prefill batch (por-set)
            # v5.15: fonte tcgcsv (funciona no CI). tcg_from_tcgcsv = cards cujo
            # preço real veio do tcgcsv; tcgcsv_prefill_sets = sets pré-carregados
            # via tcgcsv (cada um = 2 requests: /products + /prices).
            "tcg_from_tcgcsv": 0,
            "tcgcsv_prefill_sets": 0,
            "t_editions_total": 0.0,  # tempo de parede por edição, acumulado (s)
        }
        # v5.4 H1: warn-once cache pra unknown language titles
        self._unknown_lang_seen: set[str] = set()

    def _get(self, url: str, save_debug: bool = False) -> Optional[BeautifulSoup]:
        """Fetch a page and return parsed soup. M1 fix: retry com backoff."""
        _t0 = time.perf_counter()
        try:
            last_err = None
            last_status = ""
            for attempt in range(HTTP_MAX_RETRIES):
                try:
                    time.sleep(self.delay)
                    resp = self.session.get(url, timeout=TIMEOUT)
                    resp.raise_for_status()
                    self._stats["pages_fetched"] += 1

                    if save_debug:
                        # M4 fix: salva em subpasta .debug/ ao invés do CWD
                        DEBUG_DIR.mkdir(exist_ok=True)
                        debug_file = DEBUG_DIR / f"debug_{self._stats['pages_fetched']}.html"
                        debug_file.write_text(resp.text[:50000], encoding="utf-8")
                        log.info(f"  DEBUG: saved HTML to {debug_file}")

                    return BeautifulSoup(resp.text, "lxml")
                except (requests.RequestException, ConnectionError, TimeoutError, OSError) as e:
                    # v5.4 C3: catch só erros de rede. Parser bugs (lxml/bs4),
                    # AttributeError, MemoryError etc devem propagar — indicam
                    # mudança de HTML ou bug de código que merece crash, não retry.
                    last_err = e
                    if hasattr(e, 'response') and e.response is not None:
                        last_status = f" (HTTP {e.response.status_code})"
                    if attempt < HTTP_MAX_RETRIES - 1:
                        wait = (attempt + 1) * 2  # backoff 2s, 4s
                        self._stats["http_retries"] += 1
                        log.warning(f"Retry {attempt+1}/{HTTP_MAX_RETRIES} for {url}{last_status}: {e}, waiting {wait}s")
                        time.sleep(wait)
                        continue
            log.warning(f"Failed to fetch {url}{last_status} after {HTTP_MAX_RETRIES} attempts: {last_err}")
            return None
        finally:
            # loop plumbing: tempo de parede total em _get (sleep+fetch+parse)
            self._stats["t_http_total"] += time.perf_counter() - _t0

    @staticmethod
    def _parse_brl(text) -> Optional[float]:
        """Parse price string. Handles BR canonical ('R$ 1.900,00') AND US
        decimal leakage ('R$ 30.00') that MYP sometimes emits in
        `.estatistica-ultimo`. v5.8.2 fix: previously '30.00' → 3000.0 (read
        as BR thousands), broke sanity-check ratio → false negatives.

        v5.8.4 (2026-05-19): defensive against None / non-str inputs. Reviewer
        flagged that `text.strip()` raises AttributeError if a caller ever
        passes an Optional[str] that turns out None (or a numeric from a
        future refactor). Guard before stripping.
        """
        if text is None or not isinstance(text, str):
            return None
        text = text.strip()
        if not text:
            return None
        text = re.sub(r'[R$\s\xa0]', '', text)
        if not text:
            return None
        has_comma = ',' in text
        has_dot = '.' in text
        if has_comma and has_dot:
            # Both present. Whichever appears LAST is the decimal separator.
            if text.rfind(',') > text.rfind('.'):
                # BR canonical: '1.500,00' → '1500.00'
                text = text.replace('.', '').replace(',', '.')
            else:
                # US thousands: '1,500.00' → '1500.00'
                text = text.replace(',', '')
        elif has_comma:
            # Only comma → BR decimal: '30,00' → '30.00'
            text = text.replace(',', '.')
        elif has_dot:
            # Only dot → disambiguate by suffix length.
            # 2-digit suffix → decimal ('30.00' = 30.0; '1234.56' = 1234.56)
            # 3-digit suffix with single dot → BR thousands ('30.000' = 30000)
            # Multiple dots → BR thousands ('1.500.000' = 1500000)
            parts = text.split('.')
            if len(parts) > 2:
                text = text.replace('.', '')
            elif len(parts[-1]) == 3:
                text = text.replace('.', '')
            # else: keep as-is (US decimal style)
        try:
            val = float(text)
            return val if val > 0 else None
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _last_brl(text: Optional[str]) -> Optional[float]:
        """Extrai o ÚLTIMO valor R$ de um texto e parseia via _parse_brl.

        Centraliza o idiom `PRICE_RE.findall(...) + _parse_brl([-1])` que estava
        duplicado em 5 call-sites. `.estat-tcg` às vezes traz multi-preço
        ('Last R$ X | Avg R$ Y') e a referência EN é sempre o último valor.
        """
        if not text:
            return None
        matches = PRICE_RE.findall(text)
        return MYPScraper._parse_brl(matches[-1]) if matches else None

    # ── Step 1: Get all editions ─────────────────────────────────────
    def get_all_editions(self) -> list[dict]:
        """Scrape /pokemon/edicoes for all available editions."""
        editions = []
        page = 1
        # v5.4 H4: cap em MAX_EDITION_PAGES previne infinite loop se MYP
        # alguma vez retornar pages que parecem ter conteúdo novo indefinidamente.
        while page <= MAX_EDITION_PAGES:
            url = f"{BASE_URL}/pokemon/edicoes?page={page}"
            log.info(f"Fetching editions page {page}...")
            soup = self._get(url, save_debug=(page == 1))
            if not soup:
                break

            # Strategy 1: specific class selectors
            links = soup.select("a.edicao-link")

            # Strategy 2: any link inside an edicao container
            if not links:
                containers = soup.select('[class*="edicao"]')
                for c in containers:
                    a = c.select_one('a[href*="/pokemon/"]')
                    if a and a not in links:
                        links.append(a)

            # Strategy 3: broader pattern matching on all links
            if not links:
                exclude = ["produto", "edicoes", "outros", "selados",
                           "acessorios", "deck-lote", "cartas-graduadas",
                           "action-figure", "artigos-geek", "hq-livros",
                           "inscricao", "online", "pokemon?", "#"]
                for a in soup.select('a[href]'):
                    href = a.get("href", "")
                    # Match pattern: /pokemon/{slug} where slug is a valid edition
                    if re.match(r'^/pokemon/[a-z0-9][\w-]+$', href):
                        if not any(x in href for x in exclude):
                            links.append(a)

            if not links:
                if page == 1:
                    log.warning("No editions found! Check debug_1.html for page structure")
                break

            found_on_page = 0
            seen_urls = {e["url"] for e in editions}
            for link in links:
                href = link.get("href", "")
                if not href:
                    continue

                full_url = f"{BASE_URL}{href}" if href.startswith("/") else href
                if full_url in seen_urls:
                    continue

                # Get title: try parent container, then link text
                title = ""
                for parent_class in ["edicao-card", "edicao-item", "edicao"]:
                    parent = link.find_parent(class_=re.compile(parent_class))
                    if parent:
                        for title_class in ["edicao-titulo", "edicao-header", "titulo", "title"]:
                            t = parent.select_one(f'[class*="{title_class}"]')
                            if t:
                                title = t.get_text(strip=True)
                                break
                        break

                if not title:
                    title = link.get_text(strip=True)[:80]

                if not title or len(title) < 2:
                    continue

                editions.append({
                    "title": title,
                    "url": full_url,
                    "href": href,
                })
                seen_urls.add(full_url)
                found_on_page += 1

            if found_on_page == 0:
                break
            page += 1
        else:
            # v5.4 H4: hit MAX_EDITION_PAGES sem natural exit — sinal de bug
            log.warning(
                f"  ⚠️ get_all_editions hit MAX_EDITION_PAGES={MAX_EDITION_PAGES} "
                f"sem encontrar fim natural. Possível recursão de paginação no MYP."
            )

        # v5.4 C2: sanity check — catalog scrape esperado tem ~326 editions.
        # Abaixo de MIN_EDITIONS_EXPECTED é forte indicador que selectors
        # quebraram mid-catalog (Strategy 3 fallback pode silenciosamente truncar).
        if len(editions) < MIN_EDITIONS_EXPECTED:
            log.warning(
                f"  🚨 Catalog scrape suspeito: {len(editions)} editions "
                f"encontradas (esperado >={MIN_EDITIONS_EXPECTED}). "
                f"Selectors podem ter quebrado mid-catalog. Validar manualmente."
            )
        log.info(f"Found {len(editions)} editions")
        return editions

    # ── Step 2: Get product URLs from edition listing ────────────────
    def get_edition_products(self, edition_url: str) -> list[str]:
        """Get all product URLs from an edition listing page."""
        product_urls = []
        seen = set()
        page = 1
        # v5.4 H3: detecta loop de página duplicada (MYP retornando page 1
        # quando page=N overflowing). Compara primeira URL de page N vs N-1.
        prev_first_url: Optional[str] = None

        while page <= MAX_PAGES_PER_EDITION:
            url = f"{edition_url}?page={page}"
            soup = self._get(url)
            if not soup:
                break

            links = soup.select('a[href*="/pokemon/produto/"]')

            # v5.4 H3: page first-URL fingerprint
            current_first_url: Optional[str] = None
            for link in links:
                href = link.get("href", "")
                if href:
                    current_first_url = (
                        f"{BASE_URL}{href}" if href.startswith("/") else href
                    )
                    break
            if (page > 1 and prev_first_url is not None
                    and current_first_url == prev_first_url):
                log.warning(
                    f"  🚨 Pagination loop detectado em {edition_url}: "
                    f"page {page} retornou mesma primeira URL de page {page-1}. "
                    f"Stopping para evitar under-coverage silencioso."
                )
                break
            prev_first_url = current_first_url

            new_count = 0
            for link in links:
                href = link.get("href", "")
                full_url = f"{BASE_URL}{href}" if href.startswith("/") else href
                if full_url not in seen:
                    seen.add(full_url)
                    product_urls.append(full_url)
                    new_count += 1

            if new_count == 0:
                break
            page += 1

        return product_urls

    # ── Step 3 helper: parse ONE seller table element (EN-NM extraction) ─
    def _parse_seller_table(self, table) -> dict:
        """Parse a single seller-table element (or any node containing <tr>s).

        Pure parsing, no network. Extracts EN+NM prices, counts rows, tracks
        the max visible price (any language) for the truncation heuristic, and
        skips Jumbo rows. Returns a dict:
            {"rows", "en", "max_price", "en_prices": [...], "jumbo"}
        Caller aggregates across page-1 tables AND marketplace pagination
        (v5.9). Extracted from the inline loop so both paths share one parser.
        """
        rows_in_table = 0
        en_in_table = 0
        max_price_in_table = 0.0  # maior preço VISÍVEL nesta tabela
        table_en_prices: list[float] = []
        jumbo_count = 0
        for row in table.find_all("tr"):
            row_text = row.get_text()
            if "R$" not in row_text:
                continue
            rows_in_table += 1

            # Extrai preço (qualquer idioma) pra rastrear max visível.
            # 2026-05-12 v5.3: row pode ter strikethrough promo
            # ("R$ 275,00 R$ 220,00" — R$275 antigo riscado, R$220 ativo).
            # v5.4 H2: usa min() em vez de [-1] — preço ativo é sempre
            # o menor (promo); [-1] quebrava se MYP injetasse 3º R$
            # (frete, "you save", etc). Min é defensivo a layout drift.
            price_matches = PRICE_RE.findall(row_text)
            row_price = None
            if price_matches:
                parsed = [self._parse_brl(p) for p in price_matches]
                parsed = [p for p in parsed if p is not None and p > 0]
                if parsed:
                    row_price = min(parsed)
                if len(price_matches) > 2:
                    log.debug(
                        f"  Row com {len(price_matches)} R$ matches "
                        f"(esperado 1-2): {row_text[:120]}"
                    )
            if row_price and row_price > max_price_in_table:
                max_price_in_table = row_price

            # Find language from flag-icon span (specific selector)
            lang = None
            flag_el = row.select_one("span.flag-icon[title]")
            if flag_el:
                lang = flag_el.get("title", "").strip()
            else:
                # Fallback: check any [title] that matches a known language
                for el in row.select("[title]"):
                    title_val = el.get("title", "").strip()
                    if title_val in KNOWN_LANGUAGES:
                        lang = title_val
                        break

            # v5.4 H1: lang não-vazio mas fora do conhecido = drift potencial
            # (ex.: MYP normalizar "Inglês" → "Ingles" sem acento, ou novo
            # idioma adicionado). Counter + warn-once previne silent zero.
            if lang and lang not in KNOWN_LANGUAGES:
                self._stats["skipped_unknown_lang_titles"] += 1
                if lang not in self._unknown_lang_seen:
                    self._unknown_lang_seen.add(lang)
                    log.warning(
                        f"  ⚠️ Idioma desconhecido detectado: '{lang}' "
                        f"(não está em KNOWN_LANGUAGES). Pode ser drift de "
                        f"título flag-icon do MYP. Adicionar à constante "
                        f"se for mapeamento legítimo."
                    )

            if lang not in EN_LANGUAGES:
                continue

            # v5.8.3 (2026-05-18): skip rows com foil="Jumbo" (oversized).
            # MYP agrupa standard + jumbo na mesma página de produto; a
            # coluna `td.estoque-lista-nomeenfoil` indica a variante. TCG
            # Player price refere-se à standard, então jumbo rows inflam
            # `min(en_prices)` artificialmente (caso M-Rayquaza-EX 098/98
            # XY 7: 5 sellers Jumbo a R$650 enquanto TCG standard era
            # R$4801 → margin fictícia de 638%).
            foil_el = row.select_one("td.estoque-lista-nomeenfoil")
            foil_txt = foil_el.get_text(strip=True) if foil_el else ""
            if OVERSIZED_FOIL_RE.search(foil_txt):
                jumbo_count += 1
                continue

            # Filter: NM (Near Mint) only — skip Played, Damaged, etc.
            # v5.8.7: lê a célula de condição DEDICADA
            # (td.estoque-lista-qualidadenome, ex.: "NM - Quase nova",
            # "SP - Pouco jogada") e casa o código EXATO antes do " - ".
            # Antes era substring "NM" na linha inteira, que vazava não-NM
            # quando "NM" aparecia em qualquer coluna (nick de vendedor,
            # obs, etc). NM-only é invariante do scanner; sem célula de
            # qualidade confirmável (drift de layout), a linha é pulada.
            qual_el = row.select_one("td.estoque-lista-qualidadenome")
            qual_txt = qual_el.get_text(" ", strip=True) if qual_el else ""
            qual_code = qual_txt.split("-", 1)[0].strip().upper()
            if qual_code != "NM":
                continue

            # EN + NM seller — preço já extraído acima
            if row_price:
                table_en_prices.append(row_price)
                en_in_table += 1

        return {
            "rows": rows_in_table,
            "en": en_in_table,
            "max_price": max_price_in_table,
            "en_prices": table_en_prices,
            "jumbo": jumbo_count,
        }

    # ── Step 3 helper: detect marketplace pagination on a product page ───
    @staticmethod
    def _max_seller_page(soup) -> int:
        """Return the highest ?estoque-outros-page=N present in the page's
        pagination links (1 if the marketplace table is not paginated).

        v5.9: the marketplace table (#lista-anuncio-demais-vendedores) renders
        a standard <ul class="pagination"> whose links carry the page query
        param. We read N straight off the hrefs (more robust than regex on the
        whole document, which could match the param inside JS/analytics blobs).
        """
        max_page = 1
        for a in soup.select('a[href*="estoque-outros-page="]'):
            m = re.search(r'estoque-outros-page=(\d+)', a.get("href", ""))
            if m:
                max_page = max(max_page, int(m.group(1)))
        return max_page

    # ── v5.11: preço TCG REAL via pokemontcg.io (USD) ──
    def _fetch_ptcg_usd(self, cid: str) -> Optional[float]:
        """Preço real TCGplayer (USD) de um card pelo id pokemontcg.io
        (`{setcode}-{num}`). Pega o MENOR preço entre as variantes disponíveis
        (`market`, senão `mid`) — conservador, não superestima a margem.

        Retorna None se: card não existe (404), sem preço, ou erro de rede.
        Caller trata None como 'sem cobertura' → fallback no `.estat-tcg`."""
        # loop plumbing: conta round-trip REAL + cronometra. Cache-hit não chega
        # aqui (resolvido em _real_tcg_brl), então ptcg_calls = nº de fetches.
        self._stats["ptcg_calls"] += 1
        _t0 = time.perf_counter()
        try:
            headers = {"X-Api-Key": self.ptcg_api_key} if self.ptcg_api_key else {}
            # Sem API key o pokemontcg.io throttle bursts (429). Backoff robusto
            # pra NÃO cair silenciosamente no `.estat-tcg` furado por rate-limit —
            # esse fallback mascararia o preço errado justamente nos cards que
            # queremos corrigir. Definir POKEMONTCG_API_KEY elimina o throttle.
            backoffs = (5, 15, 30)
            for attempt in range(len(backoffs) + 1):
                try:
                    r = self.session.get(PTCG_API_BASE + cid, headers=headers, timeout=20)
                except Exception as e:  # noqa: BLE001
                    log.debug(f"pokemontcg.io {cid} falhou: {e!r}")
                    return None
                if r.status_code == 429:
                    if attempt < len(backoffs):
                        time.sleep(backoffs[attempt])
                        continue
                    log.warning(f"  ⚠️ pokemontcg.io rate-limited em {cid} após retries "
                                f"— fallback `.estat-tcg` (defina POKEMONTCG_API_KEY p/ evitar)")
                    return None
                if r.status_code != 200:   # 404 = sem cobertura (fallback legítimo)
                    return None
                try:
                    prices = ((r.json().get("data") or {}).get("tcgplayer") or {}).get("prices") or {}
                except Exception:  # noqa: BLE001
                    return None
                break
            return self._min_tcg_usd(prices)
        finally:
            self._stats["t_ptcg_total"] += time.perf_counter() - _t0

    def _real_tcg_brl(self, card_name: str, edition_name: str) -> Optional[float]:
        """Preço TCG real em BRL (USD do pokemontcg.io × câmbio), com cache por
        card id. Retorna None (→ fallback `.estat-tcg`) quando: sem câmbio,
        edição não mapeada, sem (NNN/MMM) no nome, ou sem cobertura."""
        if not self.fx_usd_brl:
            return None
        setcode = myp_edition_to_ptcg_setcode(edition_name)
        if not setcode:
            return None
        m = _COLLECTOR_NUM_RE.search(card_name or "")
        if not m:
            return None
        num = m.group(1).lstrip("0") or "0"
        cid = f"{setcode}-{num}"
        if cid in self._ptcg_cache:
            usd = self._ptcg_cache[cid]
        elif self.tcg_source_mode == "tcgcsv":
            # v5.15: modo só-tcgcsv — cache miss = sem cobertura tcgcsv. NÃO
            # consulta a pokemontcg.io (que de toda forma 404a no CI). Cai no
            # fallback `.estat-tcg` honesto.
            return None
        else:
            usd = self._fetch_ptcg_usd(cid)
            self._ptcg_cache[cid] = usd
            # v5.11.2: sleep adaptativo — com POKEMONTCG_API_KEY o limite é
            # 20k req/dia, então 0.3s basta. SÓ vale pra esta chamada
            # (pokemontcg.io); o delay anti-CF das páginas MYP fica intacto.
            time.sleep(min(self.delay, 0.3) if self.ptcg_api_key else self.delay)
        if usd is None:
            return None
        return usd * self.fx_usd_brl

    def _cid_for(self, card_name: str, edition_name: str) -> Optional[str]:
        """v5.15: deriva o cid `{setcode}-{num}` (mesma cascata de `_real_tcg_brl`)
        SEM tocar a rede/cache — usado só p/ rotular a proveniência do preço."""
        setcode = myp_edition_to_ptcg_setcode(edition_name)
        if not setcode:
            return None
        m = _COLLECTOR_NUM_RE.search(card_name or "")
        if not m:
            return None
        return f"{setcode}-{m.group(1).lstrip('0') or '0'}"

    def _real_tcg_source_label(self, card_name: str, edition_name: str) -> str:
        """v5.15: 'tcgcsv' se o preço real veio do dump tcgcsv, senão
        'pokemontcg.io'. Ambos são REAIS — rótulo só p/ auditoria/honestidade."""
        cid = self._cid_for(card_name, edition_name)
        if cid and cid in self._tcgcsv_cids:
            return "tcgcsv"
        return "pokemontcg.io"

    def _attribute_fallback(self, card_name: Optional[str], edition_name: str) -> None:
        """v5.13 (Iteração #2): classifica POR QUE o preço real não resolveu (→
        fallback `.estat-tcg`) num dos 4 baldes de cobertura. Cobertura faltante
        é a raiz dos falso-positivos (tcg_suspect/supranumerário sobrevivem só
        sem preço real), então saber qual balde domina aponta a menor mudança de
        maior efeito (ex.: muitos `unmapped_set` numa mesma era → adicionar 1
        setcode cobre o set inteiro).

        Re-deriva o motivo das MESMAS checagens-em-cascata de `_real_tcg_brl`
        (puro, barato, independe do cache — vale mesmo em cache-hit de None)."""
        if not self.fx_usd_brl:
            self._stats["fallback_no_fx"] += 1
        elif not myp_edition_to_ptcg_setcode(edition_name):
            self._stats["fallback_unmapped_set"] += 1
        elif not _COLLECTOR_NUM_RE.search(card_name or ""):
            self._stats["fallback_no_collector_num"] += 1
        else:
            self._stats["fallback_no_coverage"] += 1

    @staticmethod
    def _min_tcg_usd(prices: dict) -> Optional[float]:
        """Menor preço USD entre as variantes TCGplayer (`market`, senão `mid`);
        ignora ≤0. Conservador — não superestima a margem. Fonte ÚNICA da seleção
        de preço, usada pelo fetch por-card (`_fetch_ptcg_usd`) E pelo prefill
        por-set (`_prefill_ptcg_set`) — garante resultado idêntico nos 2 caminhos."""
        vals = []
        for p in (prices or {}).values():
            v = p.get("market") or p.get("mid")
            if v and float(v) > 0:
                vals.append(float(v))
        return min(vals) if vals else None

    def _prefill_ptcg_set(self, setcode: str) -> None:
        """v5.12: pré-carrega os preços TCGplayer do SET inteiro numa tacada
        (`GET /v2/cards?q=set.id:<setcode>`, paginado), populando `_ptcg_cache`.
        Troca N round-trips `/v2/cards/{id}` por ~1 request por set — o grosso do
        tempo (e do risco de 429) num scan largo.

        Cache POSITIVO: só popula cards que existem; cids ausentes caem no
        `_fetch_ptcg_usd` normal (preserva o 404→fallback `.estat-tcg` EXATO, e
        mantém os testes que mockam só `_fetch_ptcg_usd` válidos). Roda no máx.
        1× por setcode/run; falha de rede aborta em silêncio (fallback por-card
        assume)."""
        if setcode in self._prefilled_sets:
            return
        self._prefilled_sets.add(setcode)
        headers = {"X-Api-Key": self.ptcg_api_key} if self.ptcg_api_key else {}
        base = PTCG_API_BASE.rstrip("/")  # endpoint de busca não usa o /{id}
        backoffs = (5, 15, 30)
        page, page_size, cached = 1, 250, 0
        while True:
            url = f"{base}?q=set.id:{setcode}&page={page}&pageSize={page_size}"
            payload = None
            for attempt in range(len(backoffs) + 1):
                try:
                    r = self.session.get(url, headers=headers, timeout=20)
                except Exception as e:  # noqa: BLE001
                    log.debug(f"prefill {setcode} p{page} falhou: {e!r}")
                    return
                if r.status_code == 429 and attempt < len(backoffs):
                    time.sleep(backoffs[attempt])
                    continue
                if r.status_code != 200:
                    return
                try:
                    payload = r.json()
                except Exception:  # noqa: BLE001
                    return
                break
            if payload is None:
                return
            self._stats["ptcg_prefill_calls"] += 1
            data = payload.get("data") or []
            for card in data:
                num_raw = str(card.get("number") or "")
                if not num_raw.isdigit():
                    continue  # número não-numérico (TG/GG/promo) não casa o (NNN/MMM)
                cid = f"{setcode}-{num_raw.lstrip('0') or '0'}"
                if cid not in self._ptcg_cache:
                    tp = (card.get("tcgplayer") or {}).get("prices") or {}
                    self._ptcg_cache[cid] = self._min_tcg_usd(tp)
                    cached += 1
            total = payload.get("totalCount") or 0
            if not data or page * page_size >= total:
                break
            page += 1
        if cached:
            log.info(f"  💾 prefill {setcode}: {cached} preços TCG em cache (batch)")

    # ── v5.15: preço TCG REAL via tcgcsv.com (funciona no CI) ──
    def _tcgcsv_get_json(self, path: str) -> Optional[dict]:
        """GET tcgcsv + parse JSON, com o User-Agent obrigatório. None se falhar."""
        try:
            r = self.session.get(f"{TCGCSV_BASE}/{path}",
                                  headers={"User-Agent": TCGCSV_USER_AGENT}, timeout=20)
            if r.status_code != 200:
                log.debug(f"tcgcsv {path} status {r.status_code}")
                return None
            return r.json()
        except Exception as e:  # noqa: BLE001
            log.debug(f"tcgcsv {path} falhou: {e!r}")
            return None

    def _prefill_tcgcsv_set(self, setcode: str, edition: str) -> bool:
        """v5.15: pré-carrega os preços TCGplayer do SET via tcgcsv, populando o
        MESMO `_ptcg_cache` (keyed por `{setcode}-{num}`) que o caminho
        pokemontcg.io. Assim TODO o caminho de margem a jusante é reusado.

        Faz 2 requests: `/{groupId}/products` (productId→Number do colecionador)
        e `/{groupId}/prices` (productId→preço por subtype). Junta por productId,
        aplica `_min_tcg_usd` (menor market/mid entre subtypes — IDÊNTICO ao
        pokemontcg.io) e grava o USD no cache. Marca o cid em `_tcgcsv_cids` p/
        rotular a proveniência. Roda no máx. 1× por setcode/run (compartilha
        `_prefilled_sets` com o caminho pokemontcg.io — fontes mutuamente
        exclusivas por set). Retorna True se preencheu ≥1 preço; False se sem
        groupId/rede falhou (caller decide se cai no pokemontcg.io)."""
        if setcode in self._prefilled_sets:
            # já pré-carregado (por qualquer fonte) → True se há cids tcgcsv dele
            return any(c.startswith(f"{setcode}-") for c in self._tcgcsv_cids)
        if not self._tcgcsv_groups_fetched:
            self._tcgcsv_groups = tcgcsv_fetch_groups(self.session)
            self._tcgcsv_groups_fetched = True
        if not self._tcgcsv_groups:
            return False
        group_id = resolve_tcgcsv_group_id(setcode, edition, self._tcgcsv_groups)
        if not group_id:
            log.debug(f"tcgcsv: sem groupId p/ setcode {setcode} ({edition!r})")
            return False

        products = self._tcgcsv_get_json(f"{group_id}/products")
        prices = self._tcgcsv_get_json(f"{group_id}/prices")
        if not products or not prices:
            return False
        self._stats["tcgcsv_prefill_sets"] += 1

        # productId → número do colecionador (do extendedData "Number" = "NNN/MMM")
        num_by_pid: dict[int, str] = {}
        for p in products.get("results") or []:
            pid = p.get("productId")
            if pid is None:
                continue
            for ed in p.get("extendedData") or []:
                if ed.get("name") == "Number" and ed.get("value"):
                    num_by_pid[pid] = str(ed["value"])
                    break

        # productId → [preços USD] (1 por subtype) → _min_tcg_usd reusa a seleção
        from collections import defaultdict
        by_pid: dict[int, dict] = defaultdict(dict)
        for r in prices.get("results") or []:
            pid = r.get("productId")
            if pid is None:
                continue
            # adapta a forma tcgcsv {marketPrice, midPrice} p/ a forma esperada
            # por _min_tcg_usd ({market, mid}) — fonte ÚNICA de seleção de preço.
            sub = r.get("subTypeName") or f"_{len(by_pid[pid])}"
            by_pid[pid][sub] = {"market": r.get("marketPrice"),
                                "mid": r.get("midPrice")}

        cached = 0
        for pid, num_raw in num_by_pid.items():
            # "001/142" → numerador "1" (sem leading zeros), igual ao cid
            # pokemontcg.io ({setcode}-{num}); ignora não-numérico (TG/GG/promo).
            numerator = num_raw.split("/")[0].strip()
            if not numerator.isdigit():
                continue
            cid = f"{setcode}-{numerator.lstrip('0') or '0'}"
            usd = self._min_tcg_usd(by_pid.get(pid))
            if usd is None:
                continue
            if cid not in self._ptcg_cache or self._ptcg_cache[cid] is None:
                self._ptcg_cache[cid] = usd
                self._tcgcsv_cids.add(cid)
                cached += 1
        if cached:
            self._prefilled_sets.add(setcode)
            log.info(f"  💾 tcgcsv {setcode} (group {group_id}): {cached} preços "
                     f"TCG REAIS em cache (funciona no CI)")
            return True
        return False

    # ── Step 3: Scrape product detail page (v2 — per-seller language) ─
    def scrape_product(self, url: str, edition_name: str) -> Optional[CardData]:
        """Extract card data from product page, filtering sellers by language.

        v2 logic: The language is determined per SELLER ROW, not per product.
        Each <tr> in the seller table has an element with title="Português",
        title="Inglês", etc. We extract only EN seller prices and compare
        against the TCG Player reference price (which is always EN).
        """
        soup = self._get(url)
        if not soup:
            return None

        card = CardData()
        card.product_url = url
        card.edition = edition_name
        card.last_updated = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Name. v5.8.2: defensive fallback chain. h1 ausente acontece quando
        # MYP rota retorna template diferente (ex.: erro JS injetado) e o XLSX
        # 2026-05-17 saiu com Card Name=None em 1252/1252 rows. Backup: <title>
        # ou slug da URL.
        h1 = soup.select_one("h1")
        card.name = h1.get_text(strip=True) if h1 else ""
        if not card.name:
            title_tag = soup.find("title")
            if title_tag and title_tag.text:
                card.name = title_tag.text.split("|")[0].strip()
        if not card.name:
            slug = url.rstrip("/").split("/")[-1].replace("-", " ").strip()
            if slug and not slug.isdigit():
                card.name = slug.title()
        if not card.name:
            log.warning(f"  No name extractable from {url}")

        # v5.8.3 (2026-05-18): SKIP Jumbo (oversized) cards. `.estat-tcg` no MYP
        # reflete preço da carta standard, gerando deals fictícios com margem
        # gigante (ex.: M-Rayquaza-EX 098/98 XY 7 Jumbo). Skip ANTES de fetch
        # de tabela de seller pra economizar processamento e evitar contaminação.
        # v5.8.7: checa contra o nome RAW (antes do clean_card_name), pois o
        # keyword "Jumbo"/"oversized" costuma vir DEPOIS do "(NNN/MMM)" — limpar
        # primeiro removeria o keyword e burlaria o skip.
        if card.name and OVERSIZED_TITLE_RE.search(card.name):
            self._stats["skipped_jumbo"] += 1
            log.info(f"  ⏭️  Skipping oversized card: {card.name}")
            return None

        # v5.8.7: remove o nome EN duplicado que o <h1> concatena após
        # "(NNN/MMM)" (ex.: "Heatran-EX (109/116)Heatran-EX"). Copy-paste
        # limpo + casa o NUM_IN_NAME_RE ancorado do merge_myp_ct.py. DEPOIS do
        # skip de jumbo pra não engolir o keyword "Jumbo" trailing.
        card.name = clean_card_name(card.name)

        # Product code
        page_text = soup.get_text()
        code_match = re.search(r'pokemon_[a-z]{2,3}_[\w/]+', page_text)
        card.product_code = code_match.group(0) if code_match else ""

        # ── TCG Player price (always EN reference) ──
        # 2026-05-14 v5.3: usa findall + [-1] (mesma defensive pattern do
        # strikethrough fix). Cobre o caso de .estat-tcg ter múltiplos R$
        # (ex.: "Last R$ X | Avg R$ Y") — pega o último valor numérico em
        # vez de falhar parse com texto multi-preço.
        tcg_el = soup.select_one(".estat-tcg")
        if tcg_el:
            card.myp_declared_tcg_brl = self._last_brl(tcg_el.get_text())
        # v5.11: provisório = MYP declarado (.estat-tcg). É sobreposto pelo preço
        # REAL do TCGplayer (pokemontcg.io) depois do parse de sellers (candidatos
        # ≥ min_price).
        # v5.11.3 (A2, resgatado do PR #25): NÃO skipa mais aqui quando falta o
        # `.estat-tcg`. Com o preço real, um card sem declarado ainda pode ser
        # precificado na fonte. O skip por "sem TCG nenhum" acontece após o
        # override (declarado E real ausentes). O suspect-check abaixo só roda
        # quando há declarado.
        card.tcg_player_price = card.myp_declared_tcg_brl

        # v5.8 H2 (2026-05-16): capturar última venda real MYP pra sanity check.
        # MYP às vezes infla `.estat-tcg` (Jirachi PR-SM_SM161: declarava R$1499
        # mas TCGPlayer real $26=R$132 e última venda MYP foi R$19,99 — diff 75x).
        # Se TCG declarado >> última venda, provavelmente bug do MYP.
        last_sale_el = soup.select_one(".estatistica-ultimo")
        if last_sale_el:
            card.myp_last_sale_brl = self._last_brl(last_sale_el.get_text())

        # Sanity check: ratio TCG declarado / última venda real.
        # v5.11.3 (A2): guard `card.tcg_player_price` — sem declarado não há
        # ratio a checar (e evita TypeError sobre None).
        if card.tcg_player_price and card.myp_last_sale_brl and card.myp_last_sale_brl > 0:
            ratio = card.tcg_player_price / card.myp_last_sale_brl
            # v5.14.4: boundary INCLUSIVO (`>=`). Um ratio EXATAMENTE 10x (ex.
            # declarado R$1000 / última venda R$100 — números redondos comuns no
            # `.estat-tcg`) é anomalia tão forte quanto 10,01x; com `>` ele escapava
            # e virava deal "limpo" com margem possivelmente falsa (FP = erro caro
            # num scanner precision-first). FN é barato (vai pra "validar manual" e
            # o clear-on-real desfaz o suspect quando há preço real). Decisão
            # conjunta de 2 revisores; alinha com o eval asi-evolve (caso 10x = gold
            # suspect, PRECISION_FLOOR=1.0).
            if ratio >= TCG_SUSPECT_RATIO_THRESHOLD:
                # TCG declarado é ≥Nx última venda → MYP bug provável
                card.tcg_suspect = True
                self._stats["tcg_suspects"] += 1
                log.warning(
                    f"  🚨 TCG suspect: {card.name or url} | "
                    f"TCG declarado R${card.tcg_player_price:.2f} é "
                    f"{ratio:.1f}x última venda R${card.myp_last_sale_brl:.2f}. "
                    f"Provável inflação do .estat-tcg — excluído da sheet 🔥 Deals."
                )

        # ── Parse seller tables: extract EN sellers only ──
        # 2026-05-12: itera por tabela individualmente (não plano em tr)
        # pra detectar truncamento de EN. Padrão MYP: Tabela 0 (lojistas, cap ~15)
        # + Tabela 1 (marketplace, cap ~20). Quando uma tabela bate o cap E não
        # tem EN visível, há risco de listing EN-NM real mais barato escondido
        # (caso bartsimpson Psyduck R$300 EN sendo truncado por 20 listings PT/JP).
        en_prices = []
        en_sellers = 0
        jumbo_rows_seen = 0  # v5.8.3: rows com foil="Jumbo" (caso M-Rayquaza-EX XY 7)
        TABLE_CAP_THRESHOLD = 15   # tabela com >= 15 rows sem EN visível → candidato a truncamento

        seller_tables = soup.select("table.table-striped.table-bordered")
        if not seller_tables:
            # fallback: trata o documento inteiro como uma "tabela"
            seller_tables = [soup]

        # Coleta estatísticas por tabela primeiro; decisão de truncation_risk
        # acontece depois quando temos lowest_en pra comparar.
        per_table_stats = []
        for table in seller_tables:
            st = self._parse_seller_table(table)
            per_table_stats.append({
                "rows": st["rows"],
                "en": st["en"],
                "max_price": st["max_price"],
            })
            en_prices.extend(st["en_prices"])
            en_sellers += st["en"]
            jumbo_rows_seen += st["jumbo"]

        # Heurística de truncamento refinada (v5.3+): só dispara quando há
        # evidência de que listings escondidos PODEM ser mais baratos que o EN
        # reportado. Caso clássico Psyduck: Table 1 com 20 PT/JP, todos abaixo
        # de R$415 (lowest EN reportado de Table 0). Hidden listings (sorted asc
        # acima do visível) podem incluir EN entre [visible_max, lowest_en),
        # exatamente o que aconteceu com bartsimpson R$300. Quando max visível
        # já é >= lowest_en reportado, hidden listings começam acima disso e
        # não podem ser EN mais barato → não flag.
        #
        # v5.9 (2026-06-03): este sinal agora é o GATE de paginação, não o
        # veredito final. Quando dispara, os listings "escondidos" NÃO são
        # inacessíveis — estão nas páginas 2+ da tabela marketplace
        # (?estoque-outros-page=N). Seguimos essas páginas abaixo e só então
        # decidimos o truncation_risk final.
        page1_truncation_gate = False
        lowest_en_seen = min(en_prices) if en_prices else None
        if lowest_en_seen is not None:
            for ts in per_table_stats:
                if (ts["rows"] >= TABLE_CAP_THRESHOLD
                        and ts["en"] == 0
                        and ts["max_price"] > 0
                        and ts["max_price"] < lowest_en_seen):
                    page1_truncation_gate = True
                    break

        # v5.9: seguir a paginação da tabela marketplace quando a página 1
        # sinaliza truncation. Custo: cada produto truncado vira 1+N requests,
        # então só paginamos sob o gate (não em todo produto). Single-session
        # sequencial respeitando self.delay — NÃO paralelizar (CloudFlare 403).
        truncation_risk = False
        max_seller_page = self._max_seller_page(soup)
        # v5.10.1 (2026-06-07): cost gate. Um card só vira deal se MYP-EN ≥
        # min_price E margem ≥ threshold ⟹ TCG ≥ (1+threshold)·min_price >
        # min_price. Logo, se TCG < min_price o card NUNCA é deal e seria
        # filtrado adiante de qualquer forma — paginar pra resolver truncation
        # aqui é puro desperdício (medido: ~85% das paginações caíam em commons
        # < R$80). Só paginamos quando o card ainda pode ser deal.
        # v5.11.5 (A3): a premissa "TCG < min_price ⟹ nunca deal" só vale com o
        # preço REAL. O `card.tcg_player_price` aqui ainda é o `.estat-tcg`
        # DECLARADO (o real só é buscado adiante), e o MYP às vezes SUBdeclara
        # (mapeia a carta errada) — base-086 é o caso. Se a trava fosse decidir
        # com o declarado baixo, pularia a paginação e perderia o EN-NM barato
        # das páginas 2+ de uma carta que é deal de verdade. Então, quando a
        # trava está prestes a pular por TCG baixo, consulta o preço real ANTES
        # (cacheado por card-id → reusado no override adiante, sem request extra).
        gate_tcg = card.tcg_player_price or 0
        if page1_truncation_gate and max_seller_page >= 2 and gate_tcg < self.min_price:
            real_brl = self._real_tcg_brl(card.name, edition_name)
            if real_brl is not None:
                gate_tcg = max(gate_tcg, real_brl)
        can_be_deal = gate_tcg >= self.min_price
        if page1_truncation_gate and max_seller_page >= 2 and not can_be_deal:
            self._stats["pagination_skipped_low_tcg"] += 1
        elif page1_truncation_gate and max_seller_page >= 2:
            pages_to_fetch = min(max_seller_page, MAX_SELLER_PAGES)
            log.info(
                f"  📄 Truncation gate: paginando marketplace de {card.name or url} "
                f"(pág 2..{pages_to_fetch} de {max_seller_page})"
            )
            for pg in range(2, pages_to_fetch + 1):
                page_url = f"{url}?estoque-outros-page={pg}"
                page_soup = self._get(page_url)
                if page_soup is None:
                    # fetch falhou (rede/CF) → não conseguimos resolver: mantém
                    # o sinal de risco pra validação manual.
                    truncation_risk = True
                    self._stats["seller_page_fetch_failures"] += 1
                    log.warning(
                        f"  ⚠️ Falha ao buscar página {pg} da marketplace "
                        f"de {card.name or url} — EN-NM pode seguir truncado."
                    )
                    break
                # Só a tabela marketplace pagina; a de lojistas não muda entre
                # páginas, então parseamos APENAS o container marketplace pra não
                # recontar lojistas (que apareceriam de novo no reload completo).
                mkt = page_soup.select_one("#lista-anuncio-demais-vendedores")
                if mkt is None:
                    # sem container marketplace nesta página → fim natural
                    break
                pst = self._parse_seller_table(mkt)
                en_prices.extend(pst["en_prices"])
                en_sellers += pst["en"]
                jumbo_rows_seen += pst["jumbo"]
                self._stats["seller_pages_followed"] += 1
            else:
                # loop terminou sem break: se havia mais páginas além do cap,
                # o resto fica não-lido → ainda há risco residual de truncation.
                if max_seller_page > MAX_SELLER_PAGES:
                    truncation_risk = True
                    log.warning(
                        f"  ⚠️ {card.name or url}: {max_seller_page} páginas de "
                        f"marketplace > cap {MAX_SELLER_PAGES} — páginas extras "
                        f"não lidas, EN-NM pode seguir truncado."
                    )

        # v5.8.3: log se rows Jumbo foram filtradas
        if jumbo_rows_seen > 0:
            self._stats["jumbo_rows_filtered"] += jumbo_rows_seen
            log.info(
                f"  ⏭️  Skipped {jumbo_rows_seen} Jumbo seller row(s) "
                f"em {card.name or url}"
            )

        # If no EN+NM sellers found, skip
        if not en_prices:
            self._stats["skipped_no_en_sellers"] += 1
            return None

        # Filter: minimum price threshold
        lowest_en = min(en_prices)
        if lowest_en < self.min_price:
            self._stats["skipped_low_price"] += 1
            return None

        card.language = "EN"
        card.condition = "NM"
        card.myp_lowest_en_nm = min(en_prices)
        card.en_nm_sellers = en_sellers
        card.en_truncation_risk = truncation_risk
        # v5.8.3 (2026-05-18): 1 seller EN só = risco de mislabeling
        # (caso Flareon VMAX 018/203 Prize Pack: seller único listava como EN
        # carta sem print EN). Flag pra Validate Manually em vez de skip,
        # pra não suprimir deals legítimos de cards realmente raros.
        # v5.8.4 (2026-05-19): threshold agora configurável via CLI
        # (--min-en-sellers). `en_sellers < self.min_en_sellers` = flag.
        # Default 2 reproduz comportamento v5.8.3 (que era ≤1).
        if en_sellers < self.min_en_sellers:
            card.single_en_seller_risk = True
            self._stats["single_en_seller_risks"] += 1
            log.warning(
                f"  ⚠️ Low EN seller count: {card.name or url} | "
                f"{en_sellers} seller(s) EN-NM visível (< {self.min_en_sellers}) — "
                f"possível mislabeling de idioma. Validar manualmente."
            )
        if truncation_risk:
            self._stats["en_truncation_risks"] += 1
            log.warning(
                f"  🚨 EN truncation risk: {card.name} | "
                f"paginação da marketplace falhou ou excedeu cap "
                f"({MAX_SELLER_PAGES} págs) → lowest EN-NM R${card.myp_lowest_en_nm:.2f} "
                f"pode estar superestimado. Validar manualmente."
            )

        # ── Rarity ──
        # 2026-05-14 v5.3: page_text.lower() precomputed (era chamado N vezes
        # no loop, ~50µs cada para um page_text típico de 300KB).
        rarity_keywords = [
            "Illustration Rare", "Special Art Rare", "Hyper Rare",
            "Ultra Rare", "Secret Rare", "Art Rare", "Double Rare",
            "Rara Hiper", "Rara Ultra", "Rara Secreta", "Rara",
            "Incomum", "Comum",
        ]
        page_text_lower = page_text.lower()
        for rarity in rarity_keywords:
            if rarity.lower() in page_text_lower:
                card.rarity = rarity
                break

        # ── v5.11: preço TCG REAL (pokemontcg.io) p/ candidatos ──
        # Só busca o preço real pra cards que podem virar deal (EN-NM ≥ min_price)
        # — limita as requisições ao pokemontcg.io aos relevantes. Onde houver
        # cobertura, sobrepõe o `.estat-tcg` do MYP (que mapeia a carta errada em
        # base-086 etc.); onde não houver, mantém o declarado (fallback).
        if card.myp_lowest_en_nm and card.myp_lowest_en_nm >= self.min_price:
            real_brl = self._real_tcg_brl(card.name, edition_name)
            if real_brl is not None:
                card.tcg_player_price = real_brl
                card.tcg_real_usd = real_brl / self.fx_usd_brl
                # v5.15: rótulo de proveniência — 'tcgcsv' ou 'pokemontcg.io'
                # (ambos REAIS). `tcg_from_real` segue contando TODO preço real
                # (a métrica de honestidade não distingue a rota); `tcg_from_tcgcsv`
                # é o sub-contador da rota tcgcsv (a que funciona no CI).
                card.tcg_source = self._real_tcg_source_label(card.name, edition_name)
                self._stats["tcg_from_real"] += 1
                if card.tcg_source == "tcgcsv":
                    self._stats["tcg_from_tcgcsv"] += 1
                # v5.11.3 (A1, resgatado do PR #25): o preço agora é o REAL do
                # TCGplayer, não o `.estat-tcg` declarado. A flag de inflação do
                # declarado (tcg_suspect) não se aplica mais — limpa pra não
                # excluir indevidamente o card da sheet 🔥 Deals (a margem
                # agora é real).
                if card.tcg_suspect:
                    card.tcg_suspect = False
                    self._stats["tcg_suspects"] -= 1
            else:
                card.tcg_source = "myp_estat"
                self._stats["tcg_from_myp_fallback"] += 1
                self._attribute_fallback(card.name, edition_name)

        # v5.11.3 (A2): skip final — descarta só se NÃO há preço TCG nenhum
        # (nem `.estat-tcg` declarado nem real do pokemontcg.io). Antes o skip
        # era prematuro (antes do real), descartando cards que a fonte cobre.
        if not card.tcg_player_price:
            self._stats["skipped_no_tcg_price"] += 1
            return None

        # ── Calculate margin: lowest EN NM on MYP vs TCG Player EN ──
        # MARGEM BRUTA PURA (política cross-scanner 2026-06-06): só diferença de
        # preço entre produtos, SEM taxa/fee/markup embutido. O operador calcula
        # custos (frete, câmbio, taxas) por fora. NÃO adicionar multiplicador de
        # custo aqui (oposto do CardTrader scanner, que usa custo = preço × 1.06).
        if card.myp_lowest_en_nm and card.tcg_player_price and card.myp_lowest_en_nm > 0:
            card.margin_brl = card.tcg_player_price - card.myp_lowest_en_nm
            card.margin_pct = card.margin_brl / card.myp_lowest_en_nm

        # H3 fix (2026-05-12, refinado 2026-05-12 v5.3): heurística SIR/HR/SAR/IR misclassificado.
        # Bug documentado: cards supranumeráros (#>set_total) aparecem como rarity="Comum"
        # no MYP. Refinamento: extrai card_num/set_total do nome pra evitar falso
        # alarm em commons in-set genuinamente caros (raro mas possível).
        card_num_match = re.search(r"\((\d+)/(\d+)\)", card.name or "")
        is_supranumerary = False
        if card_num_match:
            try:
                num = int(card_num_match.group(1))
                total = int(card_num_match.group(2))
                is_supranumerary = num > total
                # v5.8.5 (2026-05-19): mesma extração serve pra
                # oversized_collector_risk. Quando numerator > denominator, o
                # card é variant fora do set numerado (SIR/HR/promo extra/
                # special illustration rare), frequentemente JP-only e com
                # preço TCG inflado em USD. Casos: Darumaka 097/086 (Black
                # Bolt SIR), Mew ex 232/091 (151 SIR), Charizard ex 234/091.
                # Sinaliza pra triagem visual; combina com single_en_seller
                # pra escalar pra Validate Manually.
                # Rarity-confidence gate (operador 2026-06-19): uma carta
                # supranumerária é MUITAS VEZES REAL — o problema é que a MYP às
                # vezes erra a RARIDADE dela, marcando como "Comum". Então o flag
                # significa "não confie nesse rótulo 'Comum', valide", NÃO "deal
                # falso". Uma supranumerária com raridade real (Rara/Hiper/etc.) é
                # carta normal e NÃO deve ser flagada só por ser supranumerária —
                # antes, flagar todas inundava o operador com 352 falsos flags em
                # cartas reais (de 486 supranumerárias, só 134 são "Comum"). Match
                # EXATO em "Comum" (lição NM-only: nunca substring). É flag/review,
                # nunca bloqueio: a carta segue aparecendo.
                if num > total and (card.rarity or "").strip() == "Comum":
                    card.oversized_collector_risk = True
                    self._stats["oversized_collector_risks"] += 1
                    log.info(
                        f"  ⚠️ Supranumerário + rarity='Comum': {card.name} "
                        f"({num}>{total}) — RARIDADE provavelmente mal-rotulada "
                        f"(real SIR/HR/SAR/ex marcado 'Comum'). Validar manual; "
                        f"não confiar no label 'Comum'."
                    )
            except (ValueError, TypeError):
                is_supranumerary = False  # unparseable, default to safe (no alarm)
        else:
            # Sem (X/Y) extraível, mantém heurística antiga como fallback
            is_supranumerary = True

        should_warn = (
            card.rarity in ("Comum", "Incomum")
            and card.tcg_player_price > SUPRANUMERARY_PRICE_THRESHOLD
            and is_supranumerary
        )
        if should_warn:
            self._stats["supranumerary_warnings"] += 1
            log.warning(
                f"  ⚠️ Possível SIR/HR/SAR misclassificado: {card.name} | "
                f"rarity='{card.rarity}' mas TCG R${card.tcg_player_price:.0f} é alto. "
                f"Validar manualmente antes de operar."
            )

        return card

    # ── Main scan ────────────────────────────────────────────────────
    # ── v5.11.4: checkpoint/resume ──
    def _save_checkpoint(self, path: str, done_editions: set) -> None:
        """Dump cards + edições já feitas + stats num JSON (escrita atômica via
        os.replace, pra não corromper se o processo morrer no meio do write)."""
        import json
        try:
            payload = {
                "version": CHECKPOINT_VERSION,
                "cards": [asdict(c) for c in self.cards],
                "done_editions": sorted(done_editions),
                "stats": self._stats,
            }
            tmp = f"{path}.tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(payload, f)
            os.replace(tmp, path)   # atômico
        except Exception as e:  # noqa: BLE001 — checkpoint é best-effort, nunca derruba o scan
            log.warning(f"  ⚠️ Falha ao salvar checkpoint {path}: {e!r}")

    def _load_checkpoint(self, path: str) -> set:
        """Restaura self.cards + self._stats e retorna o set de edições já feitas.
        Tolerante a checkpoint corrompido/versão diferente → começa do zero."""
        import json
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:  # noqa: BLE001
            log.warning(f"  ⚠️ Checkpoint {path} ilegível ({e!r}) — ignorando, scan do zero.")
            return set()
        if data.get("version") != CHECKPOINT_VERSION:
            log.warning(f"  ⚠️ Checkpoint {path} é de versão antiga — ignorando.")
            return set()
        fields = CardData.__dataclass_fields__
        # filtra chaves desconhecidas (defensivo a mudança de schema do CardData)
        self.cards = [
            CardData(**{k: v for k, v in c.items() if k in fields})
            for c in data.get("cards", [])
        ]
        st = data.get("stats")
        if isinstance(st, dict):
            self._stats.update(st)
        return set(data.get("done_editions", []))

    def scan(self, max_editions: int = 0, max_products: int = 0,
             edition_filter: list[str] = None,
             chunk_index: int = 0, chunk_total: int = 1,
             resume: bool = False, checkpoint_path: Optional[str] = None) -> list[CardData]:
        log.info("═" * 60)
        log.info("  MYP Cards Arbitrage Scanner")
        log.info(f"  Threshold: {self.margin_threshold*100:.0f}% | Language: EN only | Condition: NM")
        log.info(f"  Min price: R${self.min_price:.0f}")
        if edition_filter:
            log.info(f"  Edition filter: {', '.join(edition_filter)}")
        if chunk_total > 1:
            log.info(f"  Chunk: {chunk_index}/{chunk_total} (interleaved)")
        log.info("═" * 60)

        # v5.11: câmbio USD→BRL buscado UMA vez por run (preço real do TCGplayer
        # vem em USD). Se falhar, fx_usd_brl=None → todos os cards caem no
        # `.estat-tcg` do MYP (degrada pro comportamento ≤v5.10.1, com warning).
        self.fx_usd_brl = fetch_usd_brl(self.session)
        if self.fx_usd_brl:
            # v5.15: fonte do preço real conforme o modo.
            if self.tcg_source_mode == "tcgcsv":
                log.info("  💲 Preço TCG real via tcgcsv.com (ATIVO; funciona no CI); "
                         "fallback `.estat-tcg` onde não houver cobertura.")
            elif self.tcg_source_mode == "auto":
                log.info("  💲 Preço TCG real ATIVO (auto: tcgcsv primeiro — funciona "
                         "no CI; pokemontcg.io complementa); fallback `.estat-tcg` "
                         "onde não houver cobertura.")
            else:
                src = "key" if self.ptcg_api_key else "sem key (rate-limit menor)"
                log.info(f"  💲 Preço TCG real via pokemontcg.io ATIVO ({src}); "
                         f"fallback `.estat-tcg` onde não houver cobertura.")
            if not self.ptcg_api_key:
                # v5.11.2: a key grátis (dev.pokemontcg.io) elimina o throttle
                # 429 (backoff 5/15/30s) E ativa o sleep adaptativo de 0.3s —
                # em scan largo a diferença passa de 15 min.
                log.warning("  ⚠️ POKEMONTCG_API_KEY não definida — scans largos "
                            "sofrem throttle 429 + sleep cheio por cache miss. "
                            "Key grátis em dev.pokemontcg.io.")
        else:
            log.warning("  ⚠️ Sem câmbio USD→BRL — usando `.estat-tcg` (MYP) pra "
                        "todos os cards nesta run (preço real desativado).")

        editions = self.get_all_editions()

        # Filter by specific edition names (case-insensitive substring match)
        if edition_filter:
            filtered = []
            filter_lower = [f.lower().strip() for f in edition_filter]
            for ed in editions:
                title_lower = ed["title"].lower()
                for f in filter_lower:
                    if f in title_lower:
                        filtered.append(ed)
                        log.info(f"  ✅ Matched: '{ed['title']}' (filter: '{f}')")
                        break
            editions = filtered
            if not editions:
                log.warning("No editions matched the filter! Check edition names.")
                return []

        if max_editions:
            editions = editions[:max_editions]

        # v5.5: chunk slicing interleaved (load balanceado vs sequential blocks).
        # editions[N::M] garante distribuição equilibrada quando edition sizes
        # variam (sequential blocks colocariam todas as massivas num único chunk).
        if chunk_total > 1:
            if not (0 <= chunk_index < chunk_total):
                raise ValueError(
                    f"chunk_index={chunk_index} fora do range [0,{chunk_total})"
                )
            total_before = len(editions)
            editions = editions[chunk_index::chunk_total]
            log.info(
                f"  Chunk slicing: {total_before} editions → {len(editions)} "
                f"(chunk {chunk_index}/{chunk_total})"
            )

        # v5.11.4: resume — carrega progresso salvo e pula edições já feitas.
        done_editions: set = set()
        if resume and checkpoint_path and os.path.exists(checkpoint_path):
            done_editions = self._load_checkpoint(checkpoint_path)
            log.info(
                f"  ⏯️ Resume de {checkpoint_path}: {len(done_editions)} edição(ões) "
                f"já feitas, {len(self.cards)} cards restaurados."
            )

        for i, ed in enumerate(editions):
            if ed["url"] in done_editions:
                log.info(f"\n[{i+1}/{len(editions)}] ⏭️ (resume) já feita: {ed['title']}")
                continue
            log.info(f"\n[{i+1}/{len(editions)}] 📦 {ed['title']}")
            _ed_t0 = time.perf_counter()  # loop plumbing: tempo por edição

            product_urls = self.get_edition_products(ed["url"])
            if max_products:
                product_urls = product_urls[:max_products]
            log.info(f"  → {len(product_urls)} products found")

            # v5.12: prefill batch dos preços TCG do set (1 request no lugar de N
            # por-card). Só quando há câmbio e a edição mapeia a um setcode; cards
            # fora do batch caem no fetch por-card normal (fallback intacto).
            # v5.15: ordem de fontes conforme tcg_source_mode:
            #   - "auto"     → tcgcsv primeiro (funciona no CI); pokemontcg.io
            #                  cobre o set se o tcgcsv não tiver groupId/preço.
            #   - "tcgcsv"   → só tcgcsv (sem fallback pra pokemontcg.io).
            #   - "pokemontcg" → comportamento ≤v5.14 (só pokemontcg.io).
            # Ambos populam o mesmo _ptcg_cache; cards fora do batch caem no fetch
            # por-card pokemontcg.io normal (que no CI 404a → fallback honesto).
            if self.fx_usd_brl:
                _setcode = myp_edition_to_ptcg_setcode(ed["title"])
                if _setcode:
                    used_tcgcsv = False
                    if self.tcg_source_mode in ("auto", "tcgcsv"):
                        used_tcgcsv = self._prefill_tcgcsv_set(_setcode, ed["title"])
                    # pokemontcg.io complementa só em "auto" (quando o tcgcsv não
                    # cobriu o set) e em "pokemontcg".
                    if self.tcg_source_mode == "pokemontcg" or (
                        self.tcg_source_mode == "auto" and not used_tcgcsv
                    ):
                        self._prefill_ptcg_set(_setcode)

            for j, purl in enumerate(product_urls):
                self._stats["products_scanned"] += 1
                if (j + 1) % 10 == 0:
                    log.info(f"  Scanning {j+1}/{len(product_urls)}...")

                card = self.scrape_product(purl, ed["title"])
                if not card:
                    continue

                self._stats["en_found"] += 1
                card.edition_url = ed["url"]
                self.cards.append(card)

                if card.margin_pct is not None and card.margin_pct >= self.margin_threshold:
                    log.info(
                        f"  🔥 DEAL: {card.name} | "
                        f"EN NM lowest: R${card.myp_lowest_en_nm:,.2f} | "
                        f"TCG: R${card.tcg_player_price:,.2f} | "
                        f"Margin: {card.margin_pct*100:.1f}%"
                    )
                elif card.margin_pct is not None and card.margin_pct < 0:
                    log.debug(
                        f"  ⬇️ {card.name} | EN NM: R${card.myp_lowest_en_nm:,.2f} "
                        f"> TCG: R${card.tcg_player_price:,.2f} (negative)"
                    )

            # loop plumbing: tempo de parede da edição (medição do loop, mede
            # trabalho de scan; fica antes do checkpoint I/O de propósito)
            self._stats["t_editions_total"] += time.perf_counter() - _ed_t0

            # v5.11.4: checkpoint após cada edição concluída (escrita atômica).
            # Se o container reiniciar/morrer, `--resume` retoma daqui.
            if checkpoint_path:
                done_editions.add(ed["url"])
                self._save_checkpoint(checkpoint_path, done_editions)

        # Summary
        deals = [c for c in self.cards if c.margin_pct and c.margin_pct >= self.margin_threshold]
        log.info("\n" + "═" * 60)
        log.info(f"  Pages fetched: {self._stats['pages_fetched']}")
        log.info(f"  Products scanned: {self._stats['products_scanned']}")
        log.info(f"  EN cards found: {self._stats['en_found']}")
        log.info(f"  Cards with prices: {len(self.cards)}")
        log.info(f"  🔥 Deals (≥{self.margin_threshold*100:.0f}%): {len(deals)}")
        # M5 fix: funnel stats pra auditoria
        log.info(f"  ── Skipped breakdown (M5):")
        log.info(f"      No TCG price: {self._stats['skipped_no_tcg_price']}")
        log.info(f"      No EN sellers: {self._stats['skipped_no_en_sellers']}")
        log.info(f"      Low price (<R${self.min_price:.0f}): {self._stats['skipped_low_price']}")
        log.info(f"      Unknown lang titles (v5.4 H1): {self._stats['skipped_unknown_lang_titles']}")
        log.info(f"      Jumbo cards (title, v5.8.3): {self._stats['skipped_jumbo']}")
        log.info(f"      Jumbo seller rows filtered (v5.8.3): {self._stats['jumbo_rows_filtered']}")
        log.info(f"  ── Other diagnostics:")
        log.info(f"      Supranumerary warnings (H3): {self._stats['supranumerary_warnings']}")
        log.info(f"      EN truncation risks (T1): {self._stats['en_truncation_risks']}")
        log.info(f"      TCG suspects (H2 v5.8): {self._stats['tcg_suspects']}")
        log.info(f"      Single EN seller risks (v5.8.3): {self._stats['single_en_seller_risks']}")
        log.info(f"      Oversized collector# risks (v5.8.5): {self._stats['oversized_collector_risks']}")
        log.info(f"      Seller pages followed (v5.9 pagination): {self._stats['seller_pages_followed']}")
        log.info(f"      Seller page fetch failures (v5.9): {self._stats['seller_page_fetch_failures']}")
        log.info(f"      Pagination skipped (cost gate TCG<min, v5.10.1): {self._stats['pagination_skipped_low_tcg']}")
        log.info(f"      TCG real (v5.11): {self._stats['tcg_from_real']} "
                 f"(dos quais {self._stats['tcg_from_tcgcsv']} via tcgcsv — v5.15, "
                 f"funciona no CI; em {self._stats['tcgcsv_prefill_sets']} set(s))")
        log.info(f"      TCG fallback .estat-tcg MYP (v5.11): {self._stats['tcg_from_myp_fallback']}")
        # v5.13 (Iteração #2): por que caiu no fallback (raiz dos falso-positivos).
        # unmapped_set é o balde mais fixável: 1 setcode cobre o set inteiro.
        log.info(f"        ↳ fallback por motivo (v5.13): "
                 f"sem câmbio={self._stats['fallback_no_fx']} | "
                 f"set não-mapeado={self._stats['fallback_unmapped_set']} | "
                 f"sem nº colecionador={self._stats['fallback_no_collector_num']} | "
                 f"sem cobertura(404/sem preço)={self._stats['fallback_no_coverage']}")
        log.info(f"      HTTP retries (M1): {self._stats['http_retries']}")
        # loop plumbing: timings (perf_counter) pro loop de otimização iterativo
        log.info(f"  ── Timing (perf_counter):")
        log.info(f"      HTTP total (_get): {self._stats['t_http_total']:.1f}s")
        log.info(f"      pokemontcg.io: {self._stats['t_ptcg_total']:.1f}s em "
                 f"{self._stats['ptcg_calls']} chamadas por-card "
                 f"+ {self._stats['ptcg_prefill_calls']} prefill batch (v5.12)")
        log.info(f"      Editions wall-time: {self._stats['t_editions_total']:.1f}s")
        log.info("═" * 60)

        # v5.11.4: run completo → checkpoint não é mais necessário.
        if checkpoint_path and os.path.exists(checkpoint_path):
            try:
                os.remove(checkpoint_path)
            except OSError:
                pass

        return self.cards


# ══════════════════════════════════════════════════════════════════════
# XLSX GENERATOR
# ══════════════════════════════════════════════════════════════════════
def generate_xlsx(cards: list[CardData], output_path: str, threshold: float):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side("thin", "D9D9D9"), right=Side("thin", "D9D9D9"),
        top=Side("thin", "D9D9D9"), bottom=Side("thin", "D9D9D9"),
    )
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    normal = Font(name="Arial", size=10)
    bold_green = Font(name="Arial", size=10, bold=True, color="006100")
    # v5.8.8 (2026-05-29): célula de preço clicável. Mesmo azul/sublinhado que
    # add_card_hyperlinks.py / revalidate_deals.py usam no Card Name (0563C1),
    # mantido como Arial 10 pra casar o corpo das sheets de cards.
    HYPERLINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")

    # v5.8 (2026-05-16): 2 colunas novas pra surfaçar o sanity check H2:
    #   - "MYP Last Sale (R$)" entre TCG Player e Margin %
    #   - "⚠️ TCG Suspect" depois de EN Trunc
    # Sem isso, o operador via Jirachi PR-SM_SM161 como deal #1 a 1400% mesmo
    # com TCG inflado 75x vs última venda real. Aggregate lê via dict-by-name,
    # então a ordem das colunas não quebra o pipeline.
    # v5.8.5 (2026-05-19): nova coluna `⚠️ COLLECTOR#` depois de Single Seller.
    # Sinaliza cards onde collector_number > set_size (variant SIR/HR/promo
    # extra, frequentemente JP-only). Casos Darumaka 097/086, Mew ex 232/091.
    # Aggregate lê via dict-by-name, então ordem não quebra o pipeline.
    # v5.11.1 (2026-06-09): coluna "TCG US$" (preço real em USD via pokemontcg.io,
    # card.tcg_real_usd) exposta no XLSX pra alimentar a tabela de ENTREGA do
    # myp_summary.py (formato aprovado pelo operador: links clicáveis MYP + TCG).
    # Lida por nome de header (dict-by-name) → não quebra aggregate nem chunks
    # antigos (.get() retorna None). NÃO é taxa nem altera o cálculo de margem
    # (margem segue em BRL, BRUTA pura).
    # v5.11.2 (2026-06-10): coluna "TCG URL" (texto plano, fim da lista) — o
    # link TCGplayer já era computado pro hyperlink da célula "TCG Player (R$)",
    # mas hyperlink de célula não sobrevive a leitores dict-by-name (pandas/
    # openpyxl values-only). O scanner integrado consome esta coluna.
    # v5.14 (2026-06-20): coluna "TCG Source" EXPLÍCITA logo após o preço.
    # POR QUE: a distinção real-vs-fallback era IMPLÍCITA (presença de "TCG US$").
    # Isso mascarava a degradação silenciosa do CI (runners do GitHub não
    # alcançam api.pokemontcg.io → toda chamada cai no fallback `.estat-tcg`,
    # mas o XLSX não dizia "fallback" em lugar nenhum). Agora cada card declara
    # a fonte do preço usado na margem: `pokemontcg.io` (TCGplayer REAL) ou
    # `myp_estat` (fallback do campo MYP, margem NÃO-confiável). É a fonte de
    # verdade do sinal de honestidade — não inferir por presença de outra coluna.
    headers = [
        "Card Name", "Edition", "Rarity",
        "MYP EN NM (R$)", "TCG Player (R$)", "TCG US$", "TCG Source", "MYP Last Sale (R$)",
        "Margin %", "Diff (R$)", "NM Sellers",
        "⚠️ EN Trunc", "⚠️ TCG Suspect", "⚠️ Single Seller", "⚠️ COLLECTOR#",
        "URL", "Updated", "TCG URL",
    ]
    widths = [38, 32, 16, 16, 16, 12, 14, 17, 11, 13, 10, 11, 14, 14, 14, 55, 16, 55]
    PRICE_COLS = {4, 5, 8, 10}      # MYP EN NM, TCG Player, Last Sale, Diff
    MYP_PRICE_COL = 4               # v5.8.8: hyperlink → página produto MYP
    TCG_PRICE_COL = 5               # v5.8.8: hyperlink → busca TCGplayer por nome
    TCG_SOURCE_COL = 7              # v5.14: fonte do preço (real/fallback)
    MARGIN_COL = 9
    EN_TRUNC_COL = 12
    TCG_SUSPECT_COL = 13
    SINGLE_SELLER_COL = 14
    COLLECTOR_COL = 15

    def write_headers(ws):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    def write_card_row(ws, row, card):
        diff = (card.tcg_player_price or 0) - (card.myp_lowest_en_nm or 0)
        trunc_flag = "⚠️ MAYBE" if card.en_truncation_risk else ""
        suspect_flag = "🚨 SUSPECT" if card.tcg_suspect else ""
        single_flag = "⚠️ 1 SELLER" if card.single_en_seller_risk else ""
        collector_flag = "⚠️ VARIANT" if card.oversized_collector_risk else ""
        # v5.8.8: links das células de preço. MYP EN NM → página do produto
        # (card.product_url, populado em 100% das rows verificadas no XLSX
        # 2026-05-27). TCG Player → DIRETA via pokemontcg.io redirect quando
        # MYP edition é mapeada + collector# está in-range; caso contrário,
        # cai pra busca-por-nome (v5.8.8 original behavior).
        # v5.8.9: tcg_direct_url tenta o redirect; None → fallback ao search.
        # Cobertura honesta em scan vintage-heavy weekly 2026-05-27 ≈ 2% de
        # rows pegam link direto; em daily-scan de SV moderno (8 substrings),
        # cobertura é alta. Fallback de busca cobre o restante.
        tcg_link = (
            tcg_direct_url(
                card.name, card.edition,
                oversized_collector_risk=card.oversized_collector_risk,
            )
            or tcg_search_url(card.name)
        )
        # v5.14: rótulo legível da fonte do preço. REAL = preço verificável do
        # TCGplayer (via pokemontcg.io OU, v5.15, via tcgcsv — a rota que funciona
        # no CI); `myp_estat` (ou vazio) = FALLBACK (.estat-tcg do MYP, margem
        # suspeita). v5.15: ambas as rotas reais saem rotuladas `real (<fonte>)`.
        _REAL_SOURCES = {"pokemontcg.io": "real (pokemontcg.io)", "tcgcsv": "real (tcgcsv)"}
        tcg_source_label = _REAL_SOURCES.get(card.tcg_source, "fallback (.estat-tcg)")
        vals = [
            card.name, card.edition, card.rarity,
            card.myp_lowest_en_nm, card.tcg_player_price, card.tcg_real_usd,
            tcg_source_label,
            card.myp_last_sale_brl,
            card.margin_pct, diff, card.en_nm_sellers,
            trunc_flag, suspect_flag, single_flag, collector_flag,
            card.product_url, card.last_updated,
            tcg_link or "",  # v5.11.2: "TCG URL" texto plano (mesmo link do hyperlink)
        ]
        USD_COL = 6  # v5.11.1: "TCG US$" — formato USD, não BRL
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = normal
            c.border = border
            if col in PRICE_COLS:
                c.number_format = '#,##0.00'
            if col == USD_COL and v is not None:
                c.number_format = '"US$"#,##0.00'
            if col == MYP_PRICE_COL and v is not None and card.product_url:
                c.hyperlink = card.product_url
                c.font = HYPERLINK_FONT
            if col == TCG_PRICE_COL and v is not None and tcg_link:
                c.hyperlink = tcg_link
                c.font = HYPERLINK_FONT
            if col == MARGIN_COL:
                # v5.8.6 bug #5: standardize on 2-decimal % across the
                # pipeline (revalidate_deals.py also uses "0.00%"). Header
                # is "Margin %" — value is stored as fraction (e.g. 0.483)
                # so format must render as percentage to match semantics.
                c.number_format = '0.00%'
                if v and v >= 0.50:
                    c.font = bold_green
                    c.fill = green_fill
                elif v and v >= threshold:
                    c.fill = yellow_fill
                elif v and v < 0:
                    c.fill = red_fill
            if col == TCG_SOURCE_COL and card.tcg_source not in ("pokemontcg.io", "tcgcsv"):
                # v5.15: só FALLBACK (.estat-tcg) é destacado em amarelo — as duas
                # rotas REAIS (pokemontcg.io / tcgcsv) ficam sem destaque.
                c.fill = yellow_fill
                c.alignment = Alignment(horizontal="center")
            if col == EN_TRUNC_COL and card.en_truncation_risk:
                c.fill = red_fill
                c.alignment = Alignment(horizontal="center")
            if col == TCG_SUSPECT_COL and card.tcg_suspect:
                c.fill = red_fill
                c.font = Font(bold=True, color="9C0006", name="Arial", size=10)
                c.alignment = Alignment(horizontal="center")
            if col == SINGLE_SELLER_COL and card.single_en_seller_risk:
                c.fill = yellow_fill
                c.alignment = Alignment(horizontal="center")
            if col == COLLECTOR_COL and card.oversized_collector_risk:
                c.fill = yellow_fill
                c.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Deals ──
    # v5.8 (2026-05-16): exclui cards com tcg_suspect (TCG declarado >10x última
    # venda real). Jirachi PR-SM_SM161 era #1 a 1400% com TCG=R$1499 fictício;
    # ratio 75x da última venda. Suspects ainda aparecem em `All EN Cards` e na
    # sheet dedicada `🚨 TCG Suspect` pra inspeção.
    ws1 = wb.active
    ws1.title = "🔥 Deals"
    write_headers(ws1)

    # v5.8.3 (2026-05-18): excluía single_en_seller_risk de Deals.
    # v5.8.4 (2026-05-19): refinamento — single-seller SOZINHO mantém em
    # Deals (com coluna visual `⚠️ 1 SELLER`). Só vira Validate-Manually se
    # acompanhado de tcg_suspect OU en_truncation_risk.
    # v5.8.5 (2026-05-19): oversized_collector_risk segue mesma lógica:
    # sozinho mantém em Deals (coluna `⚠️ COLLECTOR#`), combinado com
    # single_en_seller_risk escala pra Validate Manually (ambos sinais
    # complementares — variant + idioma duvidoso = JP-mislabeled-as-EN).
    def _combined_single_seller_risk(c) -> bool:
        return c.single_en_seller_risk and (
            c.tcg_suspect
            or c.en_truncation_risk
            or c.oversized_collector_risk
        )
    deals = sorted(
        [c for c in cards
         if c.margin_pct and c.margin_pct >= threshold
         and not c.tcg_suspect
         and not _combined_single_seller_risk(c)],
        key=lambda x: x.margin_pct or 0, reverse=True,
    )
    for i, card in enumerate(deals, 2):
        write_card_row(ws1, i, card)
        if card.margin_pct and card.margin_pct >= 0.50:
            for col in range(1, 4):
                ws1.cell(row=i, column=col).fill = green_fill

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(deals)+1, 2)}"

    # ── Sheet 2: All EN Cards ──
    ws2 = wb.create_sheet("All EN Cards")
    write_headers(ws2)
    all_sorted = sorted(cards, key=lambda x: x.margin_pct or -999, reverse=True)
    for i, card in enumerate(all_sorted, 2):
        write_card_row(ws2, i, card)

    # ── Sheet 3: Top 50 by Margin (visual review pool) ──
    # Operador inspeciona visualmente pra decidir se é chase (pokémon
    # icônico, arte bonita, etc.) — não filtra por threshold.
    # v5.4 M3: filtra None-margin antes de slice (evita padding visual em
    # runs com <50 cards válidos).
    ws_top = wb.create_sheet("🏆 Top 50 Margin")
    write_headers(ws_top)
    top50 = [c for c in all_sorted if c.margin_pct is not None][:50]
    for i, card in enumerate(top50, 2):
        write_card_row(ws_top, i, card)
    ws_top.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(top50)+1, 2)}"

    # ── Sheet 4: 🚨 Validate Manually ──
    # Inclui cards com qualquer flag de risco de detecção:
    #   - en_truncation_risk (2026-05-12 v5.3): seller table no cap sem EN visível
    #   - single_en_seller_risk (v5.8.3 2026-05-18): 1 seller EN → possível mislabeling
    #   - v5.8.5 (2026-05-19): oversized_collector_risk SOZINHO permanece em Deals,
    #     mas combinado com single_en_seller_risk aparece aqui (variant +
    #     idioma duvidoso = JP-mislabeled-as-EN). Mantém escopo enxuto.
    ws_val = wb.create_sheet("🚨 Validate Manually")
    write_headers(ws_val)
    validate = sorted(
        [c for c in cards
         if c.en_truncation_risk
         or c.single_en_seller_risk
         or (c.oversized_collector_risk and c.single_en_seller_risk)],
        key=lambda x: x.margin_pct or -999, reverse=True,
    )
    for i, card in enumerate(validate, 2):
        write_card_row(ws_val, i, card)
    if validate:
        ws_val.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(validate)+1}"

    # ── Sheet 5: 🚨 TCG Suspect (v5.8) ──
    # Cards com TCG declarado >10x última venda real do MYP — provável bug do
    # campo .estat-tcg (caso Jirachi PR-SM_SM161). Excluídos de `🔥 Deals` mas
    # exibidos aqui pra inspeção manual antes de descartar definitivamente.
    ws_susp = wb.create_sheet("🚨 TCG Suspect")
    write_headers(ws_susp)
    suspects = sorted(
        [c for c in cards if c.tcg_suspect],
        key=lambda x: x.margin_pct or -999, reverse=True,
    )
    for i, card in enumerate(suspects, 2):
        write_card_row(ws_susp, i, card)
    if suspects:
        ws_susp.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(suspects)+1}"

    # ── Sheet 6: Summary ──
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions['A'].width = 32
    ws3.column_dimensions['B'].width = 25
    title_font = Font(bold=True, size=16, name="Arial")
    label_font = Font(bold=True, name="Arial", size=11)

    ws3.cell(row=1, column=1, value="MYP Arbitrage Scanner").font = title_font
    ws3.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = normal
    ws3.cell(row=4, column=1, value="Margin Threshold").font = label_font
    ws3.cell(row=4, column=2, value=f"{threshold*100:.0f}%").font = normal
    ws3.cell(row=5, column=1, value="Language Filter").font = label_font
    ws3.cell(row=5, column=2, value="English (EN)").font = normal
    ws3.cell(row=6, column=1, value="Total EN Cards").font = label_font
    ws3.cell(row=6, column=2, value=len(cards)).font = normal
    ws3.cell(row=7, column=1, value="Deals Found (clean)").font = label_font
    ws3.cell(row=7, column=2, value=len(deals)).font = bold_green
    # v5.8: surface TCG suspects + truncation risks no Summary
    ws3.cell(row=8, column=1, value="🚨 TCG Suspects").font = label_font
    ws3.cell(row=8, column=2, value=len(suspects)).font = normal

    ws3.cell(row=10, column=1, value="Top 10 Deals:").font = Font(bold=True, size=12, name="Arial")
    for i, d in enumerate(deals[:10], 11):
        ws3.cell(row=i, column=1, value=d.name).font = normal
        margin_cell = ws3.cell(row=i, column=2, value=f"{d.margin_pct*100:.1f}% — R${d.margin_brl:,.2f}")
        margin_cell.font = bold_green

    wb.save(output_path)
    log.info(f"📊 Spreadsheet saved: {output_path}")
    return output_path


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="MYP Cards Arbitrage Scanner — Pokémon TCG Singles",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python myp_arbitrage_scanner.py                           # Scan completo
  python myp_arbitrage_scanner.py --max-editions 3          # Teste com 3 edições
  python myp_arbitrage_scanner.py --threshold 40 --delay 2  # 40% margin, 2s delay
  python myp_arbitrage_scanner.py -o deals.xlsx             # Output customizado
  python myp_arbitrage_scanner.py --editions "Ascended Heroes" "Prismáticas"  # Edições específicas
        """,
    )
    parser.add_argument("--max-editions", type=int, default=0,
                       help="Limite de edições (0 = todas, ~326 total)")
    parser.add_argument("--max-products", type=int, default=0,
                       help="Limite de produtos por edição (0 = todos)")
    parser.add_argument("--threshold", type=float, default=30,
                       help="Margem BRUTA mínima %% para alerta (default: 30). "
                            "Percent integer (30 = 30%%); valor <1.0 auto-converte. "
                            "Margem bruta = (preço_alvo TCG − preço_BR)/preço_BR, SEM taxa embutida.")
    parser.add_argument("--min-price", type=float, default=50,
                       help="Preço mínimo EN em R$ (default: 50)")
    parser.add_argument("--delay", type=float, default=1.5,
                       help="Delay entre requests em segundos (default: 1.5)")
    parser.add_argument("--min-en-sellers", type=int,
                       default=MIN_EN_SELLERS_FOR_DEALS_DEFAULT,
                       help=f"Min EN-NM sellers for Deals inclusion (default: "
                            f"{MIN_EN_SELLERS_FOR_DEALS_DEFAULT}; was hardcoded "
                            f"in v5.8.3). Cards abaixo são flagged como "
                            f"single_en_seller_risk e podem cair em Validate "
                            f"Manually conforme outras flags.")
    parser.add_argument("--editions", nargs="+", type=str, default=None,
                       help="Filtrar por edições específicas (substring match). Ex: --editions \"Ascended Heroes\" \"Prismáticas\"")
    # v5.15: fonte do preço TCG real.
    parser.add_argument("--tcg-source", choices=["auto", "tcgcsv", "pokemontcg"],
                       default="auto",
                       help="Fonte do preço TCG REAL (default: auto). "
                            "'tcgcsv' = dump diário do TCGplayer via tcgcsv.com "
                            "(ÚNICA fonte que funciona nos runners do GitHub "
                            "Actions); 'pokemontcg' = só pokemontcg.io "
                            "(comportamento ≤v5.14; falha no CI); 'auto' = tcgcsv "
                            "primeiro, pokemontcg.io complementa por set sem "
                            "groupId tcgcsv. Em qualquer modo, sem fonte real = "
                            "fallback `.estat-tcg` honesto.")
    parser.add_argument("-o", "--output", type=str, default="",
                       help="Caminho do arquivo .xlsx de saída")
    # v5.5: chunk slicing pra GH Actions matrix job
    parser.add_argument("--chunk-index", type=int, default=0,
                       help="Índice do chunk (0-based). Usado com --chunk-total pra dividir scan em jobs paralelos.")
    parser.add_argument("--chunk-total", type=int, default=1,
                       help="Total de chunks (1 = sem chunking). Editions são fatiadas via slicing interleaved.")
    # v5.11.4: resume após kill do container. Salva progresso por edição em
    # `<output>.resume.json`; com --resume, retoma de onde parou.
    parser.add_argument("--resume", action="store_true",
                       help="Retoma de um checkpoint `<output>.resume.json` se existir "
                            "(pula edições já feitas). Útil quando o container reinicia.")
    args = parser.parse_args()

    # C1 fix (2026-05-12): MYP usa percent integer (35 = 35%), oposto do CT
    # scanner que usa fração (0.35). Se o operador passar < 1.0, é provável
    # que tenha confundido as convenções. Auto-converte com warning.
    if args.threshold < 1.0:
        log.warning(
            f"--threshold {args.threshold} < 1.0 parece fração (convenção CT scanner), "
            f"mas MYP usa percent. Convertendo para {args.threshold * 100}."
        )
        args.threshold = args.threshold * 100

    threshold_frac = args.threshold / 100
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_path = args.output or f"myp_arbitrage_{timestamp}.xlsx"

    scraper = MYPScraper(
        delay=args.delay, min_en_sellers=args.min_en_sellers,
        threshold=threshold_frac, min_price=args.min_price,
        tcg_source=args.tcg_source,
    )
    log.info(
        f"Config: threshold={args.threshold}%, min_price=R${args.min_price}, "
        f"delay={args.delay}s, min_en_sellers={args.min_en_sellers}, "
        f"tcg_source={args.tcg_source}"
    )
    # v5.11.4: checkpoint vive ao lado do XLSX de saída (sobrevive ao reciclo
    # do container; o XLSX em si só é escrito no fim).
    checkpoint_path = f"{output_path}.resume.json"
    cards = scraper.scan(max_editions=args.max_editions, max_products=args.max_products,
                         edition_filter=args.editions,
                         chunk_index=args.chunk_index, chunk_total=args.chunk_total,
                         resume=args.resume, checkpoint_path=checkpoint_path)

    # v5.4 M1 + invariant check: cron precisa distinguir "scan saudável com
    # zero deals" de "scraper quebrado". Exit codes:
    #   0 = healthy run (com ou sem deals — tem cards, OU chunk vazio legítimo)
    #   1 = scraper provavelmente quebrado (funnel collapsou OU sem cards)
    #   2 = filter user-error (--editions não casou nada)
    import sys as _sys
    stats = scraper._stats

    # v5.5 fix: chunk vazio legítimo. Quando chunk_total > 1 e editions[N::M]
    # retorna lista vazia (ex.: --editions casou 1 edição mas chunk_total=6 →
    # só chunk 0 tem trabalho), o chunk deve sair limpo com exit 0, NÃO
    # marcar o job como falha. Aggregate ignora chunks que não produziram XLSX.
    is_empty_chunk = (
        args.chunk_total > 1
        and stats["products_scanned"] == 0
        and stats["pages_fetched"] > 0  # catalog scrape rodou OK
    )
    if is_empty_chunk:
        log.info(
            f"✓ Chunk {args.chunk_index}/{args.chunk_total} vazio após slicing "
            f"(zero editions atribuídas a este chunk). Saindo limpo, "
            f"aggregate ignora chunks sem XLSX."
        )
        _sys.exit(0)

    if not cards:
        # Distinção: filter typo vs site/scraper broken
        if args.editions:
            log.error(
                f"❌ --editions filter ({', '.join(args.editions)}) não casou "
                f"nenhuma edição. Verificar nomes (substring match contra title MYP)."
            )
            _sys.exit(2)
        # Sem filter, ou filter casou mas processou zero — likely broken
        log.error(
            f"❌ Scan retornou zero cards. Funnel: "
            f"pages={stats['pages_fetched']}, products={stats['products_scanned']}, "
            f"en_found={stats['en_found']}. Check .debug/ HTML samples."
        )
        _sys.exit(1)

    # v5.4 invariant: muita página fetchada mas zero EN encontrado = scraper broken
    if stats["pages_fetched"] > 100 and stats["en_found"] == 0:
        log.error(
            f"❌ Invariant violation: {stats['pages_fetched']} páginas baixadas "
            f"mas {stats['en_found']} EN cards encontrados. Provável: selector "
            f"break, language detector quebrado, ou MYP rebuild. Check warnings."
        )
        _sys.exit(1)

    generate_xlsx(cards, output_path, scraper.margin_threshold)
    print(f"\nDone! Open: {output_path}")
