"""Offline rate-limit recovery and product checkpoint regressions."""
import json
from datetime import datetime, timezone, timedelta
from email.utils import format_datetime
from types import SimpleNamespace

import pytest
import requests
import myp_arbitrage_scanner as m


def test_retry_after_seconds_and_http_date():
    assert m.rate_limit_wait(SimpleNamespace(headers={'Retry-After': '180'}), 0) == 180
    future = format_datetime(datetime.now(timezone.utc) + timedelta(seconds=240))
    assert 238 < m.rate_limit_wait(SimpleNamespace(headers={'Retry-After': future}), 0) <= 240
    assert m.rate_limit_wait(SimpleNamespace(headers={'Retry-After': 'nan'}), 1) == 120


def test_repeated_429_waits_and_stops_instead_of_skipping(monkeypatch):
    sc = m.MYPScraper(delay=0)
    sleeps = []
    monkeypatch.setattr(m.time, 'sleep', sleeps.append)
    response = requests.Response()
    response.status_code = 429
    response.headers['Retry-After'] = '90'
    response.url = 'https://example.org/card'
    calls = []
    def get(*a, **kw):
        calls.append(1)
        return response
    monkeypatch.setattr(sc.session, 'get', get)
    with pytest.raises(m.ScanInterrupted, match='429'):
        sc._get(response.url)
    assert len(calls) == 3
    assert [s for s in sleeps if s] == [90, 120]


def prepare(sc, monkeypatch):
    monkeypatch.setattr(m, 'fetch_usd_brl', lambda *a: None)
    monkeypatch.setattr(sc, 'get_all_editions', lambda: [{'title': 'Test edition', 'url': 'ed'}])
    monkeypatch.setattr(sc, 'get_edition_products', lambda *a: ['p1', 'p2', 'p3'])


def test_resume_mid_edition_does_not_duplicate_or_skip_failed_product(monkeypatch, tmp_path):
    path = str(tmp_path/'nested'/'run.resume.json')
    sc = m.MYPScraper(delay=0)
    prepare(sc, monkeypatch)
    def scrape(url, edition):
        if url == 'p2':
            raise m.ScanInterrupted('429')
        return m.CardData(name=url, product_url=url, myp_lowest_en_nm=100, tcg_player_price=200)
    monkeypatch.setattr(sc, 'scrape_product', scrape)
    with pytest.raises(m.ScanInterrupted):
        sc.scan(checkpoint_path=path)
    saved = json.loads(open(path, encoding='utf-8').read())
    assert saved['done_products'] == ['ed|p1']
    assert saved['done_editions'] == []
    assert len(saved['cards']) == 1
    resumed = m.MYPScraper(delay=0)
    prepare(resumed, monkeypatch)
    seen = []
    def complete(url, edition):
        seen.append(url)
        return m.CardData(name=url, product_url=url)
    monkeypatch.setattr(resumed, 'scrape_product', complete)
    cards = resumed.scan(resume=True, checkpoint_path=path)
    assert seen == ['p2', 'p3']
    assert [c.name for c in cards] == ['p1', 'p2', 'p3']
    assert not __import__('pathlib').Path(path).exists()


def test_checkpoint_wrong_scope_is_ignored(tmp_path):
    sc = m.MYPScraper(delay=0)
    sc._resume_context = {'editions': ['A']}
    sc.cards = [m.CardData(name='old')]
    path = str(tmp_path/'resume.json')
    sc._save_checkpoint(path, {'old-edition'})
    fresh = m.MYPScraper(delay=0)
    fresh._resume_context = {'editions': ['B']}
    assert fresh._load_checkpoint(path) == set()
    assert fresh.cards == []


def test_cli_exports_partial_xlsx_and_status(monkeypatch, tmp_path):
    import sys
    from openpyxl import load_workbook
    out = tmp_path/'partial.xlsx'
    def interrupted(self, **kwargs):
        self.cards = [m.CardData(name='Captured card', myp_lowest_en_nm=60, tcg_player_price=100)]
        raise m.ScanInterrupted('HTTP 429 test')
    monkeypatch.setattr(m.MYPScraper, 'scan', interrupted)
    monkeypatch.setattr(sys, 'argv', ['scan', '-o', str(out)])
    with pytest.raises(SystemExit) as exc:
        m.main()
    assert exc.value.code == 2
    assert out.exists() and out.with_suffix('.xlsx.resume.json').exists()
    status = json.loads(out.with_suffix('.xlsx.status.json').read_text(encoding='utf-8'))
    assert status['status'] == 'partial' and status['cards'] == 1
    wb = load_workbook(out)
    assert wb['All EN Cards'].max_row == 2
    assert wb['All EN Cards'].cell(2, 1).value == 'Captured card'
    wb.close()
