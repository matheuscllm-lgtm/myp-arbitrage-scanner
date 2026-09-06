"""Entrega canônica do scan MYP → DRAGON BALL: XLSX → markdown.

Espelho do `myp_summary.py` (mesmo contrato de entrega do repo: rodar a
ferramenta e colar o markdown VERBATIM no chat — nunca remontar tabela à mão,
nunca dropar o link [TCG]). Lê o XLSX gerado pelo `myp_dbz_scanner.py`.

Buckets (todos com 2 links por linha):
  🟢 Top 50 deals limpos      — margem ≥ threshold, referência tcgcsv real,
                                 sem nenhuma flag;
  🚨 REVISAR                  — deals com flag (possível lixo <50% da ref |
                                 referência volátil | 1 seller EN) — validar
                                 manualmente, com o motivo por linha;
  ⚠️ Sem referência TCG       — oferta EN NM viva ≥ piso SEM produto tcgcsv
                                 casado (join determinístico não casou) —
                                 nada some em silêncio;
  🚨 EN truncation risk       — diagnóstico (menor EN NM pode estar
                                 superestimado), igual ao fluxo Pokémon.

Uso:
    python myp_dbz_summary.py results/myp_dbz_<ts>.xlsx \
        -o results/dbz-2026-08-03.md [--run-id N]
"""
from __future__ import annotations

from chat_format import reference_price

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

if (sys.stdout.encoding or "").lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from openpyxl import load_workbook

# Reusa o link de busca do scanner DBZ (linhas sem produto tcgcsv casado).
# Import tolerante (mesmo padrão do myp_summary.py): sem o módulo, o link de
# busca some mas a tabela segue.
try:
    from myp_dbz_scanner import tcg_dbz_search_url
except Exception:  # pragma: no cover - fallback defensivo
    def tcg_dbz_search_url(name, code=None):
        return None


def md_cell(s) -> str:
    """Escapa `|` e quebras de linha (célula de tabela markdown)."""
    if s is None:
        return ""
    return str(s).replace("\r", " ").replace("\n", " ").replace("|", "\\|")


def fmt_brl(v) -> str:
    if v is None or v == "":
        return "—"
    try:
        return f"R${float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "—"


def fmt_usd(v) -> str:
    if v is None or v == "":
        return "—"
    try:
        return f"US${float(v):,.2f}"
    except (ValueError, TypeError):
        return "—"


def fmt_pct(v) -> str:
    if v is None or v == "":
        return "—"
    try:
        return f"{float(v) * 100:.1f}%"
    except (ValueError, TypeError):
        return "—"


def carta_label(name, code) -> str:
    """Coluna `Carta` = nome + código numa string só ('Son Goku FB01-001').
    Não duplica quando o código já está no nome (padrão da frota)."""
    n = str(name or "").strip()
    c = str(code or "").strip()
    if c and c.lower() not in n.lower():
        return f"{n} {c}".strip()
    return n


def delivery_links(myp_url, tcg_url, name=None, code=None) -> str:
    """`[oferta](MYP) · [TCG](tcgplayer)`. Sem tcg_url (linha sem ref), cai no
    link de BUSCA do TCGplayer — nunca inventa URL de produto."""
    parts = []
    if myp_url:
        parts.append(f"[oferta]({myp_url})")
    tcg = tcg_url or tcg_dbz_search_url(str(name or ""), str(code or "") or None)
    if tcg:
        parts.append(f"[TCG]({tcg})")
    return " · ".join(parts) if parts else "—"


def _sheet_records(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    out = []
    for r in rows[1:]:
        rec = dict(zip(headers, r))
        if rec.get("Card Name"):
            out.append(rec)
    return out


def row_flags(c: dict) -> list[str]:
    """Motivos de REVISAR de uma linha (na ordem de força do sinal)."""
    flags = []
    if c.get("⚠️ Possível Lixo"):
        flags.append("possível lixo (<50% da ref)")
    if c.get("⚠️ Ref Volátil"):
        low, mkt = c.get("TCG Low US$"), c.get("TCG US$")
        try:
            flags.append(f"ref volátil (menor anúncio {fmt_usd(low)} vs "
                         f"market {fmt_usd(mkt)})")
        except Exception:
            flags.append("ref volátil")
    if c.get("⚠️ 1 Seller"):
        flags.append("1 seller EN (possível mislabel de idioma)")
    return flags


def build_markdown(xlsx: str, output: str, run_id: str = "",
                   repo: str = "matheuscllm-lgtm/myp-arbitrage-scanner") -> int:
    xlsx_path = Path(xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: XLSX não encontrado: {xlsx_path}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    summary_data = {}
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                summary_data[str(row[0]).strip().rstrip(":")] = row[1]
    all_cards = _sheet_records(wb, "All EN Cards")
    semref = _sheet_records(wb, "Sem Ref TCG")
    wb.close()

    # threshold REAL do scan (mesma lição v5.14.1 do fluxo Pokémon: o piso do
    # summary tem de casar o do XLSX, nunca um default hardcoded divergente).
    _thr_raw = summary_data.get("Margin Threshold", "30%")
    try:
        deal_floor = float(str(_thr_raw).strip().rstrip("%")) / 100.0
    except (TypeError, ValueError):
        deal_floor = 0.30

    def margin(c):
        try:
            return float(c.get("Margin %") or 0)
        except (TypeError, ValueError):
            return 0.0

    deals = sorted((c for c in all_cards if margin(c) >= deal_floor),
                   key=margin, reverse=True)
    deals_clean = [c for c in deals if not row_flags(c)]
    deals_review = [c for c in deals if row_flags(c)]
    truncations = [c for c in all_cards if c.get("⚠️ EN Trunc")]

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    thr_label = f"{deal_floor * 100:.0f}%"

    lines = []
    lines.append("---")
    lines.append("tags: [tcg, scanner, myp, dbz, dragonball, arbitrage]")
    lines.append(f"date: {today}")
    lines.append(f"source: GH Actions run {run_id}" if run_id else "source: local scan")
    lines.append("---")
    lines.append("")
    lines.append(f"# MYP Scan DRAGON BALL — {today}")
    lines.append("")
    fx = summary_data.get("FX USD→BRL")
    lines.append(
        f"**Cards EN NM ≥ piso:** {len(all_cards) + len(semref)} "
        f"({len(all_cards)} com referência + {len(semref)} sem) | "
        f"**Deals (≥{thr_label}):** {len(deals)} | "
        f"**Limpos:** {len(deals_clean)} | **REVISAR:** {len(deals_review)} | "
        f"**Truncation:** {len(truncations)}")
    lines.append("")
    lines.append(
        f"Margem BRUTA base compra `(TCG R$ − MYP R$) ÷ MYP R$`, sem taxas · "
        f"Referência: TCGplayer market via tcgcsv.com (cats 80 Fusion World + "
        f"27 Masters) · Câmbio US$1 = R${float(fx):.4f} · Ofertas: menor EN NM "
        f"ao vivo no MYP (dbsfusion + dbsmasters)."
        if fx else
        "Margem BRUTA base compra; referência TCGplayer market via tcgcsv.com.")
    lines.append("")

    # ── honestidade de referência (espelho da "Cobertura de preço TCG real") ──
    total = len(all_cards) + len(semref)
    if total == 0:
        cov = "⚠️ **Nenhuma carta EN NM ≥ piso nesta run** — nada a reportar."
    elif not all_cards:
        cov = (f"🛑 **ZERO referência tcgcsv** — 0/{total} cartas casaram no "
               f"índice (tcgcsv indisponível ou drift no código de carta?). "
               f"Investigue antes de operar; as {len(semref)} linhas estão na "
               f"seção 'Sem referência'.")
    elif semref:
        cov = (f"⚠️ **{len(all_cards)}/{total} cartas com referência tcgcsv "
               f"REAL**; {len(semref)} sem referência (join determinístico "
               f"não casou — seção própria abaixo, nunca margem inventada).")
    else:
        cov = f"✅ **{len(all_cards)}/{total} cartas com referência tcgcsv REAL**."
    lines.append(f"**Cobertura de referência real:** {cov}")
    lines.append("")
    if run_id:
        lines.append(f"**Artifact XLSX:** [`myp-dbz-consolidated-{run_id}`]"
                     f"(https://github.com/{repo}/actions/runs/{run_id})")
        lines.append("")

    # ── 🟢 deals limpos ──
    lines.append(f"## 🟢 Top 50 deals limpos (referência real, sem flags)")
    lines.append("")
    if not deals_clean:
        lines.append("> Nenhum deal limpo nesta run.")
    else:
        lines.append("| # | Margem % | MYP R$ | TCG US$ | Dif | Carta | Set | Raridade | Cond | Qtd | Links |")
        lines.append("|---|---:|---:|---:|---:|---|---|---|---|---:|---|")
        for i, c in enumerate(deals_clean[:50], 1):
            carta = md_cell(carta_label(c.get("EN Name") or c.get("Card Name"),
                                        c.get("Código")))
            links = delivery_links(c.get("URL"), c.get("TCG URL"),
                                   c.get("EN Name"), c.get("Código"))
            lines.append(
                f"| {i} | **{fmt_pct(c.get('Margin %'))}** | "
                f"{fmt_brl(c.get('MYP EN NM (R$)'))} | "
                f"{reference_price(fmt_usd(c.get('TCG US$')), c.get('TCG URL'))} | {fmt_brl(c.get('Diff (R$)'))} | "
                f"{carta} | {md_cell(c.get('Edition'))} | "
                f"{md_cell(c.get('Rarity')) or '—'} | NM | "
                f"{c.get('NM Sellers') or 0} | {links} |")
    lines.append("")

    # ── 🚨 REVISAR ──
    if deals_review:
        lines.append("## 🚨 REVISAR — deals com flag (validar manualmente)")
        lines.append("")
        lines.append("> Margem ≥ threshold mas com sinal de cautela por linha: "
                     "**possível lixo** (oferta <50% da referência — anúncio "
                     "errado/scam/condição mislabelada), **referência volátil** "
                     "(market e menor anúncio TCG divergem >2× — preço fino) ou "
                     "**1 seller EN** (possível mislabel de idioma). A margem "
                     "NÃO muda — valide no link TCG antes de operar.")
        lines.append("")
        lines.append("| # | Margem % | MYP R$ | TCG US$ | Dif | Carta | Set | Raridade | Cond | Qtd | Flag | Links |")
        lines.append("|---|---:|---:|---:|---:|---|---|---|---|---:|---|---|")
        for i, c in enumerate(deals_review[:50], 1):
            carta = md_cell(carta_label(c.get("EN Name") or c.get("Card Name"),
                                        c.get("Código")))
            links = delivery_links(c.get("URL"), c.get("TCG URL"),
                                   c.get("EN Name"), c.get("Código"))
            flag = md_cell("; ".join(row_flags(c)))
            lines.append(
                f"| {i} | {fmt_pct(c.get('Margin %'))} | "
                f"{fmt_brl(c.get('MYP EN NM (R$)'))} | "
                f"{reference_price(fmt_usd(c.get('TCG US$')), c.get('TCG URL'))} | {fmt_brl(c.get('Diff (R$)'))} | "
                f"{carta} | {md_cell(c.get('Edition'))} | "
                f"{md_cell(c.get('Rarity')) or '—'} | NM | "
                f"{c.get('NM Sellers') or 0} | {flag} | {links} |")
        lines.append("")

    # ── ⚠️ sem referência ──
    if semref:
        lines.append("## ⚠️ Sem referência TCG (join não casou — validar manualmente)")
        lines.append("")
        lines.append("> Ofertas EN NM vivas ≥ piso cujo código não casou um "
                     "produto único no tcgcsv (sem código no título, número "
                     "fora do índice, sem market price ou ambíguo). **Nunca "
                     "inventamos preço**: sem referência = sem margem. O link "
                     "TCG é de BUSCA, pra conferência manual.")
        lines.append("")
        lines.append("| # | Carta | Set | MYP R$ | Qtd | Motivo | Links |")
        lines.append("|---|---|---|---:|---:|---|---|")
        for i, s in enumerate(semref[:50], 1):
            carta = md_cell(carta_label(s.get("EN Name") or s.get("Card Name"),
                                        s.get("Código")))
            links = delivery_links(s.get("URL"), None,
                                   s.get("EN Name") or s.get("Card Name"),
                                   s.get("Código"))
            lines.append(
                f"| {i} | {carta} | {md_cell(s.get('Edition'))} | "
                f"{fmt_brl(s.get('MYP EN NM (R$)'))} | "
                f"{s.get('NM Sellers') or 0} | {md_cell(s.get('Motivo'))} | "
                f"{links} |")
        if len(semref) > 50:
            lines.append("")
            lines.append(f"> …e mais {len(semref) - 50} linha(s) na aba "
                         f"`Sem Ref TCG` do XLSX.")
        lines.append("")

    # ── truncation (diagnóstico) ──
    if truncations:
        lines.append("## 🚨 EN truncation risk (preço pode estar superestimado)")
        lines.append("")
        lines.append("> Seller table bateu o cap sem EN visível / paginação "
                     "incompleta — pode existir EN NM mais barato não visto. "
                     "Validar na página do produto.")
        lines.append("")
        lines.append("| Carta | Set | MYP R$ reportado | Links |")
        lines.append("|---|---|---:|---|")
        for c in truncations[:50]:
            carta = md_cell(carta_label(c.get("EN Name") or c.get("Card Name"),
                                        c.get("Código")))
            links = delivery_links(c.get("URL"), c.get("TCG URL"),
                                   c.get("EN Name"), c.get("Código"))
            lines.append(f"| {carta} | {md_cell(c.get('Edition'))} | "
                         f"{fmt_brl(c.get('MYP EN NM (R$)'))} | {links} |")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append(f"*Gerado em {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} "
                 f"via `myp_dbz_summary.py` (single source: XLSX do scan DBZ). "
                 f"Sem recomendação de compra — decisão de capital é do operador.*")

    out_path = Path(output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"OK markdown summary: {out_path}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Gera o markdown de ENTREGA do scan MYP Dragon Ball")
    parser.add_argument("xlsx", help="XLSX do myp_dbz_scanner.py")
    parser.add_argument("-o", "--output", required=True, help=".md de saída")
    parser.add_argument("--run-id", default="", help="GH Actions run ID")
    parser.add_argument("--repo", default="matheuscllm-lgtm/myp-arbitrage-scanner")
    args = parser.parse_args()
    return build_markdown(xlsx=args.xlsx, output=args.output,
                          run_id=args.run_id, repo=args.repo)


if __name__ == "__main__":
    sys.exit(main())
