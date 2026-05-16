"""
v0.1 EV scanner — grade-weighted EV usando Pop Report como input.

Mudanças vs v0:
  - SAFE_AT_PSA9 (assumia P(≥9)=100%, errado) → STRONG_POSITIVE_EV / MARGINAL_POSITIVE_EV / NEGATIVE_EV usando E[realização] real
  - Pop input por carta (P_10, P_9, P_8, ..., P_le_6) — vazio = PEND_POP_DATA
  - Jolteon Star hardcoded como worked example
  - Adiciona colunas: P_10..P_le_6 (inputs), E_realiz, EV_BRL, EV_%, EV_%_anual_7m, verdict
"""
import re
import time
import json
import hashlib
import sys
from pathlib import Path
from urllib.parse import quote_plus

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

XLSX = r"C:\Users\mathe\myp-arbitrage-scanner\myp_arbitrage_20260515_weekly.xlsx"
CACHE_DIR = Path(r"C:\Users\mathe\AppData\Local\Temp\pc_cache")
CACHE_DIR.mkdir(exist_ok=True)

# ─── Config (Codex-aligned) ───────────────────────────────────────────────────
CONFIG = dict(
    psa_fee_usd=24.99,
    membership_amort_usd=1.00,
    intermediario_brl=80.00,
    courier_usd_per_batch=160.00,
    batch_size=20,
    insurance_usd=2.00,
    declared_value_buffer_pct=0.05,
    sell_fee_pct=0.20,
    fx_brl_per_usd=5.05,         # FALLBACK only — overridden by fetch_fx_rate() at startup
    fx_source="fallback",         # populated by fetch_fx_rate
    fx_timestamp="hardcoded",
    fail_grade_haircut=0.50,    # <=PSA6 → liquidate raw at 50% (slab destroys raw market)
    holding_months=7,            # PSA Bulk Value turnaround
    discount_annual=0.13,        # CDB ~13% a.a. = oportunidade
)

# ─── Live FX fetch ────────────────────────────────────────────────────────────
def fetch_fx_rate():
    """Get USD/BRL bid from AwesomeAPI (preferred) or open.er-api (fallback).
    Returns (rate, source, timestamp) or (None, None, None) on total failure."""
    import urllib.request, json
    # Try AwesomeAPI first (gives bid/ask)
    try:
        with urllib.request.urlopen(
            "https://economia.awesomeapi.com.br/json/last/USD-BRL", timeout=8
        ) as resp:
            d = json.loads(resp.read())
        v = d["USDBRL"]
        bid = float(v["bid"])
        ts = v.get("create_date", "")
        return bid, "awesomeapi-bid", ts
    except Exception as e:
        print(f"  AwesomeAPI failed ({e}); trying open.er-api...", file=sys.stderr)
    # Fallback: open.er-api (mid rate only)
    try:
        with urllib.request.urlopen(
            "https://open.er-api.com/v6/latest/USD", timeout=8
        ) as resp:
            d = json.loads(resp.read())
        mid = float(d["rates"]["BRL"])
        ts = d.get("time_last_update_utc", "")
        return mid, "open-er-api-mid", ts
    except Exception as e:
        print(f"  open.er-api failed ({e}); using hardcoded fallback", file=sys.stderr)
        return None, None, None

# Resolve FX on module load
_fx_rate, _fx_src, _fx_ts = fetch_fx_rate()
if _fx_rate is not None:
    CONFIG["fx_brl_per_usd"] = _fx_rate
    CONFIG["fx_source"] = _fx_src
    CONFIG["fx_timestamp"] = _fx_ts
    print(f"[FX] R$ {_fx_rate:.4f}/USD  source={_fx_src}  ts={_fx_ts}")
else:
    print(f"[FX] WARNING: live fetch failed, using fallback R$ {CONFIG['fx_brl_per_usd']:.2f}/USD")

# Pop data hardcoded for cards we've manually verified.
# Operator can extend by reading PSA Pop online and adding entries.
KNOWN_POP = {
    # key = (card_name_normalized, number, set_hint)
    ("jolteon star", "101", "power keepers"): {
        "10": 92, "9": 450, "8": 433, "7": 339, "6": 330, "5": 303,
        "4": 220, "3": 129, "2": 93, "1.5": 6, "1": 171,
    },
}

def norm_key(name, num, set_hint):
    return (name.lower().strip(), str(num) if num else "", (set_hint or "").lower().strip())

# ─── PriceCharting (reuse from v0) ────────────────────────────────────────────
session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})

def cache_key(query):
    return hashlib.md5(query.encode("utf-8")).hexdigest()

def pc_search(query):
    ck = cache_key(query)
    cf = CACHE_DIR / f"{ck}.html"; uf = CACHE_DIR / f"{ck}.url"
    if cf.exists() and uf.exists():
        return uf.read_text(encoding="utf-8"), cf.read_text(encoding="utf-8", errors="ignore")
    url = f"https://www.pricecharting.com/search-products?q={quote_plus(query)}&type=prices"
    try:
        r = session.get(url, timeout=20, allow_redirects=True)
        time.sleep(1.0)
        if r.status_code != 200: return None, None
        cf.write_text(r.text, encoding="utf-8"); uf.write_text(r.url, encoding="utf-8")
        return r.url, r.text
    except Exception:
        return None, None

def follow_list_top(html):
    """If response is a list page, follow the 1st /game/ link and return that page."""
    soup = BeautifulSoup(html, "lxml")
    if soup.find(id="price_data"):
        return None, html  # already a product page
    first = soup.select_one('table a[href^="/game/"]')
    if not first:
        first = soup.select_one('a[href^="/game/"]')
    if not first:
        return None, None
    href = first.get("href")
    full = "https://www.pricecharting.com" + href
    ck = cache_key(full)
    cf = CACHE_DIR / f"{ck}.html"
    if cf.exists():
        return full, cf.read_text(encoding="utf-8", errors="ignore")
    try:
        r = session.get(full, timeout=20)
        time.sleep(1.0)
        if r.status_code != 200: return None, None
        cf.write_text(r.text, encoding="utf-8")
        return full, r.text
    except Exception:
        return None, None

def extract_prices(html):
    soup = BeautifulSoup(html, "lxml")
    table = soup.find(id="price_data")
    if not table: return {}
    headers = [th.get_text(strip=True) for th in table.find_all("th")]
    tbody = table.find("tbody")
    if not tbody: return {}
    first_row = tbody.find("tr")
    if not first_row: return {}
    cells = first_row.find_all("td")
    out = {}
    for label, cell in zip(headers, cells):
        span = cell.find("span", class_=re.compile(r"\bprice\b")) or cell.find("span", class_="js-price")
        if not span: continue
        m = re.search(r"\$([\d,]+\.\d{2})", span.get_text(strip=True))
        if not m: continue
        price = float(m.group(1).replace(",", ""))
        L = label.lower()
        if "ungraded" in L: out["raw"] = price
        elif "psa 10" in L: out["psa10"] = price
        elif "grade 9.5" in L: out["psa9_5"] = price
        elif "grade 9" in L: out["psa9"] = price
        elif "grade 8" in L: out["psa8"] = price
        elif "grade 7" in L: out["psa7"] = price
    return out

def lookup_pricecharting(card_name, num, set_hint):
    """Try multiple query strategies, fallback to list-page follow."""
    clean = re.sub(r"\([^)]*\)", "", str(card_name)).strip()
    queries = []
    if num:
        queries.append(f"{clean} {num}")  # simplest: name + num
        if set_hint:
            queries.append(f"{clean} {num} {set_hint}")
    queries.append(clean)

    for q in queries:
        url, html = pc_search(q)
        if not html: continue
        if BeautifulSoup(html, "lxml").find(id="price_data"):
            return url, extract_prices(html), q, "DIRECT"
        # follow list 1st
        url2, html2 = follow_list_top(html)
        if html2:
            prices = extract_prices(html2)
            if prices:
                return url2, prices, q, "LIST_FOLLOWED"
    return None, {}, queries[-1] if queries else "", "NO_RESULT"

# ─── EV math: grade-weighted ──────────────────────────────────────────────────
def compute_submission_cost_brl(declared_value_brl, cfg=CONFIG):
    fx = cfg["fx_brl_per_usd"]
    psa = cfg["psa_fee_usd"] * fx
    memb = cfg["membership_amort_usd"] * fx
    interm = cfg["intermediario_brl"]
    courier = (cfg["courier_usd_per_batch"] / cfg["batch_size"]) * fx
    ins = cfg["insurance_usd"] * fx
    tax_buf = declared_value_brl * cfg["declared_value_buffer_pct"]
    return psa + memb + interm + courier + ins + tax_buf

def net_realization_brl(prices_usd, cfg=CONFIG):
    """Per-grade net BRL after sell-side fees + FX. Returns dict by grade label."""
    fx = cfg["fx_brl_per_usd"]
    fee = 1 - cfg["sell_fee_pct"]
    out = {}
    raw = prices_usd.get("raw")
    fail_value = (raw * fx * fee * cfg["fail_grade_haircut"]) if raw else None

    for grade_label, key in [("10", "psa10"), ("9.5", "psa9_5"), ("9", "psa9"),
                              ("8", "psa8"), ("7", "psa7")]:
        usd = prices_usd.get(key)
        if usd:
            out[grade_label] = round(usd * fx * fee, 2)
    # Aggregate ≤6 as fail_value (haircut)
    if fail_value is not None:
        out["le_6"] = round(fail_value, 2)
    return out

def grade_weighted_ev(myp_brl, prices_usd, pop_dist, cfg=CONFIG):
    """
    pop_dist: dict like {"10":92, "9":450, "8":433, "7":339, "6":330, "5":303,
                         "4":220, "3":129, "2":93, "1.5":6, "1":171}
    Returns dict with ev_brl, ev_pct, ev_pct_anual, e_realiz, verdict, breakdown.
    """
    buy_cost = myp_brl * 1.06
    submission = compute_submission_cost_brl(buy_cost, cfg)
    total_invested = buy_cost + submission

    nets = net_realization_brl(prices_usd, cfg)

    # Normalize pop into probabilities
    if not pop_dist:
        return {
            "verdict": "PEND_POP_DATA",
            "buy_cost_brl": round(buy_cost, 2),
            "submission_cost_brl": round(submission, 2),
            "total_invested_brl": round(total_invested, 2),
            "nets": nets,
        }
    total_pop = sum(pop_dist.values())
    if total_pop == 0:
        return {"verdict": "ZERO_POP", "total_invested_brl": round(total_invested, 2), "nets": nets}

    probs = {g: c / total_pop for g, c in pop_dist.items()}

    # Compute E[realiz]: sum of P(g) × net(g)
    # Buckets: {"10","9.5","9","8","7","le_6"}
    breakdown = {}
    e_realiz = 0.0
    p_psa10_contrib = 0.0
    for g, p in probs.items():
        if g == "10":
            net = nets.get("10", nets.get("le_6", 0))
            if g == "10": p_psa10_contrib = p * net
        elif g in ("9", "9.5"):
            net = nets.get(g, nets.get("9", nets.get("le_6", 0)))
        elif g == "8":
            net = nets.get("8", nets.get("le_6", 0))
        elif g == "7":
            net = nets.get("7", nets.get("le_6", 0))
        else:  # 6, 5, 4, 3, 2, 1.5, 1
            net = nets.get("le_6", 0)
        contrib = p * net
        breakdown[g] = {"p": round(p, 4), "net": round(net, 2), "contrib": round(contrib, 2)}
        e_realiz += contrib

    ev_brl = e_realiz - total_invested
    ev_pct = ev_brl / total_invested if total_invested else 0
    months = cfg["holding_months"]
    discount = cfg["discount_annual"] * (months / 12)
    ev_pct_anual = ((1 + ev_pct) ** (12 / months) - 1) if ev_pct > -1 else None
    # Opportunity cost: subtract risk-free return over holding period
    ev_pct_after_oppcost = ev_pct - discount

    # Verdict
    fat_tail_dependency = (p_psa10_contrib / e_realiz) if e_realiz > 0 else 0
    if ev_pct < 0:
        verdict = "NEGATIVE_EV"
    elif fat_tail_dependency > 0.50:
        verdict = "NEEDS_FAT_TAIL"  # >50% do EV vem do PSA 10
    elif ev_pct_after_oppcost >= 0.30:
        verdict = "STRONG_POSITIVE_EV"
    elif ev_pct_after_oppcost >= 0.05:
        verdict = "MARGINAL_POSITIVE_EV"
    else:
        verdict = "LOW_EV_AFTER_OPPCOST"

    return {
        "verdict": verdict,
        "buy_cost_brl": round(buy_cost, 2),
        "submission_cost_brl": round(submission, 2),
        "total_invested_brl": round(total_invested, 2),
        "e_realiz_brl": round(e_realiz, 2),
        "ev_brl": round(ev_brl, 2),
        "ev_pct": round(ev_pct, 4),
        "ev_pct_anual": round(ev_pct_anual, 4) if ev_pct_anual else None,
        "ev_pct_after_oppcost": round(ev_pct_after_oppcost, 4),
        "fat_tail_dependency": round(fat_tail_dependency, 3),
        "breakdown": breakdown,
        "nets": nets,
        "total_pop": total_pop,
    }

# ─── Helpers ──────────────────────────────────────────────────────────────────
def extract_card_number(name):
    m = re.search(r"\((\d+)/\d+\)", str(name))
    return m.group(1) if m else None

def clean_card_name(name):
    s = str(name).strip()
    s = re.sub(r"\([^)]*\)", "", s)
    return s.strip()[:50]

EN_KEYWORDS = [
    "Sun & Moon", "Sword & Shield", "Scarlet & Violet", "Black & White",
    "Diamond & Pearl", "Heart Gold", "Promos", "Ancient Origins",
    "Delta Species", "Dragon Frontiers", "Power Keepers", "Aquapolis",
    "Skyridge", "Expedition", "Sandstorm", "Phantom Forces",
    "Primal Clash", "Roaring Skies", "Crown Zenith", "Paldea Evolved",
    "Obsidian Flames", "151", "Paradox Rift", "Paldean Fates",
    "Temporal Forces", "Twilight Masquerade", "Shrouded Fable",
    "Stellar Crown", "Surging Sparks", "Prismatic Evolutions",
    "Journey Together", "Destined Rivals", "Black Bolt", "White Flare",
    "Ascended Heroes", "Mega Evolution", "Arceus", "Platinum",
    "Neo Revelation", "Base Set", "Unseen Forces", "Dark Explorers",
    "Plasma Freeze", "Plasma Storm", "Boundaries Crossed",
    "Fusion Strike", "Lost Origin", "Silver Tempest",
]

def extract_en_set_hint(s):
    if not s: return ""
    s = str(s)
    for kw in EN_KEYWORDS:
        if kw.lower() in s.lower():
            return kw
    if ":" in s: return s.split(":")[-1].strip()
    return s[:30]

# ─── Main pipeline ────────────────────────────────────────────────────────────
def run():
    wb = load_workbook(XLSX)
    src = wb["All EN Cards"]
    rows = list(src.iter_rows(values_only=True))
    hdr = rows[0]; data = rows[1:]
    I = dict(name=0, edit=1, rar=2, myp=3, tcg=4, marg=5, diff=6, sellers=7, trunc=8, url=9)

    clean_rar = {"Rara", "Incomum", "Rara Hiper", "Rara Secreta"}
    candidates = [r for r in data
                  if isinstance(r[I["marg"]], (int, float)) and r[I["marg"]] >= 0.25
                  and str(r[I["rar"]] or "") in clean_rar
                  and not r[I["trunc"]]]
    candidates.sort(key=lambda r: r[I["marg"]], reverse=True)
    candidates = candidates[:40]

    print(f"Processing {len(candidates)} candidates...")
    results = []
    for i, r in enumerate(candidates, 1):
        name = r[I["name"]]; edit = r[I["edit"]]; myp = r[I["myp"]]; marg = r[I["marg"]]
        clean = clean_card_name(name)
        num = extract_card_number(name)
        set_hint = extract_en_set_hint(edit)
        print(f"[{i:2d}/{len(candidates)}] {clean[:35]:35s} ({num or '?'})  margem={marg:.0%}")

        url, prices, q_used, match_type = lookup_pricecharting(clean, num, set_hint)
        print(f"        {match_type}  query={q_used!r}  prices={prices}")

        # Look up pop
        pop_key = norm_key(clean, num, set_hint)
        pop = KNOWN_POP.get(pop_key)
        # also try with just card name
        if not pop:
            for k, v in KNOWN_POP.items():
                if k[0] in clean.lower() and k[1] == (num or ""):
                    pop = v
                    break

        if not prices:
            ev = {"verdict": "PEND_PRICE_DATA"}
        else:
            ev = grade_weighted_ev(myp, prices, pop)
        ev["pc_url"] = url
        ev["prices_usd"] = prices
        ev["query_used"] = q_used
        ev["match_type"] = match_type
        ev["pop"] = pop
        ev["row"] = r
        results.append(ev)

    # Build output sheet
    if "💰 EV v0.1" in wb.sheetnames:
        del wb["💰 EV v0.1"]
    out = wb.create_sheet("💰 EV v0.1", index=0)  # primeira sheet

    fx = CONFIG["fx_brl_per_usd"]
    out_hdr = [
        "Card Name", "Edition", "Rarity", "MYP R$",
        "Raw R$ (gross)", "PSA 9 R$ (gross)", "PSA 10 R$ (gross)",
        "Pop 10", "Pop 9", "Pop 8", "Pop 7", "Pop ≤6", "Pop total",
        "Buy cost R$", "Submission R$", "Invest total R$",
        "E[realiz] R$", "EV R$", "EV %", "EV % anual (7m)",
        "P(10) contrib %", "Verdict", "Match", "PriceCharting"
    ]
    out.append(out_hdr)

    link_font = Font(color="0563C1", underline="single")
    verdict_fills = {
        "STRONG_POSITIVE_EV": PatternFill("solid", fgColor="63BE7B"),
        "MARGINAL_POSITIVE_EV": PatternFill("solid", fgColor="C6EFCE"),
        "NEEDS_FAT_TAIL": PatternFill("solid", fgColor="FFEB9C"),
        "LOW_EV_AFTER_OPPCOST": PatternFill("solid", fgColor="FFE699"),
        "NEGATIVE_EV": PatternFill("solid", fgColor="F4B084"),
        "PEND_POP_DATA": PatternFill("solid", fgColor="DDEBF7"),
        "PEND_PRICE_DATA": PatternFill("solid", fgColor="EDEDED"),
    }

    # Sort: positive first, by EV % desc; pending last
    def sort_key(res):
        v = res.get("verdict")
        ev_pct = res.get("ev_pct") or -99
        v_order = {"STRONG_POSITIVE_EV": 0, "MARGINAL_POSITIVE_EV": 1,
                   "NEEDS_FAT_TAIL": 2, "LOW_EV_AFTER_OPPCOST": 3,
                   "PEND_POP_DATA": 4, "NEGATIVE_EV": 5, "PEND_PRICE_DATA": 6}
        return (v_order.get(v, 9), -ev_pct)
    results.sort(key=sort_key)

    for res in results:
        r = res["row"]
        prices = res.get("prices_usd", {})
        pop = res.get("pop") or {}
        breakdown = res.get("breakdown", {})
        # Convert USD reference prices to BRL gross using live FX
        raw_brl = round(prices["raw"] * fx, 2) if prices.get("raw") else None
        psa9_brl = round(prices["psa9"] * fx, 2) if prices.get("psa9") else None
        psa10_brl = round(prices["psa10"] * fx, 2) if prices.get("psa10") else None
        row_data = [
            r[I["name"]], r[I["edit"]], r[I["rar"]], r[I["myp"]],
            raw_brl, psa9_brl, psa10_brl,
            pop.get("10"), pop.get("9"), pop.get("8"), pop.get("7"),
            sum(int(pop.get(g, 0)) for g in ("6","5","4","3","2","1.5","1")) if pop else None,
            res.get("total_pop"),
            res.get("buy_cost_brl"), res.get("submission_cost_brl"),
            res.get("total_invested_brl"),
            res.get("e_realiz_brl"), res.get("ev_brl"),
            res.get("ev_pct"), res.get("ev_pct_anual"),
            (breakdown.get("10", {}).get("contrib", 0) / res["e_realiz_brl"] if res.get("e_realiz_brl") else None),
            res.get("verdict"),
            res.get("match_type"),
            res.get("pc_url"),
        ]
        out.append(row_data)
        ri = out.max_row
        # hyperlinks
        if r[I["url"]]:
            c = out.cell(row=ri, column=1); c.hyperlink = r[I["url"]]; c.font = link_font
        if res.get("pc_url"):
            c = out.cell(row=ri, column=len(out_hdr)); c.hyperlink = res["pc_url"]
            c.font = link_font; c.value = "PriceCharting →"
        # verdict fill
        v = res.get("verdict")
        if v in verdict_fills:
            out.cell(row=ri, column=22).fill = verdict_fills[v]
        # percent formats
        out.cell(row=ri, column=19).number_format = "0.0%"
        out.cell(row=ri, column=20).number_format = "0.0%"
        out.cell(row=ri, column=21).number_format = "0.0%"
        # BRL formats: cols 4 (MYP), 5-7 (Raw/PSA gross), 14-18 (buy/sub/inv/realiz/EV)
        for col_idx in (4, 5, 6, 7, 14, 15, 16, 17, 18):
            out.cell(row=ri, column=col_idx).number_format = '"R$ "#,##0.00'

    # widths
    widths = [30, 24, 12, 8, 9, 9, 9, 7, 7, 7, 7, 7, 8, 11, 12, 13, 13, 10, 9, 11, 12, 18, 14, 16]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        out.column_dimensions[get_column_letter(i)].width = w

    # bold header
    bold = Font(bold=True)
    for c in out[1]: c.font = bold
    out.freeze_panes = "B2"

    # Notes block — clean ALL stale notes sheets first
    for sn in list(wb.sheetnames):
        if sn.startswith("📝 EV Notes") or sn == "💰 EV Analysis":
            del wb[sn]
    notes = wb.create_sheet("📝 EV Notes")
    nr = [
        ["v0.1 grade-weighted EV — notes & methodology"],
        [""],
        ["FÓRMULA:"],
        ["  E[realiz] = Σ P(grade) × net_realiz(grade)"],
        ["  net_realiz(grade) = price_USD × FX × (1 - sell_fee_pct)"],
        ["  fail (≤PSA6): raw × FX × (1 - sell_fee) × fail_grade_haircut"],
        ["  EV = E[realiz] - (buy_cost + submission_cost)"],
        [""],
        ["CONFIG (Codex-aligned):"],
        [f"  PSA fee tier: Bulk Value ${CONFIG['psa_fee_usd']}/card"],
        [f"  Intermediário BR: R${CONFIG['intermediario_brl']}/card"],
        [f"  Courier per card (amortizado): ${CONFIG['courier_usd_per_batch']/CONFIG['batch_size']:.2f}"],
        [f"  Sell-side haircut: {CONFIG['sell_fee_pct']:.0%} (eBay+payment+FX+ship)"],
        [f"  FX: R$ {CONFIG['fx_brl_per_usd']:.4f}/USD  (source: {CONFIG.get('fx_source','?')}; ts: {CONFIG.get('fx_timestamp','?')})"],
        [f"  Fail-grade haircut: {CONFIG['fail_grade_haircut']:.0%} of raw"],
        [f"  Holding: {CONFIG['holding_months']} meses"],
        [f"  Custo oportunidade anual: {CONFIG['discount_annual']:.0%} (CDB ballpark)"],
        [""],
        ["VERDICTS:"],
        ["  STRONG_POSITIVE_EV     EV pct after opp cost ≥ 30%"],
        ["  MARGINAL_POSITIVE_EV   EV pct after opp cost ≥ 5%"],
        ["  NEEDS_FAT_TAIL         >50% do EV vem de PSA 10 (3-15% prob) — risco binário"],
        ["  LOW_EV_AFTER_OPPCOST   positivo mas < CDB"],
        ["  NEGATIVE_EV            perde dinheiro em expectativa"],
        ["  PEND_POP_DATA          preço OK, falta Pop Report (preencha Pop 10/9/8/7/≤6)"],
        ["  PEND_PRICE_DATA        PriceCharting não tem dados"],
        [""],
        ["COMO PREENCHER POP MANUALMENTE:"],
        ["  1. Abre psacard.com/pop/search"],
        ["  2. Procura por '<card_name> <number>'"],
        ["  3. Clica em 'Show Population' no result"],
        ["  4. Copia counts pras colunas Pop 10/9/8/7 (≤6 = soma de 6,5,4,3,2,1.5,1)"],
        ["  5. Re-roda o script (XLSX recalcula)"],
        [""],
        ["LIMITAÇÕES CONHECIDAS:"],
        ["  - PriceCharting search falha ~40% sem set canonical layer (Codex critique #4)"],
        ["  - Sell-side haircut 20% é estimativa; spread bid/ask real desconhecido"],
        ["  - Fail haircut 50% sobre raw é conservador; PSA 1-2 podem vender pior ainda"],
        ["  - Turnaround 7 meses fixo (PSA Bulk pode variar 140-200 dias)"],
        ["  - Pop ratio de submissões ≠ Pop ratio que sua carta específica vai gradar"],
        ["    (viés de seleção: pessoas pré-screenam antes de submeter)"],
    ]
    for row in nr:
        notes.append(row)
    notes.column_dimensions['A'].width = 80
    notes.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Summary
    print("\n" + "=" * 80 + "\nRESUMO\n" + "=" * 80)
    from collections import Counter
    vc = Counter(res.get("verdict") for res in results)
    order = ["STRONG_POSITIVE_EV","MARGINAL_POSITIVE_EV","NEEDS_FAT_TAIL",
             "LOW_EV_AFTER_OPPCOST","PEND_POP_DATA","NEGATIVE_EV","PEND_PRICE_DATA"]
    for v in order:
        if v in vc: print(f"  {v:25s}  {vc[v]}")

    wb.save(XLSX)
    print(f"\nSaved → {XLSX}")
    print(f"Sheets: '💰 EV v0.1' + '📝 EV Notes'")

if __name__ == "__main__":
    run()
