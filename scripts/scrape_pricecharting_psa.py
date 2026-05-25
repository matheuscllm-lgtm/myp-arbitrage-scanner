"""PSA price triage via PriceCharting scrape pros deals do MYP scanner.

Lê o xlsx do scanner, extrai cartas do sheet `🔥 Deals` + opcionalmente
`🚨 Validate Manually`, faz lookup em pricecharting.com via cloudscraper
(bypassa anti-bot Cloudflare), parseia grade matrix (Ungraded / PSA 8 / 9 /
10) + recent eBay sales por grade, e gera markdown com tabela de arbitragem
MYP→PSA 9 + PSA 10 jackpot.

Uso:
    python3 scripts/scrape_pricecharting_psa.py results/scan-XXXX.xlsx
    python3 scripts/scrape_pricecharting_psa.py results/scan-XXXX.xlsx --include-validate
    python3 scripts/scrape_pricecharting_psa.py results/scan-XXXX.xlsx -o results/psa-prices-YYYY-MM-DD.md
    python3 scripts/scrape_pricecharting_psa.py results/scan-XXXX.xlsx --fx 5.30

Histórico: criado 2026-05-25 após validar que cloudscraper passa o anti-bot
do PriceCharting (mesmo firefox/windows fingerprint do MYP scanner). Workflow
canônico pós-scan = MYP xlsx → este script → relatório PSA pra decidir quais
deals justificam grading.

Caveats nos preços:
- Médias do PriceCharting (info_box) podem estar voláteis em sets novos
  (ex.: ME: Ascended Heroes 2026 — pop pequena, flips iniciais inflam).
- Nenhum custo incluído: grading (~$25-40), shipping BR↔US (~$30-50),
  eBay fees (~13%), tax. Net realistic = bruto × ~0.6-0.7.
- P(PSA 10) NÃO estimado aqui. Pra decisão final, alimentar pop counts
  no PSA-Arbitrage-Scanner (`psa-arb analyze-live`).
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Optional
from urllib.parse import quote

import cloudscraper
from bs4 import BeautifulSoup
from openpyxl import load_workbook


PC_BASE = "https://www.pricecharting.com"

# IDs dos TD que o PriceCharting usa pra renderizar a grade matrix. Espelha
# constantes em psa_arb.adapters.fetchers.price_charting (PSA-Arbitrage-Scanner).
TD_MAP = {
    "ungraded": "used_price",
    "psa8": "new_price",
    "psa9": "graded_price",
    "psa10": "manual_only_price",
}

GRADE_RE = re.compile(r"\bPSA\s*(?:Grade\s*)?(\d+(?:\.\d)?)\b", re.IGNORECASE)
OTHER_GRADER_RE = re.compile(r"\b(BGS|CGC|SGC|TAG)\b", re.IGNORECASE)
MONEY_RE = re.compile(r"\$([\d,]+\.?\d*)")

# Tradução PT→EN comum em nomes do MYP. O MYP frequentemente concatena
# `<PT name> (NNN/MMM)<EN name>` — quando o EN já vem no nome, usar ele.
# Esta tabela é fallback pros raros casos onde só vem PT.
PT_EN_MAP = [
    (r"da Equipe Rocket", "Team Rocket"),
    (r"do N\b", "N's"),
    (r"da Érica", "Erika's"),
    (r"da Marine", "Marnie's"),
    (r"da Lílian", "Lillie's"),
    (r"da Misty", "Misty's"),
    (r"da Kissera", "Iono's"),
    (r"do Steven", "Steven's"),
    (r"da Iris", "Iris's"),
    (r"\bdo\b", "the"),  # last-resort
]

# Mapping conhecido de edição MYP (PT) → keyword EN pra search PriceCharting.
# Lista crescente conforme descobrimos novos sets. Sets sem mapping caem no
# fallback `<card_name> <number>` (PriceCharting search é fuzzy o bastante
# pra geralmente acertar mesmo sem nome do set).
EDITION_KEYWORDS = [
    ("ascended heroes", "Ascended Heroes"),
    ("black bolt", "Black Bolt"),
    ("white flare", "White Flare"),
    ("destined rivals", "Destined Rivals"),
    ("rivais predestinados", "Destined Rivals"),
    ("journey together", "Journey Together"),
    ("amigos de jornada", "Journey Together"),
    ("prismatic evolutions", "Prismatic Evolutions"),
    ("evoluções prismáticas", "Prismatic Evolutions"),
    ("stellar crown", "Stellar Crown"),
    ("coroa estelar", "Stellar Crown"),
    ("surging sparks", "Surging Sparks"),
    ("fagulhas impetuosas", "Surging Sparks"),
    ("equilíbrio perfeito", "Perfect Order"),
    ("perfect order", "Perfect Order"),
    ("fogo fantasmagórico", "Phantasmal Flames"),
    ("phantasmal flames", "Phantasmal Flames"),
    ("series black star", "Mega Promo"),
    ("mega evolution promo", "Mega Promo"),
]


def parse_money(text: Optional[str]) -> Optional[float]:
    if not text:
        return None
    m = MONEY_RE.search(text)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except (ValueError, TypeError):
        return None


def extract_name_and_number(card_name: str) -> tuple[str, str]:
    """Extrai nome EN-friendly + número da carta do `Card Name` do MYP xlsx.

    MYP frequentemente formata como `<PT> (NNN/MMM)<EN>`; usamos EN quando
    presente, senão aplicamos PT_EN_MAP no PT.
    """
    m = re.search(r"^(.+?)\s*\((\d+)(?:/\d+)?\)(.*)$", card_name)
    if not m:
        # No collector number — fallback ao nome inteiro.
        return _pt_to_en(card_name.strip()), ""
    pt_name = m.group(1).strip()
    number = m.group(2)
    en_name = m.group(3).strip()
    if en_name:
        return en_name, number
    return _pt_to_en(pt_name), number


def _pt_to_en(name: str) -> str:
    out = name
    for pt, en in PT_EN_MAP:
        out = re.sub(pt, en, out)
    return out


def edition_keyword(edition: str) -> str:
    lo = edition.lower()
    for needle, kw in EDITION_KEYWORDS:
        if needle in lo:
            return kw
    return ""


class PriceChartingClient:
    def __init__(self, delay: float = 1.0):
        self.scraper = cloudscraper.create_scraper(
            browser={"browser": "firefox", "platform": "windows", "desktop": True}
        )
        self.delay = delay

    def _get(self, url: str):
        time.sleep(self.delay)
        return self.scraper.get(url, timeout=30)

    def search_for_slug(self, query: str, want_number: str) -> Optional[str]:
        """Retorna slug `/game/...` cujo tail `-NNN` casa com want_number.

        Sem fallback pro primeiro hit — Search ranking do PriceCharting
        prioriza popularidade, não exatidão. Sets novos onde múltiplas cartas
        compartilham keywords (ex.: Ascended Heroes) sempre caem no Mega
        Gengar 284 sem o filtro estrito.
        """
        url = f"{PC_BASE}/search-products?q={quote(query)}&type=prices"
        r = self._get(url)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "lxml")
        want = want_number.lstrip("0") or want_number
        for a in soup.select('a[href*="/game/"]'):
            href = a.get("href") or ""
            if "pokemon" not in href.lower():
                continue
            if href.startswith("http"):
                href = href.split("pricecharting.com", 1)[-1]
            m = re.search(r"-(\d+)$", href)
            if m and m.group(1).lstrip("0") == want:
                return href
        return None

    def fetch_card_grades(self, slug: str) -> tuple[dict[str, float], dict[int, list[float]]]:
        """Retorna ({grade_key: median_price_usd}, {psa_grade: [recent_sale_prices]})."""
        r = self._get(f"{PC_BASE}{slug}")
        if r.status_code != 200:
            return {}, {}
        soup = BeautifulSoup(r.text, "lxml")
        grades: dict[str, float] = {}
        for field, td_id in TD_MAP.items():
            td = soup.select_one(f"td#{td_id}")
            if not td:
                continue
            span = td.select_one("span.price.js-price") or td.select_one(".price")
            text = span.get_text(strip=True) if span else td.get_text(strip=True)
            price = parse_money(text)
            if price is not None:
                grades[field] = price
        sales: dict[int, list[float]] = {}
        for row in soup.select('tr[id^="ebay-"]'):
            title_node = row.select_one("td.title a")
            if not title_node:
                continue
            title = title_node.get_text(strip=True)
            if OTHER_GRADER_RE.search(title):
                continue
            m = GRADE_RE.search(title)
            if not m:
                continue
            try:
                grade_f = float(m.group(1))
            except ValueError:
                continue
            if grade_f != int(grade_f):
                continue
            grade = int(grade_f)
            if grade not in (8, 9, 10):
                continue
            price_node = row.select_one("td.numeric span.js-price") or row.select_one("td.numeric")
            money = parse_money(price_node.get_text(strip=True) if price_node else "")
            if money is None:
                continue
            sales.setdefault(grade, []).append(money)
        return grades, sales


def read_deals(xlsx_path: Path, include_validate: bool) -> list[dict]:
    """Lê 🔥 Deals (e opcionalmente 🚨 Validate Manually) do xlsx do MYP scanner."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = ["🔥 Deals"]
    if include_validate and "🚨 Validate Manually" in wb.sheetnames:
        sheets.append("🚨 Validate Manually")
    deals: list[dict] = []
    seen_urls: set[str] = set()
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h}
        required = ["Card Name", "Edition", "MYP EN NM (R$)", "URL"]
        if not all(k in idx for k in required):
            print(f"  ⚠️  sheet {sheet_name!r} missing columns; skipping")
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            url = row[idx["URL"]]
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            deals.append(
                {
                    "card_name": row[idx["Card Name"]] or "",
                    "edition": row[idx["Edition"]] or "",
                    "myp_brl": row[idx["MYP EN NM (R$)"]] or 0.0,
                    "url": url,
                    "source_sheet": sheet_name,
                }
            )
    return deals


def render_markdown(
    rows: list[dict], xlsx_path: Path, fx: float,
    grading_cost: float, myp_freight_brl: float = 0.0,
) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    myp_freight_usd = myp_freight_brl / fx if myp_freight_brl else 0.0
    freight_note = (
        f"Frete MYP→comprador assumido **R${myp_freight_brl:.0f}/carta** "
        f"(~US${myp_freight_usd:.2f}, ajustar via `--myp-freight-brl`). "
    ) if myp_freight_brl else ""
    lines: list[str] = [
        f"# PriceCharting PSA Prices — {today}",
        "",
        f"Spot-check de preços graded pros deals em `{xlsx_path.name}`. "
        f"Fonte: PriceCharting (cloudscraper bypassa anti-bot). "
        f"Médias = `info_box`; PSA9 eBay = mediana das vendas recentes filtradas por título `PSA 9`. "
        f"FX assumido **BRL/USD = {fx:.2f}** (ajustar via `--fx`). "
        f"Custo grading+freight US assumido **US${grading_cost:.0f}/carta** (ajustar via `--grading-cost`). "
        f"{freight_note}"
        f"**Outros custos não incluídos:** eBay fees (~13%), tax — net realistic líquido ≈ valor da tabela × 0.87.",
        "",
        "## Preços por grade (USD)",
        "",
        "| # | Carta | Slug PC | Ungraded | PSA 8 | PSA 9 | PSA 10 | PSA9 eBay (mediana / N) |",
        "|---|---|---|---:|---:|---:|---:|---:|",
    ]
    enriched = []
    for i, row in enumerate(rows, 1):
        grades = row["grades"]
        sales = row["sales"]
        psa9_sales = sales.get(9, [])
        psa9_med = median(psa9_sales) if psa9_sales else None
        slug = row["slug"] or "_not found_"
        slug_md = f"`{slug.replace('/game/', '')}`" if slug != "_not found_" else slug
        ug = grades.get("ungraded")
        p8 = grades.get("psa8")
        p9 = grades.get("psa9")
        p10 = grades.get("psa10")
        lines.append(
            f"| {i} | {row['card_name']} | {slug_md} | "
            f"{'$' + format(ug, '.2f') if ug else '—'} | "
            f"{'$' + format(p8, '.2f') if p8 else '—'} | "
            f"{'**$' + format(p9, '.2f') + '**' if p9 else '—'} | "
            f"{'$' + format(p10, '.2f') if p10 else '—'} | "
            f"{('$' + format(psa9_med, '.2f') if psa9_med else '—') + '/' + str(len(psa9_sales))} |"
        )
        if p9 and row["myp_brl"]:
            myp_usd = row["myp_brl"] / fx
            mult9 = p9 / myp_usd
            mult10 = (p10 / myp_usd) if p10 else None
            profit9 = p9 - myp_usd
            cost_base = myp_usd + myp_freight_usd + grading_cost
            diff9 = p9 - cost_base
            diff10 = (p10 - cost_base) if p10 else None
            enriched.append({
                **row,
                "p9": p9, "p10": p10,
                "myp_usd": myp_usd,
                "mult9": mult9, "mult10": mult10,
                "profit9": profit9,
                "diff9": diff9, "diff10": diff10,
                "cost_base": cost_base,
            })

    # Ranking de arbitragem (PSA 9 e PSA 10 lado a lado, líquido após grading)
    lines.extend(
        [
            "",
            f"## Arbitragem MYP → PSA (líquido após +US${grading_cost:.0f} grading/freight)",
            "",
            "Diff = PSA grade − (MYP USD + grading). Sort por **Diff PSA 9** (piso conservador). "
            "Diff PSA 10 é o jackpot upside se a carta entrar 10 em vez de 9.",
            "",
            "| Carta | MYP R$ | MYP US$ | +grading | PSA 9 US$ | Diff PSA 9 | PSA 10 US$ | Diff PSA 10 |",
            "|---|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )
    def _fmt_diff(v: Optional[float], bold: bool = False) -> str:
        if v is None:
            return "—"
        marker = "**" if bold else ""
        sign = "+" if v >= 0 else "-"
        return f"{marker}{sign}${abs(v):.2f}{marker}"

    for r in sorted(enriched, key=lambda x: -x["diff9"]):
        diff9_str = _fmt_diff(r["diff9"], bold=r["diff9"] > 0)
        diff10_str = _fmt_diff(r["diff10"], bold=bool(r["diff10"]) and r["diff10"] > 0)
        p10_str = f"${r['p10']:.2f}" if r["p10"] else "—"
        lines.append(
            f"| {r['card_name']} | R${r['myp_brl']:.2f} | "
            f"${r['myp_usd']:.2f} | ${r['cost_base']:.2f} | "
            f"${r['p9']:.2f} | {diff9_str} | "
            f"{p10_str} | {diff10_str} |"
        )

    # Top picks profit-positive: diff9 > 0
    profitable = [r for r in enriched if r["diff9"] > 0]
    lines.extend(
        [
            "",
            f"## Top picks net-positive (Diff PSA 9 > 0 com US${grading_cost:.0f} grading)",
            "",
        ]
    )
    if not profitable:
        lines.append("> Nenhum card cobre o custo de grading mesmo no piso PSA 9.")
    else:
        for r in sorted(profitable, key=lambda x: -x["diff9"])[:10]:
            p10_extra = f", PSA 10 jackpot +${r['diff10']:.2f}" if r["diff10"] and r["diff10"] > 0 else ""
            lines.append(
                f"- **{r['card_name']}** — MYP R${r['myp_brl']:.2f} → "
                f"PSA 9 líquido **+${r['diff9']:.2f}**{p10_extra}"
            )

    lines.extend(
        [
            "",
            "---",
            "",
            f"*Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC via `scripts/scrape_pricecharting_psa.py`. "
            f"Próximo passo: rodar `psa-arb analyze-live` nos top picks com pop counts manuais "
            f"(psacard.com/pop) pra decisão final com P(PSA 10).*",
        ]
    )
    return "\n".join(lines) + "\n"


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="MYP scanner xlsx (results/scan-*.xlsx)")
    parser.add_argument(
        "-o", "--output", type=Path,
        help="Markdown de saída (default: results/psa-prices-<today>.md)"
    )
    parser.add_argument(
        "--fx", type=float, default=5.30,
        help="BRL/USD pra conversão (default 5.30)"
    )
    parser.add_argument(
        "--grading-cost", type=float, default=40.0,
        help=(
            "Custo de grading+freight USD por carta pra coluna Diff (default 40). "
            "Piso conservador BR→US: ~$25 PSA Value + ~$15 shipping ida+volta. "
            "Realista pra Brasil ~$55-70 (PSA Regular + freight asseg.)."
        )
    )
    parser.add_argument(
        "--myp-freight-brl", type=float, default=0.0,
        help=(
            "Frete MYP→comprador BRL por carta (default 0). Cards no MYP "
            "frequentemente têm frete por listing (~R$15-30 normal, R$100+ "
            "em insured/SEDEX pra cartas caras). Conferido manualmente pela "
            "Psyduck 226/217 = R$100/carta — usar 100 pra cenário realista."
        )
    )
    parser.add_argument(
        "--include-validate", action="store_true",
        help="Inclui também 🚨 Validate Manually (cards com flag)"
    )
    parser.add_argument(
        "--delay", type=float, default=1.0,
        help="Segundos entre requests PriceCharting (default 1.0)"
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"❌ xlsx não encontrado: {args.xlsx}")

    out_path = args.output or (
        args.xlsx.parent / f"psa-prices-{datetime.now().strftime('%Y-%m-%d')}.md"
    )

    deals = read_deals(args.xlsx, args.include_validate)
    print(f"Lidos {len(deals)} deals únicos de {args.xlsx.name}")
    if not deals:
        sys.exit("❌ nenhum deal pra lookup. (sheets vazias ou colunas faltando)")

    client = PriceChartingClient(delay=args.delay)
    rows = []
    for i, deal in enumerate(deals, 1):
        en_name, number = extract_name_and_number(deal["card_name"])
        set_kw = edition_keyword(deal["edition"])
        query = f"{en_name} {set_kw}".strip()
        print(f"\n[{i}/{len(deals)}] {deal['card_name']}")
        print(f"  query={query!r} | want #{number}")
        slug = client.search_for_slug(query, number) if number else None
        if not slug:
            print(f"  ❌ slug not found")
            rows.append({**deal, "slug": None, "grades": {}, "sales": {}})
            continue
        print(f"  → slug={slug}")
        grades, sales = client.fetch_card_grades(slug)
        p9 = grades.get("psa9")
        n9 = len(sales.get(9, []))
        print(f"  PSA9=${p9} (eBay sales={n9}) | PSA10=${grades.get('psa10')}")
        rows.append({**deal, "slug": slug, "grades": grades, "sales": sales})

    md = render_markdown(rows, args.xlsx, args.fx, args.grading_cost, args.myp_freight_brl)
    out_path.write_text(md, encoding="utf-8")
    print(f"\n✅ Saved: {out_path}")
    matched = sum(1 for r in rows if r["slug"])
    print(f"   {matched}/{len(rows)} cards casadas no PriceCharting")


if __name__ == "__main__":
    main()
