"""Cross-check MYP HTML scanner output against MYP public REST API tcg_price.

v5.8.5 (2026-05-19) — Heurística #B do plano source-direct-validation.

Background
----------
MYP scanner v5.8 H2 (commit 0ca15d2) já detecta TCG declarado >>> última venda
real via sanity check `.estat-tcg` vs `.estatistica-ultimo`. Esse postprocess
adiciona uma SEGUNDA camada: compara o valor que o scanner extraiu de
`.estat-tcg` contra o `tcg_price` (USD) que a API pública do MYP retorna.

Caveat importante: HTML e API consultam a MESMA fonte upstream (TCGplayer).
Esse cross-check pega:
  - Bugs de parser HTML (units, locale, regex drift)
  - Snapshots fora de sincronia entre cache HTML e API
  - Drift de unidade (BRL × FX vs USD direto)
NÃO pega inflação upstream do MYP — pra isso, v5.8 H2 (sanity vs last_sale)
e v5.8.5 #A (oversized collector#) continuam sendo as defesas primárias.

Pipeline
--------
1. Carrega XLSX existente (output do scanner).
2. Pra cada card em '🔥 Deals' + '🚨 Validate Manually':
   a. Fetch HTML page (cloudscraper) pra extrair `product_code`
      (`pokemon_<ed>_<num>/<set>`) — scanner faz isso via regex no page_text.
   b. Base64-encode o code e chama `GET /api/v1/pokemon/precos/{enc}`.
   c. Lê `tcg_price` (USD). Multiplica por FX live (open.er-api.com) →
      `tcg_api_brl_estimate`.
   d. Compara contra `TCG Player (R$)` da row do XLSX.
   e. Se ratio > 2x → marca tcg_api_mismatch=True.
3. Adiciona nova sheet `❌ TCG API Mismatch` ao XLSX original (write back).

Restrições do plano:
- Rate limit ~5 req/s (delay 0.2s entre requests)
- Sem cache (operador pediu explicitamente)
- Se API falhar pra um card: skip silenciosamente (não bloqueia o pipeline)

Uso:
    python scripts/cross_check_myp_api.py <xlsx_path>
"""
from __future__ import annotations

import argparse
import base64
import logging
import re
import sys
import time
from pathlib import Path
from typing import Optional

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import cloudscraper

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger(__name__)

API_BASE = "https://mypcards.com/api/v1"
API_DELAY = 0.2   # ~5 req/s
HTML_DELAY = 1.5  # delay extra antes do fetch HTML inicial (mesma cadência do scanner)
TIMEOUT = 20
PRODUCT_CODE_RE = re.compile(r"pokemon_[a-z0-9-]{2,15}_[\w/.-]+")
MISMATCH_RATIO_THRESHOLD = 2.0
FX_API = "https://open.er-api.com/v6/latest/USD"
FX_FALLBACK = 5.05  # 2026-05-16 bid (memory fx_brl_usd_2026_05_16)


def fetch_fx_usd_brl(scraper) -> float:
    """Pull live FX USD→BRL. Fallback hardcoded se API down."""
    try:
        r = scraper.get(FX_API, timeout=10)
        if r.status_code == 200:
            j = r.json()
            rate = j.get("rates", {}).get("BRL")
            if rate and rate > 0:
                log.info(f"FX live USD→BRL = {rate:.4f}")
                return float(rate)
    except Exception as e:
        log.warning(f"FX fetch failed: {e}")
    log.warning(f"Using FX fallback R${FX_FALLBACK}/USD (open.er-api unreachable)")
    return FX_FALLBACK


def extract_product_code(html: str) -> Optional[str]:
    """Extrai pokemon_xx_yyy/zzz do HTML do produto."""
    m = PRODUCT_CODE_RE.search(html)
    return m.group(0) if m else None


def fetch_api_tcg_price_usd(scraper, product_code: str) -> Optional[float]:
    """Chama GET /api/v1/pokemon/precos/{base64} e retorna tcg_price USD."""
    enc = base64.b64encode(product_code.encode()).decode()
    url = f"{API_BASE}/pokemon/precos/{enc}"
    try:
        r = scraper.get(url, timeout=TIMEOUT)
        if r.status_code != 200:
            log.debug(f"  API {product_code} → HTTP {r.status_code}")
            return None
        j = r.json()
        if "error" in j:
            log.debug(f"  API {product_code} → {j['error']}")
            return None
        tcg = j.get("tcg_price")
        if tcg is None:
            return None
        return float(tcg)
    except Exception as e:
        log.debug(f"  API fetch err {product_code}: {e}")
        return None


def fetch_html_and_extract_code(scraper, product_url: str) -> Optional[str]:
    """Fetch HTML do produto e extrai product_code via regex."""
    try:
        time.sleep(HTML_DELAY)
        r = scraper.get(product_url, timeout=TIMEOUT)
        if r.status_code != 200:
            log.debug(f"  HTML {product_url} → HTTP {r.status_code}")
            return None
        return extract_product_code(r.text)
    except Exception as e:
        log.debug(f"  HTML fetch err: {e}")
        return None


def load_target_rows(xlsx_path: Path) -> list[dict]:
    """Coleta rows únicas (por URL) de Deals + Validate Manually."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    targets: dict[str, dict] = {}
    for sheet_name in ("🔥 Deals", "🚨 Validate Manually"):
        if sheet_name not in wb.sheetnames:
            log.warning(f"Sheet '{sheet_name}' não existe — skip")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = list(rows[0])
        for r in rows[1:]:
            rec = dict(zip(headers, r))
            url = rec.get("URL")
            if not url or url in targets:
                continue
            targets[url] = {
                "name": rec.get("Card Name"),
                "edition": rec.get("Edition"),
                "url": url,
                "myp_en_nm_brl": rec.get("MYP EN NM (R$)"),
                "tcg_html_brl": rec.get("TCG Player (R$)"),
                "margin": rec.get("Margin %"),
                "source_sheet": sheet_name,
            }
    wb.close()
    return list(targets.values())


def write_mismatch_sheet(xlsx_path: Path, mismatches: list[dict], fx_rate: float):
    """Adiciona sheet '❌ TCG API Mismatch' ao XLSX existente."""
    wb = load_workbook(xlsx_path)
    sheet_name = "❌ TCG API Mismatch"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Card Name", "Edition", "TCG HTML (R$)", "TCG API USD",
        "TCG API → R$", "Ratio HTML/API", "Source Sheet", "URL",
    ]
    widths = [38, 32, 16, 14, 14, 14, 22, 55]
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="9C0006")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side("thin", "D9D9D9"), right=Side("thin", "D9D9D9"),
        top=Side("thin", "D9D9D9"), bottom=Side("thin", "D9D9D9"),
    )
    normal = Font(name="Arial", size=10)
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # FX note
    ws.cell(row=1, column=9, value=f"FX USD→BRL = {fx_rate:.4f}")

    mismatches_sorted = sorted(
        mismatches, key=lambda x: x.get("ratio", 0), reverse=True,
    )
    for i, m in enumerate(mismatches_sorted, 2):
        vals = [
            m["name"], m["edition"], m["tcg_html_brl"], m["tcg_usd"],
            m["tcg_api_brl"], m["ratio"], m["source_sheet"], m["url"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = normal; c.border = border
            if col in (3, 4, 5):
                c.number_format = '#,##0.00'
            elif col == 6:
                c.number_format = '0.00'
                c.fill = red_fill
                c.font = Font(bold=True, color="9C0006", name="Arial", size=10)

    if mismatches_sorted:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(mismatches_sorted)+1}"

    wb.save(xlsx_path)
    log.info(f"❌ Sheet '{sheet_name}' adicionada ({len(mismatches_sorted)} candidatos)")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Cross-check MYP HTML XLSX contra MYP API tcg_price USD",
    )
    parser.add_argument("xlsx", help="Caminho do XLSX gerado pelo scanner")
    parser.add_argument("--threshold", type=float, default=MISMATCH_RATIO_THRESHOLD,
                        help=f"Ratio mínimo HTML/API_BRL pra flagar (default: {MISMATCH_RATIO_THRESHOLD}x)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        log.error(f"Arquivo não existe: {xlsx_path}")
        return 1

    log.info(f"Cross-checking {xlsx_path}")
    scraper = cloudscraper.create_scraper(
        browser={"browser": "firefox", "platform": "windows", "desktop": True},
    )

    fx_rate = fetch_fx_usd_brl(scraper)

    targets = load_target_rows(xlsx_path)
    log.info(f"Cards alvo (Deals + Validate Manually, dedup URL): {len(targets)}")
    if not targets:
        log.warning("Sem cards pra processar.")
        return 0

    mismatches: list[dict] = []
    api_hits = 0
    api_misses = 0
    code_misses = 0

    for i, t in enumerate(targets, 1):
        if i % 10 == 0 or i == 1:
            log.info(f"  [{i}/{len(targets)}] {t['name']}")

        if not t.get("tcg_html_brl"):
            continue

        # Step 1: HTML fetch para descobrir product_code
        code = fetch_html_and_extract_code(scraper, t["url"])
        if not code:
            code_misses += 1
            continue

        # Step 2: API lookup
        time.sleep(API_DELAY)
        tcg_usd = fetch_api_tcg_price_usd(scraper, code)
        if tcg_usd is None:
            api_misses += 1
            continue
        api_hits += 1

        tcg_api_brl = tcg_usd * fx_rate
        if tcg_api_brl <= 0:
            continue

        ratio = float(t["tcg_html_brl"]) / tcg_api_brl
        if ratio >= args.threshold:
            mismatches.append({
                "name": t["name"],
                "edition": t["edition"],
                "url": t["url"],
                "tcg_html_brl": t["tcg_html_brl"],
                "tcg_usd": tcg_usd,
                "tcg_api_brl": tcg_api_brl,
                "ratio": ratio,
                "source_sheet": t["source_sheet"],
            })
            log.warning(
                f"  ❌ MISMATCH: {t['name']} | HTML R${t['tcg_html_brl']:.2f} "
                f"vs API ${tcg_usd}×{fx_rate:.2f}=R${tcg_api_brl:.2f} "
                f"(ratio {ratio:.1f}x)"
            )

    log.info("─" * 60)
    log.info(f"  API hits: {api_hits} / API misses: {api_misses} / code misses: {code_misses}")
    log.info(f"  Mismatches (ratio >= {args.threshold}x): {len(mismatches)}")

    write_mismatch_sheet(xlsx_path, mismatches, fx_rate)
    return 0


if __name__ == "__main__":
    sys.exit(main())
