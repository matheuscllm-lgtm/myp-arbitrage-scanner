"""Detalha sheets do smoke v5.8.3: 🔥 Deals vs 🚨 Validate Manually."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

import sys as _sys
path = _sys.argv[1] if len(_sys.argv) > 1 else r'C:\Users\mathe\Downloads\myp_smoke_v583.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)

for sheet_name in ['🔥 Deals', '🚨 Validate Manually', '🚨 TCG Suspect']:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0]) if rows else []
    print('  hdr:', hdr)
    print(f'  total data rows: {len(rows)-1 if rows else 0}')
    for r in rows[1:]:
        name = r[0] if r else None
        ed = r[1] if r and len(r) > 1 else None
        myp = r[3] if r and len(r) > 3 else None
        tcg = r[4] if r and len(r) > 4 else None
        margin = r[6] if r and len(r) > 6 else None
        sellers = r[8] if r and len(r) > 8 else None
        trunc = r[9] if r and len(r) > 9 else None
        susp = r[10] if r and len(r) > 10 else None
        single = r[11] if r and len(r) > 11 else None
        url = r[12] if r and len(r) > 12 else None
        margin_pct = f"{margin*100:.1f}%" if isinstance(margin, (int, float)) else margin
        print(f"  {name!s:50.50s} | {ed!s:15.15s} | MYP={myp} TCG={tcg} M={margin_pct} sellers={sellers} | trunc={trunc!r} susp={susp!r} single={single!r}")
        if url:
            print(f"    URL: {url}")
