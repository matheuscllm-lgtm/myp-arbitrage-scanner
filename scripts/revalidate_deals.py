"""Revalidate deals em XLSX existente fazendo re-scrape leve.

Pra cada deal no sheet '🔥 Deals' do XLSX input:
  1. Scrape URL do produto MYP
  2. Extrai `.estat-tcg` (TCG declarado) E `.estatistica-ultimo` (última venda real)
  3. Calcula ratio. Se > 10x = tcg_suspect.
  4. Gera novo XLSX com 3 sheets:
     - 🔥 Deals Limpos (sobreviventes)
     - 🚨 TCG Suspect (filtrados por sanity check)
     - All EN Cards (cópia inalterada)

Uso:
    python scripts/revalidate_deals.py <input.xlsx> [--output <output.xlsx>]
"""
from __future__ import annotations

import argparse
import logging
import sys
import time
from pathlib import Path

if sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

# v5.8.6 bug #3: line buffering so long-running scrape progress streams to
# the harness/operator in real time instead of being held in a 4KB block
# buffer until the script ends. Without this, runs that take 5-30min look
# "stuck" with zero output.
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup

from myp_arbitrage_scanner import MYPScraper, TCG_SUSPECT_RATIO_THRESHOLD

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger(__name__)

# v5.8 (2026-05-16): single source of truth, importa do scanner
SUSPECT_RATIO = TCG_SUSPECT_RATIO_THRESHOLD


def revalidate_url(scraper: MYPScraper, url: str) -> dict:
    """Re-scrape URL e retorna {tcg_declared, last_sale, ratio, suspect}."""
    try:
        r = scraper.session.get(url, timeout=30)
        if r.status_code != 200:
            return {"error": f"HTTP {r.status_code}"}
        soup = BeautifulSoup(r.text, "lxml")

        # _last_brl: single source of truth no scanner (mesmo idiom que
        # scrape_product usa pra .estat-tcg / .estatistica-ultimo).
        tcg_el = soup.select_one(".estat-tcg")
        tcg_decl = MYPScraper._last_brl(tcg_el.get_text()) if tcg_el else None

        ls_el = soup.select_one(".estatistica-ultimo")
        last_sale = MYPScraper._last_brl(ls_el.get_text()) if ls_el else None

        ratio = None
        suspect = False
        if tcg_decl and last_sale and last_sale > 0:
            ratio = tcg_decl / last_sale
            suspect = ratio > SUSPECT_RATIO

        return {
            "tcg_declared": tcg_decl,
            "last_sale": last_sale,
            "ratio": ratio,
            "suspect": suspect,
        }
    except Exception as e:
        return {"error": str(e)}


def revalidate(input_xlsx: Path, output_xlsx: Path, delay: float = 1.5):
    log.info(f"Loading: {input_xlsx}")
    wb_in = load_workbook(input_xlsx, data_only=True)

    deals_ws = wb_in["🔥 Deals"]
    headers = [c.value for c in next(deals_ws.iter_rows(min_row=1, max_row=1))]
    url_idx = headers.index("URL")
    name_idx = headers.index("Card Name")

    rows = list(deals_ws.iter_rows(min_row=2, values_only=True))
    log.info(f"Total deals to revalidate: {len(rows)}")

    scraper = MYPScraper()

    clean = []
    suspect = []
    skipped_no_name = 0

    for i, row in enumerate(rows, 1):
        if not row or not row[name_idx]:
            skipped_no_name += 1
            continue
        url = row[url_idx]
        name = row[name_idx]
        if not url:
            clean.append(row)
            continue

        result = revalidate_url(scraper, url)
        if "error" in result:
            log.warning(f"  [{i}/{len(rows)}] {name[:50]}: error={result['error']} — keeping in clean")
            clean.append(row)
            continue

        if result["suspect"]:
            log.warning(
                f"  [{i}/{len(rows)}] 🚨 SUSPECT {name[:40]} | "
                f"TCG decl={result['tcg_declared']} vs last sale={result['last_sale']} "
                f"(ratio={result['ratio']:.1f}x)"
            )
            suspect.append((row, result))
        else:
            ratio_str = f"{result['ratio']:.1f}x" if result['ratio'] else "n/a"
            log.info(f"  [{i}/{len(rows)}] OK {name[:50]} (ratio={ratio_str})")
            clean.append(row)

        # v5.8.6 bug #3: heartbeat every 50 rows so long runs (200+ deals,
        # 1.5s delay each ~= 8min minimum) show progress in tools that
        # buffer stdout beyond line granularity.
        if i % 50 == 0:
            print(
                f"  ... progress: {i}/{len(rows)} processed "
                f"({len(clean)} clean, {len(suspect)} suspect)",
                flush=True,
            )

        time.sleep(delay)

    log.info(f"\n=== Resultado: {len(clean)} limpos | {len(suspect)} suspeitos ===\n")

    # v5.8.6 bug #2: detect silent failure where 100% rows were skipped because
    # Card Name was None (e.g. upstream wrote =HYPERLINK formulas instead of
    # plain strings — see add_card_hyperlinks.py v5.8.6 bug #1 fix). Without
    # this guard the script exits 0 with empty output and the operator only
    # notices when they open the XLSX and find empty Deals sheet.
    processed = len(clean) + len(suspect)
    if processed == 0 and len(rows) > 0:
        log.error(
            f"REVALIDATE FAILED: 0/{len(rows)} rows processed "
            f"(skipped_no_name={skipped_no_name}). Card Name likely None - "
            f"check XLSX format (formula cells viewed with data_only=True "
            f"return None; rerun add_card_hyperlinks.py if it predates v5.8.6)."
        )
        sys.exit(2)

    # Write output XLSX
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # Clean Deals
    ws_clean = wb_out.create_sheet("🔥 Deals Limpos")
    ws_clean.append(headers)
    for r in clean:
        ws_clean.append(r)
    _style_header(ws_clean, headers, fill="C6EFCE")
    _apply_data_formats(ws_clean, headers)

    # Suspect Deals
    susp_headers = headers + ["TCG Decl Re-check", "Last Sale", "Ratio"]
    ws_susp = wb_out.create_sheet("🚨 TCG Suspect")
    ws_susp.append(susp_headers)
    for r, result in suspect:
        ws_susp.append(list(r) + [result["tcg_declared"], result["last_sale"], result["ratio"]])
    _style_header(ws_susp, susp_headers, fill="FFC7CE")
    _apply_data_formats(ws_susp, susp_headers)

    # Copy All EN Cards
    if "All EN Cards" in wb_in.sheetnames:
        src = wb_in["All EN Cards"]
        ws_all = wb_out.create_sheet("All EN Cards")
        for row in src.iter_rows(values_only=True):
            ws_all.append(row)
        _style_header(ws_all, headers, fill="DCE6F1")
        _apply_data_formats(ws_all, headers)

    # Summary
    ws_sum = wb_out.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Revalidation timestamp", time.strftime("%Y-%m-%d %H:%M UTC", time.gmtime())])
    ws_sum.append(["Input file", str(input_xlsx.name)])
    ws_sum.append(["Total deals revalidated", len(rows)])
    ws_sum.append(["Deals limpos", len(clean)])
    ws_sum.append(["Deals suspeitos (filtrados)", len(suspect)])
    ws_sum.append(["Suspect ratio threshold", f">{SUSPECT_RATIO}x"])

    wb_out.save(output_xlsx)
    log.info(f"OK output saved: {output_xlsx}")


def _style_header(ws, headers, fill="4472C4"):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        c.alignment = Alignment(horizontal="center")


def _apply_data_formats(ws, headers):
    """v5.8.6 bug #5: apply Excel number_format to columns whose header
    semantics imply a format (Margin → 0.00%, R$ prices → #,##0.00).
    v5.8.6 bug #1 follow-up: also re-attach Card Name hyperlinks using
    the URL column, since `ws.append(values_only_row)` strips hyperlink
    metadata. Without this, running revalidate on a hyperlinked XLSX
    silently drops every click-link.
    """
    from openpyxl.styles import Font
    margin_cols = [i + 1 for i, h in enumerate(headers)
                   if isinstance(h, str) and "margin" in h.lower()
                   and "%" in h]
    price_cols = [i + 1 for i, h in enumerate(headers)
                  if isinstance(h, str) and "(r$)" in h.lower()]
    name_col = None
    url_col = None
    for i, h in enumerate(headers, 1):
        if h == "Card Name":
            name_col = i
        elif h == "URL":
            url_col = i
    blue = Font(color="0563C1", underline="single", name="Arial", size=10)
    for r in range(2, ws.max_row + 1):
        for col in margin_cols:
            ws.cell(row=r, column=col).number_format = "0.00%"
        for col in price_cols:
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        if name_col and url_col:
            url_val = ws.cell(row=r, column=url_col).value
            name_cell = ws.cell(row=r, column=name_col)
            if (isinstance(url_val, str) and url_val.startswith("http")
                    and name_cell.value and not name_cell.hyperlink):
                name_cell.hyperlink = url_val
                name_cell.font = blue


def main():
    p = argparse.ArgumentParser()
    p.add_argument("input", type=Path)
    p.add_argument("--output", type=Path, default=None)
    p.add_argument("--delay", type=float, default=1.5)
    args = p.parse_args()

    if args.output is None:
        stem = args.input.stem
        args.output = args.input.parent / f"{stem}_REVALIDATED.xlsx"

    revalidate(args.input, args.output, args.delay)


if __name__ == "__main__":
    main()
