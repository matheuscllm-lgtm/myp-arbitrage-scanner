"""Quick audit: list deals from contaminated XLSX, find Jumbo + Flareon 018 entries, dump editions list."""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

path = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\mathe\Downloads\myp_weekly_20260517_1519_REVALIDATED.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)
print('SHEETS:', wb.sheetnames)

# Try Deals sheet first
target_sheet = None
for s in wb.sheetnames:
    if 'All EN' in s:
        target_sheet = s
        break
if not target_sheet:
    for s in wb.sheetnames:
        if 'Deals' in s:
            target_sheet = s
            break

print('Using sheet:', target_sheet)
ws = wb[target_sheet]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print('HDR:', hdr)

def idx_of(name):
    for i, h in enumerate(hdr):
        if h and name.lower() in str(h).lower():
            return i
    return -1

idx_name = idx_of('card name')
idx_ed = idx_of('edition')
idx_url = idx_of('product url')
idx_lang = idx_of('language')
print('cols name/ed/url/lang:', idx_name, idx_ed, idx_url, idx_lang)

flareon_rows = []
jumbo_rows = []
editions_set = set()
total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    total += 1
    nm = str(row[idx_name]) if idx_name >= 0 and row[idx_name] else ''
    ed = str(row[idx_ed]) if idx_ed >= 0 and row[idx_ed] else ''
    editions_set.add(ed)
    nm_l = nm.lower()
    if 'flareon' in nm_l and ('018' in nm or '18/203' in nm):
        flareon_rows.append((nm, ed, row[idx_url] if idx_url >= 0 else None))
    if 'jumbo' in nm_l:
        jumbo_rows.append((nm, ed, row[idx_url] if idx_url >= 0 else None))

print('Total rows:', total)
print('Editions with deals:', len(editions_set))
print('--- Jumbo rows ---')
for r in jumbo_rows[:20]:
    print(' ', r)
print('--- Flareon 018 rows ---')
for r in flareon_rows[:5]:
    print(' ', r)
print('--- ALL EDITIONS ---')
for e in sorted(editions_set):
    print('  ', repr(e))
