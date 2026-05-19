"""Post-hoc inject Card Name hyperlinks into a MYP scanner XLSX.

Looks for a column that resembles "Card" or "Card Name" plus a sibling URL column
(URL / Card URL / MYP URL / Product URL / Link) and attaches a native openpyxl
hyperlink to the Card Name cell. Edits in-place when --inplace is passed.

v5.8.6 (2026-05-19) — bugs #1 + #4: migrated from `=HYPERLINK(...)` formula to
native `cell.hyperlink`. Formulas only evaluate when Excel opens the file;
downstream scripts that read with `openpyxl.load_workbook(data_only=True)` see
formula cells as None (no cached calc value), which broke revalidate_deals.py
and cross_check_myp_api.py (both rely on Card Name being a plain string).

Use case: MYP weekly XLSX outputs occasionally ship without inline hyperlinks
on the Card Name column. See memory `feedback_xlsx_card_name_hyperlink.md`.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

CARD_HEADER_CANDIDATES = (
    "card name",
    "card",
    "name",
    "carta",
    "nome",
)

URL_HEADER_CANDIDATES = (
    "card url",
    "myp url",
    "product url",
    "url",
    "link",
    "myp_link",
    "card_url",
    "product_link",
)


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _find_header_col(header_row, candidates):
    """Return the 1-based column index of the first matching header, or None."""
    for idx, cell in enumerate(header_row, start=1):
        if _norm(cell.value) in candidates:
            return idx
    return None


def inject_sheet(ws) -> int:
    """Inject hyperlinks on a single worksheet. Returns count injected."""
    if ws.max_row < 2:
        return 0

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    card_col = _find_header_col(header_row, CARD_HEADER_CANDIDATES)
    url_col = _find_header_col(header_row, URL_HEADER_CANDIDATES)
    if card_col is None or url_col is None:
        return 0

    blue = Font(color="0563C1", underline="single")
    injected = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        card_cell = row[card_col - 1]
        url_cell = row[url_col - 1]
        url = url_cell.value
        if not url or not isinstance(url, str) or not url.startswith("http"):
            continue
        text = card_cell.value
        # Legacy: strip leftover =HYPERLINK(...) formula so we can replace
        # it with a native hyperlink (downstream readers with data_only=True
        # see formula cells as None — see module docstring).
        if isinstance(text, str) and text.startswith("=HYPERLINK"):
            import re as _re
            m = _re.match(r'=HYPERLINK\("([^"]*)",\s*"([^"]*)"\)', text)
            text = m.group(2).replace('""', '"') if m else None
        if text is None or str(text).strip() == "":
            continue
        # Native openpyxl hyperlink: value stays a plain string, .hyperlink
        # carries the URL. Excel renders it as a clickable link, and
        # load_workbook(data_only=True) returns the string (not None).
        card_cell.value = str(text)
        card_cell.hyperlink = url
        card_cell.font = blue
        injected += 1
    return injected


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--inplace", action="store_true")
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: file not found: {args.xlsx}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(args.xlsx)
    totals = {}
    for ws in wb.worksheets:
        n = inject_sheet(ws)
        totals[ws.title] = n

    if args.inplace:
        target = args.xlsx
    else:
        target = args.out or args.xlsx.with_name(args.xlsx.stem + "_linked.xlsx")
    wb.save(target)

    print(f"Saved: {target}")
    print("Hyperlinks injected per sheet:")
    grand = 0
    for sheet, n in totals.items():
        print(f"  {sheet}: {n}")
        grand += n
    print(f"TOTAL: {grand}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
